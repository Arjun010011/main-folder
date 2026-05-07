import logging
import os
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.db.models import Sum, F, Case, When, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rest_framework import exceptions

from apps.asset.models import AssetGroup, AssetDepreciationSnapshot, AssetCostMovement, AssetDisposal
from apps.institutes.models.financialyear import FinancialYear
from apps.institutes.models import Institute
from apps.shared.services import SharedService, PDFService, UploadTypeService
from apps.shared.services_shared.common import get_selected_template
from apps.shared.services_shared.store_api_result import store_long_running_process


# ━━━━━━━━━━━━━━━ UNIFIED EXCEL BUILDER ━━━━━━━━━━━━━━━

def build_asset_excel(title, headers, rows, totals=None, total_label='TOTAL',
                      status_text=None, filename='report.xlsx', institute_name=''):
    """
    Generic function that builds an Excel workbook for any asset report.

    Parameters:
        title          – str:  Title row text (e.g. "Fixed Asset Register - FY 2024-2025")
        headers        – list of dicts:
                           [{'key': 'asset_code', 'label': 'Asset Code', 'is_amount': False}, ...]
        rows           – list of dicts:  Data rows
        totals         – dict or None:  Keys matching header keys → total values
        total_label    – str:  Label for the total row (default "TOTAL")
        status_text    – str or None:  Optional status line (e.g. "LOCKED (Read-Only)")
        filename       – str:  Output filename for Content-Disposition
        institute_name – str:  School/institute name shown at the top
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet titles max 31 chars

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    col_count = len(headers)
    current_row = 1

    # Institute name row
    if institute_name:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
        ws['A1'] = institute_name
        ws['A1'].font = Font(bold=True, size=16, color='1F4E79')
        ws['A1'].alignment = Alignment(horizontal='center')
        current_row = 2

    # Title row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_count)
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    if status_text:
        status_cell = ws.cell(row=current_row, column=1, value=f"Status: {status_text}")
        status_cell.font = Font(italic=True, color='555555')
        current_row += 1

    # Blank row before headers
    data_start_row = current_row + 1

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=h['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        if h.get('is_amount'):
            cell.alignment = Alignment(horizontal='right')

    # Data rows
    for row_idx, row_data in enumerate(rows, data_start_row + 1):
        for col_idx, h in enumerate(headers, 1):
            value = row_data.get(h['key'], '')
            if h.get('is_amount') and value is not None:
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            elif h.get('is_date') and value is not None:
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if h.get('is_amount'):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Total row
    if totals:
        total_row = data_start_row + 1 + len(rows)
        total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        for col_idx, h in enumerate(headers, 1):
            if col_idx == 1:
                cell = ws.cell(row=total_row, column=col_idx, value=total_label)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = border
            elif h['key'] in totals:
                try:
                    val = float(totals[h['key']])
                except (TypeError, ValueError):
                    val = totals[h['key']]
                cell = ws.cell(row=total_row, column=col_idx, value=val)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = border
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell = ws.cell(row=total_row, column=col_idx, value='')
                cell.fill = total_fill
                cell.border = border

    # Auto-fit column widths
    for col_idx, h in enumerate(headers, 1):
        width = max(len(h['label']) + 4, 15)
        if h.get('is_amount'):
            width = max(width, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ━━━━━━━━━━━━━━━ DATA FUNCTIONS ━━━━━━━━━━━━━━━

def get_fixed_asset_register(self, financial_year_id, request=None):

    try:
        fy = FinancialYear.objects.get(id=financial_year_id)
        fy_name = f"{fy.start_date.year}-{fy.end_date.year}"
    except FinancialYear.DoesNotExist:
        raise exceptions.NotFound("Financial year not found.")

    snapshots_qs = AssetDepreciationSnapshot.objects.filter(
        financial_year=financial_year_id
    ).select_related(
        'asset', 'asset__asset_group'
    ).order_by(
        'asset__asset_group__display_order',
        'asset__asset_code'
    )

    # If request provided, paginate; otherwise return ALL rows (for excel/pdf)
    if request:
        limit = int(request.GET.get('limit', 10))
        page_no = int(request.GET.get('pageno', 1))
        page_qs, count, next_page, previous_page = SharedService.custom_pagination(
            self, snapshots_qs, limit, page_no
        )
    else:
        page_qs = snapshots_qs
        count = snapshots_qs.count()
        next_page = None
        previous_page = None

    register = [{
        'asset_id': s.asset.id,
        'asset_code': s.asset.asset_code,
        'asset_name': s.asset.asset_name,
        'asset_group_name': s.asset.asset_group.name,
        'purchase_date': s.asset.purchase_date,
        'put_to_use_date': s.asset.put_to_use_date,
        'original_cost': s.asset.original_cost,
        'opening_value': s.opening_value,
        'additions': s.additions,
        'depreciation': s.depreciation_amount,
        'closing_value': s.closing_value,
        'location': getattr(s.asset, 'location', '') or '',
        'status': s.asset.status
    } for s in page_qs]

    totals = snapshots_qs.aggregate(
        original_cost=Sum('asset__original_cost'),
        opening_value=Sum('opening_value'),
        additions=Sum('additions'),
        depreciation=Sum('depreciation_amount'),
        closing_value=Sum('closing_value')
    )

    return {
        'data': {
            'count': count,
            'next': next_page,
            'previous': previous_page,
            'register': register,
            'totals': totals,
            'financial_year': {
                'id': fy.id,
                'name': fy_name
            }
        }
    }


def get_asset_group_summary(self, financial_year_id):

    try:
        financial_year = FinancialYear.objects.get(id=financial_year_id)
    except FinancialYear.DoesNotExist:
        raise exceptions.NotFound("Financial year not found.")
    
    group_totals = AssetDepreciationSnapshot.objects.filter(
        financial_year=financial_year
    ).values(
        'asset__asset_group_id'
    ).annotate(
        opening_value=Sum('opening_value'),
        additions=Sum('additions'),
        depreciation=Sum('depreciation_amount'),
        closing_value=Sum('closing_value')
    )
    
    group_data = {}
    for total in group_totals:
        group_data[total['asset__asset_group_id']] = {
            'opening_value': total['opening_value'] or Decimal('0.00'),
            'additions': total['additions'] or Decimal('0.00'),
            'depreciation': total['depreciation'] or Decimal('0.00'),
            'closing_value': total['closing_value'] or Decimal('0.00')
        }
    
    groups = AssetGroup.objects.filter(
        is_active=True,
        financial_year=financial_year
    ).select_related('parent_group').order_by('display_order', 'name')
    
    summary = []
    for group in groups:
        if group.id in group_data:
            data = group_data[group.id]
            summary.append({
                'asset_group_id': group.id,
                'asset_group_name': group.name,
                'parent_group_name': group.parent_group.name if group.parent_group else None,
                'opening_value': data['opening_value'],
                'additions': data['additions'],
                'depreciation': data['depreciation'],
                'closing_value': data['closing_value']
            })
    
    grand_totals = {
        'opening_value': sum(s['opening_value'] for s in summary),
        'additions': sum(s['additions'] for s in summary),
        'depreciation': sum(s['depreciation'] for s in summary),
        'closing_value': sum(s['closing_value'] for s in summary)
    }
    
    return {
        'data': {
            'financial_year': {
                'id': financial_year.id,
                'name': f"{financial_year.start_date.year}-{financial_year.end_date.year}"
            },
            'summary': summary,
            'grand_totals': grand_totals
        }
    }


def get_depreciation_schedule(self, financial_year_id=None, asset_id=None):

    queryset = AssetDepreciationSnapshot.objects.select_related(
        'asset', 'financial_year'
    )
    
    if financial_year_id:
        queryset = queryset.filter(financial_year_id=financial_year_id)
    
    if asset_id:
        queryset = queryset.filter(asset_id=asset_id)
    
    queryset = queryset.order_by('asset__asset_code', 'financial_year__start_date')
    
    schedule = []
    for snapshot in queryset:
        schedule.append({
            'asset_id': snapshot.asset.id,
            'asset_code': snapshot.asset.asset_code,
            'asset_name': snapshot.asset.asset_name,
            'financial_year_name': f"{snapshot.financial_year.start_date.year}-{snapshot.financial_year.end_date.year}",
            'opening_wdv': snapshot.opening_value,
            'depreciation': snapshot.depreciation_amount,
            'closing_wdv': snapshot.closing_value,
            'calculation_method': snapshot.calculation_method
        })
    
    return {'data': {'schedule': schedule}}


def is_fy_cost_locked(financial_year_id):
    return AssetDepreciationSnapshot.objects.filter(
        financial_year_id=financial_year_id,
        is_locked=True
    ).exists()

def get_fixed_asset_cost_register(self, financial_year_id, request=None):

    try:
        fy = FinancialYear.objects.get(id=financial_year_id)
        fy_name = f"{fy.start_date.year}-{fy.end_date.year}"
    except FinancialYear.DoesNotExist:
        raise exceptions.NotFound("Financial year not found.")

    is_locked = is_fy_cost_locked(financial_year_id)

    # Check if the previous FY is locked
    previous_fys = FinancialYear.objects.filter(
        end_date__lt=fy.start_date
    ).order_by('-end_date')
    previous_year_locked = True  # No previous FY = treat as locked
    if previous_fys.exists():
        previous_year_locked = is_fy_cost_locked(previous_fys.first().id)

    # We need to find all movements for assets up to this FY's end date
    # This allows us to calculate the carry-forward opening balance from previous FYs
    movements = AssetCostMovement.objects.filter(
        financial_year__start_date__lte=fy.start_date,
        asset__status__in=['ACTIVE', 'DISPOSED'],  # we'll filter out completely disposed ones later
        asset__asset_group__financial_year_id=financial_year_id
    ).select_related(
        'asset', 'asset__asset_group', 'financial_year'
    )

    asset_data = {}

    for m in movements:
        asset_id = m.asset.id

        if asset_id not in asset_data:
            asset_data[asset_id] = {
                'asset_id': asset_id,
                'asset_code': m.asset.asset_code,
                'asset_name': m.asset.asset_name,
                'asset_group_name': m.asset.asset_group.name,
                'purchase_date': m.asset.purchase_date,
                'opening_cost': Decimal('0.00'),
                'additions': Decimal('0.00'),
                'disposals': Decimal('0.00'),
                'status': m.asset.status,
                'has_current_fy_movement': False
            }

        is_current_fy = m.financial_year_id == financial_year_id

        if is_current_fy:
            asset_data[asset_id]['has_current_fy_movement'] = True
            if m.movement_type == 'OPENING':
                asset_data[asset_id]['opening_cost'] += m.amount
            elif m.movement_type == 'ADDITION':
                asset_data[asset_id]['additions'] += m.amount
            elif m.movement_type == 'DISPOSAL':
                asset_data[asset_id]['disposals'] += m.amount
        else:
            # Previous FY movements: they all contribute to the current FY's opening balance
            if m.movement_type in ('OPENING', 'ADDITION'):
                asset_data[asset_id]['opening_cost'] += m.amount
            elif m.movement_type == 'DISPOSAL':
                asset_data[asset_id]['opening_cost'] -= m.amount

    register_data = []
    for data in asset_data.values():
        if data['status'] == 'DISPOSED':
            # If it was disposed in a PREVIOUS FY, it shouldn't appear in the current FY's register
            # unless it has a movement in the current FY (e.g., it was disposed *this* year)
            if not data['has_current_fy_movement'] and data['opening_cost'] <= 0:
                continue
            
            # For assets disposed this year, force closing cost to 0 and fix disposals if they don't match
            data['closing_cost'] = Decimal('0.00')
            data['disposals'] = data['opening_cost'] + data['additions']
        else:
            data['closing_cost'] = (
                data['opening_cost'] + data['additions'] - data['disposals']
            )
            
            # Skip fully depreciated/disposed ghosts that have no current activity
            if data['closing_cost'] <= 0 and not data['has_current_fy_movement']:
                continue

        # Remove the internal flag before sending to frontend
        data.pop('has_current_fy_movement', None)
        register_data.append(data)

    register_data.sort(
        key=lambda x: (x['asset_group_name'], x['asset_code'])
    )

    totals = {
        'opening_cost': sum(r['opening_cost'] for r in register_data),
        'additions': sum(r['additions'] for r in register_data),
        'disposals': sum(r['disposals'] for r in register_data),
        'closing_cost': sum(r['closing_cost'] for r in register_data),
    }

    # If request provided, paginate; otherwise return ALL rows (for excel/pdf)
    if request:
        limit = int(request.GET.get('limit', 10))
        page_no = int(request.GET.get('pageno', 1))
        page_data, count, next_page, previous_page = SharedService.custom_pagination(
            self,
            register_data,
            limit,
            page_no
        )
    else:
        page_data = register_data
        count = len(register_data)
        next_page = None
        previous_page = None

    return {
        'data': {
            'count': count,
            'next': next_page,
            'previous': previous_page,
            'is_locked': is_locked,
            'previous_year_locked': previous_year_locked,
            'register': page_data,
            'totals': totals,
            'financial_year': {
                'id': fy.id,
                'name': fy_name
            }
        }
    }

def get_asset_group_cost_summary(self, financial_year_id):

    try:
        financial_year = FinancialYear.objects.get(id=financial_year_id)
    except FinancialYear.DoesNotExist:
        raise exceptions.NotFound("Financial year not found.")
    
    is_locked = is_fy_cost_locked(financial_year_id)
    
    movements = AssetCostMovement.objects.filter(
        financial_year__start_date__lte=financial_year.start_date,
        asset__status__in=['ACTIVE', 'DISPOSED']
    ).select_related('asset__asset_group', 'financial_year')
    
    # We need to calculate per-asset first, then aggregate to group
    group_totals = {}
    for m in movements:
        group_id = m.asset.asset_group.id
        if group_id not in group_totals:
            group = m.asset.asset_group
            group_totals[group_id] = {
                'asset_group_id': group.id,
                'asset_group_name': group.name,
                'parent_group_name': group.parent_group.name if group.parent_group else None,
                'hierarchy_path': ' → '.join(group.get_hierarchy_path()),
                'opening_cost': Decimal('0.00'),
                'additions': Decimal('0.00'),
                'disposals': Decimal('0.00'),
                'is_fy_locked': is_locked,
                'assets': {} # track per-asset to handle carry-forward and DISPOSED legacy
            }
            
        asset_id = m.asset.id
        if asset_id not in group_totals[group_id]['assets']:
            group_totals[group_id]['assets'][asset_id] = {
                'opening_cost': Decimal('0.00'),
                'additions': Decimal('0.00'),
                'disposals': Decimal('0.00'),
                'status': m.asset.status,
                'has_current_fy_movement': False
            }
        
        is_current_fy = m.financial_year_id == financial_year_id

        if is_current_fy:
            group_totals[group_id]['assets'][asset_id]['has_current_fy_movement'] = True
            if m.movement_type == 'OPENING':
                group_totals[group_id]['assets'][asset_id]['opening_cost'] += m.amount
            elif m.movement_type == 'ADDITION':
                group_totals[group_id]['assets'][asset_id]['additions'] += m.amount
            elif m.movement_type == 'DISPOSAL':
                group_totals[group_id]['assets'][asset_id]['disposals'] += m.amount
        else:
            if m.movement_type in ('OPENING', 'ADDITION'):
                group_totals[group_id]['assets'][asset_id]['opening_cost'] += m.amount
            elif m.movement_type == 'DISPOSAL':
                group_totals[group_id]['assets'][asset_id]['opening_cost'] -= m.amount
    
    summary = []
    for group_id, data in group_totals.items():
        # Aggregate assets dynamically
        for asset_data in data['assets'].values():
            if asset_data['status'] == 'DISPOSED':
                if not asset_data['has_current_fy_movement'] and asset_data['opening_cost'] <= 0:
                    continue
                # For assets disposed this year, fix disposals
                asset_data['disposals'] = asset_data['opening_cost'] + asset_data['additions']
                asset_data['closing_cost'] = Decimal('0.00')
            else:
                asset_data['closing_cost'] = asset_data['opening_cost'] + asset_data['additions'] - asset_data['disposals']
                if asset_data['closing_cost'] <= 0 and not asset_data['has_current_fy_movement']:
                    continue
                    
            data['opening_cost'] += asset_data['opening_cost']
            data['additions'] += asset_data['additions']
            data['disposals'] += asset_data['disposals']
                
        data['closing_cost'] = data['opening_cost'] + data['additions'] - data['disposals']
        del data['assets']
        
        # Only include groups that have non-zero activity or balances
        if data['opening_cost'] > 0 or data['additions'] > 0 or data['disposals'] > 0 or data['closing_cost'] > 0:
            summary.append(data)
    
    summary.sort(key=lambda x: x['hierarchy_path'])
    
    grand_totals = {
        'opening_cost': sum(s['opening_cost'] for s in summary),
        'additions': sum(s['additions'] for s in summary),
        'disposals': sum(s['disposals'] for s in summary),
        'closing_cost': sum(s['closing_cost'] for s in summary)
    }
    
    return {
        'data': {
            'financial_year': {
                'id': financial_year.id,
                'name': f"{financial_year.start_date.year}-{financial_year.end_date.year}"
            },
            'is_locked': is_locked,
            'summary': summary,
            'grand_totals': grand_totals
        }
    }


def get_disposal_list(self, financial_year_id=None):
    """Get list of all disposed assets."""
    disposals_qs = AssetDisposal.objects.select_related(
        'asset', 'asset__asset_group'
    ).order_by('-disposal_date')

    if financial_year_id:
        try:
            fy = FinancialYear.objects.get(id=financial_year_id)
            disposals_qs = disposals_qs.filter(
                disposal_date__gte=fy.start_date,
                disposal_date__lte=fy.end_date
            )
            fy_name = f"{fy.start_date.year}-{fy.end_date.year}"
        except FinancialYear.DoesNotExist:
            raise exceptions.NotFound("Financial year not found.")
    else:
        fy_name = 'All'

    REASON_DISPLAY = dict(AssetDisposal.DISPOSAL_REASON_CHOICES)

    disposals = []
    total_disposal_value = Decimal('0.00')
    total_wdv = Decimal('0.00')
    total_gain_loss = Decimal('0.00')

    for d in disposals_qs:
        disposal_value = d.disposal_value or Decimal('0.00')
        wdv = d.wdv_at_disposal or Decimal('0.00')
        gain_loss = d.gain_loss or Decimal('0.00')

        total_disposal_value += disposal_value
        total_wdv += wdv
        total_gain_loss += gain_loss

        disposals.append({
            'asset_code': d.asset.asset_code,
            'asset_name': d.asset.asset_name,
            'asset_group_name': d.asset.asset_group.name,
            'disposal_date': d.disposal_date,
            'reason': REASON_DISPLAY.get(d.reason, d.reason),
            'original_cost': d.asset.original_cost,
            'wdv_at_disposal': wdv,
            'disposal_value': disposal_value,
            'gain_loss': gain_loss,
            'remarks': d.remarks or '',
        })

    totals = {
        'original_cost': sum(d['original_cost'] for d in disposals),
        'wdv_at_disposal': total_wdv,
        'disposal_value': total_disposal_value,
        'gain_loss': total_gain_loss,
    }

    return {
        'data': {
            'financial_year_name': fy_name,
            'disposals': disposals,
            'totals': totals,
            'count': len(disposals),
        }
    }


# ━━━━━━━━━━━━━━━ EXCEL DOWNLOADS ━━━━━━━━━━━━━━━

FAR_HEADERS = [
    {'key': 'asset_code', 'label': 'Asset Code'},
    {'key': 'asset_name', 'label': 'Asset Name'},
    {'key': 'asset_group_name', 'label': 'Asset Group'},
    {'key': 'purchase_date', 'label': 'Purchase Date', 'is_date': True},
    {'key': 'put_to_use_date', 'label': 'Put to Use Date', 'is_date': True},
    {'key': 'original_cost', 'label': 'Original Cost (₹)', 'is_amount': True},
    {'key': 'opening_value', 'label': 'Opening Value (₹)', 'is_amount': True},
    {'key': 'additions', 'label': 'Additions (₹)', 'is_amount': True},
    {'key': 'depreciation', 'label': 'Depreciation (₹)', 'is_amount': True},
    {'key': 'closing_value', 'label': 'Closing Value (₹)', 'is_amount': True},
    {'key': 'location', 'label': 'Location'},
]

AGS_HEADERS = [
    {'key': 'asset_group_name', 'label': 'Asset Group'},
    {'key': 'opening_value', 'label': 'Opening Value (₹)', 'is_amount': True},
    {'key': 'additions', 'label': 'Additions (₹)', 'is_amount': True},
    {'key': 'depreciation', 'label': 'Depreciation (₹)', 'is_amount': True},
    {'key': 'closing_value', 'label': 'Closing Value (₹)', 'is_amount': True},
]

CR_HEADERS = [
    {'key': 'asset_code', 'label': 'Asset Code'},
    {'key': 'asset_name', 'label': 'Asset Name'},
    {'key': 'asset_group_name', 'label': 'Asset Group'},
    {'key': 'purchase_date', 'label': 'Purchase Date', 'is_date': True},
    {'key': 'opening_cost', 'label': 'Opening Balance (₹)', 'is_amount': True},
    {'key': 'additions', 'label': 'Additions (Debit) (₹)', 'is_amount': True},
    {'key': 'disposals', 'label': 'Disposals (Credit) (₹)', 'is_amount': True},
    {'key': 'closing_cost', 'label': 'Closing Balance (₹)', 'is_amount': True},
]

CGS_HEADERS = [
    {'key': 'hierarchy_path', 'label': 'Asset Group'},
    {'key': 'opening_cost', 'label': 'Opening Balance (₹)', 'is_amount': True},
    {'key': 'additions', 'label': 'Additions (Debit) (₹)', 'is_amount': True},
    {'key': 'disposals', 'label': 'Disposals (Credit) (₹)', 'is_amount': True},
    {'key': 'closing_cost', 'label': 'Closing Balance (₹)', 'is_amount': True},
]

DS_HEADERS = [
    {'key': 'asset_code', 'label': 'Asset Code'},
    {'key': 'asset_name', 'label': 'Asset Name'},
    {'key': 'financial_year_name', 'label': 'Financial Year'},
    {'key': 'calculation_method', 'label': 'Method'},
    {'key': 'opening_wdv', 'label': 'Opening Value (₹)', 'is_amount': True},
    {'key': 'depreciation', 'label': 'Depreciation (₹)', 'is_amount': True},
    {'key': 'closing_wdv', 'label': 'Closing Value (₹)', 'is_amount': True},
]

DISPOSAL_HEADERS = [
    {'key': 'asset_code', 'label': 'Asset Code'},
    {'key': 'asset_name', 'label': 'Asset Name'},
    {'key': 'asset_group_name', 'label': 'Asset Group'},
    {'key': 'disposal_date', 'label': 'Disposal Date', 'is_date': True},
    {'key': 'reason', 'label': 'Reason'},
    {'key': 'original_cost', 'label': 'Original Cost (₹)', 'is_amount': True},
    {'key': 'wdv_at_disposal', 'label': 'Book Value (₹)', 'is_amount': True},
    {'key': 'disposal_value', 'label': 'Disposal Value (₹)', 'is_amount': True},
    {'key': 'gain_loss', 'label': 'Gain / Loss (₹)', 'is_amount': True},
    {'key': 'remarks', 'label': 'Remarks'},
]


def download_fixed_asset_register_excel(self, financial_year_id):
    report_data = get_fixed_asset_register(self, financial_year_id)['data']
    fy_name = report_data['financial_year']['name']
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''
    return build_asset_excel(
        title=f"Fixed Asset Register - FY {fy_name}",
        headers=FAR_HEADERS,
        rows=report_data['register'],
        totals=report_data['totals'],
        filename=f"fixed_asset_register_{fy_name}.xlsx",
        institute_name=inst_name,
    )


def download_asset_group_summary_excel(self, financial_year_id):
    report_data = get_asset_group_summary(self, financial_year_id)['data']
    fy_name = report_data['financial_year']['name']
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''
    return build_asset_excel(
        title=f"Asset Group Summary - FY {fy_name}",
        headers=AGS_HEADERS,
        rows=report_data['summary'],
        totals=report_data['grand_totals'],
        total_label='GRAND TOTAL',
        filename=f"asset_group_summary_{fy_name}.xlsx",
        institute_name=inst_name,
    )


def download_fixed_asset_cost_register_excel(self, financial_year_id):
    report_data = get_fixed_asset_cost_register(self, financial_year_id)['data']
    fy_name = report_data['financial_year']['name']
    status_text = "LOCKED (Read-Only)" if report_data['is_locked'] else "Draft"
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''
    return build_asset_excel(
        title=f"Fixed Asset Cost Register - FY {fy_name}",
        headers=CR_HEADERS,
        rows=report_data['register'],
        totals=report_data['totals'],
        status_text=status_text,
        filename=f"asset_cost_register_{fy_name}.xlsx",
        institute_name=inst_name,
    )


def download_asset_group_cost_summary_excel(self, financial_year_id):
    report_data = get_asset_group_cost_summary(self, financial_year_id)['data']
    fy_name = report_data['financial_year']['name']
    status_text = "LOCKED (Read-Only)" if report_data['is_locked'] else "Draft"
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''
    return build_asset_excel(
        title=f"Asset Group Cost Summary - FY {fy_name}",
        headers=CGS_HEADERS,
        rows=report_data['summary'],
        totals=report_data['grand_totals'],
        total_label='GRAND TOTAL',
        status_text=status_text,
        filename=f"asset_group_cost_summary_{fy_name}.xlsx",
        institute_name=inst_name,
    )


def download_depreciation_schedule_excel(self, asset_id):
    report_data = get_depreciation_schedule(self, asset_id=asset_id)['data']
    asset_label = report_data['schedule'][0]['asset_name'] if report_data['schedule'] else 'asset'
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''
    return build_asset_excel(
        title=f"Depreciation Schedule - {asset_label}",
        headers=DS_HEADERS,
        rows=report_data['schedule'],
        filename=f"depreciation_schedule_{asset_id}.xlsx",
        institute_name=inst_name,
    )


def download_disposal_list_excel(self, financial_year_id=None):
    report_data = get_disposal_list(self, financial_year_id)['data']
    fy_name = report_data['financial_year_name']
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''
    return build_asset_excel(
        title=f"Asset Disposal List - FY {fy_name}",
        headers=DISPOSAL_HEADERS,
        rows=report_data['disposals'],
        totals=report_data['totals'],
        filename=f"disposal_list_{fy_name}.xlsx",
        institute_name=inst_name,
    )


# ━━━━━━━━━━━━━━━ PDF DOWNLOADS ━━━━━━━━━━━━━━━

def download_fixed_asset_register_pdf(view_self):
    """Generate PDF for Fixed Asset Register using PDFService.receipt_new()."""
    financial_year_id = view_self.request.GET.get('financial_year')
    if not financial_year_id:
        return HttpResponse('financial_year is required', status=400)

    # Fetch ALL rows (no pagination) for PDF
    snapshots_qs = AssetDepreciationSnapshot.objects.filter(
        financial_year=financial_year_id
    ).select_related('asset', 'asset__asset_group').order_by(
        'asset__asset_group__display_order', 'asset__asset_code'
    )

    register = [{
        'asset_code': s.asset.asset_code,
        'asset_name': s.asset.asset_name,
        'asset_group_name': s.asset.asset_group.name,
        'purchase_date': s.asset.purchase_date,
        'original_cost': s.asset.original_cost,
        'opening_value': s.opening_value,
        'additions': s.additions,
        'depreciation': s.depreciation_amount,
        'closing_value': s.closing_value,
        'status': s.asset.status,
    } for s in snapshots_qs]

    totals = snapshots_qs.aggregate(
        original_cost=Sum('asset__original_cost'),
        opening_value=Sum('opening_value'),
        additions=Sum('additions'),
        depreciation=Sum('depreciation_amount'),
        closing_value=Sum('closing_value'),
    )

    try:
        fy = FinancialYear.objects.get(id=financial_year_id)
        fy_name = f"{fy.start_date.year}-{fy.end_date.year}"
    except FinancialYear.DoesNotExist:
        fy_name = ''

    data = {
        'register': register,
        'totals': totals,
        'financial_year_name': fy_name,
        'institute': Institute.get_institute(view_self),
    }

    default = 'default_fixed_asset_register.html'
    selected_template, _ = get_selected_template(view_self, 'asset', 'pdf', default)
    path = 'asset/' + selected_template
    return PDFService.receipt_new(view_self, data, 'fixed_asset_register', path, False)


def download_asset_group_summary_pdf(view_self):
    """Generate PDF for Asset Group Summary using PDFService.receipt_new()."""
    financial_year_id = view_self.request.GET.get('financial_year')
    if not financial_year_id:
        return HttpResponse('financial_year is required', status=400)

    report = get_asset_group_summary(view_self, financial_year_id)
    data = report['data']
    data['institute'] = Institute.get_institute(view_self)
    data['financial_year_name'] = data['financial_year']['name']

    default = 'default_asset_group_summary.html'
    selected_template, _ = get_selected_template(view_self, 'asset', 'pdf', default)
    path = 'asset/' + selected_template
    return PDFService.receipt_new(view_self, data, 'asset_group_summary', path, False)


def download_depreciation_schedule_pdf(view_self):
    """Generate PDF for Depreciation Schedule using PDFService.receipt_new()."""
    financial_year_id = view_self.request.GET.get('financial_year')
    asset_id = view_self.request.GET.get('asset')

    report = get_depreciation_schedule(view_self, financial_year_id, asset_id)
    data = report['data']
    data['institute'] = Institute.get_institute(view_self)

    default = 'default_depreciation_schedule.html'
    selected_template, _ = get_selected_template(view_self, 'asset', 'pdf', default)
    path = 'asset/' + selected_template
    return PDFService.receipt_new(view_self, data, 'depreciation_schedule', path, False)


def download_fixed_asset_cost_register_pdf(view_self):
    """Generate PDF for Fixed Asset Cost Register using PDFService.receipt_new()."""
    financial_year_id = view_self.request.GET.get('financial_year')
    if not financial_year_id:
        return HttpResponse('financial_year is required', status=400)

    report = get_fixed_asset_cost_register(view_self, financial_year_id)
    data = report['data']

    try:
        fy = FinancialYear.objects.get(id=financial_year_id)
        fy_name = f"{fy.start_date.year}-{fy.end_date.year}"
    except FinancialYear.DoesNotExist:
        fy_name = ''

    data['financial_year_name'] = fy_name
    data['institute'] = Institute.get_institute(view_self)

    default = 'default_fixed_asset_cost_register.html'
    selected_template, _ = get_selected_template(view_self, 'asset', 'pdf', default)
    path = 'asset/' + selected_template
    return PDFService.receipt_new(view_self, data, 'fixed_asset_cost_register', path, False)


def download_asset_group_cost_summary_pdf(view_self):
    """Generate PDF for Asset Group Cost Summary using PDFService.receipt_new()."""
    financial_year_id = view_self.request.GET.get('financial_year')
    if not financial_year_id:
        return HttpResponse('financial_year is required', status=400)

    report = get_asset_group_cost_summary(view_self, financial_year_id)
    data = report['data']

    data['financial_year_name'] = data['financial_year']['name']
    data['institute'] = Institute.get_institute(view_self)

    default = 'default_asset_group_cost_summary.html'
    selected_template, _ = get_selected_template(view_self, 'asset', 'pdf', default)
    path = 'asset/' + selected_template
    return PDFService.receipt_new(view_self, data, 'asset_group_cost_summary', path, False)


# ━━━━━━━━━━━━━━━ LONG-RUNNING PROCESS (LRP) WRAPPERS ━━━━━━━━━━━━━━━

logger = logging.getLogger(__name__)


def _run_lrp_excel(view_self, download_fn, report_label):
    """
    Generic LRP wrapper for any download-excel function.
    Calls the download function, saves to a temp file, uploads, and stores result.
    """

    transaction_id = view_self.request.GET.get('transaction_id')
    financial_year_id = view_self.request.GET.get('financial_year')
    try:
        response = download_fn(view_self, financial_year_id)
        if response.status_code == 200:
            file_name = f'{report_label}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            with open(file_name, 'wb') as f:
                f.write(response.content)
            url = UploadTypeService.upload_local_file(file_name, path='AssetReports')
            if os.path.exists(file_name):
                os.remove(file_name)
            store_long_running_process(view_self, transaction_id, {'url': url})
        else:
            store_long_running_process(
                view_self, transaction_id,
                {'error': f'Error with status code {response.status_code}'},
            )
    except Exception as e:
        logger.error(f'Error in {report_label} LRP: {e}', exc_info=True)
        store_long_running_process(
            view_self, transaction_id, {'error': str(e)[:250]},
        )


def download_fixed_asset_register_excel_lrp(view_self):
    """LRP wrapper for Fixed Asset Register Excel."""
    _run_lrp_excel(view_self, download_fixed_asset_register_excel, 'fixed_asset_register')


def download_asset_group_summary_excel_lrp(view_self):
    """LRP wrapper for Asset Group Summary Excel."""
    _run_lrp_excel(view_self, download_asset_group_summary_excel, 'asset_group_summary')


def download_fixed_asset_cost_register_excel_lrp(view_self):
    """LRP wrapper for Fixed Asset Cost Register Excel."""
    _run_lrp_excel(view_self, download_fixed_asset_cost_register_excel, 'asset_cost_register')


def download_asset_group_cost_summary_excel_lrp(view_self):
    """LRP wrapper for Asset Group Cost Summary Excel."""
    _run_lrp_excel(view_self, download_asset_group_cost_summary_excel, 'asset_group_cost_summary')


def _run_lrp_pdf(view_self, download_fn, report_label):
    """Generic LRP wrapper for PDF download functions."""
    transaction_id = view_self.request.GET.get('transaction_id')
    try:
        response = download_fn(view_self)
        if response.status_code == 200:
            file_name = f'{report_label}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            with open(file_name, 'wb') as f:
                f.write(response.content)
            url = UploadTypeService.upload_local_file(file_name, path='AssetReports')
            if os.path.exists(file_name):
                os.remove(file_name)
            store_long_running_process(view_self, transaction_id, {'url': url})
        else:
            store_long_running_process(
                view_self, transaction_id,
                {'error': f'Error with status code {response.status_code}'},
            )
    except Exception as e:
        logger.error(f'Error in {report_label} PDF LRP: {e}', exc_info=True)
        store_long_running_process(
            view_self, transaction_id, {'error': str(e)[:250]},
        )


def download_fixed_asset_register_pdf_lrp(view_self):
    """LRP wrapper for Fixed Asset Register PDF."""
    _run_lrp_pdf(view_self, download_fixed_asset_register_pdf, 'fixed_asset_register')


def download_asset_group_summary_pdf_lrp(view_self):
    """LRP wrapper for Asset Group Summary PDF."""
    _run_lrp_pdf(view_self, download_asset_group_summary_pdf, 'asset_group_summary')


def download_depreciation_schedule_pdf_lrp(view_self):
    """LRP wrapper for Depreciation Schedule PDF."""
    _run_lrp_pdf(view_self, download_depreciation_schedule_pdf, 'depreciation_schedule')


def download_fixed_asset_cost_register_pdf_lrp(view_self):
    """LRP wrapper for Fixed Asset Cost Register PDF."""
    _run_lrp_pdf(view_self, download_fixed_asset_cost_register_pdf, 'fixed_asset_cost_register')


def download_asset_group_cost_summary_pdf_lrp(view_self):
    """LRP wrapper for Asset Group Cost Summary PDF."""
    _run_lrp_pdf(view_self, download_asset_group_cost_summary_pdf, 'asset_group_cost_summary')


def download_disposal_list_pdf(view_self):
    """Generate PDF for Disposal List using PDFService.receipt_new()."""
    financial_year_id = view_self.request.GET.get('financial_year')

    report = get_disposal_list(view_self, financial_year_id)
    data = report['data']
    data['institute'] = Institute.get_institute(view_self)

    default = 'default_disposal_list.html'
    selected_template, _ = get_selected_template(view_self, 'asset', 'pdf', default)
    path = 'asset/' + selected_template
    return PDFService.receipt_new(view_self, data, 'disposal_list', path, False)


def download_disposal_list_excel_lrp(view_self):
    """LRP wrapper for Disposal List Excel."""
    _run_lrp_excel(view_self, download_disposal_list_excel, 'disposal_list')


def download_disposal_list_pdf_lrp(view_self):
    """LRP wrapper for Disposal List PDF."""
    _run_lrp_pdf(view_self, download_disposal_list_pdf, 'disposal_list')
