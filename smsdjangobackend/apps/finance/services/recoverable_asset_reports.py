import io
import logging
import os
from datetime import date, datetime as dt
from decimal import Decimal

import openpyxl
from django.db.models import Sum, Case, When, DecimalField, F, Q
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill

from apps.finance.models.recoverable_asset import RecoverableAsset, RecoverableAssetTransaction
from apps.finance.models.recoverable_asset_category import RecoverableAssetCategory
from apps.institutes.models import FinancialYear, Institute
from apps.shared.services import PDFService, UploadTypeService
from apps.shared.services_shared.common import get_selected_template
from apps.shared.services_shared.store_api_result import store_long_running_process

ZERO = Decimal('0.00')


def get_ledger_report(asset_id, from_date=None, to_date=None):
    asset = RecoverableAsset.objects.select_related('category').get(id=asset_id)

    txn_qs = RecoverableAssetTransaction.objects.filter(
        recoverable_asset=asset, is_active=True
    ).order_by('transaction_date', 'created_at')

    if from_date:
        txn_qs = txn_qs.filter(transaction_date__gte=from_date)
    if to_date:
        txn_qs = txn_qs.filter(transaction_date__lte=to_date)

    transactions = []
    running_balance = asset.opening_balance

    for txn in txn_qs:
        debit = txn.amount if txn.is_debit_type() else ZERO
        credit = txn.amount if txn.is_credit_type() else ZERO
        running_balance = running_balance + debit - credit

        transactions.append({
            'id': txn.id,
            'date': txn.transaction_date.isoformat(),
            'type': txn.get_transaction_type_display(),
            'type_code': txn.transaction_type,
            'remarks': txn.remarks or '',
            'debit': debit,
            'credit': credit,
            'balance': running_balance,
            'source': txn.get_source_type_display(),
        })

    return {
        'asset_id': asset.id,
        'asset_name': asset.get_particulars(),
        'category': asset.category.name if asset.category else asset.get_asset_type_display(),
        'opening_balance': asset.opening_balance,
        'closing_balance': asset.closing_balance,
        'transactions': transactions,
    }


def download_ledger_pdf(view_self):
    asset_id = view_self.request.GET.get('asset_id')
    from_date = view_self.request.GET.get('from_date')
    to_date = view_self.request.GET.get('to_date')

    if not asset_id:
        return HttpResponse('asset_id is required', status=400)

    data = get_ledger_report(asset_id, from_date, to_date)

    default = 'default_recoverable_ledger_report.html'
    selected_template, number_of_copies = get_selected_template(
        view_self, 'recoverable_asset', 'pdf', default
    )
    data['institute'] = Institute.get_institute(view_self)
    path = 'recoverable_asset/' + selected_template
    response = PDFService.receipt_new(view_self, data, f'ledger_{asset_id}', path, False)
    return response


def download_ledger_excel(view_self):
    """Generate Excel ledger for a single asset."""

    asset_id = view_self.request.GET.get('asset_id')
    from_date = view_self.request.GET.get('from_date')
    to_date = view_self.request.GET.get('to_date')

    if not asset_id:
        return HttpResponse('asset_id is required', status=400)

    data = get_ledger_report(asset_id, from_date, to_date)
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ledger'

    ws.merge_cells('A1:F1')
    ws['A1'] = inst_name
    ws['A1'].font = Font(size=16, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Asset Ledger — {data['asset_name']}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:F3')
    ws['A3'] = f"Opening: ₹{data['opening_balance']:,.2f} | Closing: ₹{data['closing_balance']:,.2f}"
    ws['A3'].font = Font(size=10, italic=True, color='4472C4')
    ws['A3'].alignment = Alignment(horizontal='center')

    headers = ['Date', 'Type', 'Remarks', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)']
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for i, txn in enumerate(data['transactions'], 6):
        ws.cell(row=i, column=1, value=txn['date'])
        ws.cell(row=i, column=2, value=txn['type'])
        ws.cell(row=i, column=3, value=txn['remarks'])
        ws.cell(row=i, column=4, value=float(txn['debit'])).number_format = '#,##0.00'
        ws.cell(row=i, column=5, value=float(txn['credit'])).number_format = '#,##0.00'
        ws.cell(row=i, column=6, value=float(txn['balance'])).number_format = '#,##0.00'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ledger_{asset_id}.xlsx'
    return response


def get_category_summary(financial_year_id=None):
    categories = RecoverableAssetCategory.objects.filter(is_active=True).order_by('display_order')

    summary = []
    grand_opening = ZERO
    grand_closing = ZERO

    for cat in categories:
        assets = RecoverableAsset.objects.filter(
            is_active=True,
            category=cat,
            status__in=['APPROVED', 'CLOSED']
        )
        if financial_year_id:
            assets = assets.filter(category__financial_year_id=financial_year_id)

        total_opening = assets.aggregate(t=Sum('opening_balance'))['t'] or ZERO
        total_closing = assets.aggregate(t=Sum('closing_balance'))['t'] or ZERO
        asset_count = assets.count()

        summary.append({
            'category_id': cat.id,
            'category_code': cat.code,
            'category_name': cat.name,
            'asset_count': asset_count,
            'total_opening': total_opening,
            'total_closing': total_closing,
            'net_movement': total_closing - total_opening,
        })
        grand_opening += total_opening
        grand_closing += total_closing

    return {
        'categories': summary,
        'grand_total_opening': grand_opening,
        'grand_total_closing': grand_closing,
        'grand_net_movement': grand_closing - grand_opening,
    }


def download_category_summary_pdf(view_self):
    financial_year_id = view_self.request.GET.get('financial_year_id')
    data = get_category_summary(financial_year_id)

    default = 'default_recoverable_category_summary.html'
    selected_template, number_of_copies = get_selected_template(
        view_self, 'recoverable_asset', 'pdf', default
    )
    data['institute'] = Institute.get_institute(view_self)
    path = 'recoverable_asset/' + selected_template
    response = PDFService.receipt_new(view_self, data, 'category_summary', path, False)
    return response


def download_category_summary_excel(view_self):

    financial_year_id = view_self.request.GET.get('financial_year_id')
    data = get_category_summary(financial_year_id)
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Category Summary'

    ws.merge_cells('A1:E1')
    ws['A1'] = inst_name
    ws['A1'].font = Font(size=16, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = 'Recoverable Assets — Category Summary'
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Category', 'Assets', 'Opening (₹)', 'Closing (₹)', 'Net Movement (₹)']
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for i, cat in enumerate(data['categories'], 5):
        ws.cell(row=i, column=1, value=cat['category_name'])
        ws.cell(row=i, column=2, value=cat['asset_count'])
        ws.cell(row=i, column=3, value=float(cat['total_opening'])).number_format = '#,##0.00'
        ws.cell(row=i, column=4, value=float(cat['total_closing'])).number_format = '#,##0.00'
        ws.cell(row=i, column=5, value=float(cat['net_movement'])).number_format = '#,##0.00'

    row = 5 + len(data['categories'])
    ws.cell(row=row, column=1, value='GRAND TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(data['grand_total_opening'])).number_format = '#,##0.00'
    ws.cell(row=row, column=4, value=float(data['grand_total_closing'])).number_format = '#,##0.00'
    ws.cell(row=row, column=5, value=float(data['grand_net_movement'])).number_format = '#,##0.00'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=category_summary.xlsx'
    return response

def get_period_report(from_date, to_date, category_id=None):
    txn_qs = RecoverableAssetTransaction.objects.filter(
        is_active=True,
        transaction_date__gte=from_date,
        transaction_date__lte=to_date,
    ).select_related('recoverable_asset', 'recoverable_asset__category').order_by('transaction_date')

    if category_id:
        txn_qs = txn_qs.filter(recoverable_asset__category_id=category_id)

    transactions = []
    total_debits = ZERO
    total_credits = ZERO

    for txn in txn_qs:
        debit = txn.amount if txn.is_debit_type() else ZERO
        credit = txn.amount if txn.is_credit_type() else ZERO
        total_debits += debit
        total_credits += credit

        transactions.append({
            'date': txn.transaction_date.isoformat(),
            'asset_name': txn.recoverable_asset.get_particulars(),
            'category': txn.recoverable_asset.category.name if txn.recoverable_asset.category else '',
            'type': txn.get_transaction_type_display(),
            'debit': debit,
            'credit': credit,
            'remarks': txn.remarks or '',
        })

    return {
        'from_date': from_date,
        'to_date': to_date,
        'transactions': transactions,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'net_movement': total_debits - total_credits,
    }


def download_period_report_pdf(view_self):
    from_date = view_self.request.GET.get('from_date')
    to_date = view_self.request.GET.get('to_date')
    category_id = view_self.request.GET.get('category_id')

    if not from_date or not to_date:
        return HttpResponse('from_date and to_date are required', status=400)

    data = get_period_report(from_date, to_date, category_id)

    default = 'default_recoverable_period_report.html'
    selected_template, number_of_copies = get_selected_template(
        view_self, 'recoverable_asset', 'pdf', default
    )
    data['institute'] = Institute.get_institute(view_self)
    path = 'recoverable_asset/' + selected_template
    response = PDFService.receipt_new(
        view_self, data, f'period_report_{from_date}_{to_date}', path, False
    )
    return response


def download_period_report_excel(view_self):

    from_date = view_self.request.GET.get('from_date')
    to_date = view_self.request.GET.get('to_date')
    category_id = view_self.request.GET.get('category_id')

    if not from_date or not to_date:
        return HttpResponse('from_date and to_date are required', status=400)

    data = get_period_report(from_date, to_date, category_id)
    inst_obj = Institute.objects.first()
    inst_name = inst_obj.name if inst_obj else ''

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Period Report'

    ws.merge_cells('A1:G1')
    ws['A1'] = inst_name
    ws['A1'].font = Font(size=16, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Recoverable Assets — Period Report ({from_date} to {to_date})'
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Date', 'Asset', 'Category', 'Type', 'Debit (₹)', 'Credit (₹)', 'Remarks']
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for i, txn in enumerate(data['transactions'], 5):
        ws.cell(row=i, column=1, value=txn['date'])
        ws.cell(row=i, column=2, value=txn['asset_name'])
        ws.cell(row=i, column=3, value=txn['category'])
        ws.cell(row=i, column=4, value=txn['type'])
        ws.cell(row=i, column=5, value=float(txn['debit'])).number_format = '#,##0.00'
        ws.cell(row=i, column=6, value=float(txn['credit'])).number_format = '#,##0.00'
        ws.cell(row=i, column=7, value=txn['remarks'])

    row = 5 + len(data['transactions'])
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(data['total_debits'])).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=float(data['total_credits'])).number_format = '#,##0.00'

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=period_report_{from_date}_{to_date}.xlsx'
    return response

logger = logging.getLogger(__name__)


def _run_recoverable_lrp(view_self, download_fn, report_label, ext='xlsx'):
    
    transaction_id = view_self.request.GET.get('transaction_id')
    try:
        response = download_fn(view_self)
        if response.status_code == 200:
            file_name = f'{report_label}_{dt.now().strftime("%Y%m%d_%H%M%S")}.{ext}'
            with open(file_name, 'wb') as f:
                f.write(response.content)
            url = UploadTypeService.upload_local_file(file_name, path='RecoverableAssetReports')
            if os.path.exists(file_name):
                os.remove(file_name)
            store_long_running_process(view_self, transaction_id, {'url': url})
        else:
            store_long_running_process(
                view_self, transaction_id,
                {'error': f'Error with status code {response.status_code}'},
            )
    except Exception as e:
        logger.error(f'Error in {report_label} LRP: {e}', exc_info=True)
        store_long_running_process(
            view_self, transaction_id, {'error': str(e)[:250]},
        )


def download_ledger_pdf_lrp(view_self):
    _run_recoverable_lrp(view_self, download_ledger_pdf, 'ledger_report', 'pdf')


def download_ledger_excel_lrp(view_self):
    _run_recoverable_lrp(view_self, download_ledger_excel, 'ledger_report', 'xlsx')


def download_category_summary_pdf_lrp(view_self):
    _run_recoverable_lrp(view_self, download_category_summary_pdf, 'category_summary', 'pdf')


def download_category_summary_excel_lrp(view_self):
    _run_recoverable_lrp(view_self, download_category_summary_excel, 'category_summary', 'xlsx')


def download_period_report_pdf_lrp(view_self):
    _run_recoverable_lrp(view_self, download_period_report_pdf, 'period_report', 'pdf')


def download_period_report_excel_lrp(view_self):
    _run_recoverable_lrp(view_self, download_period_report_excel, 'period_report', 'xlsx')
