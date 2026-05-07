import io
import os
import logging
from decimal import Decimal
from datetime import datetime, timedelta

from django.db.models import Sum, Case, When, DecimalField, F, Q
from rest_framework import exceptions

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from apps.asset.models import Asset, AssetGroup, AssetCostMovement
from apps.finance.models.recoverable_asset import RecoverableAsset
from apps.institutes.models import FinancialYear, Institute
from apps.shared.services import PDFService, UploadTypeService
from apps.shared.services_shared.common import get_selected_template
from apps.shared.services_shared.store_api_result import store_long_running_process
from apps.finance.models.recoverable_asset_category import RecoverableAssetCategory
from apps.finance.services.advance_fee_calculator import compute_advance_fee_balance
from apps.finance.serializers import _compute_actual_bank_balance

ZERO = Decimal('0.00')


def _get_financial_year(financial_year_id):
    fy = FinancialYear.objects.get(id=financial_year_id)
    return fy, fy.start_date, fy.end_date


def _build_fixed_assets(from_date, to_date, financial_year_id=None, group_type='FIXED_ASSET'):
    
    parent_groups = AssetGroup.objects.filter(
        is_active=True, parent_group__isnull=True, group_type=group_type
    ).order_by('display_order', 'name')

    all_movements = AssetCostMovement.objects.filter(movement_date__lte=to_date)

    opening_agg = all_movements.filter(
        Q(movement_date__lt=from_date) | Q(movement_date=from_date, movement_type='OPENING')
    ).values('asset_id').annotate(
        opening=Sum(
            Case(
                When(movement_type__in=['OPENING', 'ADDITION'], then=F('amount')),
                When(movement_type='DISPOSAL', then=-F('amount')),
                default=ZERO, output_field=DecimalField()
            )
        )
    )
    opening_map = {r['asset_id']: r['opening'] or ZERO for r in opening_agg}

    period_add = all_movements.filter(
        movement_date__gte=from_date,
        movement_date__lte=to_date,
        movement_type='ADDITION'
    ).values('asset_id').annotate(total=Sum('amount'))
    add_map = {r['asset_id']: r['total'] or ZERO for r in period_add}

    period_disp = all_movements.filter(
        movement_date__gte=from_date,
        movement_date__lte=to_date,
        movement_type='DISPOSAL'
    ).values('asset_id').annotate(total=Sum('amount'))
    disp_map = {r['asset_id']: r['total'] or ZERO for r in period_disp}

    group_ids = list(AssetGroup.objects.filter(
        is_active=True, group_type=group_type
    ).values_list('id', flat=True))

    asset_ids_with_movements = set(all_movements.values_list('asset_id', flat=True))
    all_assets = Asset.objects.filter(
        is_active=True, asset_group_id__in=group_ids, id__in=asset_ids_with_movements
    ).select_related('asset_group').order_by('asset_name')

    assets_by_group = {}
    for asset in all_assets:
        gid = asset.asset_group_id
        if gid not in assets_by_group:
            assets_by_group[gid] = []
        assets_by_group[gid].append(asset)

    def _build_asset_items(group_id):
        items = []
        group_total = ZERO
        group_opening_total = ZERO
        for asset in assets_by_group.get(group_id, []):
            opening = opening_map.get(asset.id, ZERO)
            additions = add_map.get(asset.id, ZERO)
            disposals = disp_map.get(asset.id, ZERO)
            closing = opening + additions - disposals
            if closing == ZERO and opening == ZERO:
                continue
            items.append({
                'id': f'FA_{asset.id}',
                'name': asset.asset_name,
                'opening_balance': opening,
                'additions': additions,
                'deductions': disposals,
                'closing_balance': closing,
                'source': 'FIXED_ASSET',
            })
            group_total += closing
            group_opening_total += opening
        return items, group_total, group_opening_total

    groups = []
    grand_total = ZERO
    grand_opening_total = ZERO

    for parent in parent_groups:
        child_groups = AssetGroup.objects.filter(
            parent_group=parent, is_active=True
        ).order_by('display_order', 'name')

        sub_groups = []
        for child in child_groups:
            child_items, child_total, child_opening_total = _build_asset_items(child.id)
            if child_items or child_total != ZERO or child_opening_total != ZERO:
                sub_groups.append({
                    'group_name': child.name,
                    'group_code': child.code if hasattr(child, 'code') else f'FA_SG_{child.id}',
                    'items': child_items,
                    'total': child_total,
                    'opening_total': child_opening_total,
                })

        direct_items, direct_total, direct_opening_total = _build_asset_items(parent.id)

        group_total = direct_total + sum(sg['total'] for sg in sub_groups)
        group_opening_total = direct_opening_total + sum(sg['opening_total'] for sg in sub_groups)

        if group_total == ZERO and group_opening_total == ZERO and not direct_items and not sub_groups:
            continue

        group_entry = {
            'group_name': parent.name,
            'group_code': parent.code if hasattr(parent, 'code') else f'FA_{parent.id}',
            'total': group_total,
            'opening_total': group_opening_total,
        }

        if sub_groups:
            group_entry['sub_groups'] = sub_groups
        if direct_items:
            group_entry['items'] = direct_items

        groups.append(group_entry)
        grand_total += group_total
        grand_opening_total += group_opening_total

    label = 'Fixed Assets' if group_type == 'FIXED_ASSET' else 'Assets (Liability Side)'
    code = 'FIXED_ASSETS' if group_type == 'FIXED_ASSET' else 'LIABILITY_ASSETS'

    return {
        'group_name': label,
        'group_code': code,
        'sub_groups': groups,
        'total': grand_total,
        'opening_total': grand_opening_total,
    }

def _build_recoverable_group(group_name, group_code, from_date, to_date,
                              financial_year_id=None, fc_code=None,
                              category_codes=None,
                              balance_sheet_classification=None,
                              category_ids=None):

    queryset = RecoverableAsset.objects.filter(
        is_active=True,
        status__in=['APPROVED', 'CLOSED']
    ).select_related('category', 'category__financial_year', 'salary_advance')

    if financial_year_id:
        queryset = queryset.filter(
            Q(category__financial_year_id=financial_year_id) |
            Q(category__financial_year__isnull=True)
        )

    if balance_sheet_classification:
        queryset = queryset.filter(category__balance_sheet_classification=balance_sheet_classification)

    if category_ids:
        queryset = queryset.filter(category_id__in=category_ids)
    elif fc_code:
        queryset = queryset.filter(category__financial_category__code=fc_code)
    elif category_codes:
        queryset = queryset.filter(category__code__in=category_codes)
    elif not balance_sheet_classification:
        return {'group_name': group_name, 'group_code': group_code,
                'items': [], 'total': ZERO, 'opening_total': ZERO}

    items = []
    total_closing_debit = ZERO
    total_closing_credit = ZERO
    total_opening_debit = ZERO
    total_opening_credit = ZERO

    for asset in queryset:
        ob_type = getattr(asset, 'opening_balance_type', 'DEBIT') or 'DEBIT'

        if asset.linked_module == 'BANK_ACCOUNT' and asset.bank:
            bank_opening = _compute_actual_bank_balance(asset.bank, up_to_date=from_date - timedelta(days=1))
            bank_closing = _compute_actual_bank_balance(asset.bank, up_to_date=to_date)

            if ob_type == 'DEBIT':
                o_debit, o_credit = bank_opening, ZERO
                c_debit, c_credit = bank_closing, ZERO
            else:
                o_debit, o_credit = ZERO, bank_opening
                c_debit, c_credit = ZERO, bank_closing

        elif asset.linked_module == 'ADVANCE_FEE' and asset.advance_fee_config:
            adv_opening = Decimal(str(asset.opening_balance or 0))
            adv_closing = compute_advance_fee_balance(asset, up_to_date=to_date)

            o_debit, o_credit = ZERO, max(adv_opening, ZERO)
            c_debit, c_credit = ZERO, max(adv_closing, ZERO)

        elif asset.linked_module == 'STAFF_SALARY_ADVANCE' and asset.salary_advance_id:
            sa = asset.salary_advance
            sa_opening = Decimal(str(sa.opening_balance or 0))
            sa_closing = Decimal(str(sa.closing_balance or 0))

            if ob_type == 'DEBIT':
                o_debit, o_credit = max(sa_opening, ZERO), ZERO
                c_debit, c_credit = max(sa_closing, ZERO), ZERO
            else:
                o_debit, o_credit = ZERO, max(sa_opening, ZERO)
                c_debit, c_credit = ZERO, max(sa_closing, ZERO)

        else:
            pre_txns = asset.recoverable_asset_transaction_recoverable_asset.filter(
                is_active=True, transaction_date__lt=from_date
            ).aggregate(
                debits=Sum(Case(
                    When(transaction_type__in=['DEBIT', 'ADVANCE', 'INTEREST', 'PENALTY'], then='amount'),
                    default=ZERO, output_field=DecimalField()
                )),
                credits=Sum(Case(
                    When(transaction_type__in=['CREDIT', 'RECOVERY', 'ADJUSTMENT', 'REVERSAL'], then='amount'),
                    default=ZERO, output_field=DecimalField()
                ))
            )

            pre_debits = pre_txns['debits'] or ZERO
            pre_credits = pre_txns['credits'] or ZERO

            if ob_type == 'DEBIT':
                net_opening = asset.opening_balance + pre_debits - pre_credits
                o_debit = max(net_opening, ZERO)
                o_credit = max(-net_opening, ZERO)
            else:
                net_opening = asset.opening_balance + pre_credits - pre_debits
                o_debit = max(-net_opening, ZERO)
                o_credit = max(net_opening, ZERO)

            period = asset.recoverable_asset_transaction_recoverable_asset.filter(
                is_active=True,
                transaction_date__gte=from_date,
                transaction_date__lte=to_date
            ).aggregate(
                debits=Sum(Case(
                    When(transaction_type__in=['DEBIT', 'ADVANCE', 'INTEREST', 'PENALTY'], then='amount'),
                    default=ZERO, output_field=DecimalField()
                )),
                credits=Sum(Case(
                    When(transaction_type__in=['CREDIT', 'RECOVERY', 'ADJUSTMENT', 'REVERSAL'], then='amount'),
                    default=ZERO, output_field=DecimalField()
                ))
            )
            p_debits = period['debits'] or ZERO
            p_credits = period['credits'] or ZERO

            c_debit = o_debit + p_debits
            c_credit = o_credit + p_credits

        amount = c_debit - c_credit

        if c_debit == ZERO and c_credit == ZERO and o_debit == ZERO and o_credit == ZERO:
            continue

        items.append({
            'id': f'RA_{asset.id}',
            'name': asset.get_particulars(),
            'opening_debit': o_debit if o_debit else None,
            'opening_credit': o_credit if o_credit else None,
            'closing_debit': c_debit if c_debit else None,
            'closing_credit': c_credit if c_credit else None,
            'amount': float(abs(amount)) if amount != ZERO else None,
            'source': 'RECOVERABLE_ASSET',
        })
        total_closing_debit += c_debit
        total_closing_credit += c_credit
        total_opening_debit += o_debit
        total_opening_credit += o_credit

    total = total_closing_debit - total_closing_credit
    opening_total = total_opening_debit - total_opening_credit

    return {
        'group_name': group_name,
        'group_code': group_code,
        'items': items,
        'total': total,
        'opening_total': opening_total,
        'total_closing_debit': total_closing_debit,
        'total_closing_credit': total_closing_credit,
        'total_opening_debit': total_opening_debit,
        'total_opening_credit': total_opening_credit,
    }

def _build_ra_groups_by_classification(classification, from_date, to_date, financial_year_id):
    categories = RecoverableAssetCategory.objects.filter(
        is_active=True,
        balance_sheet_classification=classification,
    ).order_by('display_order', 'name')

    if financial_year_id:
        categories = categories.filter(
            Q(financial_year_id=financial_year_id) | Q(financial_year__isnull=True)
        )

    groups = []
    for cat in categories:
        group = _build_recoverable_group(
            cat.name, f'RA_CAT_{cat.id}',
            from_date, to_date, financial_year_id,
            balance_sheet_classification=classification,
            category_ids=[cat.id],
        )
        if group['total'] != ZERO or group.get('items'):
            groups.append(group)

    return groups

def get_balance_sheet(financial_year_id):

    fy, from_date, to_date = _get_financial_year(financial_year_id)
    fy_label = f"{from_date.year}-{to_date.year}"

    fixed_assets = _build_fixed_assets(from_date, to_date, financial_year_id, group_type='FIXED_ASSET')

    ra_right_groups = _build_ra_groups_by_classification(
        'FIXED_ASSET', from_date, to_date, financial_year_id
    )

    right_groups = [fixed_assets] + ra_right_groups
    right_groups = [g for g in right_groups if g.get('total', ZERO) != ZERO or g.get('items') or g.get('sub_groups')]

    total_right = sum(g.get('total', ZERO) for g in right_groups)
    total_opening_right = sum(g.get('opening_total', ZERO) for g in right_groups)

    right_side = {
        'title': 'Fixed Assets',
        'classification': 'FIXED_ASSET',
        'groups': right_groups,
        'total': total_right,
        'opening_total': total_opening_right,
    }

    liability_assets = _build_fixed_assets(from_date, to_date, financial_year_id, group_type='LIABILITY')

    ra_left_groups = _build_ra_groups_by_classification(
        'LIABILITY', from_date, to_date, financial_year_id
    )

    left_groups = [liability_assets] + ra_left_groups
    left_groups = [g for g in left_groups if g.get('total', ZERO) != ZERO or g.get('items') or g.get('sub_groups')]

    total_left_debit = sum(g.get('total_closing_debit', ZERO) for g in left_groups)
    total_left_credit = sum(g.get('total_closing_credit', ZERO) for g in left_groups)
    opening_left_debit = sum(g.get('total_opening_debit', ZERO) for g in left_groups)
    opening_left_credit = sum(g.get('total_opening_credit', ZERO) for g in left_groups)

    total_left_net = sum(g.get('total', ZERO) for g in left_groups)
    opening_left_net = sum(g.get('opening_total', ZERO) for g in left_groups)

    left_side = {
        'title': 'Liabilities',
        'classification': 'LIABILITY',
        'groups': left_groups,
        'total': total_left_net,
        'opening_total': opening_left_net,
    }

    display_total_liabilities = total_left_credit or abs(total_left_net)
    display_total_assets = total_right
    difference = display_total_assets - display_total_liabilities

    return {
        'financial_year_id': financial_year_id,
        'financial_year_label': fy_label,
        'as_of_date': to_date.isoformat(),
        'liabilities': left_side,
        'assets': right_side,
        'total_liabilities': display_total_liabilities,
        'total_assets': display_total_assets,
        'opening_total_liabilities': opening_left_credit or abs(opening_left_net),
        'opening_total_assets': total_opening_right,
        'is_balanced': difference == ZERO,
        'difference': difference,
    }

def _get_institute_info():
    inst = Institute.objects.first()
    if not inst:
        return {'name': '', 'trust_name': '', 'address': '', 'phone': '', 'email': ''}

    address = ''
    try:
        inst_data = inst.get_institute()
        if inst_data:
            address = getattr(inst_data, 'address', '') or ''
    except Exception:
        pass

    return {
        'name': inst.name or '',
        'trust_name': inst.trust_name or '',
        'phone': inst.tel_num or '',
        'email': inst.email or '',
        'address': address,
    }


def _flatten_balance_sheet(data):
    
    rows = []

    def _split_debit_credit(val, is_liability_side):
        
        v = float(val) if val is not None else 0
        if v == 0:
            return None, None
        if is_liability_side:
            return (abs(v), None) if v < 0 else (None, v)
        else:
            return (v, None) if v > 0 else (None, abs(v))

    def _add_items(group_items, is_liability):
        for item in group_items:
            if 'opening_debit' in item:
                rows.append({
                    'type': 'item',
                    'name': item['name'],
                    'opening_debit': _to_float(item.get('opening_debit')),
                    'opening_credit': _to_float(item.get('opening_credit')),
                    'closing_debit': _to_float(item.get('closing_debit')),
                    'closing_credit': _to_float(item.get('closing_credit')),
                    'amount': _to_float(item.get('amount')),
                })
            else:
                op = item.get('opening_balance', ZERO)
                cl = item.get('closing_balance', ZERO)
                op_d, op_c = _split_debit_credit(op, is_liability)
                cl_d, cl_c = _split_debit_credit(cl, is_liability)
                amt = float(cl) if float(cl) != 0 else None
                rows.append({
                    'type': 'item',
                    'name': item['name'],
                    'opening_debit': op_d, 'opening_credit': op_c,
                    'closing_debit': cl_d, 'closing_credit': cl_c,
                    'amount': amt,
                })

    def _to_float(val):
        if val is None:
            return None
        v = float(val)
        return v if v != 0 else None

    def _add_side(side_data, is_liability):
        rows.append({
            'type': 'section_header',
            'name': side_data.get('title', ''),
            'opening_debit': None, 'opening_credit': None,
            'closing_debit': None, 'closing_credit': None,
            'amount': None,
        })
        for group in side_data.get('groups', []):
            rows.append({
                'type': 'category',
                'name': group['group_name'],
                'opening_debit': None, 'opening_credit': None,
                'closing_debit': None, 'closing_credit': None,
                'amount': None,
            })

            if 'sub_groups' in group:
                for sg in group['sub_groups']:
                    sg_total = float(sg.get('total', ZERO))
                    sg_op_total = float(sg.get('opening_total', ZERO))
                    if sg_total == 0 and sg_op_total == 0 and not sg.get('items'):
                        continue
                    sg_op_d, sg_op_c = _split_debit_credit(sg_op_total, is_liability)
                    sg_cl_d, sg_cl_c = _split_debit_credit(sg_total, is_liability)
                    rows.append({
                        'type': 'subcategory',
                        'name': sg.get('group_name', ''),
                        'opening_debit': sg_op_d, 'opening_credit': sg_op_c,
                        'closing_debit': sg_cl_d, 'closing_credit': sg_cl_c,
                        'amount': float(sg_total) if sg_total != 0 else None,
                    })
                    if 'items' in sg:
                        _add_items(sg['items'], is_liability)

            if 'items' in group:
                _add_items(group['items'], is_liability)

            if 'total_closing_debit' in group:
                cat_op_d = _to_float(group.get('total_opening_debit', ZERO))
                cat_op_c = _to_float(group.get('total_opening_credit', ZERO))
                cat_cl_d = _to_float(group.get('total_closing_debit', ZERO))
                cat_cl_c = _to_float(group.get('total_closing_credit', ZERO))
                cat_amount = _to_float(abs(float(group.get('total', ZERO))))
            else:
                cat_op_total = float(group.get('opening_total', ZERO))
                cat_cl_total = float(group.get('total', ZERO))
                cat_op_d, cat_op_c = _split_debit_credit(cat_op_total, is_liability)
                cat_cl_d, cat_cl_c = _split_debit_credit(cat_cl_total, is_liability)
                cat_amount = float(cat_cl_total) if cat_cl_total != 0 else None

            rows.append({
                'type': 'category_total',
                'name': f"Total {group['group_name']}",
                'opening_debit': cat_op_d, 'opening_credit': cat_op_c,
                'closing_debit': cat_cl_d, 'closing_credit': cat_cl_c,
                'amount': cat_amount,
            })

        side_op = float(side_data.get('opening_total', ZERO))
        side_cl = float(side_data.get('total', ZERO))
        side_op_d, side_op_c = _split_debit_credit(side_op, is_liability)
        side_cl_d, side_cl_c = _split_debit_credit(side_cl, is_liability)
        rows.append({
            'type': 'section_total',
            'name': f"TOTAL {side_data.get('title', '').upper()}",
            'opening_debit': side_op_d, 'opening_credit': side_op_c,
            'closing_debit': side_cl_d, 'closing_credit': side_cl_c,
            'amount': float(side_cl) if side_cl != 0 else None,
        })

    _add_side(data.get('assets', {}), is_liability=False)
    _add_side(data.get('liabilities', {}), is_liability=True)

    return rows


def build_balance_sheet_excel(data, inst):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Balance Sheet'

    school_font = Font(name='Calibri', size=14, bold=True)
    addr_font = Font(name='Calibri', size=10)
    title_font = Font(name='Calibri', size=12, bold=True)
    date_font = Font(name='Calibri', size=11, bold=True)
    section_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=10, bold=True)
    cat_font = Font(name='Calibri', size=10, bold=True)
    cat_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    subcat_font = Font(name='Calibri', size=10, bold=True)
    subcat_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    item_font = Font(name='Calibri', size=10)
    total_font = Font(name='Calibri', size=10, bold=True)
    grand_font = Font(name='Calibri', size=11, bold=True)
    num_fmt = '#,##0.00'
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')
    thin_border = Border(
        bottom=Side(style='thin'),
    )
    double_border = Border(
        top=Side(style='double'),
        bottom=Side(style='double'),
    )

    row = 1
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = inst.get('name', '')
    c.font = school_font
    c.alignment = center

    if inst.get('trust_name'):
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row=row, column=1, value=inst['trust_name'])
        c.font = addr_font
        c.alignment = center

    info_parts = [p for p in [inst.get('address', ''), inst.get('phone', ''), inst.get('email', '')] if p]
    if info_parts:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row=row, column=1, value=', '.join(info_parts))
        c.font = addr_font
        c.alignment = center

    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row=row, column=1, value='Balance Sheet')
    c.font = title_font
    c.alignment = center

    row += 1
    as_of = data.get('as_of_date', '')
    try:
        dt = datetime.strptime(as_of, '%Y-%m-%d') if isinstance(as_of, str) else as_of
        formatted_date = dt.strftime('For %d-%b-%Y')
    except Exception:
        formatted_date = f'For {as_of}'
    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row=row, column=1, value=formatted_date)
    c.font = date_font
    c.alignment = center

    row += 1
    ws.cell(row=row, column=1, value='Particulars').font = section_font
    ws.merge_cells(f'B{row}:C{row}')
    ws.cell(row=row, column=2, value='Opening Balance').font = section_font
    ws.cell(row=row, column=2).alignment = center
    ws.merge_cells(f'D{row}:E{row}')
    ws.cell(row=row, column=4, value='Closing Balance').font = section_font
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=6, value='Amount').font = section_font
    ws.cell(row=row, column=6).alignment = center
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = section_fill

    row += 1
    ws.cell(row=row, column=1, value='').font = section_font
    ws.cell(row=row, column=2, value='Debit').font = section_font
    ws.cell(row=row, column=2).alignment = right
    ws.cell(row=row, column=3, value='Credit').font = section_font
    ws.cell(row=row, column=3).alignment = right
    ws.cell(row=row, column=4, value='Debit').font = section_font
    ws.cell(row=row, column=4).alignment = right
    ws.cell(row=row, column=5, value='Credit').font = section_font
    ws.cell(row=row, column=5).alignment = right
    ws.cell(row=row, column=6, value='').font = section_font
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = section_fill

    data_rows = _flatten_balance_sheet(data)

    data_start = row + 1

    for i, row_dict in enumerate(data_rows):
        r = data_start + i
        rtype = row_dict['type']
        name = row_dict['name']
        op_deb = row_dict['opening_debit']
        op_cred = row_dict['opening_credit']
        cl_deb = row_dict['closing_debit']
        cl_cred = row_dict['closing_credit']
        amt = row_dict.get('amount')

        if rtype == 'section_header':
            cell = ws.cell(row=r, column=1, value=name)
            cell.font = Font(name='Calibri', size=11, bold=True)
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        elif rtype == 'category':
            cell = ws.cell(row=r, column=1, value=name)
            cell.font = cat_font
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = cat_fill

        elif rtype == 'subcategory':
            subcat_font = Font(name='Calibri', size=10, bold=True, color='2E7D32')
            ws.cell(row=r, column=1, value=f'  {name}').font = subcat_font
            fields = [(2, op_deb), (3, op_cred), (4, cl_deb), (5, cl_cred), (6, amt)]
            for col_idx, val in fields:
                if val is not None and val != 0:
                    c = ws.cell(row=r, column=col_idx, value=val)
                    c.number_format = num_fmt
                    c.font = subcat_font
                    c.alignment = right

        elif rtype == 'item':
            ws.cell(row=r, column=1, value=f'    {name}').font = item_font
            fields = [(2, op_deb), (3, op_cred), (4, cl_deb), (5, cl_cred), (6, amt)]
            for col_idx, val in fields:
                if val is not None and val != 0:
                    c = ws.cell(row=r, column=col_idx, value=val)
                    c.number_format = num_fmt
                    c.font = item_font
                    c.alignment = right

        elif rtype == 'category_total' or rtype == 'section_total':
            fnt = total_font if rtype == 'category_total' else grand_font
            brd = thin_border if rtype == 'category_total' else double_border
            ws.cell(row=r, column=1, value=name).font = fnt
            fields = [(2, op_deb), (3, op_cred), (4, cl_deb), (5, cl_cred), (6, amt)]
            for col_idx, val in fields:
                if val is not None and val != 0:
                    c = ws.cell(row=r, column=col_idx, value=val)
                    c.number_format = num_fmt
                    c.font = fnt
                    c.alignment = right
            for c_idx in range(1, 7):
                ws.cell(row=r, column=c_idx).border = brd

    check_row = data_start + len(data_rows) + 1
    if data.get('is_balanced'):
        ws.cell(row=check_row, column=1, value='✓ Balance Sheet is balanced').font = Font(color='006100', bold=True)
    else:
        diff = data.get('difference', ZERO)
        ws.cell(row=check_row, column=1, value=f'⚠ Difference: ₹{diff:,.2f}').font = Font(color='9C0006', bold=True)

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    return wb


def download_balance_sheet_excel(view_self):
    financial_year_id = view_self.request.GET.get('financial_year_id')
    if not financial_year_id:
        return exceptions.ValidationError('financial_year_id is required')

    data = get_balance_sheet(financial_year_id)
    inst = _get_institute_info()
    wb = build_balance_sheet_excel(data, inst)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=balance_sheet_{data["financial_year_label"]}.xlsx'
    return response


def _fmt_amount(val):
    if val is None or val == ZERO or val == 0:
        return ''
    if isinstance(val, float):
        return f'{val:,.2f}'
    return f'{val:,.2f}'


def _flatten_side_rows(side_data, is_liability=False):
    
    rows = []

    def _to_debit_credit(val):
        v = float(val) if val is not None else 0
        if v == 0:
            return 0, 0
        if is_liability:
            return (abs(v), 0) if v < 0 else (0, v)
        else:
            return (v, 0) if v > 0 else (0, abs(v))

    for group in side_data.get('groups', []):
        has_items = group.get('items') or group.get('sub_groups')
        if not has_items and float(group.get('total', 0)) == 0:
            continue

        rows.append({'type': 'category', 'name': group['group_name'], 'debit': 0, 'credit': 0})

        if 'sub_groups' in group:
            for sg in group['sub_groups']:
                sg_total = float(sg.get('total', ZERO))
                if sg_total == 0 and not sg.get('items'):
                    continue
                d, c = _to_debit_credit(sg_total)
                rows.append({'type': 'subcategory', 'name': sg.get('group_name', ''), 'debit': d, 'credit': c})
                for item in sg.get('items', []):
                    cl = item.get('closing_balance', ZERO)
                    d, c = _to_debit_credit(cl)
                    rows.append({'type': 'item', 'name': item['name'], 'debit': d, 'credit': c})

        for item in group.get('items', []):
            cl = item.get('closing_balance', ZERO)
            d, c = _to_debit_credit(cl)
            rows.append({'type': 'item', 'name': item['name'], 'debit': d, 'credit': c})

        cat_total = float(group.get('total', ZERO))
        d, c = _to_debit_credit(cat_total)
        rows.append({'type': 'category_total', 'name': f"Total {group['group_name']}", 'debit': d, 'credit': c})

    side_total = float(side_data.get('total', ZERO))
    d, c = _to_debit_credit(side_total)
    rows.append({'type': 'grand_total', 'name': f"TOTAL {side_data.get('title', '').upper()}", 'debit': d, 'credit': c})

    return rows


def _add_fmt_to_rows(rows):
    for r in rows:
        r['debit_fmt'] = _fmt_amount(r.get('debit'))
        r['credit_fmt'] = _fmt_amount(r.get('credit'))
    return rows


def download_balance_sheet_pdf(view_self):
    financial_year_id = view_self.request.GET.get('financial_year_id')
    if not financial_year_id:
        return exceptions.ValidationError('financial_year_id is required')

    data = get_balance_sheet(financial_year_id)
    inst = _get_institute_info()

    as_of = data.get('as_of_date', '')
    try:
        dt = datetime.strptime(as_of, '%Y-%m-%d') if isinstance(as_of, str) else as_of
        formatted_date = dt.strftime('For %d-%b-%Y')
    except Exception:
        formatted_date = f'For {as_of}'

    info_parts = [p for p in [inst.get('address', ''), inst.get('phone', ''), inst.get('email', '')] if p]
    info_line = ' | '.join(info_parts)

    liab_rows = _add_fmt_to_rows(_flatten_side_rows(data.get('liabilities', {}), is_liability=True))
    asset_rows = _add_fmt_to_rows(_flatten_side_rows(data.get('assets', {}), is_liability=False))

    diff = data.get('difference', ZERO)
    difference_fmt = _fmt_amount(diff) if diff else '0.00'

    context = {
        'data': data,
        'inst': inst,
        'info_line': info_line,
        'formatted_date': formatted_date,
        'liab_rows': liab_rows,
        'asset_rows': asset_rows,
        'difference_fmt': difference_fmt,
        'institute': Institute.get_institute(view_self),
    }

    default = 'default_balance_sheet_report.html'
    selected_template, number_of_copies = get_selected_template(
        view_self, 'balance_sheet', 'pdf', default
    )
    path = 'balance_sheet/' + selected_template
    response = PDFService.receipt_new(
        view_self, context, f'balance_sheet_{data["financial_year_label"]}', path, False
    )
    return response

logger = logging.getLogger(__name__)


def download_balance_sheet_excel_lrp(view_self):
    transaction_id = view_self.request.GET.get('transaction_id')
    try:
        response = download_balance_sheet_excel(view_self)
        if response.status_code == 200:
            file_name = f'balance_sheet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            with open(file_name, 'wb') as f:
                f.write(response.content)
            url = UploadTypeService.upload_local_file(file_name, path='BalanceSheetReports')
            if os.path.exists(file_name):
                os.remove(file_name)
            store_long_running_process(view_self, transaction_id, {'url': url})
        else:
            store_long_running_process(
                view_self, transaction_id,
                {'error': f"Error with status code {response.status_code}"},
            )
    except Exception as e:
        logger.error(f"Error in balance sheet Excel LRP: {e}", exc_info=True)
        store_long_running_process(
            view_self, transaction_id, {'error': str(e)[:250]},
        )


def download_balance_sheet_pdf_lrp(view_self):
    transaction_id = view_self.request.GET.get('transaction_id')
    try:
        response = download_balance_sheet_pdf(view_self)
        if response.status_code == 200:
            file_name = f'balance_sheet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            with open(file_name, 'wb') as f:
                f.write(response.content)
            url = UploadTypeService.upload_local_file(file_name, path='BalanceSheetReports')
            if os.path.exists(file_name):
                os.remove(file_name)
            store_long_running_process(view_self, transaction_id, {'url': url})
        else:
            store_long_running_process(
                view_self, transaction_id,
                {'error': f"Error with status code {response.status_code}"},
            )
    except Exception as e:
        logger.error(f"Error in balance sheet PDF LRP: {e}", exc_info=True)
        store_long_running_process(
            view_self, transaction_id, {'error': str(e)[:250]},
        )
