import copy
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")   # MUST be before pyplot

import matplotlib.pyplot as plt

from datetime import datetime,date
from num2words import num2words
from collections import defaultdict
from django.db.models import F, Value as V, IntegerField
from django.db.models.functions import Concat, Cast
from apps.exams.models.marks import Grade, GradeExamScheduleMapping, GradePlan, StudentCumulativeMark, StudentMarkAuditLog, StudentMarkQuestionWise,StudentMarkSectionSubjectWiseApproval
from rest_framework.exceptions import ValidationError
from apps.exams.models.result_configuration import ExamResultConfiguration
from apps.exams.models.schedule import ExamScheduleCumulativeMapping, ExamScheduleQuestionmapping
from apps.institutes.models import AcademicYear
from apps.institutes.models.institute import Institute
from apps.institutes.serializers import InstituteSerializer
from apps.shared.services import ConfigurationService, FormdefinitionService, PDFService, SharedService, UploadTypeService
from apps.shared.services_shared.store_api_result import store_long_running_process
from apps.students.models.studentDetail import StudentParentMapping

from apps.exams.models.exam import Exam, ExamTerm
from apps.classes.models.subject import SubjectDetails, CumulativeType, StandardSectionMapping, SubjectPartType, SubjectCourseOutcomeMapping, SubjectProgramOutcomeMapping, SubjectProgramSpecificOutcomeMapping ,SubjectCourseOutcomeProgramMappingMatrix, SubjectCourseOutcomeProgramSpecificOutcomeMappingMatrix
from apps.classes.models.staff_subject import StaffSubjectDetails
from apps.shared.services_shared.common import get_full_name, get_selected_template
from apps.staffs.models.staff_standard import StaffStandardMapping
from apps.students.services.student import get_student_admission_form_details
from apps.tenants.services.middlewares import get_current_db_name
from apps.users.models.user import User
from apps.students.models.student import Student
from apps.students.models.studentDetail import StudentDetails
from apps.shared.serializers import DocumentSerializer
from apps.classes.serializers import SubjectDetailsReadSerializer
from apps.shared.models import Document
from apps.exams.models import ExamSchedule, StudentMark, StudentMarkSectionWiseApproval, StudentExamFinalResult 
from apps.classes.models import Enrollment, SubjectStudent, Standard, Subject
from apps.exams.models.result import ResultConfigurationMerge
from apps.exams.models.final_result import FinalResultConfigurationMerge,FinalResultMarksConfiguration
from apps.shared.services import ApprovalService, SharedService
from apps.students.serializers import StudentDocumentMappingSerializer
from apps.exams.serializers import (ExamScheduleCumulativeMappingReadSerializer, GradeSerializer, StudentCumulativeMarkSerializer, StudentMarkReadSerializer, ExamScheduleReadSerilaizer,
                                    StudentMarkSectionWiseApprovalSerializer, StudentExamFinalResultSerializer, StudentMarkSerializer)
from apps.classes.services.subject import check_subject_assigned_to_student, read_subject_course_outcome_program_mapping_matrix
from operator import itemgetter
from itertools import groupby
from django.db import transaction

from django.db.models import F, Count, Value as V
from django.db.models.functions import Concat
from apps.bdu.services.write_to_excel import write_to_excel_new_consolidation,write_to_excel_multiple_tabs,write_to_excel_multiple_tabs_amrita
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import os
from openpyxl.styles import Alignment,PatternFill
from apps.shared.models.approval import ApproveStatus
from apps.hr.models.staffTeachingHour import StaffHourSubjectMapping
import math
from apps.notification.services.notification_service import (send_notification,get_users_for_notification_list,get_user_contact_details, 
                                                             handle_email_notification_bulk,handle_sms_notification_bulk,handle_push_notification_bulk,
                                                             add_to_notification_users,post_to_notification
                                                             )
from apps.notification.models.notification import NotificationMedium,BulkNotification
from apps.students.models.studentDetail import StudentAddress

def add_update_mark(self, return_data):
    response = {'Result': True, 'Reason': 'Data Added Successfully'}
    deletable_list = return_data['deletable_list']
    standard_section_id = return_data['standard_section']
    data = return_data['mark_details']
    schedule_ids = set()
    exam_ids = set()
    grade_plan_ids = set()
    student_subject_data = []
    student_unique_check = []
    student_cumulative_mark = []
    deletable_cumulative_mark_ids = []
    student_mark_dict = {}
    deletabe_data = StudentMark.objects.filter(id__in=deletable_list)
    deletable_student_mark_data = list(
        deletabe_data.values_list('exam_schedule__exam', flat=True))
    for mark_data in data:
        for subject_data in mark_data['subject_list']:
            schedule_ids.add(subject_data['schedule'])
        if mark_data['student'] in student_unique_check:
            raise ValidationError('Duplicate Student Data Found')
        student_unique_check.append(mark_data['student'])
    student_mark_data = StudentMark.objects.filter(exam_schedule__in=list(schedule_ids)).values()
    for student in student_mark_data:
        if student['student_id'] not in student_mark_dict:
            student_mark_dict[student['student_id']]={}
        if student['id'] not in student_mark_dict[student['student_id']]:
            student_mark_dict[student['student_id']][student['id']]=student
    exam_schedule_details = ExamSchedule.objects.filter(id__in=list(schedule_ids)).values('id', 'exam', 'subject',
                                                                                        'start_time',
                                                                                       'exam__academic_year',
                                                                                       'max_marks', 'min_marks',
                                                                                       'fordate', 'standard_section', 
                                                                                       'is_marks', 'grade_plan')
    schedule_cummulative_mapping = {}
    cumulative_data = ExamScheduleCumulativeMapping.objects.filter(exam_schedule__in=list(schedule_ids)).values()
    for cumulative in cumulative_data:
        if cumulative['exam_schedule_id'] not in schedule_cummulative_mapping:
            schedule_cummulative_mapping[cumulative['exam_schedule_id']] = {}
        schedule_cummulative_mapping[cumulative['exam_schedule_id']][cumulative['id']] = cumulative
    exam_schedule_id_mapping = {}
    nowdate = datetime.now().date()
    nowtime = datetime.now().time()
    for exam_obj in exam_schedule_details:
        if exam_obj['grade_plan']:
            grade_plan_ids.add(exam_obj['grade_plan'])
        exam_ids.add(exam_obj['exam'])
        # if nowdate <= exam_obj['fordate']:
        #     if nowdate != exam_obj['fordate'] or nowtime <= exam_obj['start_time']:
        #         raise ValidationError('Exam not yet started/ended')
    grade_plan_data = Grade.objects.filter(
        grade_plan__grade_type=2, grade_plan__in=grade_plan_ids
    ).values('name', 'grade_plan')
    grade_plan_mapping = {}
    for grade_plan in grade_plan_data:
        if grade_plan['grade_plan'] not in grade_plan_mapping:
            grade_plan_mapping[grade_plan['grade_plan']] = []
        grade_plan_mapping[grade_plan['grade_plan']].append(grade_plan['name'])
    exam_ids = set(list(exam_ids) + deletable_student_mark_data)
    if len(exam_ids) > 1:
        raise ValidationError('You can schedule only one exam at a time')
    if len(exam_ids) <= 0:
        response['Reason'] = 'No changes to update'
        return response
    exam_obj_d = Exam.objects.get(id=list(exam_ids)[0])
    ApprovalService.get_approval_status(self, exam_obj_d)
    for schedule in exam_schedule_details:
        exam_ids.add(schedule['exam'])
        exam_schedule_id_mapping[schedule['id']] = schedule
    student_data = Student.get_student_for_standard(exam_obj_d.academic_year, \
                                                       [], [standard_section_id],
                                                       ['first_name', 'middle_name', 'last_name', 'id'])
    approved_standard_mark_list = get_approved_standard_section_list(exam_obj_d.id)
    if standard_section_id in approved_standard_mark_list:
        raise ValidationError('The Student marks is finalized you cant change.')
    student_data = {data['id']: data for data in student_data}
    data_to_save = []
    staff_id = User.get_my_staff_id(self, False)
    for mark_data in data:
        subject_list = []
        if mark_data['subject_list']:
            for subject_data in mark_data['subject_list']:
                schedule_id = subject_data['schedule']
                if schedule_id not in exam_schedule_id_mapping:
                    raise ValidationError('Invalid exam schedule id')
                is_marks = True
                available_grades = []
                if not exam_schedule_id_mapping[schedule_id]['is_marks']:
                    is_marks = False
                    selected_grade_plan = exam_schedule_id_mapping[schedule_id]['grade_plan']
                    available_grades = grade_plan_mapping[selected_grade_plan]
                if is_marks and not exam_schedule_id_mapping[schedule_id]['max_marks']:
                    raise ValidationError("Please update max_marks before entering marks.")
                # if is_marks and subject_data['attendance_status'] != "Absent" and (subject_data['marks'] == None or ('grade' in subject_data and subject_data['grade'])):
                #     raise ValidationError("Subject is not Grade only")
                if 'id' in subject_data and (mark_data['student'] in student_mark_dict and subject_data['id'] in student_mark_dict[mark_data['student']] and student_mark_dict[mark_data['student']][subject_data['id']]['marked_attendance_days']) and not mark_data['marked_attendance_days']:
                    raise ValidationError("Attendance is already updated can't be zero")
                if 'id' in subject_data and (mark_data['student'] in student_mark_dict and subject_data['id'] in student_mark_dict[mark_data['student']] and student_mark_dict[mark_data['student']][subject_data['id']]['remark_id']) and ('remark' not in mark_data or not mark_data['remark']):
                    raise ValidationError("Remark is already updated can't be NUll")
                if subject_data['attendance_status'] == "Absent":
                    subject_data['marks'] = None
                elif is_marks and subject_data['marks'] and float(subject_data['marks']) > exam_schedule_id_mapping[schedule_id]['max_marks']:
                    raise ValidationError(
                        f'Please enter marks lesser than {exam_schedule_id_mapping[schedule_id]["max_marks"]}')
                elif not is_marks and ('grade' not in subject_data or not subject_data['grade']):
                    raise ValidationError(f'Grade is mandatory')
                elif not is_marks and subject_data['grade'] not in available_grades:
                    raise ValidationError(f'Available Grades are {available_grades}')
                elif not is_marks and subject_data['marks']:
                    raise ValidationError('Marks should be null for the grade wise subject')
                if mark_data['student'] not in student_data:
                    raise ValidationError('Student not exist in given standard')
                subject_list.append(exam_schedule_id_mapping[subject_data['schedule']]['subject'])
                if (
                    (is_marks and subject_data['attendance_status'] == 'Present' and subject_data['marks'] is not None)
                    or (subject_data['attendance_status'] == "Absent")
                    or (not is_marks)
                ):
                    temp = {'marks': subject_data['marks'], 'exam_schedule': subject_data['schedule'],
                            'student': mark_data['student'], 'staff': staff_id,
                            'attendance_status': subject_data['attendance_status'],
                            'grade': subject_data['grade'] if 'grade' in subject_data else '',
                            'is_active': True,
                            'marked_attendance_days': mark_data['marked_attendance_days'] if 'marked_attendance_days' in mark_data else 0, #support only for global for now
                            'remark': mark_data['remark'] if 'remark' in mark_data else None
                            }
                    if 'id' in subject_data:
                        temp['id'] = subject_data['id']
                    data_to_save.append(temp)
                duplicate_cummulative_mapping = {}
                if 'cumulative_marks' in subject_data:
                    for cum_data in subject_data['cumulative_marks']:
                        attendance_status = 'Present'
                        if cum_data['examschedulecumulativemapping'] in duplicate_cummulative_mapping:
                            raise ValidationError('Duplicate cummulative mapping data')
                        duplicate_cummulative_mapping[cum_data['examschedulecumulativemapping']] = ''
                        if cum_data['examschedulecumulativemapping'] not in schedule_cummulative_mapping[subject_data['schedule']]:
                            raise ValidationError(f'Invalid cumulative type {cum_data["examschedulecumulativemapping"]}')
                        if 'attendance_status' in cum_data and cum_data['attendance_status'] == 'Absent':
                            attendance_status = 'Absent'
                        if attendance_status == 'Present' and ( not schedule_cummulative_mapping[subject_data['schedule']][cum_data['examschedulecumulativemapping']] or 
                            schedule_cummulative_mapping[subject_data['schedule']][cum_data['examschedulecumulativemapping']]['max_marks'] < float(cum_data['marks'])
                            ):
                            raise ValidationError(f'Cummaltive - {cum_data["examschedulecumulativemapping"]} marks is greater than given marks {schedule_cummulative_mapping[subjectData["schedule"]][cum_data["examschedulecumulativemapping"]]["max_marks"] }')
                        temp = {
                            'exam_cumulative': cum_data['examschedulecumulativemapping'],
                            'marks': cum_data['marks'],
                            'student': mark_data['student'],
                            'staff': staff_id,
                            'attendance_status': attendance_status,
                            'is_active': True
                        }
                        if 'id' in cum_data and cum_data['id']:
                            temp['id'] = cum_data['id']
                        student_cumulative_mark.append(temp)
                if 'deletable_cumulative_mark_ids' in subject_data and subject_data['deletable_cumulative_mark_ids']:
                    deletable_cumulative_mark_ids += subject_data['deletable_cumulative_mark_ids']
            temp_validate_data = {'student': mark_data['student'], 'subject_list': subject_list}
            student_subject_data.append(temp_validate_data)
    check_subject_assigned_to_student(student_subject_data)
    with transaction.atomic(using=get_current_db_name()):
        if len(deletable_list) > 0:
            deletabe_data.update(is_active=False)
        if data_to_save:
            existing_marks = StudentMark.objects.filter(
                id__in=[item['id'] for item in data_to_save if 'id' in item]
            )
            existing_marks_map = {obj.id: obj for obj in existing_marks}
            for temp_data in data_to_save:
                if 'id' in temp_data:
                    old_instance = existing_marks_map.get(temp_data['id'])
                    if old_instance:
                        StudentMarkAuditLog.log_student_mark_changes(
                            old_instance=old_instance,
                            new_data=temp_data,
                            user=self.request.user,
                        )
            response = SharedService.add_or_update_data(self, data_to_save)
        if deletable_cumulative_mark_ids:
            StudentCumulativeMark.objects.filter(id__in=deletable_cumulative_mark_ids).delete()
        if student_cumulative_mark:
            add_or_update_cum_data(self, student_cumulative_mark)
    return response


def add_or_update_cum_data(self, data):
    for cum_data in data:
        existing = StudentCumulativeMark.objects.filter(
            exam_cumulative=cum_data['exam_cumulative'],
            student=cum_data['student'],
            is_active=True
        ).first()
        if existing:
            cum_data['id'] = existing.id 
        if 'id' in cum_data and cum_data['id']:
            serializer = StudentCumulativeMarkSerializer(instance=StudentCumulativeMark.objects.get(id=cum_data['id']), data=cum_data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
        else:
            serializer = StudentCumulativeMarkSerializer(data=cum_data)
            serializer.is_valid(raise_exception=True)
            serializer.save()

def add_update_grade(self, data):
    exclude_query = {}
    excludeIds = []
    if 'deletable_grade_ids' in data and len(data['deletable_grade_ids']):
        excludeIds += data['deletable_grade_ids']
    if 'id' in data and data['id']:
        excludeIds.append(data['id'])
        if GradeExamScheduleMapping.objects.filter(grade_plan=data['id']).values():
            raise ValidationError('Grade plan is referred cant edit the data')
    if excludeIds:
        exclude_query = {'grade_plan__in': excludeIds}
    existing_data = Grade.objects.filter(grade_plan__name=data['plan_name']).exclude(**exclude_query).values()
    data['range_list'] += existing_data
    SharedService.duplicate_list_one_object(data['range_list'], 'name')
    if data['grade_type'] != 2:
        SharedService.checkduplicate_range_exist(self, data['range_list'], 'from_range', 'to_range', 'name')
    for grade in data['range_list']:
        if data['grade_type'] != 2:
            if grade['from_range'] < 0:
                raise ValidationError('Grade from range should be greater than 0')
            if grade['to_range'] < 0:
                raise ValidationError('Grade to range should be greater than 0')
            if grade['from_range'] > 100:
                raise ValidationError('Grade from range should be lesser than 100')
            if grade['to_range'] > 100:
                raise ValidationError('Grade to range should be lesser than 100')
    with transaction.atomic(using=get_current_db_name()):
        if 'deletable_grade_ids' in data and len(data['deletable_grade_ids']):
            Grade.objects.filter(id__in=data['deletable_grade_ids']).delete()
        data_to_save = {
            'name': data['plan_name'],
            'is_default': True if 'is_default' in data and data['is_default'] else False,
            'grade_type': data['grade_type'] if 'grade_type' in data else True
        }
        if data_to_save['is_default']:
            self.get_queryset().all().update(is_default=False)
        if 'id' in data and data['id']:
            data_to_save['id'] = data['id']
        response = SharedService.add_or_update_data(self, [data_to_save])
        for row_data in data['range_list']:
            row_data['grade_plan'] = response['data']['id']
            if 'id' in row_data and row_data['id']:
                serializer = GradeSerializer(instance=Grade.objects.get(id=row_data['id']), data=row_data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            else:
                serializer = GradeSerializer(data=row_data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
        return {'Reason': 'Data added successfully'}


""" Standard section and subject we get """

def generate_chart(labels, values, title, chart_type="bar", colors=None):
    """Generate a high-quality chart and return base64 encoded image."""
    plt.figure(figsize=(6, 4), dpi=300)  # Reduced box size for clarity
    if not colors:
        if chart_type == "bar":
            colors = ['#1f77b4', '#6baed6', '#9ecae1', '#c6dbef']  # Shades of blue
        else:
            colors = plt.cm.get_cmap("tab10", len(labels)).colors

    if chart_type == "pie":
        wedges, texts, autotexts = plt.pie(
            values, labels=labels, autopct='%1.1f%%', startangle=140,
            colors=colors, wedgeprops={'edgecolor': 'black'}
        )
        for text in texts + autotexts:
            text.set_fontsize(10)
    elif chart_type == "line":
        line_types = ["normal", "with_marker", "dashed"]
        selected_type = line_types[len(labels) % 3]  # Rotate between types
        plt.plot(labels, values, color=colors[0] if colors is not None and len(colors) > 0 else None, marker='o', linewidth=2)
        plt.xlabel("Subject", fontsize=10)
        plt.ylabel("Marks", fontsize=10)
        plt.xticks(rotation=-45, ha='left', fontsize=8)
        plt.yticks(fontsize=8)
        plt.yticks([0, 10, 20, 30, 40, 50])
        
    else:
        # bar_types = ["vertical", "horizontal", "stacked"]
        # selected_type = bar_types[len(labels) % 3]  # Rotate between types

        # if selected_type == "vertical":
        plt.bar(labels, values, color=colors, edgecolor='black')
        # elif selected_type == "horizontal":
        #     plt.barh(labels, values, color=colors, edgecolor='black')
        # elif selected_type == "stacked":
        #     plt.bar(labels, values, color=colors, edgecolor='black', width=0.6)

        plt.xlabel("Subject", fontsize=10)
        plt.ylabel("Marks", fontsize=10)
        plt.xticks(rotation=-45, ha='left', fontsize=8)
        plt.yticks(fontsize=8)
        plt.yticks([0, 20, 40, 60, 80, 100])

    plt.title(title, fontsize=12, fontweight='bold')

    # Convert plot to Base64
    img_stream = BytesIO()
    plt.savefig(img_stream, format='png', bbox_inches='tight', dpi=300, transparent=True)
    plt.close()
    img_stream.seek(0)
    return base64.b64encode(img_stream.getvalue()).decode('utf-8')

def get_standard_section_subjects(self, examId, standardSectionId, raiseErrorIfNotFinalized=False, student_ids=[], ignore_final_result_data=False):
    response = {'data': {}}
    required_form_definition = [
        {'form_name': 'exam_configurations', 'column_name': 'grade_plan'}
    ]
    temp_form_defintion = FormdefinitionService.get_formdefinition_for_multiple_data(self, required_form_definition)
    form_definition_tracking = {'exam_configurations_grade_plan': temp_form_defintion['exam_configurations']['grade_plan']}

    approval_data = StudentMarkSectionWiseApproval.objects.filter(exam=examId,
                                                                standard_section=standardSectionId).values()
    institute_objs = Institute.get_institute(self)
    if self.request.GET.get('exam_result'):
        if not approval_data or approval_data[0]['approval_status'] != 1:
            raise ValidationError('Student marks are not approved for the section')
    exam_obj = Exam.objects.get(id=examId)
    standard_sec_obj = StandardSectionMapping.objects.get(id=standardSectionId)
    values_list = ['id', 'first_name', 'middle_name','last_name', 'current_reg_num', 'student_name','dob','gender']
    custom_annotate = {'student_name': Concat('first_name', V(' '), 'middle_name',V(' '), 'last_name')}
    if student_ids:
        student_data = Student.objects.filter(
            id__in=student_ids
        ).annotate(**custom_annotate).values(*values_list)
    else:
        student_data = Student.get_student_for_standard(None, None, [standardSectionId], values_list, custom_annotate)
    student_ids = []
    studentIdDict = {}
    part_type_list = SubjectPartType.objects.all().values()
    part_types = [{'name': s['name']} for s in part_type_list]
    for i in student_data:
        student_ids.append(i['id'])
        i['student'] = i['id']
        studentIdDict[i['id']] = i
    filterQuery = {'student__in': student_ids,
                'exam_schedule__exam': examId}
    scheduleQuery = {'exam': examId}
    scheduleQuery['standard_section'] = standardSectionId
    scheduleQuery['sub_schedule_parent'] = None
    scheduleData = ExamSchedule.objects.filter(**scheduleQuery).values('id', 'subject', 'subject__name', 'min_marks',
                                                                    'max_marks', 'subject__subject_part_type__name',
                                                                        'subject__subject_part_type',
                                                                        'subject__subject_part_type__code_name',
                                                                        'subject__subject_code', 'is_marks', 'grade_plan',
                                                                        'grade_plan__name','subject__codename', 'schedule_sequence','exam__from_date','exam__to_date'
                                                                    )
    student_admission_form = get_student_admission_form_details(self, student_ids)
    scheduleIds = []
    subjectList = {}
    scheduleSubjectIds = []
    grade_plan_ids = set()
    for schedule in scheduleData:
        scheduleIds.append(schedule['id'])
        temp = {'subject': schedule['subject'], 'subject_name': schedule['subject__name'], 'subject_part_type': schedule['subject__subject_part_type__name'],
                'min_marks': schedule['min_marks'], 'max_marks': schedule['max_marks'], 'schedule': schedule['id'],
                'subject_part_type_id': schedule['subject__subject_part_type'],
                'subject_part_type_code_name': schedule['subject__subject_part_type__code_name'],
                'subject__subject_code': schedule['subject__subject_code'], 'is_marks': schedule['is_marks'],
                'grade_plan': schedule['grade_plan'], 'grade_plan_name': schedule['grade_plan__name'],'subject__codename':schedule['subject__codename'],
                'exam_from_date': schedule['exam__from_date'],'exam_to_date': schedule['exam__to_date'],
                }
        grade_plan_ids.add(schedule['grade_plan'])
        subjectList[schedule['subject']] = temp
        scheduleSubjectIds.append(schedule['subject'])
    studentSubjectData = SubjectStudent.objects.filter(student__in=student_ids,
                                                    academic_year=exam_obj.academic_year,
                                                    subject__in=scheduleSubjectIds).values('subject', 'student',
                                                                                            'subject__name', 'subject__subject_code','subject__codename',
                                                                                            subject_name=F('subject__name'),
                                                                                            subject_part_type=F('subject__subject_part_type__name'),
                                                                                            subject_part_type_id=F('subject__subject_part_type'),
                                                                                            subject_part_type_code_name=F('subject__subject_part_type__code_name')
                                                                                            )
    studentSubjectMapping = {}
    for studentSubject in studentSubjectData:
        if studentSubject['student'] in studentSubjectMapping:
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
        else:
            studentSubjectMapping[studentSubject['student']] = []
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
    filterQuery['exam_schedule__in'] = scheduleIds
    filterQuery['is_active'] = True
    student_mark_data = []
    student_list=[]
    temp_student_data = StudentMark.objects.filter(**filterQuery).values(
        'student__first_name', 'student__middle_name', 'student__last_name',
        'student__sts', 'student__current_reg_num', 'exam_schedule_id', 'marks',
        'student', 'staff', 'attendance_status', 'id', 'exam_schedule__is_marks',
        'exam_schedule__grade_plan', 'grade','student__mobile_num', 'student__dob',
        'student__gender',
        'marked_attendance_is_global','remark', 'student__profile_pic',
        'remark_is_global', 'marked_attendance_days', 'remark', 'remark__name'
    )
    exam_schedule_ids = []
    profile_photo_dict = {}
    for student_row in temp_student_data:
        exam_schedule_ids.append(student_row['exam_schedule_id'])
        student_list.append(student_row['student'])
        if student_row['student__profile_pic'] and student_row['student__profile_pic'] not in profile_photo_dict:
            profile_photo_dict[student_row['student__profile_pic']] = {
                'student_id' : student_row['student'],
                'student_profile_pic' : student_row['student__profile_pic']
            }
        student_row['first_name'] = student_row['student__first_name']
        student_row['dob'] = student_row['student__dob']
        student_row['gender'] = student_row['student__gender']
        student_row['middle_name'] = student_row['student__middle_name']
        student_row['last_name'] = student_row['student__last_name']
        student_row['mobile_num'] = student_row['student__mobile_num']
        student_row['is_marks'] = student_row['exam_schedule__is_marks']
        student_row['grade_plan'] = student_row['exam_schedule__grade_plan']
        student_row['full_name'] = get_full_name(
            student_row['student__first_name'],
            student_row['student__middle_name'],
            student_row['student__last_name']
        )
        student_row['sts'] = student_row['student__sts']
        student_row['current_reg_num'] = student_row['student__current_reg_num']
    profile_photo_list = list(profile_photo_dict.keys())
    temp_document_data = Document.objects.filter(id__in=profile_photo_list)
    document_serializer = DocumentSerializer(temp_document_data,many=True)        
    temp_parent_data = StudentParentMapping.objects.filter(student__in=student_list).values(
        'parent__father_name','parent__mother_name','student'
    )
    temp_student_detail = StudentDetails.objects.filter(student__in=student_list).values(
        'caste__name','student'
    )
    for student_row in temp_student_data:
        for parent in temp_parent_data:
            if parent['student'] == student_row['student']:
                student_row['father_name'] = parent['parent__father_name']
                student_row['mother_name'] = parent['parent__mother_name']
        for document in document_serializer.data:
            if student_row['student__profile_pic'] == document['id']:
                student_row['student_profile_pic_file'] = document['file']
        for student_detail in temp_student_detail:
            if student_detail['student'] == student_row['student']:
                student_row['caste_name'] = student_detail['caste__name']
        student_mark_data.append(student_row)
    temp_exam_schedule = ExamSchedule.objects.filter(id__in=exam_schedule_ids).values(
        'subject__name', 'subject__subject_code', 'subject__subject_part_type__name',
        'subject__subject_part_type__code_name', 'subject__subject_part_type__id',
        'subject', 'min_marks', 'max_marks', 'id', 'is_marks', 'grade_plan', 'subject__codename'
    )
    exam_schedule_data = {}
    for temp_schedule in temp_exam_schedule:
        temp_schedule['subject_name'] = temp_schedule['subject__name']
        temp_schedule['subject_code'] = temp_schedule['subject__subject_code']
        temp_schedule['codename'] = temp_schedule['subject__codename']
        temp_schedule['subject_part_type'] = temp_schedule['subject__subject_part_type__name']
        temp_schedule['subject_part_type_code_name'] = temp_schedule['subject__subject_part_type__code_name']
        temp_schedule['subject_part_type_id'] = temp_schedule['subject__subject_part_type__id']
        exam_schedule_data[temp_schedule['id']] = temp_schedule
        
    # serializer = StudentMarkReadSerializer(student_data, many=True)
    studentBasedData = {}
    enteredMarksStudentIds = []
    filter_query_cumulative = {
        'student__in': student_ids,
        'exam_cumulative__exam_schedule__in': scheduleIds,
        'is_active': True
    }
    student_cumulative_data = StudentCumulativeMark.objects.filter(
        **filter_query_cumulative
    ).values(
        'id', 'marks', 'exam_cumulative_id', 'exam_cumulative__exam_schedule', 'student', 'attendance_status', 'exam_cumulative__cumulative_type',
        'exam_cumulative__cumulative_type__name', 'exam_cumulative__max_marks','exam_cumulative__min_marks'
    )
    temp_schedule_cumulative_mapping = {}
    schedule_cumulative_mapping = {}
    schedule_cumulative_data = ExamScheduleCumulativeMapping.objects.filter(
        exam_schedule__in=scheduleIds,
    ).values(
        'exam_schedule', 'cumulative_type', 'max_marks', 'min_marks', 'cumulative_type__name',
        'cumulative_type__alias', 'id'
    )
    for schedule_cumulative in schedule_cumulative_data:
        if schedule_cumulative['id'] not in schedule_cumulative_data:
            temp_schedule_cumulative_mapping[schedule_cumulative['id']] = {
                'exam_schedule' : schedule_cumulative['exam_schedule'],
                'max_marks': schedule_cumulative['max_marks'],
                'min_marks': schedule_cumulative['min_marks'],
                'cumulative_type_data': [],
                'id': schedule_cumulative['id']
            }
        temp_schedule_cumulative_mapping[schedule_cumulative['id']]['cumulative_type_data'].append(
            {
                'id': schedule_cumulative['cumulative_type'],
                'name': schedule_cumulative['cumulative_type__name'],
                'alias': schedule_cumulative['cumulative_type__alias']
            }
        )
    for cum_row in temp_schedule_cumulative_mapping.values():
        if cum_row['exam_schedule'] not in schedule_cumulative_mapping:
            schedule_cumulative_mapping[cum_row['exam_schedule']] = []
        schedule_cumulative_mapping[cum_row['exam_schedule']].append(cum_row)
    student_cumulative_data_temp_mapping = {}
    for student_cum in student_cumulative_data: #this is to remove the duplicates
        if student_cum['id'] not in student_cumulative_data_temp_mapping:
            student_cumulative_data_temp_mapping[student_cum['id']] = {
                'cumulative_data_mapping': [], 'id' : student_cum['id'],
                'marks': student_cum['marks'], 'exam_cumulative_id': student_cum['exam_cumulative_id'],
                'exam_cumulative__exam_schedule': student_cum['exam_cumulative__exam_schedule'], 'student': student_cum['student'],
                'attendance_status': student_cum['attendance_status'], 'exam_cumulative__max_marks': student_cum['exam_cumulative__max_marks'],
                'exam_cumulative__min_marks': student_cum['exam_cumulative__min_marks']
            }
        student_cumulative_data_temp_mapping[student_cum['id']]['cumulative_data_mapping'].append({
            'cumulative_type_id': student_cum['exam_cumulative__cumulative_type'],
            'cumulative_type_name': student_cum['exam_cumulative__cumulative_type__name'],
        })
    student_cumulative_data_mapping = {}
    for cum in student_cumulative_data_temp_mapping.values():
        if cum['exam_cumulative__exam_schedule'] not in student_cumulative_data_mapping:
            student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']] = {cum['student']: []}
        elif cum['student'] not in student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']]:
            student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']] = []
        if cum['exam_cumulative__exam_schedule'] not in schedule_cumulative_mapping:
            schedule_cumulative_mapping[cum['exam_cumulative__exam_schedule']] = {}
        student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']].append(cum)
    for subject_id, schedule_row in subjectList.items():
        subjectList[subject_id]['total_max_marks'] = subjectList[subject_id]['max_marks'] if subjectList[subject_id]['max_marks'] else 0
        subjectList[subject_id]['total_min_marks'] = subjectList[subject_id]['min_marks'] if subjectList[subject_id]['min_marks'] else 0
        if subjectList[subject_id]['schedule'] in schedule_cumulative_mapping:
            subjectList[subject_id]['cumulative_data'] = schedule_cumulative_mapping[subjectList[subject_id]['schedule']]
            for temp_row in schedule_cumulative_mapping[subjectList[subject_id]['schedule']]:
                subjectList[subject_id]['total_max_marks'] += temp_row['max_marks'] if temp_row['max_marks'] else 0
                subjectList[subject_id]['total_min_marks'] += temp_row['min_marks'] if temp_row['min_marks'] else 0
    # entered student marks data
    for markData in student_mark_data:
        markData['exam_schedule'] = exam_schedule_data[markData['exam_schedule_id']]
        obtained_marks = 0
        total_marks = subjectList[markData['exam_schedule']['subject']]['total_max_marks']
        total_min_marks = subjectList[markData['exam_schedule']['subject']]['total_min_marks']
        #whatever key is added please add in unentered marks section also
        temp = {'id': markData['id'], 'marks': markData['marks'], 'subject': markData['exam_schedule']['subject'],
                'subject_name': markData['exam_schedule']['subject_name'],
                'subject__subject_code': markData['exam_schedule']['subject_code'],
                'subject__codename':markData['exam_schedule']['codename'],
                'subject_part_type': markData['exam_schedule']['subject_part_type'],
                'subject_part_type_code_name': markData['exam_schedule']['subject_part_type_code_name'],
                'subject_part_type_id': markData['exam_schedule']['subject_part_type_id'],
                'attendance_status': markData['attendance_status'], 'min_marks': markData['exam_schedule']['min_marks'],
                'max_marks': markData['exam_schedule']['max_marks'], 'result': 'pass', 'total_max_marks': total_marks,
                'total_min_marks': total_min_marks, 'cumulative_marks_data': [],
                'cumulative_data': [], 'is_marks': markData['is_marks'], 'grade' : markData['grade'],'grade_plan' :markData['exam_schedule']['grade_plan']
                }
        if markData['exam_schedule']['id'] in student_cumulative_data_mapping and markData['student'] in student_cumulative_data_mapping[markData['exam_schedule']['id']]:
            temp['cumulative_marks_data'] = student_cumulative_data_mapping[markData['exam_schedule']['id']][markData['student']]
            for cum_data in temp['cumulative_marks_data']:
                obtained_marks += cum_data['marks'] if cum_data['marks'] else 0
        if temp['attendance_status'] == 'Absent' or (temp['min_marks'] and not temp['marks']) or (temp['marks'] and temp['marks'] < temp['min_marks']):
            temp['result'] = 'fail'
        if markData['student'] in studentBasedData:
            studentBasedData[markData['student']]['obtained_marks'] += obtained_marks
            if temp['result'] == 'fail':
                studentBasedData[markData['student']]['total_result'] = 'fail'
            if temp['attendance_status'] == 'Absent':
                studentBasedData[markData['student']]['total_result'] = 'fail'
            if temp['attendance_status'] == 'Present':
                studentBasedData[markData['student']]['total_marks'] += total_marks
                studentBasedData[markData['student']]['obtained_marks'] += temp['marks'] if temp['marks'] else 0
            temp['total_marks'] = total_marks
            studentBasedData[markData['student']]['subject_list'][temp['subject']] = temp
        else:
            obtained_marks += temp['marks'] if temp['marks'] else 0
            mainTemp = {'student_name': markData['full_name'], 'student': markData['student'],
                        'sts': markData['sts'], 'father_name':markData['father_name'],'mother_name':markData['mother_name'],
                        'caste_name':markData['caste_name'],
                        'mobile_num':markData['mobile_num'],'current_reg_num':markData['current_reg_num'],
                        'first_name': markData['first_name'], 'middle_name': markData['middle_name'],
                        'last_name': markData['last_name'],
                        'total_marks': total_marks,'dob':markData['dob'],'gender':markData['gender'],
                        'obtained_marks': obtained_marks, 'total_result': temp['result'],
                        'subject_list': {temp['subject']: temp}, 'part_type_list': part_types,
                        'marked_attendance_days': markData['marked_attendance_days'],
                        'remark': markData['remark'],
                        'remark_name': markData['remark__name']
                        }
            mainTemp['profile_pic_file'] = markData['student_profile_pic_file'] if 'student_profile_pic_file' in markData and markData['student_profile_pic_file'] else None
            studentBasedData[markData['student']] = mainTemp
            enteredMarksStudentIds.append(markData['student'])
    studentList = []
    finalResultData = {}
    isAnnounced = False
    finalResultData = StudentExamFinalResult.objects.filter(student__in=enteredMarksStudentIds, exam=examId).values(
        'student', 'status', 'is_announced','section_rank','standard_rank')
    finalResultData = {temp['student']: temp for temp in finalResultData}
    # enterd marks students list -> checking for students subject unentered marks of the student
    for studentD in studentBasedData:
        if studentD not in finalResultData and raiseErrorIfNotFinalized:
            raise ValidationError('Student marks are not yet finalized')
        if studentD in studentSubjectMapping:
            if 'subject_list' not in studentBasedData[studentD]:
                studentBasedData[studentD]['subject_list'] = {}
            studentBasedData[studentD]['obtained_marks'] = 0
            for subData in studentSubjectMapping[studentD]:
                if subData['subject'] not in studentBasedData[studentD]['subject_list']:
                    if 'total_marks' not in studentBasedData[studentD]:
                        studentBasedData[studentD]['total_marks'] = 0
                    if subjectList[subData['subject']]['is_marks']:
                        studentBasedData[studentD]['total_marks'] += subjectList[subData['subject']]['max_marks'] if subjectList[subData['subject']]['max_marks'] else 0
                    if subData['subject'] in subjectList and 'cumulative_data' in subjectList[subData['subject']]:
                        subData['cumulative_data'] = subjectList[subData['subject']]['cumulative_data']
                    studentBasedData[studentD]['subject_list'][subData['subject']] = subData
                    if 'cumulative_data' in subData:
                        for cum_config_data in subData['cumulative_data']:
                            if cum_config_data['exam_schedule'] in student_cumulative_data_mapping and studentD in student_cumulative_data_mapping[cum_config_data['exam_schedule']]:
                                if 'cumulative_marks_data' not in subData:
                                    studentBasedData[studentD]['subject_list'][subData['subject']]['cumulative_marks_data'] = []
                                subData['cumulative_marks_data'] += student_cumulative_data_mapping[cum_config_data['exam_schedule']][studentD]
                                for cum_data in temp['cumulative_marks_data']:
                                    studentBasedData[studentD]['obtained_marks'] += cum_data['marks'] if cum_data['marks'] else 0
        if not ignore_final_result_data and studentD in finalResultData: #ignore fetchin from final result data when we are syncing the result again
            studentBasedData[studentD]['total_result'] = finalResultData[studentD]['status']
            studentBasedData[studentD]['section_rank'] = finalResultData[studentD]['section_rank']
            studentBasedData[studentD]['standard_rank'] = finalResultData[studentD]['standard_rank']
            if finalResultData[studentD]['is_announced']:
                studentBasedData[studentD]['is_announced'] = True
                isAnnounced = True
        studentList.append((studentBasedData[studentD]))
        enteredMarksStudentIds.append(studentD)
    # unentered subject data for student
    for studentId in studentIdDict:
        #if finalized ignoring unentered marks
        if studentId in finalResultData:
            continue
        if studentId not in studentBasedData and studentId in studentSubjectMapping:
            if raiseErrorIfNotFinalized:  # raise error if unentered student list found
                student_obj = Student.objects.get(id=studentId)
                name = get_full_name(student_obj.first_name, student_obj.middle_name, student_obj.last_name)
                raise ValidationError(f'{name} marks are not yet entered for the section')
            if 'subject_list' not in studentIdDict[studentId]:
                studentIdDict[studentId]['subject_list'] = {}
            for subData in studentSubjectMapping[studentId]:
                if subData['subject'] not in studentIdDict[studentId]['subject_list']:
                    if 'total_marks' not in studentIdDict[studentId]:
                        studentIdDict[studentId]['total_marks'] = 0
                    if subjectList[subData['subject']]['is_marks']:
                        studentIdDict[studentId]['total_marks'] += subjectList[subData['subject']]['max_marks'] if subjectList[subData['subject']]['max_marks'] else 0
                    if subData['subject'] in subjectList and 'cumulative_data' in subjectList[subData['subject']]:
                        subData['cumulative_data'] = subjectList[subData['subject']]['cumulative_data']
                    studentIdDict[studentId]['subject_list'][subData['subject']] = subData
                if subData['subject'] in subjectList and 'cumulative_data' in subjectList[subData['subject']]:
                    subData['cumulative_data'] = subjectList[subData['subject']]['cumulative_data']
                    if 'cumulative_data' in subData:
                        for cum_config_data in subData['cumulative_data']:
                            if cum_config_data['exam_schedule'] in student_cumulative_data_mapping and studentId in student_cumulative_data_mapping[cum_config_data['exam_schedule']]:
                                if 'cumulative_marks_data' not in subData:
                                    subData['cumulative_marks_data'] = []
                                subData['cumulative_marks_data'] += student_cumulative_data_mapping[cum_config_data['exam_schedule']][studentId]
                                # for cum_data in temp['cumulative_marks_data']:
                                #     studentIdDict[studentId]['obtained_marks'] += cum_data['marks'] if cum_data['marks'] else 0
            studentList.append(studentIdDict[studentId])
    grade_data_tracking_indiviual = {'grade_plan_obj': None, 'grade_data': None}
    grade_data_tracking_total = {'grade_plan_obj': None, 'grade_data': None}
    # update obtained marks for each subject
    grade_exam_schedule_mapping = GradeExamScheduleMapping.objects.filter(
            standard_section=standard_sec_obj.id, exam=exam_obj.id
        ).values(
        'max_no_of_days_attendance','grade_plan','grade_plan_for_total','attendance_from_date','attendance_to_date'
    )
    for idx, student_data in enumerate(studentList):
        part_one_subject_list = []
        part_two_subject_list = []
        is_present_in_all_subject = True
        studentList[idx]['part_type_data'] = {}
        studentList[idx]['part_type_data_code_wise'] = {}
        studentList[idx]['is_finalized'] = False
        studentList[idx]['cumulative_id_type_mapping'] = {}
        studentList[idx]['cumulative_id_type_mapping_total'] = []
        studentList[idx]['part_type_sequence'] = {}
        studentList[idx]['subject_names_part_wise']={}
        studentList[idx]['obtained_marks_list_part_wise']={}
        studentList[idx]['obtained_marks_list_part_wise_graph']={}
        studentList[idx]['obtained_grade_list_part_wise']={}
        if student_data['student'] in finalResultData:
            studentList[idx]['is_finalized'] = True
        #student_data['subject_list_data'] = sorted(student_data['subject_list'].values(), key=lambda v: (isinstance(v.get('subject__subject_code', "NA"), str), v.get('subject__subject_code', "NA"))) #used for printing marks card
        student_data['subject_list_data'] = sorted(
        student_data['subject_list'].values(),
            key=lambda v: (
                0, int(v['subject__subject_code'])  # Numeric codes
            ) if v.get('subject__subject_code') and str(v['subject__subject_code']).isdigit() else (
                1, str(v.get('subject__subject_code') or '')  # Non-numeric or None
            )
        )
        for subject_data in student_data['subject_list_data']:
            if subject_data['subject_part_type_code_name'] not in studentList[idx]['part_type_sequence']:
                studentList[idx]['part_type_sequence'][subject_data['subject_part_type_code_name']] = {'sequence': 0}
            studentList[idx]['part_type_sequence'][subject_data['subject_part_type_code_name']]['sequence'] += 1
            subject_data['sequence'] = studentList[idx]['part_type_sequence'][subject_data['subject_part_type_code_name']]['sequence']
            if subject_data['subject_part_type_code_name'] == 'part1':
                part_one_subject_list.append(subject_data)
            elif subject_data['subject_part_type_code_name'] == 'part2':
                part_two_subject_list.append(subject_data)
            if subject_data['subject_part_type_code_name'] not in studentList[idx]['subject_names_part_wise']:
                studentList[idx]['subject_names_part_wise'][subject_data['subject_part_type_code_name']]=[]
            if subject_data['subject_part_type_code_name'] not in studentList[idx]['obtained_marks_list_part_wise']:
                studentList[idx]['obtained_marks_list_part_wise'][subject_data['subject_part_type_code_name']]=[]
                studentList[idx]['obtained_grade_list_part_wise'][subject_data['subject_part_type_code_name']]=[]
                studentList[idx]['obtained_marks_list_part_wise_graph'][subject_data['subject_part_type_code_name']]=[]
            studentList[idx]['subject_names_part_wise'][subject_data['subject_part_type_code_name']].append(subject_data['subject_name'])
            if 'marks' in subject_data and subject_data['marks']:
                studentList[idx]['obtained_marks_list_part_wise'][subject_data['subject_part_type_code_name']].append(subject_data['marks'])
                studentList[idx]['obtained_marks_list_part_wise_graph'][subject_data['subject_part_type_code_name']].append(subject_data['marks'])
            else:
                studentList[idx]['obtained_marks_list_part_wise'][subject_data['subject_part_type_code_name']].append(0)
                if 'grade_plan' in subject_data and subject_data['grade_plan'] and subject_data['grade']:
                    grade_plans = Grade.objects.filter(grade_plan_id = subject_data['grade_plan'],name = subject_data['grade']).values().first()
                    studentList[idx]['obtained_marks_list_part_wise_graph'][subject_data['subject_part_type_code_name']].append(grade_plans['to_range'])
                else:
                    studentList[idx]['obtained_marks_list_part_wise_graph'][subject_data['subject_part_type_code_name']].append(0)
            if "result" in subject_data and subject_data['result'] == "fail":
                student_data['total_result'] = "fail"
        student_data['subject_list_data_part1']=part_one_subject_list
        student_data['subject_list_data_part2']=part_two_subject_list
        student_data['student_admission_form'] = student_admission_form[student_data['student']]
        student_data['student_admission_form']['admission_num'] =  student_data['student_admission_form']['admission_num'].replace('/', '-')
        studentList[idx]['total_summary'] = {
            'total_result': student_data['total_result'] if 'total_result' in student_data else None,
            'section_rank':student_data['section_rank'] if 'section_rank' in student_data else None,
            'standard_rank':student_data['standard_rank'] if 'standard_rank' in student_data else None,
            'total_marks': 0,
            'total_obtained_marks': 0,
            'total_min_marks': 0,
            'percentage': 0
        }
        for subject_id in student_data['subject_list']:
            subject_data = student_data['subject_list'][subject_id]
            student_data['subject_list'][subject_id]['cumulative_id_mark_mapping'] = {}
            studentList[idx]['subject_list'][subject_id]['cumulative_data'] = subjectList[subject_id]['cumulative_data'] if 'cumulative_data' in subjectList[subject_id] else []
            studentList[idx]['subject_list'][subject_id]['total_obtained_marks'] = 0
            studentList[idx]['subject_list'][subject_id]['grade'] = subject_data['grade'] if 'grade' in subject_data else ''
            if subject_data['subject_part_type'] not in studentList[idx]['part_type_data']:
                studentList[idx]['part_type_data'][subject_data['subject_part_type']] = {
                    'total_marks': 0,'total_min_marks':0,
                    'total_result': student_data['total_result'] if 'total_result' in student_data else None,
                    'total_obtained_marks': 0
                }
            if subject_data['subject_part_type_code_name'] not in studentList[idx]['part_type_data_code_wise']:
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']] = {
                    'total_written_marks': 0,
                    'total_result': student_data['total_result'] if 'total_result' in student_data else None,
                    'total_written_obtained_marks': 0,
                    'total_marks': 0,'total_obtained_marks': 0,
                    'total_cumulative_obtained_marks': 0,
                    'total_cumulative_max_marks': 0,
                    'total_written_min_marks':0,
                    'total_min_marks':0
                }
            studentList[idx]['subject_list'][subject_id]['total_obtained_marks'] = 0
            if 'marks' in subject_data:
                studentList[idx]['subject_list'][subject_id]['total_obtained_marks'] += subject_data['marks'] if subject_data['marks'] else 0
            if 'cumulative_data' in subject_data:
                for cumulative_row in subject_data['cumulative_data']:
                    col_name = []
                    for cumulative_type in cumulative_row['cumulative_type_data']:
                        col_name.append(cumulative_type['name'])
                    col_name = ','.join(col_name)
                    studentList[idx]['cumulative_id_type_mapping'][col_name] = {
                        'col_name': col_name
                    }
                    student_data['subject_list'][subject_id]['cumulative_id_mark_mapping'][col_name] = {
                        'marks': 0,  # default if not entered
                        'exam_cumulative__max_marks': cumulative_row['max_marks'],
                        'exam_cumulative__min_marks': cumulative_row['min_marks'],
                        'cumulative_type_data': cumulative_row['cumulative_type_data'],
                    }
            if 'cumulative_marks_data' in subject_data:
                for cum_data in subject_data['cumulative_marks_data']:
                    col_name = []
                    for cumulative_type in cum_data['cumulative_data_mapping']:
                        col_name.append(cumulative_type['cumulative_type_name'])
                    col_name = ','.join(col_name)
                    student_data['subject_list'][subject_id]['cumulative_id_mark_mapping'][col_name] = cum_data
                    studentList[idx]['subject_list'][subject_id]['total_obtained_marks'] += cum_data['marks']
            studentList[idx]['subject_list'][subject_id]['cumulative_mark_data_for_marks_card'] = [] #Indexing is imp
            if 'attendance_status' in subject_data and subject_data['attendance_status'] != 'Present':
                is_present_in_all_subject = False
            studentList[idx]['subject_list'][subject_id] = subject_data
            if 'total_obtained_marks' in subject_data and 'total_max_marks' in subject_data:
                if studentList[idx]['subject_list'][subject_id]['is_marks']:
                    studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_obtained_marks'] += subject_data['total_obtained_marks']
                    studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_marks'] += subject_data['total_max_marks']
                    studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_min_marks'] += subject_data['total_min_marks']
                    extra_params = {}
                    extra_params.update(grade_data_tracking_indiviual)
                    extra_params.update(form_definition_tracking)
                    is_fail = True if subject_data['result'] == "fail" else False
                    temp_grade_data = get_student_grade([subject_data],examId, standardSectionId, 'total_obtained_marks', 'total_max_marks', False, extra_params, is_fail)
                    grade_data = temp_grade_data['marksData']
                    grade_data_tracking_indiviual['grade_data'] = temp_grade_data['grade_data']
                    grade_data_tracking_indiviual['grade_plan_obj'] = temp_grade_data['grade_plan_obj']
                    studentList[idx]['subject_list'][subject_id]['grade'] = grade_data[0]['grade']
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_obtained_marks'] = studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_obtained_marks']
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_obtained_marks_in_word'] = num2words(studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_obtained_marks'], lang='en')
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_marks'] = studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_marks']
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_min_marks'] = studentList[idx]['part_type_data'][subject_data['subject_part_type']]['total_min_marks']
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_written_obtained_marks'] += 0 if not studentList[idx]['subject_list'][subject_id]['marks'] else studentList[idx]['subject_list'][subject_id]['marks']
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_written_marks'] += studentList[idx]['subject_list'][subject_id]['max_marks'] if studentList[idx]['subject_list'][subject_id]['max_marks'] else 0
                studentList[idx]['part_type_data_code_wise'][subject_data['subject_part_type_code_name']]['total_written_min_marks'] += studentList[idx]['subject_list'][subject_id]['min_marks'] if studentList[idx]['subject_list'][subject_id]['min_marks'] else 0

                student_data['total_summary']['total_marks'] += studentList[idx]['subject_list'][subject_id]['total_max_marks']
                student_data['total_summary']['total_min_marks'] += studentList[idx]['subject_list'][subject_id]['total_min_marks'] if studentList[idx]['subject_list'][subject_id]['total_min_marks'] else 0
                student_data['total_summary']['total_obtained_marks'] += studentList[idx]['subject_list'][subject_id]['total_obtained_marks'] if studentList[idx]['subject_list'][subject_id]['total_obtained_marks'] else 0
                student_data['obtained_marks'] = student_data['total_summary']['total_obtained_marks']
        studentList[idx]['cumulative_id_type_mapping'] = studentList[idx]['cumulative_id_type_mapping'].values()
        for cumulative_row in studentList[idx]['cumulative_id_type_mapping']:
            for subject_id in student_data['subject_list']:
                mark = 0
                max_mark = 0
                subject_data = student_data['subject_list'][subject_id]
                if cumulative_row['col_name'] in subject_data['cumulative_id_mark_mapping']:
                    mark = subject_data['cumulative_id_mark_mapping'][cumulative_row['col_name']]['marks']
                    max_mark = subject_data['cumulative_id_mark_mapping'][cumulative_row['col_name']]['exam_cumulative__max_marks']
                studentList[idx]['subject_list'][subject_id]['cumulative_mark_data_for_marks_card'].append(
                    {
                        'mark': mark,
                        'max_mark' : max_mark
                    }
                )
        for index, cummulative_type in enumerate(studentList[idx]['cumulative_id_type_mapping']):
            total_mark = 0
            total_max_mark = 0
            for subject_id in studentList[idx]['subject_list']:
                subject_data = studentList[idx]['subject_list'][subject_id]
                if subject_data['subject_part_type_code_name'] =='part1': #for now only getting part1 total
                    try:
                        mark = subject_data['cumulative_mark_data_for_marks_card'][index]['mark']
                        max_mark = subject_data['cumulative_mark_data_for_marks_card'][index]['max_mark']
                    except:
                        mark = 0
                        max_mark=0
                    total_mark += mark
                    total_max_mark +=max_mark
            studentList[idx]['cumulative_id_type_mapping_total'].append({'total_obtained_mark': total_mark,'total_max_mark':total_max_mark})
        for subject_part_type_code_name in studentList[idx]['part_type_data_code_wise']:
            extra_params = {}
            extra_params.update(grade_data_tracking_total)
            extra_params.update(form_definition_tracking)
            temp_grade = get_student_grade([studentList[idx]['part_type_data_code_wise'][subject_part_type_code_name]],examId, standardSectionId, 'total_obtained_marks', 'total_marks', True, extra_params)
            finalized_grade=StudentExamFinalResult.objects.filter(student_id=studentList[idx]['student'],exam_id=examId).values('totalgrade')
            grade = temp_grade['marksData']
            grade_data_tracking_total['grade_data'] = temp_grade['grade_data']
            grade_data_tracking_total['grade_plan_obj'] = temp_grade['grade_plan_obj']
            studentList[idx]['part_type_data_code_wise'][subject_part_type_code_name]['grade'] = grade[0]['grade']
            studentList[idx]['part_type_data_code_wise'][subject_part_type_code_name]['percentage'] = grade[0]['percentage']
            if finalized_grade:
                studentList[idx]['part_type_data_code_wise'][subject_part_type_code_name]['dynamic_grade'] = finalized_grade[0]['totalgrade'] if finalized_grade[0]['totalgrade'] else ''
            if institute_objs.code == 'nandinividyanikethana' and self.request.GET.get('print_marks_card'):
                studentList[idx]['part1_graph'] = generate_chart(studentList[idx]['subject_names_part_wise']["part1"], studentList[idx]['obtained_marks_list_part_wise_graph']["part1"] ,'Scholastic', chart_type="bar", colors=None)
                obtained_part_wise_graph = []
                subject_names_part_wise =[]
                if "parta" in studentList[idx]['obtained_marks_list_part_wise_graph']:
                    obtained_part_wise_graph+=studentList[idx]['obtained_marks_list_part_wise_graph']['parta']
                if "partb" in studentList[idx]['obtained_marks_list_part_wise_graph']:
                    obtained_part_wise_graph+=studentList[idx]['obtained_marks_list_part_wise_graph']['partb']  
                if "partc" in studentList[idx]['obtained_marks_list_part_wise_graph']:
                    obtained_part_wise_graph+=studentList[idx]['obtained_marks_list_part_wise_graph']['partc']
                if "parta" in studentList[idx]['subject_names_part_wise']:
                    subject_names_part_wise+=studentList[idx]['subject_names_part_wise']['parta']
                if "partb" in studentList[idx]['subject_names_part_wise']:
                    subject_names_part_wise+=studentList[idx]['subject_names_part_wise']['partb']  
                if "partc" in studentList[idx]['subject_names_part_wise']:
                    subject_names_part_wise+=studentList[idx]['subject_names_part_wise']['partc']  
                if sum(obtained_part_wise_graph):
                    studentList[idx]['part2_graph'] = generate_chart(subject_names_part_wise, 
                                            obtained_part_wise_graph ,'Co-Scholastic', chart_type="pie", colors=None)
                studentList[idx]['part3_graph'] = generate_chart(studentList[idx]['subject_names_part_wise']["char"], studentList[idx]['obtained_marks_list_part_wise_graph']["char"] ,'Personality And Character', chart_type="line", colors=None)
        extra_params = {}
        extra_params.update(grade_data_tracking_total)
        extra_params.update(form_definition_tracking)
        is_fail = True if student_data.get('total_result') == 'fail' else False
        temp_grade = get_student_grade([student_data],examId, standardSectionId, 'obtained_marks', 'total_marks', True, extra_params, is_fail)
        grade_data = temp_grade['marksData']
        grade_data_tracking_total['grade_data'] = temp_grade['grade_data']
        grade_data_tracking_total['grade_plan_obj'] = temp_grade['grade_plan_obj']
        studentList[idx]['grade'] = grade_data[0]['grade']
        for part_type in student_data['part_type_data']:
            extra_params = {}
            extra_params.update(grade_data_tracking_total)
            extra_params.update(form_definition_tracking)
            temp_grade = get_student_grade([student_data['part_type_data'][part_type]],examId, standardSectionId, 'total_obtained_marks', 'total_marks', True, extra_params)
            grade_data = temp_grade['marksData']
            grade_data_tracking_total['grade_data'] = temp_grade['grade_data']
            grade_data_tracking_total['grade_plan_obj'] = temp_grade['grade_plan_obj']
            studentList[idx]['part_type_data'][part_type]['grade'] = grade_data[0]['grade']

        #always this line should be in last
        studentList[idx]['total_summary']['total_obtained_marks_in_word'] = num2words(studentList[idx]['total_summary']['total_obtained_marks'], lang='en')
        if studentList[idx]['total_summary']['total_obtained_marks'] and studentList[idx]['total_summary']['total_marks']:
            studentList[idx]['total_summary']['percentage'] = (
                studentList[idx]['total_summary']['total_obtained_marks'] / studentList[idx]['total_summary']['total_marks']
            ) * 100
        else:
            studentList[idx]['total_summary']['percentage'] = 0
        marked_attendance_days = studentList[idx]['marked_attendance_days'] if 'marked_attendance_days' in studentList[idx] and studentList[idx]['marked_attendance_days'] else 0
        total_attendance = grade_exam_schedule_mapping[0]['max_no_of_days_attendance'] if len(grade_exam_schedule_mapping) > 0 and grade_exam_schedule_mapping[0]['max_no_of_days_attendance'] else 0
        studentList[idx]['marked_attendance_days_percentage'] = (marked_attendance_days / total_attendance)*100 if total_attendance else 0
    response['data']['student_list'] = sorted(studentList, key=lambda d: d['student_name'])
    response['data']['subject_list'] = sorted(
    subjectList.values(),
    key=lambda v: (
        0, int(v['subject__subject_code'])  # Numeric codes
    ) if v.get('subject__subject_code') and str(v['subject__subject_code']).isdigit() else (
        1, str(v.get('subject__subject_code') or '')  # Non-numeric or None
    ))
    response['data']['standard_name'] = standard_sec_obj.standard.name
    response['data']['section_name'] = standard_sec_obj.section.name
    response['data']['standard_section'] = standard_sec_obj.id
    response['data']['standard'] = standard_sec_obj.standard.id
    response['data']['is_announced'] = isAnnounced
    response['data']['approval_status'] = approval_data[0]['approval_status'] if approval_data else '0'
    response['data']['academic_year_details'] = {
        'start_date': exam_obj.academic_year.start_date,
        'end_date': exam_obj.academic_year.end_date,
        'academic_year': exam_obj.academic_year.id
    }
    response['data']['term_details'] = exam_obj.term.name
    response['data']['exam_details'] = exam_obj.exam_type.name
    response['data']['exam_code_name'] = exam_obj.exam_type.code
    response['data']['part_type_list'] = part_type_list
    if grade_exam_schedule_mapping:
        grade_plan_ids.add(grade_exam_schedule_mapping[0]['grade_plan'])
        grade_plan_ids.add(grade_exam_schedule_mapping[0]['grade_plan_for_total'])
        response['data']['max_no_of_days_attendance'] = grade_exam_schedule_mapping[0]['max_no_of_days_attendance'] if len(grade_exam_schedule_mapping) > 0 else None
        from_date_val = grade_exam_schedule_mapping[0]['attendance_from_date']
        to_date_val = grade_exam_schedule_mapping[0]['attendance_to_date']

        if isinstance(from_date_val, str) and from_date_val.lower() == "none":
          from_date_val = None
        if isinstance(to_date_val, str) and to_date_val.lower() == "none":
          to_date_val = None       
        response['data']['attendance_from_date'] = from_date_val
        response['data']['attendance_to_date'] = to_date_val
        grade_data = Grade.objects.filter(grade_plan__in=grade_plan_ids).values(
            'name', 'from_range', 'to_range', 'grade_plan',
            'grade_plan__name', 'grade_plan__grade_type'
        )
        response['data']['grade_type_data'] = {}
        for grade_row in grade_data:
            if grade_row['grade_plan'] not in response['data']['grade_type_data']:
                response['data']['grade_type_data'][grade_row['grade_plan']] = {
                    'grade_plan_name': grade_row['grade_plan__name'],
                    'grade_type': grade_row['grade_plan__grade_type'],
                    'grade_plan': grade_row['grade_plan'],
                    'grade_list': []
                }
            grade_data_temp={
                    'name': grade_row['name'], 'from_range': grade_row['from_range'],
                    'to_range': grade_row['to_range'], 'grade_plan': grade_row['grade_plan']
                }
            if grade_row['from_range'] and grade_row['to_range']:
                grade_data_temp['from_to_range'] = str(math.ceil(grade_row['to_range']))+'-'+str(math.ceil(grade_row['from_range']))
            else:
                grade_data_temp['from_to_range'] = ''
            response['data']['grade_type_data'][grade_row['grade_plan']]['grade_list'].append(grade_data_temp)
    else:
        response['data']['max_no_of_days_attendance'] = None
        response['data']['grade_type_data'] = {}
    try:
        response['data']['grade_plan_data_for_total'] = response['data']['grade_type_data'][grade_exam_schedule_mapping[0]['grade_plan_for_total']]['grade_list']
        response['data']['grade_plan_data_for_total'].sort(
        key=lambda x: x['to_range'] if x['to_range'] is not None else float('-inf'),
        reverse=True
        )
    except:
        response['data']['grade_plan_data_for_total'] = []
    response['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
    
    if self.request.GET.get('print_marks_card') or self.request.GET.get('type') == 'pdf':
        return print_marks_card(self, response)
    elif self.request.GET.get('print_consolidated_marks'):
        if response['institute_data']['code'] == 'sbvshr':
            if self.request.GET.get('print_consolidated_marks_new'):
                return download_consolidation_marks(self,response)
            return download_consolidation_marks_subject_wise(self,response)
        return download_consolidation_marks(self,response)
    return response

def get_multiple_standard_section_subjects(self, examId, standardSectionIds, raiseErrorIfNotFinalized=False, student_ids=[], ignore_final_result_data=False,data={}):
    """
    Supports multiple standard sections with bulk fetch. Does NOT call get_standard_section_subjects in a loop.
    """
    if not standardSectionIds:
        raise ValidationError('At least one standard section is required')
    if isinstance(standardSectionIds, (int, str)):
        standardSectionIds = [int(standardSectionIds)]
    else:
        standardSectionIds = [int(sid) for sid in standardSectionIds]
    exam_obj = Exam.objects.get(id=examId)
    form_definition_tracking = {'exam_configurations_grade_plan': FormdefinitionService.get_formdefinition_for_multiple_data(
        self, [{'form_name': 'exam_configurations', 'column_name': 'grade_plan'}]
    )['exam_configurations']['grade_plan']}
    institute_objs = Institute.get_institute(self)
    values_list = ['id', 'first_name', 'middle_name', 'last_name', 'current_reg_num', 'student_name', 'dob']
    custom_annotate = {'student_name': Concat('first_name', V(' '), 'middle_name', V(' '), 'last_name')}
    all_student_data = list(Student.get_student_for_standard(None, None, standardSectionIds, values_list, custom_annotate))
    student_to_sections = defaultdict(list)
    for e in Enrollment.objects.filter(standard_section__in=standardSectionIds).values('student', 'standard_section'):
        student_to_sections[e['student']].append(e['standard_section'])
    section_to_students = defaultdict(list)
    section_to_student_ids = defaultdict(list)
    user_list = []
    for row in all_student_data:
        row['student'] = row['id']
        for sec in student_to_sections.get(row['id'], []):
            if sec in standardSectionIds:
                section_to_students[sec].append(dict(row))
                section_to_student_ids[sec].append(row['id'])
                user_list.append(row['student__user_student'])

    all_schedules = list(ExamSchedule.objects.filter(
        exam=examId, standard_section__in=standardSectionIds, sub_schedule_parent=None
    ).values('id', 'subject', 'subject__name', 'min_marks', 'max_marks',
             'subject__subject_part_type__name', 'subject__subject_part_type',
             'subject__subject_part_type__code_name', 'subject__subject_code',
             'is_marks', 'grade_plan', 'grade_plan__name', 'subject__codename', 'schedule_sequence', 'standard_section'))
    section_to_schedules = defaultdict(list)
    all_schedule_ids = []
    for s in all_schedules:
        section_to_schedules[s['standard_section']].append(s)
        all_schedule_ids.append(s['id'])

    all_student_ids = [r['id'] for r in all_student_data]
    approvals_by_section = defaultdict(list)
    for a in StudentMarkSectionWiseApproval.objects.filter(exam=examId, standard_section__in=standardSectionIds).values():
        if a['standard_section_id'] not in approvals_by_section:
            approvals_by_section[a['standard_section_id']] = []
        approvals_by_section[a['standard_section_id']].append(a)
    if self.request.GET.get('exam_result'):
        for sec_id in standardSectionIds:
            ap = approvals_by_section.get(sec_id, [])
            if not ap or ap[0].get('approval_status') != 1:
                raise ValidationError('Student marks are not approved for the section')

    student_admission_form = get_student_admission_form_details(self, all_student_ids)
    all_marks_raw = list(StudentMark.objects.filter(
        student__in=all_student_ids, exam_schedule__in=all_schedule_ids, exam_schedule__exam=examId, is_active=True
    ).values(
        'student__first_name', 'student__middle_name', 'student__last_name', 'student__sts', 'student__current_reg_num',
        'exam_schedule_id', 'marks', 'student', 'staff', 'attendance_status', 'id', 'exam_schedule__is_marks',
        'exam_schedule__grade_plan', 'grade', 'student__mobile_num', 'student__dob', 'marked_attendance_is_global',
        'remark', 'student__profile_pic', 'remark_is_global', 'marked_attendance_days', 'remark', 'remark__name',
        'exam_schedule__standard_section','student__user_student'
    ))
    marks_by_section = defaultdict(list)
    for m in all_marks_raw:
        sec = m.get('exam_schedule__standard_section')
        if sec:
            marks_by_section[sec].append(m)

    schedule_detail = {r['id']: r for r in ExamSchedule.objects.filter(id__in=all_schedule_ids).values(
        'subject__name', 'subject__subject_code', 'subject__subject_part_type__name', 'subject__subject_part_type__code_name',
        'subject__subject_part_type__id', 'subject', 'min_marks', 'max_marks', 'id', 'is_marks', 'grade_plan', 'subject__codename'
    )}
    cum_mapping = list(ExamScheduleCumulativeMapping.objects.filter(exam_schedule__in=all_schedule_ids).values(
        'exam_schedule', 'cumulative_type', 'max_marks', 'min_marks', 'cumulative_type__name', 'cumulative_type__alias', 'id'
    ))
    student_cum = list(StudentCumulativeMark.objects.filter(
        student__in=all_student_ids, exam_cumulative__exam_schedule__in=all_schedule_ids, is_active=True
    ).values(
        'id', 'marks', 'exam_cumulative_id', 'exam_cumulative__exam_schedule', 'student', 'attendance_status',
        'exam_cumulative__cumulative_type', 'exam_cumulative__cumulative_type__name',
        'exam_cumulative__max_marks', 'exam_cumulative__min_marks'
    ))
    parents = {p['student']: p for p in StudentParentMapping.objects.filter(student__in=all_student_ids).values('parent__father_name', 'parent__mother_name', 'student')}
    details = {d['student']: d for d in StudentDetails.objects.filter(student__in=all_student_ids).values('caste__name', 'student')}
    profile_ids = list({m.get('student__profile_pic') for m in all_marks_raw if m.get('student__profile_pic')})
    doc_map = {d['id']: d for d in DocumentSerializer(Document.objects.filter(id__in=profile_ids), many=True).data}

    sections_data = {}
    standard_name = None
    part_type_list = [{'name': s['name']} for s in SubjectPartType.objects.all().values()]

    for standardSectionId in standardSectionIds:
        sec_student_ids = section_to_student_ids.get(standardSectionId, [])
        sec_schedules = section_to_schedules.get(standardSectionId, [])
        sec_schedule_ids = [s['id'] for s in sec_schedules]
        if not sec_schedule_ids or not sec_student_ids:
            continue
        approval_data = approvals_by_section.get(standardSectionId, [])
        standard_sec_obj = StandardSectionMapping.objects.select_related('standard', 'section').get(id=standardSectionId)
        studentIdDict = {r['id']: dict(r) for r in section_to_students.get(standardSectionId, [])}
        scheduleSubjectIds = [s['subject'] for s in sec_schedules]
        subjectList = {}
        grade_plan_ids = set()
        for schedule in sec_schedules:
            t = {'subject': schedule['subject'], 'subject_name': schedule['subject__name'],
                 'subject_part_type': schedule['subject__subject_part_type__name'],
                 'min_marks': schedule['min_marks'], 'max_marks': schedule['max_marks'], 'schedule': schedule['id'],
                 'subject_part_type_id': schedule['subject__subject_part_type'],
                 'subject_part_type_code_name': schedule['subject__subject_part_type__code_name'],
                 'subject__subject_code': schedule['subject__subject_code'], 'is_marks': schedule['is_marks'],
                 'grade_plan': schedule['grade_plan'], 'grade_plan_name': schedule['grade_plan__name'],
                 'subject__codename': schedule['subject__codename']}
            subjectList[schedule['subject']] = t
            if schedule['grade_plan']:
                grade_plan_ids.add(schedule['grade_plan'])

        ss_data = list(SubjectStudent.objects.filter(
            student__in=sec_student_ids, academic_year=exam_obj.academic_year, subject__in=scheduleSubjectIds
        ).values('subject', 'student', 'subject__name', 'subject__subject_code', 'subject__codename',
                 subject_name=F('subject__name'), subject_part_type=F('subject__subject_part_type__name'),
                 subject_part_type_id=F('subject__subject_part_type'), subject_part_type_code_name=F('subject__subject_part_type__code_name')))
        studentSubjectMapping = defaultdict(list)
        for ss in ss_data:
            studentSubjectMapping[ss['student']].append(dict(ss))

        sec_marks = marks_by_section.get(standardSectionId, [])
        student_mark_data = []
        for m in sec_marks:
            m = dict(m)
            m['first_name'] = m['student__first_name']
            m['middle_name'] = m['student__middle_name']
            m['last_name'] = m['student__last_name']
            m['mobile_num'] = m['student__mobile_num']
            m['dob'] = m['student__dob']
            m['is_marks'] = m['exam_schedule__is_marks']
            m['grade_plan'] = m['exam_schedule__grade_plan']
            m['full_name'] = get_full_name(m['student__first_name'], m['student__middle_name'], m['student__last_name'])
            m['sts'] = m['student__sts']
            m['current_reg_num'] = m['student__current_reg_num']
            p = parents.get(m['student'], {})
            m['father_name'] = p.get('parent__father_name')
            m['mother_name'] = p.get('parent__mother_name')
            d = details.get(m['student'], {})
            m['caste_name'] = d.get('caste__name')
            doc = doc_map.get(m.get('student__profile_pic'), {})
            m['student_profile_pic_file'] = doc.get('file') if doc else None
            student_mark_data.append(m)

        exam_schedule_data = {}
        for sid in sec_schedule_ids:
            r = schedule_detail.get(sid)
            if r:
                r = dict(r)
                r['subject_name'] = r.get('subject__name')
                r['subject_code'] = r.get('subject__subject_code')
                r['codename'] = r.get('subject__codename')
                r['subject_part_type'] = r.get('subject__subject_part_type__name')
                r['subject_part_type_code_name'] = r.get('subject__subject_part_type__code_name')
                r['subject_part_type_id'] = r.get('subject__subject_part_type__id')
                exam_schedule_data[sid] = r

        schedule_cumulative_mapping = defaultdict(list)
        for sc in cum_mapping:
            if sc['exam_schedule'] in sec_schedule_ids:
                schedule_cumulative_mapping[sc['exam_schedule']].append({
                    'exam_schedule': sc['exam_schedule'], 'max_marks': sc['max_marks'], 'min_marks': sc['min_marks'],
                    'cumulative_type_data': [{'id': sc['cumulative_type'], 'name': sc['cumulative_type__name'], 'alias': sc['cumulative_type__alias']}]})

        student_cumulative_data_mapping = defaultdict(lambda: defaultdict(list))
        for sc in student_cum:
            if sc['exam_cumulative__exam_schedule'] in sec_schedule_ids and sc['student'] in sec_student_ids:
                student_cumulative_data_mapping[sc['exam_cumulative__exam_schedule']][sc['student']].append({
                    'marks': sc['marks'], 'exam_cumulative_id': sc['exam_cumulative_id'],
                    'exam_cumulative__exam_schedule': sc['exam_cumulative__exam_schedule'], 'student': sc['student'],
                    'attendance_status': sc['attendance_status'], 'exam_cumulative__max_marks': sc['exam_cumulative__max_marks'],
                    'exam_cumulative__min_marks': sc['exam_cumulative__min_marks'],
                    'cumulative_data_mapping': [{'cumulative_type_id': sc['exam_cumulative__cumulative_type'],
                                               'cumulative_type_name': sc['exam_cumulative__cumulative_type__name']}]})

        for sub in subjectList.values():
            sub['total_max_marks'] = sub['max_marks'] or 0
            sub['total_min_marks'] = sub['min_marks'] or 0
            for r in schedule_cumulative_mapping.get(sub['schedule'], []):
                sub['total_max_marks'] += r.get('max_marks') or 0
                sub['total_min_marks'] += r.get('min_marks') or 0
            sub['cumulative_data'] = schedule_cumulative_mapping.get(sub['schedule'], [])
        section_response = _process_section_marks_bulk(
            self, examId, standardSectionId, exam_obj, standard_sec_obj, sec_student_ids, studentIdDict,
            sec_schedule_ids, subjectList, scheduleSubjectIds, dict(studentSubjectMapping), student_mark_data,
            exam_schedule_data, dict(schedule_cumulative_mapping), dict(student_cumulative_data_mapping),
            approval_data, form_definition_tracking, part_type_list,
            {sid: student_admission_form.get(sid, {'admission_num': ''}) for sid in sec_student_ids},
            grade_plan_ids, raiseErrorIfNotFinalized, ignore_final_result_data
        )
        if section_response:
            # section_response['section_id'] = standardSectionId
            # section_response['standard_section'] = standardSectionId
            # if not standard_name:
            #     standard_name = section_response.get('standard_name')
            sections_data.update(section_response)
    user_data = get_user_contact_details(user_list)
    if not user_list:
        raise ValidationError('No users to send notification')
    if 'language' not in data or not data['language']:
        raise ValidationError('Langauge is mandatory')
    if 'schedule' in data and data['schedule']:
        if data['schedule'] < datetime.now().strftime('%Y-%m-%d %H:%M:%S'):
            raise ValidationError('Schedule should be greater than now')
    else:
        data['schedule'] = None
    medium_error = ''
    if data['medium'] == 'email':
        handle_medium = handle_email_notification_bulk
        medium_error = 'Email id is not present'
    elif data['medium'] == 'sms':
        handle_medium = handle_sms_notification_bulk
        medium_error = 'Mobile num is not present'
    elif data['medium'] == 'push':
        handle_medium = handle_push_notification_bulk
        medium_error = 'Push not registered for the user'
    sent_user_ids = []
    notification_data = []
    unsendable_user_with_error = {}
    s3_objects = {'document_list': []}
    standard_suffix_data = SharedService.get_standard_suffix_data(self)
    for user_datas in user_data:
        user_datas['institute_name'] = Institute.objects.all().first().name
        user_datas['exam_name'] = exam_obj.exam_type.name
        user_datas.update(sections_data.get(user_datas['id'], {}))
    data['message_data_new']=data['message_data']
    for user in user_data:
        data['message_data'] = data['message_data_new'].format(**user)
        institue = Institute.objects.all().first()
        return_data = handle_medium(self, user, data, institue, **s3_objects)
        if return_data:
            sent_user_ids.append(user['id'])
            return_data['transaction_id'] = data['transaction_id']
            notification_data.append(return_data)
        else:
            name = ''
            if user['student']:
                name += 'Student '+ get_full_name(user['student__first_name'], user['student__middle_name'], user['student__last_name'])
            else:
                name += 'Staff '+ get_full_name(user['staff__first_name'], user['staff__middle_name'], user['staff__last_name'])
            unsendable_user_with_error[user['id']] = name + ' ' +medium_error
    sent_user_ids = list(set(sent_user_ids))
    with transaction.atomic(using=get_current_db_name()):
        data['notification_type'] = 1
        bulk_notification_data={
            'message_data':data['message_data'],
            'heading':data['heading'],
            'academic_year_id': exam_obj.academic_year.id,
            'created_by_user_id': self.request.user.id,
            'schedule':data['schedule'],
            'notification_medium_id':NotificationMedium.objects.filter(name=data['medium']).first().id,
            'notification_type':1,
            'language_id':data['language'],
            'transaction_id':data['transaction_id']
        }
        bulk_response = BulkNotification.objects.create(**bulk_notification_data)
        bulk_response.save()
        data['bulk_notification'] = bulk_response.id
        add_to_notification_users(self, data, sent_user_ids)
        if notification_data:
            SharedService.custom_thread(post_to_notification, notification_data, 'bulk_notification')
        return {"data":"Notification sent successfully"}

def _process_section_marks_bulk(self, examId, standardSectionId, exam_obj, standard_sec_obj, student_ids, studentIdDict,
    scheduleIds, subjectList, scheduleSubjectIds, studentSubjectMapping, student_mark_data, exam_schedule_data,
    schedule_cumulative_mapping, student_cumulative_data_mapping, approval_data, form_definition_tracking,
    part_type_list, student_admission_form, grade_plan_ids, raiseErrorIfNotFinalized, ignore_final_result_data):
    """Process one section from bulk-fetched data. Returns section response dict."""
    grade_exam_schedule_mapping = list(GradeExamScheduleMapping.objects.filter(
        standard_section=standardSectionId, exam=examId
    ).values('max_no_of_days_attendance', 'grade_plan', 'grade_plan_for_total', 'attendance_from_date', 'attendance_to_date'))
    if grade_exam_schedule_mapping:
        grade_plan_ids = set(grade_plan_ids)
        grade_plan_ids.add(grade_exam_schedule_mapping[0].get('grade_plan'))
        grade_plan_ids.add(grade_exam_schedule_mapping[0].get('grade_plan_for_total'))

    studentBasedData = {}
    enteredMarksStudentIds = []
    for markData in student_mark_data:
        markData['exam_schedule'] = exam_schedule_data.get(markData['exam_schedule_id']) or {}
        subj = markData['exam_schedule'].get('subject')
        if not subj or subj not in subjectList:
            continue
        total_max_marks = subjectList[subj].get('total_max_marks', 0)
        total_min_marks = subjectList[subj].get('total_min_marks', 0)
        temp = {'id': markData['id'], 'marks': markData['marks'], 'subject': subj,
                'subject_name': markData['exam_schedule'].get('subject_name'), 'subject__subject_code': markData['exam_schedule'].get('subject_code'),
                'subject__codename': markData['exam_schedule'].get('codename'),
                'subject_part_type': markData['exam_schedule'].get('subject_part_type'),
                'subject_part_type_code_name': markData['exam_schedule'].get('subject_part_type_code_name'),
                'subject_part_type_id': markData['exam_schedule'].get('subject_part_type_id'),
                'attendance_status': markData['attendance_status'], 'min_marks': markData['exam_schedule'].get('min_marks'),
                'max_marks': markData['exam_schedule'].get('max_marks'), 'result': 'pass', 'total_max_marks': total_max_marks,
                'total_min_marks': total_min_marks, 'cumulative_marks_data': [],
                'cumulative_data': subjectList[subj].get('cumulative_data', []), 'is_marks': markData.get('is_marks'),
                'grade': markData.get('grade'), 'grade_plan': markData['exam_schedule'].get('grade_plan')}
        if markData['exam_schedule'].get('id') in student_cumulative_data_mapping and markData['student'] in student_cumulative_data_mapping.get(markData['exam_schedule']['id'], {}):
            temp['cumulative_marks_data'] = student_cumulative_data_mapping[markData['exam_schedule']['id']][markData['student']]
        obtained = (temp['marks'] or 0) + sum(c.get('marks') or 0 for c in temp['cumulative_marks_data'])
        if temp['attendance_status'] == 'Absent' or (temp['min_marks'] and not temp['marks']) or (temp['marks'] and temp['marks'] < temp['min_marks']):
            temp['result'] = 'fail'
        if markData['student'] in studentBasedData:
            studentBasedData[markData['student']]['total_obtained_marks'] += obtained
            if temp['result'] == 'fail':
                studentBasedData[markData['student']]['total_result'] = 'fail'
            if temp['attendance_status'] == 'Absent':
                studentBasedData[markData['student']]['total_result'] = 'fail'
            if temp['attendance_status'] == 'Present':
                studentBasedData[markData['student']]['total_max_marks'] += total_max_marks
            studentBasedData[markData['student']]['subject_list'][subj] = temp
            studentBasedData[markData['student']]['subject_list_str']+=f'{temp["subject_name"]}: {temp["marks"]}\n'
        else:
            studentBasedData[markData['student']] = {
                'student_name': markData['full_name'], 'student': markData['student'], 'sts': markData['sts'],
                'father_name': markData.get('father_name'), 'mother_name': markData.get('mother_name'),
                'caste_name': markData.get('caste_name'), 'mobile_num': markData.get('mobile_num'),
                'current_reg_num': markData['current_reg_num'], 'first_name': markData['first_name'],
                'middle_name': markData['middle_name'], 'last_name': markData['last_name'],
                'total_max_marks': total_max_marks, 'dob': markData.get('dob'), 'total_obtained_marks': obtained,
                'total_result': temp['result'], 'subject_list': {subj: temp}, 'part_type_list': part_type_list,
                'marked_attendance_days': markData.get('marked_attendance_days'), 'remark': markData.get('remark'),'section_rank':None,'standard_rank':None,
                'remark_name': markData.get('remark__name'), 'profile_pic_file': markData.get('student_profile_pic_file'),'user_id': markData.get('student__user_student'),'subject_list_str':""}
            enteredMarksStudentIds.append(markData['student'])
    finalResultData = {r['student']: r for r in StudentExamFinalResult.objects.filter(student__in=enteredMarksStudentIds, exam=examId).values('student', 'status', 'is_announced', 'section_rank', 'standard_rank')}
    user_student_data = {}
    for studentD in studentBasedData:
        if studentD not in finalResultData and raiseErrorIfNotFinalized:
            raise ValidationError('Student marks are not yet finalized')
        if studentD in studentSubjectMapping:
            for subData in studentSubjectMapping[studentD]:
                if subData['subject'] not in studentBasedData[studentD].setdefault('subject_list', {}):
                    studentBasedData[studentD].setdefault('total_marks', 0)
                    if subjectList.get(subData['subject'], {}).get('is_marks'):
                        studentBasedData[studentD]['total_marks'] += subjectList[subData['subject']].get('max_marks') or 0
                    subData['cumulative_data'] = subjectList.get(subData['subject'], {}).get('cumulative_data', [])
                    studentBasedData[studentD]['subject_list'][subData['subject']] = subData
        if not ignore_final_result_data and studentD in finalResultData:
            studentBasedData[studentD]['total_result'] = finalResultData[studentD]['status']
            studentBasedData[studentD]['section_rank'] = finalResultData[studentD]['section_rank']
            studentBasedData[studentD]['standard_rank'] = finalResultData[studentD]['standard_rank']
        user_student_data[studentBasedData[studentD]['user_id']]=studentBasedData[studentD]
    for studentId in studentIdDict:
        if studentId not in studentBasedData:
            if raiseErrorIfNotFinalized:
                s = studentIdDict[studentId]
                raise ValidationError(f"{get_full_name(s.get('first_name'), s.get('middle_name'), s.get('last_name'))} marks are not yet entered for the section")
            if studentId in studentSubjectMapping:
                for subData in studentSubjectMapping[studentId]:
                    if subData['subject'] not in studentIdDict[studentId].setdefault('subject_list', {}):
                        studentIdDict[studentId].setdefault('total_marks', 0)
                        if subjectList.get(subData['subject'], {}).get('is_marks'):
                            studentIdDict[studentId]['total_marks'] += subjectList[subData['subject']].get('max_marks') or 0
                        subData['cumulative_data'] = subjectList.get(subData['subject'], {}).get('cumulative_data', [])
                        studentIdDict[studentId]['subject_list'][subData['subject']] = subData
            studentIdDict[studentId]['subject_list_str'] = "Students Not Attended the Exam"
            studentIdDict[studentId]['total_obtained_marks'] = "NA"
            studentIdDict[studentId]['total_max_marks'] = "NA"
            studentIdDict[studentId]['total_result'] = "NA"
            studentIdDict[studentId]['standard_rank'] = "NA"
            user_student_data[studentIdDict[studentId]['student__user_student']]=studentIdDict[studentId]
    # for idx, student_data in enumerate(user_student_data.values()):
    #     student_data['student_admission_form'] = student_admission_form.get(student_data['student'], {'admission_num': ''})
    #     student_data['student_admission_form']['admission_num'] = student_data['student_admission_form'].get('admission_num', '').replace('/', '-')
    #     student_data['total_summary'] = {'total_result': student_data.get('total_result'), 'section_rank': student_data.get('section_rank'),
    #         'standard_rank': student_data.get('standard_rank'), 'total_marks': 0, 'total_obtained_marks': 0, 'total_min_marks': 0, 'percentage': 0}
    #     for subject_id, sub in student_data['subject_list'].items():
    #         sub['total_obtained_marks'] = (sub.get('marks') or 0) + sum(c.get('marks', 0) for c in sub.get('cumulative_marks_data', []))
    #         student_data['total_summary']['total_marks'] += sub.get('total_max_marks', 0)
    #         student_data['total_summary']['total_obtained_marks'] += sub['total_obtained_marks']
    #     student_data['obtained_marks'] = student_data['total_summary']['total_obtained_marks']
    #     student_data['total_summary']['total_obtained_marks_in_word'] = num2words(student_data['total_summary']['total_obtained_marks'], lang='en')
    #     student_data['total_summary']['percentage'] = (student_data['total_summary']['total_obtained_marks'] / student_data['total_summary']['total_marks'] * 100) if student_data['total_summary']['total_marks'] else 0
    return user_student_data

def download_consolidation_marks_subject_wise(self, data):
    data['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
    consolidated_data = []
    dynamic_labels = {}
    is_cum_type = 0
    cum_marks = None
    return_data = {}
    first_subject_name = None  # Initialize to track first subject name
    for subject_new in data['data']['subject_list']:
        subject = subject_new['subject']
        subject_total = 0
        subject_max_total = 0  
        # Store first subject name for later use
        if first_subject_name is None:
            first_subject_name = subject_new['subject_name'] 
        if subject not in return_data:
            return_data[subject] = {
                'stand_name': data['data']['standard_name'],
                'section_name': data['data']['section_name'],
                'standard_name': subject_new['subject_name'],
                'subject_name':subject_new['subject_name'],
                'student_list': [],
                'summary_dict': {}
            }
        for grade in data['data']['grade_plan_data_for_total']:
            if grade['name'] not in return_data[subject]['summary_dict']:
                return_data[subject]['summary_dict'][grade['name']] = copy.deepcopy(grade)
            if 'no_of_students' not in return_data[subject]['summary_dict'][grade['name']]:
                return_data[subject]['summary_dict'][grade['name']]['no_of_students'] = 0
        for students in data['data']['student_list']:
            student_name = students['student_name']
            student_row_data = {'student_name': student_name}
            subject_data = students['subject_list'].get(subject)
            if not subject_data:
                continue
            student_row_data['subject_obtained_marks'] = 'Ab' if subject_data['attendance_status'] == 'Absent' else subject_data['total_obtained_marks']
            student_row_data['subject_obtained_grade'] = subject_data['grade']
            student_row_data['subject_obtained_percentage'] = 'Ab' if subject_data['attendance_status'] == 'Absent' else subject_data['percentage']
            student_row_data['cum_marks'] = 0
            cum_num = 0
            if subject_data.get('cumulative_mark_data_for_marks_card'):
                is_cum_type = 1
                cum_marks = subject_data['cumulative_mark_data_for_marks_card']
                for marks in cum_marks:
                    student_row_data['cum_marks'] += marks['mark']
                    cum_num += 1
            student_row_data['written_marks'] = 'Ab' if subject_data['attendance_status'] == 'Absent' else subject_data.get('marks', '')
            subject_total += subject_data['total_obtained_marks']
            subject_max_total += subject_data['total_max_marks']
            grade_key = subject_data['grade']
            if 'ab' not in return_data[subject]['summary_dict']:
                return_data[subject]['summary_dict']['ab'] = {
                    'from_to_range': 'ABSENT', 'name': '', 'no_of_students': 0
                }
            if subject_data['attendance_status'] == 'Absent':
                return_data[subject]['summary_dict']['ab']['no_of_students'] += 1
            else:
                return_data[subject]['summary_dict'][grade_key]['no_of_students'] += 1
            if 'total' not in return_data[subject]['summary_dict']:
                return_data[subject]['summary_dict']['total'] = {
                    'from_to_range': 'TOTAL', 'name': '', 'no_of_students': 0
                }
            return_data[subject]['summary_dict']['total']['no_of_students'] += 1
            return_data[subject]['student_list'].append(student_row_data)
        sorted_summary = dict(sorted(return_data[subject]['summary_dict'].items(), key=lambda x: x[0]))
        return_data[subject]['summary'] = list(sorted_summary.values())
        subject_percentage = (subject_total / subject_max_total) * 100 if subject_max_total else 0
        return_data[subject]['abstract_data'] = [
            {'name': 'Subject Total', 'value': subject_total},
            {'name': 'Subject Maximum Marks', 'value': subject_max_total},
            {'name': 'Subject Percentage', 'value': subject_percentage}
        ]
    options = {
        'title': 'Consolidated_marks',
        'description': 'marks',
        'examname': data['data']['exam_details'],
        'standardname': data['data']['standard_name'],
        'sectionname': data['data']['section_name'],
        'institute_name': data['institute_data']['name'],
        'extraWorksheet': False,
        'Data': return_data,
        'extraWorksheetData': {},
        'subjectname': first_subject_name if first_subject_name else '',
        'institute_details':{
            'name':'SRI BHAIRAVESHWARA',
            'name_two':'VIDYA SAMSTHE(R)',
            'Affiliation':'Affiliated to Cental Board of Secondary Education(CBSE)',
            'Affiliaction_num':'Delhi-Affiliation No.831117 School No:46750',
            'Address':'Kaiwara Cross, Kaiwara, Chintamani(Tq)',
            'Address_two':'Chikkaballapur(Dist)'
        },
        'columns': json_for_consolidated_marks_subject_wise(data, is_cum_type, cum_marks),
        'abstract_data': [
            {'column': '', 'required': False, 'schemacolumn': 'name'},
            {'column': '', 'required': False, 'schemacolumn': 'value'}
        ],
        'summary': [
            {'column': 'PERCENTAGE RANGE', 'required': False, 'schemacolumn': 'from_to_range'},
            {'column': 'GRADE', 'required': False, 'schemacolumn': 'name'},
            {'column': 'NUMBER OF STUDENTS', 'required': False, 'schemacolumn': 'no_of_students'}
        ]
    }
    return write_to_excel_multiple_tabs(self, options, {}, {})


def download_consolidation_marks(self, data):
    data['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
    consolidated_data=[]
    dyanmic_labels = {}
    is_cum_type=0
    cum_marks=None
    total_obtained_marks = 0
    total_max_marks = 0
    subject_marks_summary = {}
    for students in data['data']['student_list']:
        student_name = students['student_name']
        student_row_data = {'student_name': student_name}
        attendance_status = ''
        
        if data['institute_data']['code'] == 'sbvshr':
            for sub_value in students['subject_list'].values():
                # if sub_value['subject_part_type_code_name'] == 'part2':
                #     continue
                subject_name = sub_value['subject_name']
                if subject_name not in subject_marks_summary:
                    subject_marks_summary[subject_name] = {}
                if 'max_marks_conducted' not in subject_marks_summary[subject_name]:
                    subject_marks_summary[subject_name]['max_marks_conducted'] = 0
                subject_marks_summary[subject_name]['max_marks_conducted'] += sub_value['total_max_marks']
                if 'min_marks_conducted' not in subject_marks_summary[subject_name]:
                    subject_marks_summary[subject_name]['min_marks_conducted'] = 0
                subject_marks_summary[subject_name]['min_marks_conducted'] += sub_value['total_min_marks']
                if 'total_obtained_marks' not in subject_marks_summary[subject_name]:
                    subject_marks_summary[subject_name]['total_obtained_marks'] = 0
                subject_marks_summary[subject_name]['total_obtained_marks']+=sub_value['total_obtained_marks']
                if 'subject_grades_list' not in subject_marks_summary[subject_name]:
                    subject_marks_summary[subject_name]['subject_grades_list'] = {}
                if sub_value['grade'] not in subject_marks_summary[subject_name]['subject_grades_list']:
                    subject_marks_summary[subject_name]['subject_grades_list'][sub_value['grade']] = 0
                subject_marks_summary[subject_name]['subject_grades_list'][sub_value['grade']] += 1
                if 'attendance' not in subject_marks_summary[subject_name]:
                    subject_marks_summary[subject_name]['attendance'] = {'Absent_Count': 0, 'Present_Count': 0}
                if sub_value['attendance_status'].lower() == 'absent':
                    subject_marks_summary[subject_name]['attendance']['Absent_Count'] += 1
                else:
                    subject_marks_summary[subject_name]['attendance']['Present_Count'] += 1
        for subjects in students['subject_list_data']:
            if data['institute_data']['code'] == 'lourdes':
                subject = subjects['subject']
                attendance_status = subjects.get('attendance_status')
                subject_marks_obtained = 'subject_'+str(subject)+'_obtained_marks'
                student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks'] if subjects['is_marks'] else subjects['grade']
                cum_num=0
                if subjects['cumulative_mark_data_for_marks_card']:
                    is_cum_type=1
                    cum_marks=subjects['cumulative_mark_data_for_marks_card']
                    for marks in subjects['cumulative_mark_data_for_marks_card']:
                        student_row_data['marks'+str(subject)+str(cum_num)] = marks['mark']
                        cum_num+=1
                    student_row_data[str(subject)+'written_marks'] = subjects['marks'] if 'marks' in subjects else ''
            elif data['institute_data']['code'] == 'sbvshr':
                subject = subjects['subject']   
                attendance_status = subjects.get('attendance_status')
                subject_marks_obtained = 'subject_'+str(subject)+'_obtained_marks'
                subject_grade_obtained='subject_'+str(subject)+'_obtained_grade'
                student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks'] if subjects['is_marks'] else subjects['grade']
                student_row_data[subject_grade_obtained] = subjects['grade']
                cum_num=0
                if subjects['cumulative_mark_data_for_marks_card']:
                    is_cum_type=1
                    cum_marks=subjects['cumulative_mark_data_for_marks_card']
                    for marks in subjects['cumulative_mark_data_for_marks_card']:
                        student_row_data['marks'+str(subject)+str(cum_num)] = marks['mark']
                        cum_num+=1
                    student_row_data[str(subject)+'written_marks'] = subjects['marks'] if 'marks' in subjects else '' 
            elif data['institute_data']['code'] == 'jnanajyothi' and subjects['subject_part_type'] == 'Part 1':
                subject = subjects['subject']
                attendance_status = subjects.get('attendance_status')
                subject_marks_obtained = 'subject_'+str(subject)+'_obtained_marks'
                subject_grade_obtained='subject_'+str(subject)+'_obtained_grade'
                student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks']
                student_row_data[subject_grade_obtained] = subjects['grade']
                cum_num=0
                if subjects['cumulative_mark_data_for_marks_card']:
                    is_cum_type=1
                    cum_marks=subjects['cumulative_mark_data_for_marks_card']
                    for marks in subjects['cumulative_mark_data_for_marks_card']:
                        student_row_data['marks'+str(subject)+str(cum_num)] = marks['mark']
                        cum_num+=1
                    student_row_data[str(subject)+'written_marks'] = 'Ab' if attendance_status == 'Absent' else subjects['marks'] if 'marks' in subjects else ''
            else:
                subject = subjects['subject']
                attendance_status = subjects.get('attendance_status')
                subject_marks_obtained = 'subject_'+str(subject)+'_obtained_marks'
                subject_grade_obtained='subject_'+str(subject)+'_obtained_grade'
                student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks']
                student_row_data[subject_grade_obtained] = subjects['grade']
                cum_num=0
                if subjects.get('cumulative_mark_data_for_marks_card'):
                    is_cum_type=1
                    cum_marks=subjects['cumulative_mark_data_for_marks_card']
                    # Check if cumulative_marks_data exists before iterating
                    if 'cumulative_marks_data' in subjects and subjects['cumulative_marks_data']:
                        for marks in subjects['cumulative_marks_data']:
                            if marks.get('cumulative_data_mapping') and len(marks['cumulative_data_mapping']) > 0:
                                student_row_data['marks'+str(subject)+str(marks['cumulative_data_mapping'][0]['cumulative_type_id'])] = marks.get('marks', '')
                            cum_num+=1
                    student_row_data[str(subject)+'written_marks'] = 'Ab' if attendance_status == 'Absent' else subjects.get('marks', '') if 'marks' in subjects else ''
        student_row_data['section_rank'] = students['section_rank'] if 'section_rank' in students and students['section_rank'] else None
        student_row_data['standard_rank'] = students['standard_rank'] if 'standard_rank' in students and students['standard_rank'] else None
        student_row_data['total_obtained_marks'] = students['part_type_data']['Part 1']['total_obtained_marks']
        student_row_data['total_obtained_grade'] = students['part_type_data']['Part 1']['grade'] 
        student_row_data['total_obtained_percentage'] =  students['part_type_data']['Part 1']['percentage']
        student_row_data['total_obtained_percentage'] = round(student_row_data['total_obtained_percentage'],2)
        student_row_data['remark'] = students['remark_name'] if 'remark_name' in students else None
        student_row_data['marked_attendance_days'] = students['marked_attendance_days'] if 'marked_attendance_days' in students else ''
        try: #ignoring divide by zero for now 
            total_obtained_marks += student_row_data['total_obtained_marks']
            total_max_marks += students['part_type_data']['Part 1']['total_marks']
            total_percentage = (total_obtained_marks/total_max_marks)*100
        except:
            pass
        consolidated_data.append(student_row_data)
    multiple_data = []
    options={}
    options['title'] = 'Consolidated_marks'
    options['description'] = 'marks'
    options['examname']=data['data']['exam_details']
    options['standardname']=data['data']['standard_name']
    options['sectionname']=data['data']['section_name']
    options['institute_name'] = data['institute_data']['name'] 
    options['institute_code'] = data['institute_data']['code']
    options['extraWorksheet'] = False
    options['Data'] = consolidated_data
    options['total_summary_details'] = [
            {'name': 'Total Obtained Marks', 'value': total_obtained_marks},
            {'name': 'Total Maximum Marks', 'value': total_max_marks},
            # {'name': 'Total Percentage', 'value': total_percentage}
        ]
    
    if data['institute_data']['code'] == "sbvshr":
        options['total_summary_data'] = [{'column': '', 'required': False, 'schemacolumn': 'name'},
            {'column': '', 'required': False, 'schemacolumn': 'value'}]
        
        def assign_summary_percentage(subject_marks_summary):
            class_efficiency = 0.0
            total_max_marks = 0
            total_obtained_marks = 0.0

            for subject_name, sub_value in subject_marks_summary.items():
                if not isinstance(sub_value, dict):
                    continue
                max_marks = sub_value.get('max_marks_conducted', 0)
                obtained = sub_value.get('total_obtained_marks', 0)
                total_max_marks += max_marks
                total_obtained_marks += obtained
                if max_marks > 0:
                    sub_value['subject_percentage'] = (obtained / max_marks) * 100
                else:
                    sub_value['subject_percentage'] = 0.0
                    
            if total_max_marks > 0:
                class_efficiency = (total_obtained_marks / total_max_marks) * 100
            return subject_marks_summary, class_efficiency
        
        def assign_subject_ranks(subject_marks_summary):
            subject_percentages = {
                subject: summary.get('subject_percentage', 0)
                for subject, summary in subject_marks_summary.items()
            }
            sorted_subjects = sorted(subject_percentages.items(), key=lambda x: x[1], reverse=True)
            rank = 1
            last_percentage = None
            rank_map = {}

            for i, (subject, percentage) in enumerate(sorted_subjects, start=1):
                if percentage != last_percentage:  
                    rank = i
                rank_map[subject] = rank
                last_percentage = percentage

            for subject, summary in subject_marks_summary.items():
                summary['subject_rank'] = rank_map[subject]

            return subject_marks_summary

        subject_marks_summary, class_eff = assign_summary_percentage(subject_marks_summary)
        subject_marks_summary = assign_subject_ranks(subject_marks_summary)
        subject_marks_summary['class_efficiency'] = class_eff
        subject_marks_summary['subject_total_grade'] = data['data'].get('grade_plan_data_for_total','')
        options['subject_marks_summary'] = subject_marks_summary   
    options['extraWorksheetData'] = dict()
    options['columns'] = json_for_consolidated_marks(data,is_cum_type,cum_marks)
    return write_to_excel_new_consolidation(self, options, {},{})

def process_standard_consolidated_marks(self, request, standard_id, throwError, student_ids):
    try:
        section_mappings = list(StandardSectionMapping.objects.filter(
            standard=standard_id
        ).values('id', 'standard__name', 'section__name'))
        if not section_mappings:
            raise ValidationError('No active sections found for the standard')
        all_sections_data = []
        standard_name = None
        sections_debug_info = []
        original_get = request.GET.copy()
        temp_get = original_get.copy()
        temp_get.pop('print_consolidated_marks', None)
        temp_get.pop('print_marks_card', None)
        temp_get.pop('long_running_process', None)
        for section in section_mappings:
            section_debug = {
                'section_id': section['id'],
                'section_name': section['section__name'],
                'response_exists': False,
                'has_data_key': False,
                'student_count': 0,
                'error': None
            }
            try:
                request.GET = temp_get
                section_response = get_standard_section_subjects(
                    self,
                    request.GET.get('exam'),
                    section['id'],
                    throwError,
                    student_ids
                )
                request.GET = original_get
                section_debug['response_exists'] = section_response is not None
                if section_response and 'data' in section_response:
                    section_debug['has_data_key'] = True
                    student_count = len(section_response['data'].get('student_list', []))
                    section_debug['student_count'] = student_count
                    if student_count > 0:
                        section_response['data']['section_id'] = section['id']
                        section_response['data']['section_name'] = section['section__name']
                        all_sections_data.append(section_response['data'])
                        if not standard_name:
                            standard_name = section['standard__name']
                else:
                    print(f"DEBUG: Section {section['section__name']} - No data in response or response is None")
            except Exception as e:
                import traceback
                section_debug['error'] = str(e)
                traceback.print_exc()
                request.GET = original_get
            sections_debug_info.append(section_debug)
        file_name = f"consolidated_marks_{standard_name}.xlsx"
        if all_sections_data:
            result = download_consolidation_marks_for_standard(self, all_sections_data, standard_name)
            if result.status_code == 200:
                with open(file_name, 'wb') as file:
                    file.write(result.content)
                filename = file_name
            url = UploadTypeService.upload_local_file(filename, path='exam')
            if os.path.exists(filename):
                os.remove(filename)
            transaction_id = self.request.GET.get('transaction_id')
            store_long_running_process(self, transaction_id, {'url': url})
        else:
            error_details = f"No marks data found for any section in the standard. Sections checked: {len(section_mappings)}. "
            error_details += "Details: " + "; ".join([
                f"{s['section_name']}: students={s['student_count']}, has_data={s['has_data_key']}, error={s['error'] or 'None'}"
                for s in sections_debug_info
            ])
            raise ValidationError(error_details)

    except Exception as e:
        import traceback
        traceback.print_exc()
        store_long_running_process(self, {'error': str(e)})

def download_consolidation_marks_for_standard(self, all_sections_data, standard_name):
    """
    Download consolidated marks for all sections within a standard
    """
    institute_data = InstituteSerializer(Institute.get_institute(self)).data
    consolidated_data = []
    is_cum_type = 0
    cum_marks = None
    total_obtained_marks = 0
    total_max_marks = 0
    total_percentage = 0
    subject_marks_summary = {}
    # Aggregate data from all sections
    for section_data in all_sections_data:
        for student in section_data.get('student_list', []):
            student_name = student['student_name']
            student_row_data = {'student_name': student_name}
            # Add section name to distinguish students from different sections
            student_row_data['section'] = section_data.get('section_name', '')
            attendance_status = ''
            if institute_data['code'] == 'sbvshr':
                for sub_value in student['subject_list'].values():
                    subject_name = sub_value['subject_name']
                    if subject_name not in subject_marks_summary:
                        subject_marks_summary[subject_name] = {}
                    if 'max_marks_conducted' not in subject_marks_summary[subject_name]:
                        subject_marks_summary[subject_name]['max_marks_conducted'] = 0
                    subject_marks_summary[subject_name]['max_marks_conducted'] += sub_value['total_max_marks']
                    if 'min_marks_conducted' not in subject_marks_summary[subject_name]:
                        subject_marks_summary[subject_name]['min_marks_conducted'] = 0
                    subject_marks_summary[subject_name]['min_marks_conducted'] += sub_value['total_min_marks']
                    if 'total_obtained_marks' not in subject_marks_summary[subject_name]:
                        subject_marks_summary[subject_name]['total_obtained_marks'] = 0
                    subject_marks_summary[subject_name]['total_obtained_marks'] += sub_value['total_obtained_marks']
                    if 'subject_grades_list' not in subject_marks_summary[subject_name]:
                        subject_marks_summary[subject_name]['subject_grades_list'] = {}
                    if sub_value['grade'] not in subject_marks_summary[subject_name]['subject_grades_list']:
                        subject_marks_summary[subject_name]['subject_grades_list'][sub_value['grade']] = 0
                    subject_marks_summary[subject_name]['subject_grades_list'][sub_value['grade']] += 1
                    if 'attendance' not in subject_marks_summary[subject_name]:
                        subject_marks_summary[subject_name]['attendance'] = {'Absent_Count': 0, 'Present_Count': 0}
                    if sub_value['attendance_status'].lower() == 'absent':
                        subject_marks_summary[subject_name]['attendance']['Absent_Count'] += 1
                    else:
                        subject_marks_summary[subject_name]['attendance']['Present_Count'] += 1
            for subjects in student['subject_list_data']:
                if institute_data['code'] == 'lourdes':
                    subject = subjects['subject']
                    attendance_status = subjects.get('attendance_status')
                    subject_marks_obtained = 'subject_' + str(subject) + '_obtained_marks'
                    student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks'] if subjects['is_marks'] else subjects['grade']
                    cum_num = 0
                    if subjects['cumulative_mark_data_for_marks_card']:
                        is_cum_type = 1
                        cum_marks = subjects['cumulative_mark_data_for_marks_card']
                        for marks in subjects['cumulative_mark_data_for_marks_card']:
                            student_row_data['marks' + str(subject) + str(cum_num)] = marks['mark']
                            cum_num += 1
                        student_row_data[str(subject) + 'written_marks'] = subjects['marks'] if 'marks' in subjects else ''
                elif institute_data['code'] == 'sbvshr':
                    subject = subjects['subject']
                    attendance_status = subjects.get('attendance_status')
                    subject_marks_obtained = 'subject_' + str(subject) + '_obtained_marks'
                    subject_grade_obtained = 'subject_' + str(subject) + '_obtained_grade'
                    student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks'] if subjects['is_marks'] else subjects['grade']
                    student_row_data[subject_grade_obtained] = subjects['grade']
                    cum_num = 0
                    if subjects['cumulative_mark_data_for_marks_card']:
                        is_cum_type = 1
                        cum_marks = subjects['cumulative_mark_data_for_marks_card']
                        for marks in subjects['cumulative_mark_data_for_marks_card']:
                            student_row_data['marks' + str(subject) + str(cum_num)] = marks['mark']
                            cum_num += 1
                        student_row_data[str(subject) + 'written_marks'] = subjects['marks'] if 'marks' in subjects else ''
                elif institute_data['code'] == 'jnanajyothi' and subjects['subject_part_type'] == 'Part 1':
                    subject = subjects['subject']
                    attendance_status = subjects.get('attendance_status')
                    subject_marks_obtained = 'subject_' + str(subject) + '_obtained_marks'
                    subject_grade_obtained = 'subject_' + str(subject) + '_obtained_grade'
                    student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks']
                    student_row_data[subject_grade_obtained] = subjects['grade']
                    cum_num = 0
                    if subjects['cumulative_mark_data_for_marks_card']:
                        is_cum_type = 1
                        cum_marks = subjects['cumulative_mark_data_for_marks_card']
                        for marks in subjects['cumulative_mark_data_for_marks_card']:
                            student_row_data['marks' + str(subject) + str(cum_num)] = marks['mark']
                            cum_num += 1
                        student_row_data[str(subject) + 'written_marks'] = 'Ab' if attendance_status == 'Absent' else subjects['marks'] if 'marks' in subjects else ''
                else:
                    subject = subjects['subject']
                    attendance_status = subjects.get('attendance_status')
                    subject_marks_obtained = 'subject_' + str(subject) + '_obtained_marks'
                    subject_grade_obtained = 'subject_' + str(subject) + '_obtained_grade'
                    student_row_data[subject_marks_obtained] = 'Ab' if attendance_status == 'Absent' else subjects['total_obtained_marks']
                    student_row_data[subject_grade_obtained] = subjects['grade']
                    cum_num = 0
                    if subjects.get('cumulative_mark_data_for_marks_card'):
                        is_cum_type = 1
                        cum_marks = subjects['cumulative_mark_data_for_marks_card']
                        if 'cumulative_marks_data' in subjects and subjects['cumulative_marks_data']:
                            for marks in subjects['cumulative_marks_data']:
                                if marks.get('cumulative_data_mapping') and len(marks['cumulative_data_mapping']) > 0:
                                    student_row_data['marks' + str(subject) + str(marks['cumulative_data_mapping'][0]['cumulative_type_id'])] = marks.get('marks', '')
                                cum_num += 1
                        student_row_data[str(subject) + 'written_marks'] = 'Ab' if attendance_status == 'Absent' else subjects.get('marks', '') if 'marks' in subjects else ''                 
            student_row_data['section_rank'] = student['section_rank'] if 'section_rank' in student and student['section_rank'] else None
            student_row_data['standard_rank'] = student['standard_rank'] if 'standard_rank' in student and student['standard_rank'] else None
            student_row_data['total_obtained_marks'] = student['part_type_data']['Part 1']['total_obtained_marks']
            student_row_data['total_obtained_grade'] = student['part_type_data']['Part 1']['grade']
            student_row_data['total_obtained_percentage'] = student['part_type_data']['Part 1']['percentage']
            student_row_data['total_obtained_percentage'] = round(student_row_data['total_obtained_percentage'], 2)
            student_row_data['remark'] = student['remark_name'] if 'remark_name' in student else None
            student_row_data['marked_attendance_days'] = student['marked_attendance_days'] if 'marked_attendance_days' in student else ''       
            try:
                total_obtained_marks += student_row_data['total_obtained_marks']
                total_max_marks += student['part_type_data']['Part 1']['total_marks']
                total_percentage = (total_obtained_marks / total_max_marks) * 100
            except:
                pass         
            consolidated_data.append(student_row_data)

    consolidated_data.sort(key=lambda x: (x['standard_rank'] is None, x['standard_rank'] if x['standard_rank'] is not None else 0))
    first_section_data = {'data': all_sections_data[0], 'institute_data': institute_data} 
    options = {}
    options['title'] = 'Consolidated_marks'
    options['description'] = 'marks'
    options['examname'] = all_sections_data[0]['exam_details']
    options['standardname'] = standard_name
    options['sectionname'] = 'All Sections'
    options['institute_name'] = institute_data['name']
    options['institute_code'] = institute_data['code']
    options['extraWorksheet'] = False
    options['Data'] = consolidated_data
    options['total_summary_details'] = [
        {'name': 'Total Obtained Marks', 'value': total_obtained_marks},
        {'name': 'Total Maximum Marks', 'value': total_max_marks},
        {'name': 'Total Percentage', 'value': total_percentage}
    ]   
    if institute_data['code'] == "sbvshr":
        options['total_summary_data'] = [
            {'column': '', 'required': False, 'schemacolumn': 'name'},
            {'column': '', 'required': False, 'schemacolumn': 'value'}
        ]  
        def assign_summary_percentage(subject_marks_summary):
            class_efficiency = 0.0
            total_max_marks = 0
            total_obtained_marks = 0.0          
            for subject_name, sub_value in subject_marks_summary.items():
                if not isinstance(sub_value, dict):
                    continue
                max_marks = sub_value.get('max_marks_conducted', 0)
                obtained = sub_value.get('total_obtained_marks', 0)
                total_max_marks += max_marks
                total_obtained_marks += obtained
                if max_marks > 0:
                    sub_value['subject_percentage'] = (obtained / max_marks) * 100
                else:
                    sub_value['subject_percentage'] = 0.0                   
            if total_max_marks > 0:
                class_efficiency = (total_obtained_marks / total_max_marks) * 100
            return subject_marks_summary, class_efficiency       
        def assign_subject_ranks(subject_marks_summary):
            subject_percentages = {
                subject: summary.get('subject_percentage', 0)
                for subject, summary in subject_marks_summary.items()
            }
            sorted_subjects = sorted(subject_percentages.items(), key=lambda x: x[1], reverse=True)
            rank = 1
            last_percentage = None
            rank_map = {}
            
            for i, (subject, percentage) in enumerate(sorted_subjects, start=1):
                if percentage != last_percentage:
                    rank = i
                rank_map[subject] = rank
                last_percentage = percentage
            
            for subject, summary in subject_marks_summary.items():
                summary['subject_rank'] = rank_map[subject]
            
            return subject_marks_summary  
        subject_marks_summary, class_eff = assign_summary_percentage(subject_marks_summary)
        subject_marks_summary = assign_subject_ranks(subject_marks_summary)
        subject_marks_summary['class_efficiency'] = class_eff
        subject_marks_summary['subject_total_grade'] = first_section_data['data'].get('grade_plan_data_for_total', '')
        options['subject_marks_summary'] = subject_marks_summary        
    options['extraWorksheetData'] = dict()

    columns = [{'column': 'SECTION', 'required': False, 'schemacolumn': 'section'}]
    columns.extend(json_for_consolidated_marks(first_section_data, is_cum_type, cum_marks))
    options['columns'] = columns
    
    return write_to_excel_new_consolidation(self, options, {}, {})

def json_for_consolidated_marks_subject_wise(data,is_cum_type,cum_marks):
    inst_obj = Institute.objects.all().first()
    column_data=[
        {
            'column': 'SL No', 'required': False, 'schemacolumn': 'sl_no'
        },
        {
            'column': 'NAME OF THE STUDENT', 'required': False, 'schemacolumn': 'student_name'
        },
        {
            'column':'THEORY', 'required' : False, 'schemacolumn' : 'written_marks'
        },
        {
            'column':'INTERNAL', 'required' : False, 'schemacolumn' : 'cum_marks',
        },
        {
            'column':'TOTAL', 'required' : False, 'schemacolumn' : 'subject_obtained_marks',
        },
        {
            'column':'PERCENTAGE 100%', 'required' : False, 'schemacolumn' : 'subject_obtained_percentage',
        },
        {
            'column':'GRADE', 'required' : False, 'schemacolumn' : 'subject_obtained_grade',
        }]
    return column_data

def json_for_consolidated_marks(data,is_cum_type,cum_marks):
    inst_obj = Institute.objects.all().first()
    if inst_obj.code == 'jnanajyothi':
        column_data_jnanajyothi=[
            {
                'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
            }]
        cum_num=0
        for subjects in data['data']['subject_list']:
            if subjects['subject_part_type'] == 'Part 1':
                subject=subjects['subject']
                key_written_marks=str(subject)+'written_marks'                             
                key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
                key_grade = 'subject_'+str(subjects['subject'])+'_obtained_grade'
                cum_num=0
                if is_cum_type==1:
                    cumulative_data_length = len(subjects['cumulative_data']) if 'cumulative_data' in subjects else 0
                    for index,marks in enumerate(cum_marks):
                        key_cum='marks'+str(subjects['subject'])+str(cum_num)
                        column_data_jnanajyothi.append(
                        {
                            'column':'A1' if index==0 else 'A2', 'required' : False, 'schemacolumn' : key_cum,
                            'parent':{'schemacolumn':key_cum, 'column': subjects['subject_name'], 'number_of_cells': cumulative_data_length+3 if index==0 else 0}
                        }
                        )
                        cum_num+=1
                    column_data_jnanajyothi.append(
                        {
                            'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                            'parent':{'schemacolumn':'written_marks', 'column': 'written_marks', 'number_of_cells': 0 }
                        }
                        )
                column_data_jnanajyothi.append(
                    {
                    'column':'TOT', 'required' : False, 'schemacolumn' : key_marks,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'], 'number_of_cells': 0 if is_cum_type else 2 }
                    }
                    )
                column_data_jnanajyothi.append(
                    {
                    'column':'GRD', 'required' : False, 'schemacolumn' : key_grade,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'] , 'number_of_cells': 0}
                    }
                    )
        column_data_jnanajyothi.append(
            {
                'column':'TOT', 'required':False,'schemacolumn':'total_obtained_marks'
            })
        column_data_jnanajyothi.append({
                'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
            }
        )
        column_data_jnanajyothi.append({
                'column':'GRD' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
            }
        )
        column_data_jnanajyothi.append({
                'column':'Attendance' , 'required':False, 'schemacolumn' : 'marked_attendance_days'
            }
        )
        column_data_jnanajyothi.append({
                'column':'Remark' , 'required':False, 'schemacolumn' : 'remark'
            }
        )
        return column_data_jnanajyothi
    if inst_obj.code == 'lourdes':
        column_data_lourdes=[
            {
                'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
            }]
        
        cum_num=0
        for subjects in data['data']['subject_list']:
            subject=subjects['subject']
            key_written_marks=str(subject)+'written_marks'
            key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
            cum_num=0
            column_data_lourdes.append(
                {
                'column':'TOT', 'required' : False, 'schemacolumn' : key_marks,
                'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'], 'number_of_cells':1 }
                }
                )
        column_data_lourdes.append(
            {
                'column':'TOT', 'required':False,'schemacolumn':'total_obtained_marks'
            })
        column_data_lourdes.append({
                'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
            }
        )
        column_data_lourdes.append({
                'column':'GRD' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
            }
        )
        column_data_lourdes.append({
                'column':'Attendance' , 'required':False, 'schemacolumn' : 'marked_attendance_days'
            }
        )
        column_data_lourdes.append({
                    'column':'Standard Rank' , 'required':False, 'schemacolumn' : 'standard_rank'
                }
            )
        column_data_lourdes.append({
                    'column':'Section Rank' , 'required':False, 'schemacolumn' : 'section_rank'
                }
            )
        column_data_lourdes.append({
                'column':'Remark' , 'required':False, 'schemacolumn' : 'remark'
            }
        )
        return column_data_lourdes
    if inst_obj.code == 'nisarga':
        column_data_nisarga=[
            {
                'column': 'SI', 'required': False, 'schemacolumn': 'sl_no'
            },
            {
                'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
            }]
        cum_num=0
        for subjects in data['data']['subject_list']:
            if subjects['subject_part_type'] == 'Part 1':
                subject=subjects['subject']
                key_written_marks=str(subject)+'written_marks'
                key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
                key_grade = 'subject_'+str(subjects['subject'])+'_obtained_grade'
                cum_num=0
                if is_cum_type==1:
                    cumulative_data_length = len(subjects['cumulative_data']) if 'cumulative_data' in subjects else 0
                    for index,marks in enumerate(cum_marks):
                        key_cum='marks'+str(subjects['subject'])+str(cum_num)
                        column_data_nisarga.append(
                        {
                            'column':'A1' if index==0 else 'A2', 'required' : False, 'schemacolumn' : key_cum,
                            'parent':{'schemacolumn':key_cum, 'column': subjects['subject_name'], 'number_of_cells': cumulative_data_length+3 if index==0 else 0}
                        }
                        )
                        cum_num+=1
                    column_data_nisarga.append(
                        {
                            'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                            'parent':{'schemacolumn':'written_marks', 'column': 'written_marks', 'number_of_cells': 0 }
                        }
                        )
                column_data_nisarga.append(
                    {
                    'column':'M', 'required' : False, 'schemacolumn' : key_marks,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'], 'number_of_cells': 0 if is_cum_type else 2 }
                    }
                    )
                column_data_nisarga.append(
                    {
                    'column':'G', 'required' : False, 'schemacolumn' : key_grade,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'] , 'number_of_cells': 0}
                    }
                    )
        column_data_nisarga.append(
            {
                'column':'Total', 'required':False,'schemacolumn':'total_obtained_marks'
            })
        column_data_nisarga.append({
                'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
            }
        )
        column_data_nisarga.append({
                'column':'Grade' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
            }
        )  
        return column_data_nisarga
    if inst_obj.code == 'sbvshr':
        column_data_sbvshr = [
        {
            'column': 'SI', 'required': False, 'schemacolumn': 'sl_no'
        },
        {
            'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
        }]
        cum_num=0
        for subjects in data['data']['subject_list']:
            # if subjects['subject_part_type'] == 'Part 1':
                subject=subjects['subject']
                key_written_marks=str(subject)+'written_marks'
                key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
                key_grade = 'subject_'+str(subjects['subject'])+'_obtained_grade'
                cum_num=0
                if is_cum_type==1:
                    cumulative_data_length = len(subjects['cumulative_data']) if 'cumulative_data' in subjects else 0
                    for index,marks in enumerate(cum_marks):
                        key_cum='marks'+str(subjects['subject'])+str(cum_num)
                        column_data_sbvshr.append(
                        {
                            'column':'I' if index==0 else 'A2', 'required' : False, 'schemacolumn' : key_cum,
                            'parent':{'schemacolumn':key_cum, 'column': subjects['subject_name'], 'number_of_cells': cumulative_data_length+3 if index==0 else 0}
                        }
                        )
                        cum_num+=1
                    column_data_sbvshr.append(
                        {
                            'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                            'parent':{'schemacolumn':'written_marks', 'column': 'written_marks', 'number_of_cells': 0 }
                        }
                        )
                column_data_sbvshr.append(
                    {
                    'column':'M', 'required' : False, 'schemacolumn' : key_marks,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'], 'number_of_cells': 0 if is_cum_type else 2 }
                    }
                    )
                column_data_sbvshr.append(
                    {
                    'column':'G', 'required' : False, 'schemacolumn' : key_grade,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'] , 'number_of_cells': 0}
                    }
                    )
        column_data_sbvshr.append(
            {
                'column':'Total', 'required':False,'schemacolumn':'total_obtained_marks'
            })
        column_data_sbvshr.append({
                'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
            }
        )
        column_data_sbvshr.append({
                'column':'Grade' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
            }
        )
        column_data_sbvshr.append({
                'column':'Section Rank' , 'required':False, 'schemacolumn' : 'section_rank'
            }
        )
        column_data_sbvshr.append({
                'column':'Standard Rank' , 'required':False, 'schemacolumn' : 'standard_rank'
            }
        )
        return column_data_sbvshr

    if inst_obj.code == 'springwell':
        column_data_springwell = [
            {
                'column': 'SI', 'required': False, 'schemacolumn': 'sl_no'
            },
            {
                'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
            }]
        cum_num=0
        for subjects in data['data']['subject_list']:
            if subjects['subject_part_type'] == 'Part 1':
                subject=subjects['subject']
                key_written_marks=str(subject)+'written_marks'
                key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
                key_grade = 'subject_'+str(subjects['subject'])+'_obtained_grade'
                cum_num=0
                if is_cum_type==1:
                    cumulative_data_length = len(subjects['cumulative_data']) if 'cumulative_data' in subjects else 0
                    for index,marks in enumerate(cum_marks):
                        key_cum='marks'+str(subjects['subject'])+str(cum_num)
                        column_data_springwell.append(
                        {
                            'column':'P' if index==0 else 'I', 'required' : False, 'schemacolumn' : key_cum,
                            'parent':{'schemacolumn':key_cum, 'column': subjects['subject_name'], 'number_of_cells': cumulative_data_length+3 if index==0 else 0}
                        }
                        )
                        cum_num+=1
                    column_data_springwell.append(
                        {
                            'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                            'parent':{'schemacolumn':'written_marks', 'column': 'written_marks', 'number_of_cells': 0 }
                        }
                        )
                column_data_springwell.append(
                    {
                    'column':'T', 'required' : False, 'schemacolumn' : key_marks,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'], 'number_of_cells': 0 if is_cum_type else 2 }
                    }
                    )
                column_data_springwell.append(
                    {
                    'column':'G', 'required' : False, 'schemacolumn' : key_grade,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'] , 'number_of_cells': 0}
                    }
                    )
        column_data_springwell.append(
            {
                'column':'Total', 'required':False,'schemacolumn':'total_obtained_marks'
            })
        column_data_springwell.append({
                'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
            }
        )
        column_data_springwell.append({
                'column':'Grade' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
            }
        )
        return column_data_springwell
    else:
        column_data_sbvshr = [
        {
            'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
        }]
        cum_num=0
        for subjects in data['data']['subject_list']:
            subject=subjects['subject']
            key_written_marks=str(subject)+'written_marks'
            key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
            key_grade = 'subject_'+str(subjects['subject'])+'_obtained_grade'
            cum_num=0
            if is_cum_type==1:
                cumulative_data_length = len(subjects['cumulative_data']) if 'cumulative_data' in subjects else 0
                if 'cumulative_data' in subjects and subjects['cumulative_data'] and subjects['is_marks']:
                    column_data_sbvshr.append(
                    {
                        'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                        'parent':{'schemacolumn':'written_marks', 'column': subjects['subject_name'], 'number_of_cells': cumulative_data_length+3}
                    }
                    )
                    for index,marks in enumerate(subjects['cumulative_data']):
                        key_cum='marks'+str(subjects['subject'])+str(marks['cumulative_type_data'][0]['id'])
                        column_data_sbvshr.append(
                        {
                            'column':subjects['cumulative_data'][index]['cumulative_type_data'][0]['alias'], 'required' : False, 'schemacolumn' : key_cum,
                            'parent':{'schemacolumn':key_cum, 'column': subjects['subject_name'], 'number_of_cells': 0}
                        }
                        )
                        cum_num+=1
                elif ('cumulative_data' not in subjects or not subjects['cumulative_data']) and subjects['is_marks']:
                    column_data_sbvshr.append(
                    {
                        'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                        'parent':{'schemacolumn':'written_marks', 'column': subjects['subject_name'], 'number_of_cells': 1 }
                    }
                    )
            if subjects['is_marks']:
                column_data_sbvshr.append(
                    {
                    'column':'M', 'required' : False, 'schemacolumn' : key_marks,
                    'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'], 'number_of_cells': 0 if is_cum_type else 2 }
                    }
                    )
            column_data_sbvshr.append(
                {
                'column':'G', 'required' : False, 'schemacolumn' : key_grade,
                'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject_name'] , 'number_of_cells': 0 if subjects['is_marks'] else 1}
                }
                )
        column_data_sbvshr.append(
            {
                'column':'Total', 'required':False,'schemacolumn':'total_obtained_marks'
            })
        column_data_sbvshr.append({
                'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
            }
        )
        column_data_sbvshr.append({
                'column':'Grade' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
            }
        )
        return column_data_sbvshr

def json_for_term_consolidated_marks(data,is_cum_type,cum_marks):
    inst_obj = Institute.objects.all().first()
    column_data=[
        {
            'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
        }]
    for subjects in data['data']['subject_list']:
        if subjects['subject_part_type'] == 'Part 1':
            subject=subjects['subject']
            key_written_marks='subject_'+str(subjects['subject'])+'_written_marks'
            key_marks = 'subject_'+str(subjects['subject'])+'_obtained_marks'
            key_grade = 'subject_'+str(subjects['subject'])+'_obtained_grade'
            key_config = 'subject_'+str(subjects['subject'])+'_config_marks'
            column_data.append(
            {
                    'column':'CONF', 'required' : False, 'schemacolumn' : key_config,
                    'parent':{'schemacolumn':key_config, 'column': subjects['subject__name'], 'number_of_cells': 4}
            }
            )
            column_data.append(
            {
                'column':'W', 'required' : False, 'schemacolumn' : key_written_marks,
                'parent':{'schemacolumn':'written_marks', 'column': 'written_marks', 'number_of_cells': 0 }
            }
            )
            column_data.append(
                {
                'column':'TOT', 'required' : False, 'schemacolumn' : key_marks,
                'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject__name'], 'number_of_cells': 0}
                }
                )
            column_data.append(
                {
                'column':'GRD', 'required' : False, 'schemacolumn' : key_grade,
                'parent':{'schemacolumn':f'parent_{subject}', 'column': subjects['subject__name'] , 'number_of_cells': 0}
                }
                )
    column_data.append(
        {
            'column':'TOT', 'required':False,'schemacolumn':'total_obtained_marks'
        })
    column_data.append({
            'column':'%' , 'required':False, 'schemacolumn' : 'total_obtained_percentage'
        }
    )
    column_data.append({
            'column':'GRD' , 'required':False, 'schemacolumn' : 'total_obtained_grade'
        }
    )
    return column_data

def get_marks_card_for_config(self,request):
    response=get_marks_for_config(self,request)
    standard = response['data']['standard']
    academic_year = response['data']['academic_year_details']['academic_year']
    selected_template, number_of_copies = get_selected_template(self, 'marks_card', 'pdf', 'jnana_jyothi_marks_card_term1.html', academic_year, [standard])
    selected_template='jnana_jyothi_marks_card_term1.html'
    path = 'marks_card/'+selected_template
    response['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
    #from django.shortcuts import render
    #return render(self.request, path, response)
    print('data',response)
    response = PDFService.receipt_new(self, response, "marks_card", path, False)
    return response

def get_consolidated_report_for_config(self,request):
    data=get_marks_for_config(self,request)
    data['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
    consolidated_data=[]
    dyanmic_labels = {}
    is_cum_type=0
    cum_marks=None
    for students in data['data']['student_list']:
        student_name = students['student_name']
        student_row_data = {'student_name': student_name}
        for subjects in students['subject_list_data']:
            if subjects['subject_part_type'] == 'Part 1':
                subject = subjects['subject']
                subject_marks_obtained = 'subject_'+str(subject)+'_obtained_marks'
                subject_grade_obtained='subject_'+str(subject)+'_obtained_grade'
                subject_config_marks_obtained='subject_'+str(subject)+'_config_marks'
                subject_written_marks_obtained='subject_'+str(subject)+'_written_marks'
                student_row_data[subject_marks_obtained] = subjects['marks_round_off'] if 'marks_round_off' in subjects and subjects['marks_round_off'] else 0
                student_row_data[subject_grade_obtained] = subjects['grade_round_off'] if 'grade_round_off' in subjects else ''
                student_row_data[subject_config_marks_obtained] = subjects['config_marks_of_test']
                student_row_data[subject_written_marks_obtained] = subjects['config_marks_of_exam'] if 'config_marks_of_exam' in subjects else 0
        student_row_data['total_obtained_marks'] = students['part1_obtained_marks_round_off']
        student_row_data['total_obtained_grade'] = students['part1_grade_round_off'] 
        student_row_data['total_obtained_percentage'] = students['part1_total_percentage_round_off']
        student_row_data['total_obtained_percentage'] = round(student_row_data['total_obtained_percentage'],2)
        consolidated_data.append(student_row_data)
    multiple_data = []
    options={}
    options['title'] = 'Consolidated_marks'
    options['description'] = 'marks'
    options['examname']= data['data']['term_alias_name']
    options['standardname']=data['data']['standard_name']
    options['sectionname']=data['data']['section_name']
    options['institute_name']=data['institute_data']['name']
    options['extraWorksheet'] = False
    options['Data'] = consolidated_data
    options['extraWorksheetData'] = dict()
    options['columns'] = json_for_term_consolidated_marks(data,is_cum_type,cum_marks)
    return write_to_excel_new_consolidation(self, options, {},{})

def json_for_excel_amrita(data):
    column_data_sad=[
        {
            'column': 'SL No', 'required': False, 'schemacolumn': 'sl_no'
        },
        {
            'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
        }]
    for exam in data['details_for_heading']['internal_details']:
        for index,co in enumerate(data['details_for_heading']['internal_details'][exam]):
            if co != 'exam_name':
                column_data_sad.append(
                {
                    'column':data['details_for_heading']['internal_details'][exam][co]['course_outcome_name'], 'required' : False, 'schemacolumn' : str(exam)+'_'+str(co),
                    'parent':{'schemacolumn':str(exam)+'_'+str(co), 'column': data['details_for_heading']['internal_details'][exam]['exam_name'], 'number_of_cells': len(data['details_for_heading']['internal_details'][exam])-1 if index == 1 else 0}
                }
                )
    column_data_sad.append({
        'column':'Total Internal', 'required' : False, 'schemacolumn' : 'total_internal',
        'parent':{'schemacolumn':'total_internal', 'column': 'External Exam', 'number_of_cells': 1}
    })
    column_data_da=[
        {
            'column': 'SL No', 'required': False, 'schemacolumn': 'sl_no'
        },
        {
            'column': 'STUDENT NAME', 'required': False, 'schemacolumn': 'student_name'
        }]
    for exam in data['details_for_heading']['co_details']:
        for index,co in enumerate(data['details_for_heading']['co_details'][exam]):
            column_data_da.append(
            {
                'column':data['details_for_heading']['co_details'][exam][co]['course_outcome_name'], 'required' : False, 'schemacolumn' : str(exam)+str(co),
                'parent':{'schemacolumn':str(exam)+str(co), 'column': data['details_for_heading']['co_details'][exam][co]['max_marks'], 'number_of_cells': 1}
            }
            )
    column_data_da2=[]
    for index,exam in enumerate(data['level']):
        if index == 0:
            for data1 in exam:
                column_data_da2.append(
                {
                    'column':'', 'required' : False, 'schemacolumn' : data1,
                }
                )
    column_data_da3=[
    {
        'column':'', 'required' : False, 'schemacolumn' : 'name',
    },
    {
        'column':'', 'required' : False, 'schemacolumn' : 'value',
    }
    ]
    column_data_ada=[
    {
        'column':'CO', 'required' : False, 'schemacolumn' : 'co_name',
    },
    {
        'column':'X1', 'required' : False, 'schemacolumn' : 'co_ia',
        'parent':{'schemacolumn':'co_ia', 'column': 'CO Wise IA Average', 'number_of_cells': 1}
    },
    {
        'column':'X2', 'required' : False, 'schemacolumn' : 'co_ass',
        'parent':{'schemacolumn':'co_ass', 'column': 'CO Wise Assessment Average', 'number_of_cells': 1}
    },
    {
        'column':'X3', 'required' : False, 'schemacolumn' : 'see',
        'parent':{'schemacolumn':'see', 'column': 'SEE', 'number_of_cells': 1}
    },
    {
        'column':'(X1+X2)/2', 'required' : False, 'schemacolumn' : 'cie_avg',
        'parent':{'schemacolumn':'cie_avg', 'column': 'CIE Average', 'number_of_cells': 1}
    },
    {
        'column':'', 'required' : False, 'schemacolumn' : 'direct',
        'parent':{'schemacolumn':'direct', 'column': 'Direct Attainment', 'number_of_cells': 1}
    }]
    column_data_co_attainment=[
    {
        'column':'CO', 'required' : False, 'schemacolumn' : 'co_name',
    },
    {
        'column':'Direct Attainment', 'required' : False, 'schemacolumn' : 'direct',
    },
    {
        'column':'Indirect Attainment', 'Indirect Appointment' : False, 'schemacolumn' : 'indirect',
    },
    {
        'column':'CO Attainment', 'required' : False, 'schemacolumn' : 'co_attainment',
    }
    ]
    column_data_ia=[{
        'column':'CO', 'required' : False, 'schemacolumn' : 'co'},
        {
        'column':'3', 'required' : False, 'schemacolumn' : '3',
        'parent':{'schemacolumn':'3', 'column':"Strongly Agree" , 'number_of_cells': 1}
    },
    {
        'column':'2', 'required' : False, 'schemacolumn' : '2',
        'parent':{'schemacolumn':'2', 'column':"Moderately Agree" , 'number_of_cells': 1}
    },
    {
        'column':'1', 'required' : False, 'schemacolumn' : '1',
        'parent':{'schemacolumn':'1', 'column':"DisAgree" , 'number_of_cells': 1}
    },
    {
        'column':'Indirect Attainment', 'required' : False, 'schemacolumn' : 'IA',
        'parent':{'schemacolumn':'IA', 'column':"Indirect Attainment" , 'number_of_cells': 1}
    },
    ]
    column_data_da3=[
    {
        'column':'', 'required' : False, 'schemacolumn' : 'name',
    },
    {
        'column':'', 'required' : False, 'schemacolumn' : 'value',
    }
    ]
    column_data_matrix=[]
    for column in data['co_po_matrix']:
       column_data_matrix.append(
        {
        'column':column['name'], 'required' : False, 'schemacolumn' : column['value'],
        }
       ) 

    return {'column_data_sad':column_data_sad,'column_data_da':column_data_da,'column_data_da2':column_data_da2,'column_data_da3':column_data_da3,'column_data_ia':column_data_ia,
            'column_data_ada':column_data_ada,'column_data_co_attainment':column_data_co_attainment,'column_data_matrix':column_data_matrix}

def get_marks_card_for_finalconfig(self,request):
    try:
        response=get_marks_for_final_config(self,request)
        standard = response['data']['standard']
        academic_year = response['data']['academic_year_details']['academic_year']

        selected_template, number_of_copies = get_selected_template(self, 'marks_card_config', 'pdf', 'jnana_jyothi_marks_card_term1.html', academic_year, [standard])
        path = 'marks_card_config/'+selected_template
        response['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
        #from django.shortcuts import render
        #return render(self.request, path, response)
        file_path = PDFService.receipt_new(self, response, "marks_card_config", path, False)
        if self.request.GET.get('long_running_process'):
            url = UploadTypeService.upload_local_file(file_path, path='markscard_pdfs')
            if os.path.exists(file_path):
                os.remove(file_path)
            transaction_id = self.request.GET.get('transaction_id')
            store_long_running_process(self, transaction_id, {'url': url})
        else:
            return file_path
    except Exception as e:
        if self.request.GET.get('long_running_process'):
            transaction_id = self.request.GET.get('transaction_id')
            store_long_running_process(self, transaction_id,{'error': e.args[:250]})
        else:
            raise e

def get_consolidated_report_for_finalconfig(self,request):
    data=get_marks_for_final_config(self,request)
    data['institute_data'] = InstituteSerializer(Institute.get_institute(self)).data
    consolidated_data=[]
    dyanmic_labels = {}
    is_cum_type=0
    cum_marks=None
    for students in data['data']['student_list']:
        student_name = students['student_name']
        student_row_data = {'student_name': student_name}
        for subjects in students['subject_list_data']:
            if subjects['subject_part_type'] == 'Part 1':
                subject = subjects['subject']
                subject_marks_obtained = 'subject_'+str(subject)+'_obtained_marks'
                subject_grade_obtained='subject_'+str(subject)+'_obtained_grade'
                subject_config_marks_obtained='subject_'+str(subject)+'_config_marks'
                subject_written_marks_obtained='subject_'+str(subject)+'_written_marks'
                student_row_data[subject_marks_obtained] = subjects['marks_round_off_to_upper_number'] if 'marks_round_off_to_upper_number' in subjects and subjects['marks_round_off_to_upper_number'] else 0
                student_row_data[subject_grade_obtained] = subjects['grade_round_off_to_upper_number'] if 'grade_round_off_to_upper_number' in subjects else ''
                student_row_data[subject_config_marks_obtained] = subjects['config_marks_of_test']
                student_row_data[subject_written_marks_obtained] = subjects['config_marks_of_exam'] if 'config_marks_of_exam' in subjects else 0
        student_row_data['total_obtained_marks'] = students['part1_obtained_marks_round_off']
        student_row_data['total_obtained_grade'] = students['part1_grade_round_off'] 
        student_row_data['total_obtained_percentage'] = students['part1_total_percentage_round_off']
        student_row_data['total_obtained_percentage'] = round(student_row_data['total_obtained_percentage'],2)
        consolidated_data.append(student_row_data)
    multiple_data = []
    options={}
    options['title'] = 'Consolidated_marks'
    options['description'] = 'marks'
    options['examname']= 'Final Exam'
    options['standardname']=data['data']['standard_name']
    options['sectionname']=data['data']['section_name']
    options['institute_name']=data['institute_data']['name']
    options['institute_code'] = data['institute_data']['code']
    options['extraWorksheet'] = False
    options['Data'] = consolidated_data
    options['extraWorksheetData'] = dict()
    # print("options",options)
    options['columns'] = json_for_term_consolidated_marks(data,is_cum_type,cum_marks)
    return write_to_excel_new_consolidation(self, options, {},{})

from apps.exams.models.result import ExamFinalResultConfiguration, ResultSectionMapping
def get_marks_for_config(self, request):
    from apps.exams.services.result import get_result_configuration_mapping
    from apps.exams.models.result import ResultConfiguration
    approvalError = ''
    isapproved = True
    result_config_id = request.GET.get('result_config', None)
    student_ids_param = request.GET.get('student_ids', None)
    if result_config_id:
        result_config_data = ResultConfiguration.objects.filter(id=result_config_id).first()
        termId = result_config_data.term_id
        academicYear = result_config_data.academic_year_id
    else:
        termId = request.GET.get('term', None) 
        academicYear = request.GET.get('academic_year', None)
        result_config_data = ResultConfiguration.objects.filter(term_id=termId,academic_year_id=academicYear).first()
        result_config_id=result_config_data.id

    standardSectionId = request.GET.get('standard_section', None)
    raiseApprovalError = request.GET.get('raise_approval_error', True)
    standard_obj = StandardSectionMapping.objects.get(id=standardSectionId)
    standard_id = standard_obj.standard_id
    if not termId or not academicYear or not standardSectionId:
        raise ValidationError('Term Academic StandardSection should be mandatory')
    examData = Exam.objects.filter(term=termId, academic_year=academicYear).values('id', 'standard_section_ids')
    examIds = []
    for exam in examData:
        standard_section_ids = exam['standard_section_ids'].split(',')
        if str(standardSectionId) in standard_section_ids:
            examIds.append(exam['id'])
    approvalData = StudentMarkSectionWiseApproval.objects.filter(exam__in=examIds, standard_section=standardSectionId,
                                                                 approval_status='1').values_list('exam', flat=True)
    values_list = ['id', 'first_name', 'middle_name', 'last_name', 'current_reg_num','student_name']
    custom_annotate = {'student_name': Concat('first_name', V(' '), 'middle_name',V(' '), 'last_name')}
    
    if set(examIds) - set(approvalData):
        examData = Exam.objects.filter(id__in=list(set(examIds) - set(approvalData))).values(
            'exam_type__name', 'term__name'
        )
        errorData = ','.join((str(e['exam_type__name'])+'('+str(e['term__name'])+')') for e in examData )
    
        if raiseApprovalError:
            raise ValidationError(
                f'{errorData} the given exams Marks of all standard are not yet approved')
        
        else:
            isapproved = False
            approvalError = errorData
    elif student_ids_param:
        student_ids_param = student_ids_param.split(',')
        studentData = Student.objects.filter(
           id__in=student_ids_param
            ).annotate(**custom_annotate).values(*values_list)
          
    else:
        studentData = Student.get_student_for_standard(None, None, [standardSectionId], values_list, custom_annotate)
    student_ids = []
    studentIdDict = {}
    configuredSubjectIds = []
    for i in studentData:
        student_ids.append(i['id'])
        i['student'] = i['id']
        studentIdDict[i['id']] = i
    sectionResultConfigData = get_result_configuration_mapping(result_config_id, termId, academicYear, standardSectionId,
                                                               True)
    for sectionData in sectionResultConfigData['result_section_data']:
        configuredSubjectIds.append(sectionData['subject'])
    studentSubjectData = SubjectStudent.objects.filter(student__in=student_ids,
                                                       academic_year=academicYear,
                                                       subject__in=configuredSubjectIds).values('subject', 'student',
                                                                                                'subject__name','subject__subject_code', subject_part_type_id=F('subject__subject_part_type__id'),
                                                                                                subject_part_type=F('subject__subject_part_type__name'),subject_part_type_code_name=F('subject__subject_part_type__code_name'))
    student_admission_form = get_student_admission_form_details(self, student_ids)
    for student_data in studentData:
        student_data['student_admission_form'] = student_admission_form[student_data['student']]
        student_data['student_admission_form']['admission_num'] =  student_data['student_admission_form']['admission_num'].replace('/', '-')
    studentSubjectMapping = {}
    for studentSubject in studentSubjectData:
        if studentSubject['student'] in studentSubjectMapping:
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
        else:
            studentSubjectMapping[studentSubject['student']] = []
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
    studentMarkData = StudentMark.objects.filter(is_active=True, student__in=student_ids,
                                                 exam_schedule__exam__in=list(examIds)).values(
        'student', 'exam_schedule__exam', 'exam_schedule__exam__exam_type__name','exam_schedule__subject', 'marks', 'exam_schedule__max_marks',
        'attendance_status','exam_schedule','grade','exam_schedule__is_marks'
    )
    scheduleIds=[]
    for schedule in studentMarkData:
        scheduleIds.append(schedule['exam_schedule'])
    filter_query_cumulative = {
        'student__in': student_ids,
        'exam_cumulative__exam_schedule__in': list(scheduleIds),
        'is_active': True
    }
    student_cumulative_data = StudentCumulativeMark.objects.filter(
        **filter_query_cumulative
    ).values(
        'id', 'marks', 'exam_cumulative_id', 'exam_cumulative__exam_schedule', 'student', 'attendance_status', 'exam_cumulative__cumulative_type',
        'exam_cumulative__cumulative_type__name', 'exam_cumulative__max_marks'
    )
    schedule_cumulative_data = ExamScheduleCumulativeMapping.objects.filter(
        exam_schedule__in = scheduleIds,
    ).values(
        'exam_schedule', 'cumulative_type', 'max_marks', 'min_marks', 'cumulative_type__name',
        'cumulative_type__alias', 'id'
    )
    exams_to_merge=[]
    try:
        standardMergeData = ResultConfigurationMerge.objects.get(result=result_config_id,standard_section=standardSectionId)
        exam_to_merge=standardMergeData.exam.all().values('id')
        merge_name=standardMergeData.name.name
    except ResultConfigurationMerge.DoesNotExist:
        standardMergeData = None
        exam_to_merge = []
        merge_name=''
    for exam_id in exam_to_merge:
        exams_to_merge.append(exam_id['id'])
    temp_schedule_cumulative_mapping = {}
    schedule_cumulative_mapping = {}
    for schedule_cumulative in schedule_cumulative_data:
        if schedule_cumulative['id'] not in schedule_cumulative_data:
            temp_schedule_cumulative_mapping[schedule_cumulative['id']] = {
                'exam_schedule' : schedule_cumulative['exam_schedule'],
                'max_marks': schedule_cumulative['max_marks'],
                'min_marks': schedule_cumulative['min_marks'],
                'cumulative_type_data': [],
                'id': schedule_cumulative['id']
            }
        temp_schedule_cumulative_mapping[schedule_cumulative['id']]['cumulative_type_data'].append(
            {
                'id': schedule_cumulative['cumulative_type'],
                'name': schedule_cumulative['cumulative_type__name'],
                'alias': schedule_cumulative['cumulative_type__alias']
            }
        )
    for cum_row in temp_schedule_cumulative_mapping.values():
        if cum_row['exam_schedule'] not in schedule_cumulative_mapping:
            schedule_cumulative_mapping[cum_row['exam_schedule']] = []
        schedule_cumulative_mapping[cum_row['exam_schedule']].append(cum_row)
    student_cumulative_data_temp_mapping = {}
    for student_cum in student_cumulative_data: #this is to remove the duplicates
        if student_cum['id'] not in student_cumulative_data_temp_mapping:
            student_cumulative_data_temp_mapping[student_cum['id']] = {
                'cumulative_data_mapping': [], 'id' : student_cum['id'],
                'marks': student_cum['marks'], 'exam_cumulative_id': student_cum['exam_cumulative_id'],
                'exam_cumulative__exam_schedule': student_cum['exam_cumulative__exam_schedule'], 'student': student_cum['student'],
                'attendance_status': student_cum['attendance_status'], 'exam_cumulative__max_marks': student_cum['exam_cumulative__max_marks']
            }
        student_cumulative_data_temp_mapping[student_cum['id']]['cumulative_data_mapping'].append({
            'cumulative_type_id': student_cum['exam_cumulative__cumulative_type'],
            'cumulative_type_name': student_cum['exam_cumulative__cumulative_type__name'],
        })
    student_cumulative_data_mapping = {}
    for cum in student_cumulative_data_temp_mapping.values():
        if cum['exam_cumulative__exam_schedule'] not in student_cumulative_data_mapping:
            student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']] = {cum['student']: []}
        elif cum['student'] not in student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']]:
            student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']] = []
        if cum['exam_cumulative__exam_schedule'] not in schedule_cumulative_mapping:
            schedule_cumulative_mapping[cum['exam_cumulative__exam_schedule']] = {}
        student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']].append(cum)
        cum_marks=student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']]
    calculateMarksForStudent = calculate_mark_for_student(self,studentData, studentSubjectMapping, studentMarkData,student_cumulative_data,
                                                          sectionResultConfigData, academicYear,standardSectionId,student_cumulative_data_temp_mapping,exams_to_merge)
    subjectList = {}
    for index, student in enumerate(calculateMarksForStudent['studentData']):
        calculateMarksForStudent['studentData'][index]['subject_list'] = {s['subject']: s for s in student['subject_list']}
        subjectList.update(calculateMarksForStudent['studentData'][index]['subject_list'])
        calculateMarksForStudent['studentData'][index]['subject_list_data'] = sorted(student['subject_list'].values(), key=lambda v: (isinstance(v.get('subject__subject_code', "NA"), str), v.get('subject__subject_code', "NA"))) #used for printing marks card
        part1_sequence=0
        part2_sequence=0
        for subject in calculateMarksForStudent['studentData'][index]['subject_list_data']:
            if subject['subject_part_type_code_name'] == 'part1':
                part1_sequence+=1
            else:
                part2_sequence+=1
            subject['part1_sequence'] = part1_sequence
            subject['part2_sequence'] = part2_sequence
    subjectList = subjectList.values()
    academic_year_obj = AcademicYear.objects.get(id=academicYear)
    del sectionResultConfigData['result_section_data']
    return {'data': {
                    'student_list': calculateMarksForStudent['studentData'],
                    'term_details': ExamTerm.objects.get(id=termId).name,
                    'term_alias_name': ExamTerm.objects.get(id=termId).alias_name,
                    'standard_name': standard_obj.standard.name,
                    'standard' : standard_id,
                    'section_name': standard_obj.section.name,
                    'merge_name': merge_name,
                    'academic_year_details': {'start_date': academic_year_obj.start_date, 'end_date': academic_year_obj.end_date,'academic_year':academicYear},
                    'configuration_details': sectionResultConfigData,
                    'subject_list': subjectList, 'is_announced': calculateMarksForStudent['isAnnounced']},
                    'isapproved': isapproved, 'approvalError': approvalError,
                    'part_type_list': SubjectPartType.objects.all().values()
                    }
import math
from apps.exams.models.result import ExamFinalResultConfiguration, ResultSectionApproval,ResultSectionMapping
def calculate_mark_for_student(self,studentData, studentSubjectMapping, studentMarkData,student_cumulative_data, configurationData, academicYear,standardSectionId,student_cumulative_data_temp_mapping,exams_to_merge):
    sectionConfiguration = configurationData['result_section_data']
    subjectConfigurationMapping = {}
    studentSubjectSchedule = {}
    studentIds = []
    studentFinalResult = {}
    studentExamFinalList = []
    isAnnounced = False
    # nikhil
    gradeData = []
    totalgrade =[]
    # gradeData = Grade.objects.filter(academic_year=academicYear).values()
    for configData in sectionConfiguration:
        if configData['subject'] not in subjectConfigurationMapping:
            subjectConfigurationMapping[configData['subject']] = {}
        subjectConfigurationMapping[configData['subject']] = configData
    for markData in studentMarkData:
        studentIds.append(markData['student'])
        if markData['student'] not in studentSubjectSchedule:
            studentSubjectSchedule[markData['student']] = {}
        if markData['exam_schedule__subject'] not in studentSubjectSchedule[markData['student']]:
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']] = {}
        if markData['exam_schedule__exam'] not in studentSubjectSchedule[markData['student']]:
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
                markData['exam_schedule__subject']] = {}
        cum_total_marks=0
        cum_total_max_marks=0
        cum_data_list=[]
        for cum_data in student_cumulative_data_temp_mapping:
            if student_cumulative_data_temp_mapping[cum_data]['student'] == markData['student'] and \
            student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__exam_schedule'] == markData['exam_schedule']:
                cum_id=str(student_cumulative_data_temp_mapping[cum_data]['cumulative_data_mapping'][0]['cumulative_type_id'])
                temp_cum_data = [{
                    'cum_id': student_cumulative_data_temp_mapping[cum_data]['exam_cumulative_id'],
                    'cum_marks': student_cumulative_data_temp_mapping[cum_data]['marks'],
                    'cum_max_marks':student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__max_marks']
                }]
                cum_data_list.append(temp_cum_data)
                cum_total_marks += student_cumulative_data_temp_mapping[cum_data]['marks']
                cum_total_max_marks += student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__max_marks']
        temp_max_marks = markData['exam_schedule__max_marks'] if markData['exam_schedule__max_marks'] else 0
        total_max_marks=cum_total_max_marks+temp_max_marks
        if markData['attendance_status']=='Present' and markData['exam_schedule__is_marks']:
            total_marks=cum_total_marks+markData['marks']
        else: # grade sub or absent in marks sub or absent in grade sub
            total_marks=cum_total_marks+0
        studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
            markData['exam_schedule__exam']] = {
            'marks': markData['marks'], 'max_marks': markData['exam_schedule__max_marks'],
            'attendance_status': markData['attendance_status'],'grade':markData['grade'],'exam':markData['exam_schedule__exam'],'cum_data':cum_data_list,'total_marks':total_marks,'total_max_marks':total_max_marks
        }
    if sectionConfiguration:
        studentExamFinalList = StudentExamFinalResult.objects.filter(student__in=list(set(studentIds)),
                                                                     result_config=configurationData['id']).values()
    for finalData in studentExamFinalList:
        studentFinalResult[finalData['student_id']] = finalData
        if finalData['is_announced']:
            isAnnounced = True
    updated_student_data = []
    for index, std_tmp in enumerate(studentData):
        student = copy.deepcopy(std_tmp)
        student['subject_list'] = []
        student['total_result'] = ''
        student['total_marks'] = 0
        student['obtained_marks'] = 0
        student['obtained_marks_round_off'] = 0
        student['part1_obtained_marks'] = 0
        student['part1_obtained_marks_round_off'] = 0
        student['part1_total_marks'] = 0
        student['part1_total_config_marks_of_test'] = 0
        student['part1_total_config_max_marks_of_test'] = 0
        student['total_config_marks_of_test'] = 0
        student['total_config_max_marks_of_test'] = 0
        student['grade'] = ''
        student['total_percentage'] = 0
        student['part1_total_config_max_marks_of_exam'] = 0
        student['part1_total_config_marks_of_exam'] = 0
        config_marks_of_test=0
        config_max_marks_of_test=0
        config_marks_of_exam=0
        config_max_marks_of_exam=0
        if student['id'] in studentSubjectMapping:
            student['subject_list'] = studentSubjectMapping[student['id']]
        for index1, subjectData in enumerate(student['subject_list']):
            subject = subjectData['subject']
            if subject in subjectConfigurationMapping:
                student['subject_list'][index1]['id'] = subject
                if not subjectConfigurationMapping[subject]['max_marks']:
                    subjectConfigurationMapping[subject]['max_marks']=0
                if not subjectConfigurationMapping[subject]['min_marks']:
                    subjectConfigurationMapping[subject]['min_marks']=0
                student['subject_list'][index1]['max_marks'] = subjectConfigurationMapping[subject][
                    'max_marks']
                student['total_marks'] += subjectConfigurationMapping[subject][
                    'max_marks']
                student['subject_list'][index1]['min_marks'] = subjectConfigurationMapping[subject][
                    'min_marks']
                student['subject_list'][index1]['result'] = 'fail'
                student['subject_list'][index1]['marks'] = 0
                student['subject_list'][index1]['attendance_status'] = 'Present'
                student['subject_list'][index1]['unattended_schedule_marks'] = []
                tempTotalObtainedMarks = 0
                config_marks_of_test=0
                config_max_marks_of_test=0
                config_marks_of_exam=0
                config_max_marks_of_exam=0
                for index2,examSubject in enumerate(subjectConfigurationMapping[subject]['subject_exam_data']):
                    if examSubject['is_only_grade_for_config']:
                        examSubject['marks']
                    if student['id'] in studentSubjectSchedule:
                        if subject not in studentSubjectSchedule[student['id']] or examSubject['exam'] not in \
                                studentSubjectSchedule[student['id']][subject]:
                            student['subject_list'][index1]['unattended_schedule_marks'].append(
                                {'subject': examSubject['exam'], 'exam': examSubject, }
                            )
                            continue
                        orignalMarkData = studentSubjectSchedule[student['id']][subject][examSubject['exam']]
                        if not orignalMarkData or orignalMarkData['attendance_status'] == 'Absent':
                            student['subject_list'][index1]['attendance_status'] = 'Absent'
                            tempTotalObtainedMarks = 0
                            orignalMarkData['marks'] =0
                            examSubject['config_marks'] = 0
                            if 'cum_data' in orignalMarkData and orignalMarkData['cum_data']:
                                orignalMarkData['total_marks']=orignalMarkData['total_marks']
                            else:
                                orignalMarkData['total_marks']=0
                            if 'cum_marks' in examSubject and examSubject['cum_marks']:
                                student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']+examSubject['cum_marks']
                            else:
                                student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']
                            if 'cum_data' in orignalMarkData and orignalMarkData['cum_data'] and 'cum_marks' in examSubject and examSubject['cum_marks']:
                                tempTotalObtainedMarks += ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                                examSubject['config_marks'] = ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                                student['subject_list'][index1]['config_marks'+str(index2)]=examSubject['config_marks']
                                student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']+examSubject['cum_marks']
                        elif not orignalMarkData['marks'] and examSubject['is_only_grade_for_config']:
                            tempTotalObtainedMarks=0
                            if not orignalMarkData['marks'] and orignalMarkData['grade']:
                                if orignalMarkData['exam'] == examSubject['exam']:
                                    student['subject_list'][index1]['grade'] = orignalMarkData['grade']
                                    examSubject['config_marks']=0
                        elif 'cum_marks' in examSubject and examSubject['cum_marks']:
                            tempTotalObtainedMarks += ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                            examSubject['config_marks'] = ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                            student['subject_list'][index1]['config_marks'+str(index2)]=examSubject['config_marks']
                            student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']+examSubject['cum_marks']
                        else:
                            tempTotalObtainedMarks += (
                                    (orignalMarkData['marks'] * examSubject['marks']) / orignalMarkData['max_marks'])
                            examSubject['config_marks']=((orignalMarkData['marks'] * examSubject['marks']) / orignalMarkData['max_marks'])
                            student['subject_list'][index1]['config_marks'+str(index2)]=examSubject['config_marks']
                            student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']
                        student['subject_list'][index1]['config_marks_of_test']=examSubject['config_marks']
                        if examSubject['exam'] in exams_to_merge:
                            config_marks_of_test+=examSubject['config_marks']
                            if examSubject['cum_marks']:
                                config_max_marks_of_test+=(examSubject['marks']+examSubject['cum_marks'])
                            else:
                                config_max_marks_of_test+=examSubject['marks']
                        else:
                            config_marks_of_exam+=examSubject['config_marks']
                            if examSubject['cum_marks']:
                                config_max_marks_of_exam+=(examSubject['marks'] if examSubject['marks'] else 0+examSubject['cum_marks'])
                            else:
                                config_max_marks_of_exam+=examSubject['marks'] if examSubject['marks'] else 0
                student['subject_list'][index1]['config_marks_of_test']=math.ceil(config_marks_of_test)
                student['subject_list'][index1]['config_max_marks_of_test']=math.ceil(config_max_marks_of_test)
                student['subject_list'][index1]['config_marks_of_exam']=config_marks_of_exam
                student['subject_list'][index1]['config_max_marks_of_exam']=config_max_marks_of_exam
                student['subject_list'][index1]['marks'] = tempTotalObtainedMarks
                student['subject_list'][index1]['marks_round_off'] = student['subject_list'][index1]['config_marks_of_test']+student['subject_list'][index1]['config_marks_of_exam']
                if student['subject_list'][index1]['marks']:
                    try:
                        grade_plan_obj_sub=ResultSectionMapping.objects.get(standard_section=standardSectionId,subject=subjectData['subject'],grade_plan__isnull=False)
                        grade_plan=grade_plan_obj_sub.grade_plan
                    except ResultSectionMapping.DoesNotExist:
                        try:
                            grade_plan_obj_sub=ResultSectionApproval.objects.get(standard_section=standardSectionId)
                            grade_plan=grade_plan_obj_sub.grade_plan
                        except ResultSectionApproval.DoesNotExist:
                            grade_plan_obj_sub=None
                            grade_plan=None
                    grade_list = Grade.objects.filter(grade_plan=grade_plan).values()
                    student['subject_list'][index1]['percentage']=0
                    if 'grade' not in student['subject_list'][index1] or not student['subject_list'][index1]['grade']:
                        student['subject_list'][index1]['grade'],student['subject_list'][index1]['percentage']=get_grade_for_marks(grade_list,student['subject_list'][index1]['marks'], student['subject_list'][index1]['max_marks'],grade_plan)
                        student['subject_list'][index1]['grade_round_off'],student['subject_list'][index1]['percentage_round_off']=get_grade_for_marks(grade_list,student['subject_list'][index1]['marks_round_off'], student['subject_list'][index1]['max_marks'],grade_plan)
                if student['subject_list'][index1]['subject_part_type_code_name']=='part1':
                    student['part1_obtained_marks']+=tempTotalObtainedMarks
                    student['part1_obtained_marks_round_off'] += (student['subject_list'][index1]['config_marks_of_test']+student['subject_list'][index1]['config_marks_of_exam'])
                    student['part1_total_marks'] += subjectConfigurationMapping[subject]['max_marks']
                    student['part1_total_config_marks_of_test']+=student['subject_list'][index1]['config_marks_of_test']
                    student['part1_total_config_max_marks_of_test']+=student['subject_list'][index1]['config_max_marks_of_test']
                    student['part1_total_config_max_marks_of_exam']+=student['subject_list'][index1]['config_max_marks_of_exam']
                    student['part1_total_config_marks_of_exam']+=student['subject_list'][index1]['config_marks_of_exam']
                    student['part1_obtained_marks_round_off_in_words']=num2words(student['part1_obtained_marks_round_off'], lang='en')
                student['obtained_marks'] += tempTotalObtainedMarks
                student['obtained_marks_round_off'] += (student['subject_list'][index1]['config_marks_of_test']+student['subject_list'][index1]['config_marks_of_exam'])
                student['total_config_marks_of_test']+=student['subject_list'][index1]['config_marks_of_test']
                student['total_config_max_marks_of_test']+=student['subject_list'][index1]['config_max_marks_of_test']
                if student['subject_list'][index1]['marks'] > \
                        student['subject_list'][index1]['min_marks']:
                    student['subject_list'][index1]['result'] = 'pass'
                if student['total_result'] != 'fail':
                    student['total_result'] = student['subject_list'][index1]['result']
        try:
            grade_plan_obj = ResultSectionApproval.objects.get(standard_section=standardSectionId)
            grade_list = Grade.objects.filter(grade_plan=grade_plan_obj.total_grade_plan).values()
            total_grade_plan = grade_plan_obj.total_grade_plan
        except ResultSectionApproval.DoesNotExist:
            grade_plan_obj = []
            grade_list = []
            total_grade_plan = None
        student['grade'],student['total_percentage'] = get_grade_for_marks(grade_list, student['obtained_marks'], student['total_marks'],total_grade_plan)
        student['part1_grade'],student['part1_total_percentage'] = get_grade_for_marks(grade_list,student['part1_obtained_marks'],student['part1_total_marks'],total_grade_plan)
        student['grade_round_off'],student['total_percentage_round_off'] = get_grade_for_marks(grade_list, student['obtained_marks_round_off'], student['total_marks'],total_grade_plan)
        student['part1_grade_round_off'],student['part1_total_percentage_round_off'] = get_grade_for_marks(grade_list,student['part1_obtained_marks_round_off'],student['part1_total_marks'],total_grade_plan)
        # student['grade'] = get_grade_for_marks(gradeData, student['obtained_marks'], student['total_marks'], grade_plan_id)
        # nikhil
        # student['subject_list'] = get_student_grade(student['subject_list'], academicYear, 'marks', 'max_marks')
        if student['id'] in studentFinalResult:
            student['total_result'] = studentFinalResult[student['id']]['status']
            if studentFinalResult[student['id']]['is_announced']:
                isAnnounced = True
                student['is_announced'] = True
            student['final_result_id'] = studentFinalResult[student['id']]['id']
        updated_student_data.append(student)
    return {'studentData': updated_student_data, 'isAnnounced': isAnnounced}

def get_all_standard_marks(self, examId=None):
    response = {'data': {}}
    standard_ids = self.request.GET.get('standard_ids')
    standard_section_ids_data = self.request.GET.get('standard_section_ids')
    if not examId:
        raise ValidationError('Please Provide exam Id')
    filter_query = {'exam':examId}
    exam_obj = Exam.objects.get(id=examId)
    approvalData = ApprovalService.get_approval_status(self, exam_obj)
    if approvalData['approval_status'] != '1':
        raise ValidationError('Exam is not yet approved.')
    setting_value = int(ConfigurationService.get_setting_value('staffstandardmapping'))
    if setting_value and not self.request.user.is_superuser:
        filter_query['standard_section__standard__in'] = StaffStandardMapping.objects.filter(staff=self.request.user.staff).values_list('standard', flat=True)
    if standard_ids:
        filter_query['standard_section__standard__in'] = standard_ids.split(',')
    if standard_section_ids_data:
        filter_query['standard_section__in'] = standard_section_ids_data.split(',')
    approvedMarkData = get_approved_standard_section_list(exam_obj.id)
    scheduleData = (
        ExamSchedule.objects
        .filter(**filter_query)
        .select_related(
            'subject',
            'standard_section',
            'standard_section__standard',
            'subject__subject_part_type'  # if needed
        )
        .values(
            'id',
            'subject',  # actual field name, no alias needed
            'standard_section_id',
            subject_sequence=F('subject__sequence'), 
            is_language=F('subject__is_language'), 
            subject_part_type_id=F('subject__subject_part_type_id'),
            subject_part_type=F('subject__subject_part_type__name'),
            standard_id=F('standard_section__standard_id'),
            subject_name=F('subject__name'),
            subject_code=F('subject__subject_code'),
        )
    )
    # serializer = ExamScheduleReadSerilaizer(scheduleData, many=True)
    examSheduledData = {}  # standard scheduled mappingg
    standardIds = set()
    standard_section_ids = set()
    examScheduleIds = []
    number_of_language = int(ConfigurationService.get_setting_value('number_of_language'))
    for schedule in scheduleData:
        examScheduleIds.append((schedule['id']))
        tmp = {
                'subject': schedule['subject'],
                'subject_name': schedule['subject_name'],
                'subject_code': schedule['subject_code'],
                'is_language': schedule['is_language'],
                'sequence': schedule['subject_sequence'],
                'subject_part_type_id': schedule['subject_part_type_id'],
                'subject_part_type': schedule['subject_part_type']
            }
        if number_of_language > 1 and schedule['is_language']:
            tmp['subject_name'] = schedule['subject_name'] + ' ' + str(tmp['sequence'])
        if schedule['standard_id'] in examSheduledData:
            examSheduledData[schedule['standard_id']][schedule['subject']] = tmp
        else:
            examSheduledData[schedule['standard_id']] = {schedule['subject']: tmp}
        standard_section_ids.add(schedule['standard_section_id'])
    for temp in examSheduledData:
        examSheduledData[temp] = examSheduledData[temp].values()
    filter_enrollment = {
        'standard_section__academic_year': exam_obj.academic_year,
        'student__is_active': True
    }
    filter_enrollment['standard_section__in'] = list(standard_section_ids)
    studentData = Enrollment.objects.filter(**filter_enrollment).values('student','standard_section',
                                                                            standardId=F(
                                                                                'standard_section__standard'),
                                                                            sectionId=F(
                                                                                'standard_section__section'),
                                                                            section_name=F(
                                                                                'standard_section__section__name'),
                                                                            standard_name=F(
                                                                                'standard_section__standard__name'
                                                                        ))  # Enrolled students for the scheduled exam data
    standardStudentMapping = {}
    studentDataNew = {}  # studentid and stduent data mapping
    for student in studentData:
        if student['standard_section'] in standardStudentMapping:
            standardStudentMapping[student['standard_section']].append(student)
        else:
            standardStudentMapping[student['standard_section']] = []
            standardStudentMapping[student['standard_section']].append(student)
        studentDataNew[student['student']] = student
    standardardSubj = {}  # student subject list mapping
    subjectStudentData = SubjectStudent.objects.filter(academic_year=exam_obj.academic_year,
                                                       student__in=studentDataNew.keys()).values(
        'subject', 'student'
    )
    for standSubj in subjectStudentData:
        if standSubj['student'] in standardardSubj:
            standardardSubj[standSubj['student']].append(standSubj['subject'])
        else:
            standardardSubj[standSubj['student']] = []
            standardardSubj[standSubj['student']].append(standSubj['subject'])
    filter_standard_section={
        'standard__is_active': True,
        'section__is_active': True,
        'academic_year': exam_obj.academic_year
    }
    filter_standard_section['id__in'] = list(standard_section_ids)
    queryset = StandardSectionMapping.objects.filter(**filter_standard_section).annotate(
        strengthEnrolled=Count('enrollments__student')).order_by(
        'standard')  # standard - section list data
    standardSectionData = queryset.values('standard', 'standard__name').distinct()
    section = queryset.values('id', 'standard', 'section', 'section__name').order_by('standard',
                                                                                     'section__name')
    rows = groupby(section, itemgetter('standard'))
    sectiondata = {key: list(items) for key, items in rows}
    studentMarkData = StudentMark.objects.filter(exam_schedule__in=examScheduleIds, is_active=True).values('student',
                                                                                                           subject=F(
                                                                                                               'exam_schedule__subject'),
                                                                                                           subject_name=F(
                                                                                                               'exam_schedule__subject__name'))
    studentEnteredMark = {}
    for enterdMark in studentMarkData:
        if enterdMark['student'] in studentEnteredMark:
            studentEnteredMark[enterdMark['student']].append(enterdMark['subject'])
        else:
            studentEnteredMark[enterdMark['student']] = []
            studentEnteredMark[enterdMark['student']].append(enterdMark['subject'])
    for items in standardSectionData:
        tempSectionData = sectiondata[items['standard']]
        for sectionData in tempSectionData:
            sectionData['approval_status'] = '1' if sectionData['id'] in approvedMarkData else '0'
            sectionData['subject_data'] = {}
            tempSubjectStudentData = {item['subject']: {'total': 0, 'entered': 0} for item in
                                      examSheduledData[items['standard']]}
            if sectionData['id'] in standardStudentMapping:
                for studentData in standardStudentMapping[sectionData['id']]:
                    student_id = studentData['student']
                    if student_id in standardardSubj:
                        # Use set to avoid duplicate subjects for the same student
                        for studentSubjectId in set(standardardSubj[student_id]):
                            if studentSubjectId in tempSubjectStudentData:
                                tempSubjectStudentData[studentSubjectId]['total'] += 1

                            if (
                                student_id in studentEnteredMark
                                and studentSubjectId in studentEnteredMark[student_id]
                                and studentSubjectId in tempSubjectStudentData
                            ):
                                tempSubjectStudentData[studentSubjectId]['entered'] += 1

            sectionData['subject_data'] = tempSubjectStudentData
        items.update(
            {'section_list': sectiondata[items['standard']], 
            'subject_list': sorted(examSheduledData[items['standard']], key=lambda d: d['subject_code'])})
    response['data'] = standardSectionData
    response['part_type_list'] = SubjectPartType.objects.all().values()
    return response


def validate_approve_mark(self, examId, standardSectionId):
    responseData = get_standard_section_subjects(self, examId, standardSectionId, ignore_final_result_data=True)
    existingData = {}
    studentIds = [s['student'] for s in responseData['data']['student_list']]
    existingData = {s['student']: s for s in StudentExamFinalResult.objects.filter(student__in=studentIds, exam=examId).values('id', 'student')}
    errorData = []
    resultData = []  # Result data save finalresult
    for studentData in responseData['data']['student_list']:
        proceed = True
        for subjectId in studentData['subject_list']:
            subjectData = studentData['subject_list'][subjectId]
            if 'attendance_status' not in subjectData:
                temp = {'student': studentData['student'], 'subject': subjectData['subject'],
                        'student_name': studentData['student_name'], 'subject_name': subjectData['subject__name']}
                errorData.append(temp)
                proceed = False
        if proceed:
            temp = {'status': studentData['total_result'], 'student': studentData['student'], 'exam': examId,
                               'changed_user': self.request.user.id}
            if studentData['student'] in existingData:
                temp['id'] = existingData[studentData['student']]['id']
            resultData.append(temp)
    if errorData:
        errorMsg = {
            'message': 'Not able to Finalize for the section. Some of the student marks status is not marked',
            'errorData': errorData
        }
        raise ValidationError(errorMsg)
    return resultData

def student_exam_final_result_obj(id):
    return StudentExamFinalResult.objects.get(id=id)

def approve_student_mark(self, examId, standardSectionId):
    resultData = validate_approve_mark(self, examId, standardSectionId)
    with transaction.atomic(using=get_current_db_name()):
        self.queryset = StudentMarkSectionWiseApproval
        self.serializer_class = StudentMarkSectionWiseApprovalSerializer
        dataToSave = {'standard_section': standardSectionId, 'exam': examId, 'approval_status': 1, 'result_config': ''}
        try:
            section_wise_approval = StudentMarkSectionWiseApproval.objects.get(
                standard_section=standardSectionId, exam=examId
            )
            dataToSave['id'] = section_wise_approval.id
        except:
            pass
        if 'id' in dataToSave:
            response = SharedService.update_data(self, dataToSave, **{'customObjectData': section_wise_approval})
        else:
            response = SharedService.add_data(self, dataToSave, False)
        self.queryset = StudentExamFinalResult
        self.serializer_class = StudentExamFinalResultSerializer
        SharedService.add_or_update_data(self, resultData, **{'partial': True, 'customObject': student_exam_final_result_obj})
        SharedService.custom_thread(update_exam_ranks, examId, standardSectionId)
    return response

def build_grade_mark_map(exam_id, standard_section_id):
    grade_plan_map = {}
    grade_value_map = {}

    schedules = ExamSchedule.objects.filter(exam_id=exam_id, standard_section_id=standard_section_id)
    for schedule in schedules:
        if schedule.grade_plan_id:
            grade_plan_map[schedule.subject_id] = schedule.grade_plan_id

    try:
        mapping = GradeExamScheduleMapping.objects.get(exam_id=exam_id, standard_section_id=standard_section_id)
        fallback_plan = mapping.grade_plan_for_total or mapping.grade_plan
    except GradeExamScheduleMapping.DoesNotExist:
        fallback_plan = None

    plans = set(grade_plan_map.values())
    if fallback_plan:
        plans.add(fallback_plan.id)

    grades = Grade.objects.filter(grade_plan_id__in=plans)
    for g in grades:
        if g.from_range is not None and g.to_range is not None:
            avg = (g.from_range + g.to_range) / 2
            grade_value_map[(g.name, g.grade_plan_id)] = avg
    return grade_plan_map, fallback_plan.id if fallback_plan else None, grade_value_map


def get_student_totals(exam_id, standard_section_id, valid_student_ids):
    grade_plan_map, fallback_plan_id, grade_value_map = build_grade_mark_map(exam_id, standard_section_id)

    marks = StudentMark.objects.filter(
        exam_schedule__exam_id=exam_id,
        is_active=True,
        student__in=valid_student_ids,
        student__enrollment__standard_section_id=standard_section_id
    ).select_related('exam_schedule', 'student')

    student_totals = defaultdict(float)
    for mark in marks:
        sid = mark.student_id
        if mark.marks is not None:
            student_totals[sid] += mark.marks
        elif mark.grade and mark.exam_schedule:
            subject_id = mark.exam_schedule.subject_id
            plan_id = grade_plan_map.get(subject_id) or fallback_plan_id
            avg = grade_value_map.get((mark.grade, plan_id), 0)
            student_totals[sid] += avg

    return {k: round(v, 2) for k, v in student_totals.items()}


def assign_continuous_ranks(student_totals):
    sorted_students = sorted(student_totals.items(), key=lambda x: x[1], reverse=True)
    ranks = {}
    for idx, (sid, _) in enumerate(sorted_students):
        ranks[sid] = idx + 1  # Continuous rank
    return ranks


def update_exam_ranks(exam_id, standard_section_id):
    try:
        section = StandardSectionMapping.objects.get(id=standard_section_id)
        academic_year_id = section.academic_year_id
        standard_id = section.standard_id

        # Only students who have StudentExamFinalResult
        result_students = StudentExamFinalResult.objects.filter(exam_id=exam_id)
        student_ids = list(result_students.values_list('student_id', flat=True))

        # Section totals & rank
        section_totals = get_student_totals(exam_id, standard_section_id, student_ids)
        section_ranks = assign_continuous_ranks(section_totals)

        # Standard totals & rank
        standard_section_ids = StandardSectionMapping.objects.filter(
            standard_id=standard_id,
            academic_year_id=academic_year_id
        ).values_list('id', flat=True)

        # Combine all totals from standard-wise
        all_standard_totals = {}
        for sec_id in standard_section_ids:
            totals = get_student_totals(exam_id, sec_id, student_ids)
            for sid, t in totals.items():
                all_standard_totals[sid] = all_standard_totals.get(sid, 0) + t

        standard_ranks = assign_continuous_ranks(all_standard_totals)

        # Update ranks in DB
        for result in result_students:
            sid = result.student_id
            result.section_rank = section_ranks.get(sid)
            result.standard_rank = standard_ranks.get(sid)
            result.save(update_fields=['section_rank', 'standard_rank'])
    except: #should not effect existing flow
        pass

def approve_student_mark_for_config(self, request, configId, standardSectionId):
    resultData = validate_approve_student_mark_config(self, request, configId, standardSectionId)
    with transaction.atomic(using=get_current_db_name()):
        self.queryset = StudentMarkSectionWiseApproval
        self.serializer_class = StudentMarkSectionWiseApprovalSerializer
        dataToSave = {'standard_section': standardSectionId, 'result_config': configId, 'approval_status': 1,
                      'exam': None}
        response = SharedService.add_data(self, dataToSave, False)
        self.queryset = StudentExamFinalResult
        self.serializer_class = StudentExamFinalResultSerializer
        SharedService.add_data(self, resultData)
    return response


def validate_approve_student_mark_config(self, request, configId, standardSectionId):
    from apps.exams.models.result import ResultConfiguration
    from apps.exams.serializers import ResultConfigurationReadSerializer
    configData = ResultConfiguration.objects.filter(id=configId).first()
    errorData = []
    dataToSave = []
    if StudentMarkSectionWiseApproval.objects.filter(result_config=configId).first():
        raise ValidationError('Already Approved')
    if not configData:
        raise ValidationError('Invalid configuration data')
    configSerializer = ResultConfigurationReadSerializer(configData)
    if not configSerializer.data['result_section_data']:
        raise ValidationError('Nothing to approve')
    examIds = set()
    for resultData in configSerializer.data['result_section_data']:
        for examData in resultData['subject_exam_data']:
            examIds.add(examData['exam'])
    approvedExamIds = StudentMarkSectionWiseApproval.objects.filter(exam__in=list(examIds), approval_status='1',
                                                                    standard_section=standardSectionId).values_list(
        'exam', flat=True)
    unapprovedExamIds = set(examIds) - set(approvedExamIds)
    if unapprovedExamIds:
        examDatas = Exam.objects.filter(id__in=unapprovedExamIds).values('exam_type__name')
        errorMsg = ''
        for exam in examDatas:
            errorMsg += exam['exam_type__name'] + ', '
        raise ValidationError(f'[ {errorMsg[:len(errorMsg) - 2]} ] - List of exams are not approved')
    request.GET._mutable = True
    request.GET['raise_approval_error'] = False
    request.GET['standard_section'] = standardSectionId
    marksData = get_marks_for_config(self, request)
    for studentData in marksData['data']['student_list']:
        for subjectData in studentData['subject_list']:
            proceed = True
            if 'attendance_status' not in subjectData or subjectData['unattended_schedule_marks']:
                temp = {'student': studentData['student'], 'subject': subjectData['subject'],
                        'student_name': studentData['student_name'], 'subject_name': subjectData['subject__name']}
                errorData.append(temp)
                proceed = False
            if proceed:
                dataToSave.append({'status': studentData['total_result'], 'student': studentData['student'],
                                   'result_config': configId, 'changed_user': self.request.user.id})
    if errorData:
        errorMsg = {
            'message': 'Not able to Finalize for the section. Some of the student marks status is not marked',
            'errorData': errorData
        }
        raise ValidationError(errorMsg)
    return dataToSave


def get_approved_standard_section_list(examId):
    return StudentMarkSectionWiseApproval.objects.filter(exam=examId, approval_status='1').values_list(
        'standard_section', flat=True)

def get_student_grade(marksData, exam_id, standard_section_id, scoredMarkKey='obtained_marks', totMarkKey='total_marks', is_for_total_marks=False, extra_params={},is_fail=False):
    grade_data = None
    grade_plan_obj = None
    exam_configurations_grade_plan = None
    is_raise_error = False
    if 'grade_data' in extra_params and extra_params['grade_data'] and 'grade_plan_obj' in extra_params and extra_params['grade_plan_obj']: #Added here for not to call database again and again
        grade_data = extra_params['grade_data']
        grade_plan_obj = extra_params['grade_plan_obj']
    if 'exam_configurations_grade_plan' in extra_params and extra_params['exam_configurations_grade_plan']:
        exam_configurations_grade_plan = extra_params['exam_configurations_grade_plan']
    else:
        exam_configurations_grade_plan = FormdefinitionService.get_formdefintion_data({}, 'exam_configurations', 'grade_plan')
    if exam_configurations_grade_plan:
        is_raise_error = True
    is_raise_error = False #for some bad reason
    grade_data, grade_plan_obj = GradeExamScheduleMapping.get_my_grade_data({}, exam_id, standard_section_id, is_raise_error, is_for_total_marks,marksData[0]['subject'] if 'subject' in marksData[0] else None)
    for data in marksData:
        if scoredMarkKey not in data or totMarkKey not in data:
            data['grade'] = ''
            data['percentage'] = 0
        else:
            data['grade'], data['percentage'] = get_grade_for_marks(grade_data, data[scoredMarkKey], data[totMarkKey], grade_plan_obj, is_fail)
            data['grade_plan'] = None
        if grade_data:
            data['grade_plan'] = grade_data[0]['grade_plan_id']
    return {
        'marksData': marksData, 'grade_data': grade_data, 'grade_plan_obj': grade_plan_obj,
        'exam_configurations_grade_plan': exam_configurations_grade_plan
    }

def get_grade_for_marks(gradeData, obtainedMarks, totalMarks, grade_plan_obj, is_fail=False):
    if grade_plan_obj:
        if grade_plan_obj.grade_type == 2:
            return '', 0
        marks_percentage = 0
        if totalMarks > 0:
            marks_percentage = (obtainedMarks/totalMarks)*100
        if is_fail:
            fail_grade = next((grade for grade in gradeData if grade.get('is_fail_grade')), None)
            if fail_grade:
                return fail_grade['name'], marks_percentage
        for grade in gradeData:
            if grade['to_range'].is_integer():
                grade['to_range'] += 0.99 #nikhil please change this logic
            if grade_plan_obj.grade_type == 1:
                if grade['from_range'] <= marks_percentage <= grade['to_range']:
                    return grade['name'], marks_percentage
            elif grade_plan_obj.grade_type == 1 or grade_plan_obj.grade_type == 0:
                if grade['from_range'] <= obtainedMarks <= grade['to_range']:
                    return grade['name'], marks_percentage
        return '', marks_percentage
    else:
        return '',0

def student_exam_mark(self, request):
    from apps.exams.services.exam import get_student_exam_schedule
    if not self.request.user.student or not self.request.user.student.id:
        raise ValidationError('Only student are allowed to access')
    exam_id = request.GET.get('exam')
    term_id = request.GET.get('term')
    exam_data = Exam.objects.filter(id=exam_id).values('exam_type__name', 'description', 'from_date', 'to_date', 'id', 'academic_year')[0]
    announced_data = StudentExamFinalResult.objects.filter(is_announced=True, exam=exam_id, student=self.request.user.student.id).values()
    if not announced_data:
        raise ValidationError('Result still not announced')
    if announced_data and not announced_data[0]['is_announced']:
        raise ValidationError('Result not Announced')
    announced_data = announced_data[0]
    student_exam_schedule = get_student_exam_schedule(self, request, exam_id, term_id)
    schedule_data = student_exam_schedule['data']['schedule_list']
    schedule_ids = []
    for schedule in schedule_data:
        schedule_ids.append(schedule['id'])
    student_mark_data = StudentMark.objects.filter(is_active=True, student=self.request.user.student.id, exam_schedule__in=schedule_ids).values()
    student_cumulative_mark_data = StudentCumulativeMark.objects.filter(is_active=True, student=self.request.user.student.id, exam_cumulative__exam_schedule__in=schedule_ids).values(
        'exam_cumulative', 'exam_cumulative__exam_schedule', 'marks', 'student'
    )
    student_cumulative_schedule_mapping = {}
    for cumulative_mark in student_cumulative_mark_data:
        if cumulative_mark['exam_cumulative__exam_schedule'] not in student_cumulative_schedule_mapping:
            student_cumulative_schedule_mapping[cumulative_mark['exam_cumulative__exam_schedule']] = {}
        student_cumulative_schedule_mapping[cumulative_mark['exam_cumulative__exam_schedule']][cumulative_mark['exam_cumulative']] = cumulative_mark
    student_mark_data = {s['exam_schedule_id'] : s for s in student_mark_data}
    total_obtained_marks = 0
    total_marks = 0
    min_date = None
    max_date = None
    grade_data = []
    grade_data, grade_plan_obj = GradeExamScheduleMapping.get_my_grade_data({}, exam_id, student_exam_schedule['data']['standard_section'], False)
    for schedule in schedule_data:
        schedule['obtained_marks'] = 0
        schedule['total_obtained_marks'] = 0
        schedule['attendance_status'] = 'Absent'
        if not min_date or schedule['fordate'] < min_date:
            min_date = schedule['fordate']
        if not max_date or schedule['fordate'] > max_date:
            max_date = schedule['fordate']
        if schedule['id'] in student_mark_data and student_mark_data[schedule['id']]['marks']:
            schedule['obtained_marks'] = student_mark_data[schedule['id']]['marks']
            schedule['total_obtained_marks'] += student_mark_data[schedule['id']]['marks']
            schedule['attendance_status'] = student_mark_data[schedule['id']]['attendance_status']
        total_obtained_marks += schedule['obtained_marks']
        total_marks += schedule['max_marks']
        schedule['grade'], schedule['percentage'] = get_grade_for_marks(grade_data, schedule['obtained_marks'], schedule['max_marks'], grade_plan_obj)
        schedule['cumulative_mark_data'] = []
        if 'cumulative_mapping' in schedule and schedule['id'] in student_cumulative_schedule_mapping:
            for cumulative_row_data in schedule['cumulative_mapping']:
                if cumulative_row_data['id'] in student_cumulative_schedule_mapping[schedule['id']]:
                    schedule['cumulative_mark_data'].append(student_cumulative_schedule_mapping[schedule['id']][cumulative_row_data['id']])
                    schedule['total_obtained_marks'] += student_cumulative_schedule_mapping[schedule['id']][cumulative_row_data['id']]['marks']
                    total_marks += student_cumulative_schedule_mapping[schedule['id']][cumulative_row_data['id']]['marks']
    exam_data['total_marks'] = total_marks
    exam_data['total_obtained_marks'] = total_obtained_marks
    exam_data['grade'], exam_data['percentage'] = get_grade_for_marks(grade_data, exam_data['total_obtained_marks'], exam_data['total_marks'], grade_plan_obj)
    exam_data['min_date'] = min_date
    exam_data['max_date'] = max_date
    termData = ExamTerm.objects.filter(id=term_id).values()[0]
    return {'schedule_data': schedule_data, 'exam_data': exam_data, 'term_data': termData, 'announcement_data': announced_data}

#common function
def get_my_grade_data(self, exam, standard_section):
        try:
            grade_plan = GradeExamScheduleMapping.objects.get(exam=exam, standard_section=standard_section)
        except:
            try:
                grade_plan = GradeExamScheduleMapping.objects.get(default=True)
            except:
                return None
        return Grade.objects.filter(grade_plan=grade_plan).values()


def print_marks_card(self, response):
    # Import store_long_running_process at function level to avoid UnboundLocalError
    # when it's used in exception handlers later in the function
    from apps.shared.services_shared.store_api_result import store_long_running_process
    
    try:
        standard = response['data']['standard']
        academic_year = response['data']['academic_year_details']['academic_year']
        institute_obj = Institute.get_institute(self)
        response['institute_data'] = InstituteSerializer(institute_obj).data
        response['institute_data']['address'] = institute_obj.address
        response['institute_data']['city'] = institute_obj.city
        response['institute_data']['pincode'] = institute_obj.pincode
        
        # Check if custom template is requested
        custom_template_id = self.request.GET.get('custom_template_id')
        if custom_template_id:
            from apps.shared.models.custom_design_template import CustomDesignTemplate
            from apps.shared.services_shared.custom_template_renderer import render_custom_template
            
            try:
                custom_template = CustomDesignTemplate.objects.get(id=custom_template_id, is_active=True, institute=institute_obj)
                # Render custom template
                html_content = render_custom_template(custom_template, response, response['institute_data'])
                
                # Generate PDF from custom HTML
                pdf_filename = "Marks_Report_Custom.pdf"
                import pdfkit
                options = {
                    'page-size': 'A4',
                    'margin-top': '0mm',
                    'margin-right': '0mm',
                    'margin-bottom': '0mm',
                    'margin-left': '0mm',
                    'encoding': "UTF-8",
                    'no-outline': None,
                    'enable-local-file-access': None,
                }
                pdfkit.from_string(html_content, pdf_filename, options=options)
                
                if self.request.GET.get('long_running_process'):
                    try:
                        url = UploadTypeService.upload_local_file(pdf_filename, path='markscard_pdfs')
                        if os.path.exists(pdf_filename):
                            os.remove(pdf_filename)
                        transaction_id = self.request.GET.get('transaction_id')
                        store_long_running_process(self, transaction_id, {'url': url})
                        return
                    except Exception as upload_error:
                        transaction_id = self.request.GET.get('transaction_id')
                        error_msg = f"Failed to upload custom template PDF: {str(upload_error)}"
                        store_long_running_process(self, transaction_id, {'error': [error_msg[:250]]})
                        if os.path.exists(pdf_filename):
                            try:
                                os.remove(pdf_filename)
                            except:
                                pass
                        return
                else:
                    return pdf_filename
            except CustomDesignTemplate.DoesNotExist:
                # Fall back to default template if custom template not found
                pass
            except Exception as e:
                # Log error and fall back to default template
                print(f"Error rendering custom template: {str(e)}")
        
        # Default template rendering
        if self.request.GET.get('type') == 'pdf':
            total_max_marks_part1=0
            total_min_marks_part1=0
            for item in response['data']['subject_list']:
                if item['subject_part_type_id'] == 1:
                    total_max_marks_part1+= item['total_max_marks']
                    total_min_marks_part1+= item['total_min_marks']
            response['data']['total_max_marks_part1']=total_max_marks_part1
            response['data']['total_min_marks_part1']=total_min_marks_part1
            for item in response['data']['student_list']:
                item['total_marks_obtained_part1']= sum(item['obtained_marks_list_part_wise']['part1']) or 0
                item['total_percentage_obtained_part1'] =(item['total_marks_obtained_part1']/response['data']['total_max_marks_part1'])*100
            
            selected_template_obj = get_selected_template(self, 'consolidated_report', None, 'template_2.html', academic_year, [standard],None,None,True)
            path = 'consolidated_report/'+selected_template_obj['path']
            #Drag and drop design template check
            designtemplatecheck = SharedService.prepare_pdf(key='consolidated_report', data=response)
            if designtemplatecheck == False:
                response =PDFService.receipt_new(self, response, "consolidated_report", path, False)
            else:
                response = designtemplatecheck
            
            return response
        selected_template_obj = get_selected_template(self, 'marks_card', None, 'template_2.html', academic_year, [standard],None,None,True)
        path = 'marks_card/'+selected_template_obj['path']
        if selected_template_obj['template_type'] == 'html':
            from django.shortcuts import render
            return render(self.request, path, response)
        else:
            if institute_obj.code == "nandinividyanikethana":
                response = PDFService.graph_receipt(self, response, "marks_card", path, False)
            else:
                #Drag and drop design template check
                designtemplatecheck = SharedService.prepare_pdf(key='marks_card', data=response)
                if designtemplatecheck == False:
                    response = PDFService.receipt_new(self, response, "marks_card", path, False)
                else:
                    response = designtemplatecheck
                
            if self.request.GET.get('long_running_process'):
                try:
                    # Check if file exists before uploading
                    if not os.path.exists(response):
                        raise FileNotFoundError(f"PDF file not found: {response}")
                    
                    # Get file size for logging
                    file_size_mb = os.path.getsize(response) / (1024 * 1024)
                    print(f"Uploading PDF file: {response}, Size: {file_size_mb:.2f} MB")
                    
                    url = UploadTypeService.upload_local_file(response, path='markscard_pdfs')
                    
                    if os.path.exists(response):
                        os.remove(response)
                    transaction_id = self.request.GET.get('transaction_id')
                    store_long_running_process(self, transaction_id, {'url': url})
                except FileNotFoundError as e:
                    transaction_id = self.request.GET.get('transaction_id')
                    error_msg = f"Failed to upload Marks_Report.pdf: PDF file not found. {str(e)}"
                    store_long_running_process(self, transaction_id, {'error': [error_msg[:250]]})
                except Exception as upload_error:
                    transaction_id = self.request.GET.get('transaction_id')
                    error_msg = f"Failed to upload Marks_Report.pdf to S3: {str(upload_error)}"
                    store_long_running_process(self, transaction_id, {'error': [error_msg[:250]]})
                    # Try to clean up the file if it exists
                    if os.path.exists(response):
                        try:
                            os.remove(response)
                        except:
                            pass
            else:
                return response
    except Exception as e:
        if self.request.GET.get('long_running_process'):
            transaction_id = self.request.GET.get('transaction_id')
            store_long_running_process(self, transaction_id,{'error': e.args[:250]})
        else:
            raise e
    
from apps.exams.models.final_result import FinalResultSectionApproval,FinalResultSectionMapping

def get_marks_for_final_config(self, request):
    from apps.exams.services.final_result import get_final_result_configuration_mapping
    from apps.exams.models.final_result import FinalResultConfiguration
    approvalError = ''
    isapproved = True
    result_config_id = request.GET.get('result_config', None)
    student_ids_param = request.GET.get('student_ids', None)
    if result_config_id:
        result_config_data = FinalResultConfiguration.objects.filter(id=result_config_id).first()
        academicYear = result_config_data.academic_year_id
        exam=result_config_data.exam
    else:
        academicYear = request.GET.get('academic_year', None)
        exam=request.GET.get('exam',None)
        result_config_data = FinalResultConfiguration.objects.filter(academic_year_id=academicYear,exam=exam).first()
        result_config_id=result_config_data.id

    standardSectionId = request.GET.get('standard_section', None)
    raiseApprovalError = request.GET.get('raise_approval_error', True)
    standard_obj = StandardSectionMapping.objects.get(id=standardSectionId)
    standard_id = standard_obj.standard_id
    if not academicYear or not standardSectionId:
        raise ValidationError('Term Academic StandardSection should be mandatory')
    examData = FinalResultMarksConfiguration.objects.filter(result_section__final_result=result_config_id,result_section__standard_section=standardSectionId).values('exam_id','exam__standard_section_ids')
    examIds = []
    for exam in examData:
        standard_section_ids = exam['exam__standard_section_ids'].split(',')
        if str(standardSectionId) in standard_section_ids:
            examIds.append(exam['exam_id'])
    approvalData = StudentMarkSectionWiseApproval.objects.filter(exam__in=examIds, standard_section=standardSectionId,
                                                                 approval_status='1').values_list('exam', flat=True)
    values_list = [
        'id', 'first_name', 'middle_name', 'last_name', 'current_reg_num', 'student_name',
        # Parent/guardian details (StudentParentMapping -> ParentDetail / GuardianDetail)
        'student_parent__parent__father_name', 'student_parent__parent__f_mobile_num',
        'student_parent__parent__mother_name', 'student_parent__parent__m_mobile_num',
        'student_parent__guardian__guardian_name', 'student_parent__guardian__g_mobile_num',
        'profile_pic', 'dob', 'mobile_num',
    ]
    custom_annotate = {'student_name': Concat('first_name', V(' '), 'middle_name',V(' '), 'last_name')}
    
    if set(examIds) - set(approvalData):
        examData = Exam.objects.filter(id__in=list(set(examIds) - set(approvalData))).values(
            'exam_type__name', 'term__name'
        )
        errorData = ','.join((str(e['exam_type__name'])+'('+str(e['term__name'])+')') for e in examData )
    
        if raiseApprovalError:
            raise ValidationError(
                f'{errorData} the given exams Marks of all standard are not yet approved')
        
        else:
            isapproved = False
            approvalError = errorData
    elif student_ids_param:
        student_ids_param = student_ids_param.split(',')
        studentData = Student.objects.filter(
           id__in=student_ids_param
            ).annotate(**custom_annotate).values(*values_list)
          
    else:
        studentData = Student.get_student_for_standard(None, None, [standardSectionId], values_list, custom_annotate)
    student_ids = []
    studentIdDict = {}
    configuredSubjectIds = []
    profile_pic=[]
    for i in studentData:
        profile_pic.append(i['profile_pic'])
        student_ids.append(i['id'])
        i['student'] = i['id']
        studentIdDict[i['id']] = i
    doc=Document.objects.filter(id__in=profile_pic)
    doc_mapping = {d['id']:d for d in DocumentSerializer(doc,many=True).data}
    
    sectionResultConfigData = get_final_result_configuration_mapping(result_config_id,None,academicYear, standardSectionId,
                                                               True)
    for sectionData in sectionResultConfigData['result_section_data']:
        configuredSubjectIds.append(sectionData['subject'])
    studentSubjectData = SubjectStudent.objects.filter(student__in=student_ids,
                                                       academic_year=academicYear,
                                                       subject__in=configuredSubjectIds).values('subject', 'student',
                                                                                                'subject__name','subject__subject_code', subject_part_type_id=F('subject__subject_part_type__id'),
                                                                                                subject_part_type=F('subject__subject_part_type__name'),subject_part_type_code_name=F('subject__subject_part_type__code_name'))
    student_admission_form = get_student_admission_form_details(self, student_ids)
    # Fetch student addresses
    student_addresses = StudentAddress.objects.filter(student__in=student_ids).values('student', 'address', 'type')
    student_address_mapping = {}
    for addr in student_addresses:
        if addr['student'] not in student_address_mapping:
            student_address_mapping[addr['student']] = addr['address']
    for student_data in studentData:
        if student_data['profile_pic']:
            student_data['profile_pic_details'] = doc_mapping[student_data['profile_pic']]
        student_data['student_admission_form'] = student_admission_form[student_data['student']]
        student_data['student_admission_form']['admission_num'] =  student_data['student_admission_form']['admission_num'].replace('/', '-')
        student_data['address'] = student_address_mapping.get(student_data['student'], '')
    studentSubjectMapping = {}
    for studentSubject in studentSubjectData:
        if studentSubject['student'] in studentSubjectMapping:
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
        else:
            studentSubjectMapping[studentSubject['student']] = []
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
    studentMarkData = StudentMark.objects.filter(is_active=True, student__in=student_ids,
                                                 exam_schedule__exam__in=list(examIds)).values(
        'student', 'exam_schedule__exam', 'exam_schedule__exam__exam_type__name','exam_schedule__subject', 'marks', 'exam_schedule__max_marks','remark','exam_schedule__min_marks','remark__name','exam_schedule__subject__subject_part_type__code_name',
        'attendance_status','exam_schedule','grade','exam_schedule__is_marks','exam_schedule__standard_section','exam_schedule__subject__subject_part_type__code_name','marked_attendance_days','exam_schedule__exam__description'
    )
    scheduleIds=[]
    for schedule in studentMarkData:
        scheduleIds.append(schedule['exam_schedule'])
    filter_query_cumulative = {
        'student__in': student_ids,
        'exam_cumulative__exam_schedule__in': list(scheduleIds),
        'is_active': True
    }
    student_cumulative_data = StudentCumulativeMark.objects.filter(
        **filter_query_cumulative
    ).values(
        'id', 'marks', 'exam_cumulative_id', 'exam_cumulative__exam_schedule', 'student', 'attendance_status', 'exam_cumulative__cumulative_type','exam_cumulative__min_marks',
        'exam_cumulative__cumulative_type__name', 'exam_cumulative__max_marks','exam_cumulative__exam_schedule__exam_id','exam_cumulative__cumulative_type__alias'
    )
    schedule_cumulative_data = ExamScheduleCumulativeMapping.objects.filter(
        exam_schedule__in = scheduleIds,
    ).values(
        'exam_schedule', 'cumulative_type', 'max_marks', 'min_marks', 'cumulative_type__name',
        'cumulative_type__alias', 'id','exam_schedule__exam_id','id'
    )
    exams_to_merge=[]
    exam_cum_type_dict = {}
    try:
        standardMergeData = FinalResultConfigurationMerge.objects.filter(final_result_config=result_config_id,standard_section=standardSectionId).values('exam','name__name','name__id')
        # exam_to_merge=standardMergeData.exam.all().values('id','name__name')
        # merge_name=standardMergeData.name.name
    except FinalResultConfigurationMerge.DoesNotExist:
        standardMergeData = None
        # exam_to_merge = []
        # merge_name=''
    # for exam_id in exam_to_merge:
    #     exams_to_merge.append(exam_id['id'])
    merge_dict={}
    for merge in standardMergeData:
        if merge['name__id'] not in merge_dict:
            merge_dict[merge['name__id']] = {}
            merge_dict[merge['name__id']]['merge_id'] = merge['name__id']
            merge_dict[merge['name__id']]['merge_name'] = merge['name__name']
            merge_dict[merge['name__id']]['exams'] = []
        merge_dict[merge['name__id']]['exams'].append(merge['exam'])
    merge_list = merge_dict.values()
    temp_schedule_cumulative_mapping = {}
    schedule_cumulative_mapping = {}
    for schedule_cumulative in schedule_cumulative_data:
        if schedule_cumulative['exam_schedule__exam_id'] not in exam_cum_type_dict:
            exam_cum_type_dict[schedule_cumulative['exam_schedule__exam_id']] = {}
        if schedule_cumulative['cumulative_type'] not in exam_cum_type_dict[schedule_cumulative['exam_schedule__exam_id']]:
            exam_cum_type_dict[schedule_cumulative['exam_schedule__exam_id']][schedule_cumulative['cumulative_type']] = schedule_cumulative
        if schedule_cumulative['id'] not in schedule_cumulative_data:
            temp_schedule_cumulative_mapping[schedule_cumulative['id']] = {
                'exam_schedule' : schedule_cumulative['exam_schedule'],
                'max_marks': schedule_cumulative['max_marks'],
                'min_marks': schedule_cumulative['min_marks'],
                'cumulative_type_data': [],
                'id': schedule_cumulative['id']
            }
        temp_schedule_cumulative_mapping[schedule_cumulative['id']]['cumulative_type_data'].append(
            {
                'id': schedule_cumulative['cumulative_type'],
                'name': schedule_cumulative['cumulative_type__name'],
                'alias': schedule_cumulative['cumulative_type__alias']
            }
        )
    for cum_row in temp_schedule_cumulative_mapping.values():
        if cum_row['exam_schedule'] not in schedule_cumulative_mapping:
            schedule_cumulative_mapping[cum_row['exam_schedule']] = []
        schedule_cumulative_mapping[cum_row['exam_schedule']].append(cum_row)
    student_cumulative_data_temp_mapping = {}
    for student_cum in student_cumulative_data: #this is to remove the duplicates
        if student_cum['id'] not in student_cumulative_data_temp_mapping:
            student_cumulative_data_temp_mapping[student_cum['id']] = {
                'cumulative_data_mapping': [], 'id' : student_cum['id'],
                'marks': student_cum['marks'], 'exam_cumulative_id': student_cum['exam_cumulative_id'],
                'exam_cumulative__exam_schedule__exam_id':student_cum['exam_cumulative__exam_schedule__exam_id'],
                'exam_cumulative__exam_schedule': student_cum['exam_cumulative__exam_schedule'], 'student': student_cum['student'],
                'attendance_status': student_cum['attendance_status'], 'exam_cumulative__max_marks': student_cum['exam_cumulative__max_marks'],'exam_cumulative__min_marks': student_cum['exam_cumulative__min_marks']
            }
        student_cumulative_data_temp_mapping[student_cum['id']]['cumulative_data_mapping'].append({
            'cumulative_type_id': student_cum['exam_cumulative__cumulative_type'],
            'cumulative_type_name': student_cum['exam_cumulative__cumulative_type__name'],
            'cumulative_type_alias': student_cum['exam_cumulative__cumulative_type__alias'],
        })
    student_cumulative_data_mapping = {}
    for cum in student_cumulative_data_temp_mapping.values():
        if cum['exam_cumulative__exam_schedule'] not in student_cumulative_data_mapping:
            student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']] = {cum['student']: []}
        elif cum['student'] not in student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']]:
            student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']] = []
        if cum['exam_cumulative__exam_schedule'] not in schedule_cumulative_mapping:
            schedule_cumulative_mapping[cum['exam_cumulative__exam_schedule']] = {}
        student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']].append(cum)
        cum_marks=student_cumulative_data_mapping[cum['exam_cumulative__exam_schedule']][cum['student']]
    calculateMarksForStudent = calculate_final_mark_for_student(self,studentData, studentSubjectMapping, studentMarkData,student_cumulative_data,
                                                          sectionResultConfigData, academicYear,standardSectionId,student_cumulative_data_temp_mapping,
                                                          merge_list,result_config_id,exam_cum_type_dict)
    subjectList = {}
    def sort_key(v):
        code = v.get('subject__subject_code', "NA")
        # Try to convert to int if possible
        try:
            return (0, int(code))   # numeric codes first, sorted by number
        except (ValueError, TypeError):
            return (1, str(code))  # non-numeric codes next, alphabetically
    for index, student in enumerate(calculateMarksForStudent['studentData']):
        calculateMarksForStudent['studentData'][index]['subject_list'] = {s['subject']: s for s in student['subject_list']}
        subjectList.update(calculateMarksForStudent['studentData'][index]['subject_list'])
        # calculateMarksForStudent['studentData'][index]['subject_list_data'] = sorted(student['subject_list'].values(), key=lambda v: (isinstance(v.get('subject__subject_code', "NA"), str), v.get('subject__subject_code', "NA"))) #used for printing marks card
        calculateMarksForStudent['studentData'][index]['subject_list_data'] = sorted(
            student['subject_list'].values(),
            key=sort_key
        )
        part1_sequence=0
        part2_sequence=0
        for subject in calculateMarksForStudent['studentData'][index]['subject_list_data']:
            if subject['subject_part_type_code_name'] == 'part1':
                part1_sequence+=1
            else:
                part2_sequence+=1
            subject['part1_sequence'] = part1_sequence
            subject['part2_sequence'] = part2_sequence
        # Create total_summary with rank data (same as working endpoint)
        calculateMarksForStudent['studentData'][index]['total_summary'] = {
            'total_result': student['total_result'] if 'total_result' in student else None,
            'section_rank': student['section_rank'] if 'section_rank' in student else None,
            'standard_rank': student['standard_rank'] if 'standard_rank' in student else None,
            'total_marks': 0,
            'total_obtained_marks': 0,
            'total_min_marks': 0,
            'percentage': 0
        }
    subjectList = subjectList.values()
    academic_year_obj = AcademicYear.objects.get(id=academicYear)
    del sectionResultConfigData['result_section_data']
    return {'data': {
                    'student_list': calculateMarksForStudent['studentData'],
                    'standard_name': standard_obj.standard.name,
                    'standard' : standard_id,
                    'section_name': standard_obj.section.name,
                    #'merge_name': merge_name,
                    'academic_year_details': {'start_date': academic_year_obj.start_date, 'end_date': academic_year_obj.end_date,'academic_year':academicYear},
                    'configuration_details': sectionResultConfigData,
                    'subject_list': subjectList, 'is_announced': calculateMarksForStudent['isAnnounced']},
                    'isapproved': isapproved, 'approvalError': approvalError,
                    'part_type_list': SubjectPartType.objects.all().values()
                    }
import math
from apps.exams.models.final_result import FinalResultSectionApproval,FinalResultSectionMapping

def calculate_final_mark_for_student(self,studentData, studentSubjectMapping, studentMarkData,student_cumulative_data, configurationData, academicYear,standardSectionId,student_cumulative_data_temp_mapping,merge_list, final_result_section_approval_id=None,exam_cum_type_dict={}):
    institute_objs = Institute.get_institute(self)
    sectionConfiguration = configurationData['result_section_data']
    subjectConfigurationMapping = {'exam_list':{}}
    studentSubjectSchedule = {}
    studentExamSchedule={}
    studentIds = []
    studentFinalResult = {}
    studentExamFinalList = []
    isAnnounced = False
    # nikhil
    gradeData = []
    totalgrade =[]
    # gradeData = Grade.objects.filter(academic_year=academicYear).values()
    for configData in sectionConfiguration:
        if configData['subject'] not in subjectConfigurationMapping:
            subjectConfigurationMapping[configData['subject']] = {}
        subjectConfigurationMapping[configData['subject']] = configData
        for exam in configData['subject_exam_data']:
            if exam['exam'] not in subjectConfigurationMapping['exam_list']:
                subjectConfigurationMapping['exam_list'][exam['exam']]={'exam_name':exam['exam_name'],'exam_description':exam['exam_description']}
    for markData in studentMarkData:
        studentIds.append(markData['student'])
        if markData['student'] not in studentSubjectSchedule:
            studentSubjectSchedule[markData['student']] = {}
        if markData['exam_schedule__subject'] not in studentSubjectSchedule[markData['student']]:
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']] = {}
        if markData['exam_schedule__exam'] not in studentSubjectSchedule[markData['student']]:
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
                markData['exam_schedule__exam']] = {}
        if markData['student'] not in studentExamSchedule:
            studentExamSchedule[markData['student']]={}
        if markData['exam_schedule__exam'] not in studentExamSchedule[markData['student']]:
            studentExamSchedule[markData['student']][
                markData['exam_schedule__exam']] = {'exam_total_marks':0,'exam_part1_total_marks':0,'exam_part1_obtained_grade':'',"exam_part1_total_max_marks":0,"exam_part1_total_min_marks":0}
        cum_total_marks=0
        total_written_marks=0
        cum_total_max_marks=0
        cum_total_min_marks=0
        cum_data_list=[]
        schedule_cum_type_list={}
        for cum_data in student_cumulative_data_temp_mapping:
            cumulative_data_list = []
            if student_cumulative_data_temp_mapping[cum_data]['student'] == markData['student'] and \
            student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__exam_schedule'] == markData['exam_schedule']:
                cum_id=(student_cumulative_data_temp_mapping[cum_data]['cumulative_data_mapping'][0]['cumulative_type_id'])
                exam_schedule_id = student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__exam_schedule']
                if exam_schedule_id not in schedule_cum_type_list:
                    schedule_cum_type_list[exam_schedule_id] = []
                schedule_cum_type_list[exam_schedule_id].append(cum_id)
        for exams in exam_cum_type_dict:
            for cum_type in exam_cum_type_dict[exams]:
                for cum_data in student_cumulative_data_temp_mapping:
                    cumulative_data_list = []
                    if student_cumulative_data_temp_mapping[cum_data]['student'] == markData['student'] and \
                    student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__exam_schedule'] == markData['exam_schedule']:
                        cum_id=(student_cumulative_data_temp_mapping[cum_data]['cumulative_data_mapping'][0]['cumulative_type_id'])
                        # exam_id = student_cumulative_data_temp_mapping[cum_data]['exam_schedule__exam_schedule__exam_id'] 
                        # exam_schedule_id = student_cumulative_data_temp_mapping[cum_data]['exam_schedule__exam_schedule']
                        if cum_type not in schedule_cum_type_list[student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__exam_schedule']]:
                            temp_cum_data = {
                                'cum_id': '',
                                'cum_marks': 0,
                                'cum_max_marks':0,
                                'cum_min_marks':0,
                                'cum_type_name':exam_cum_type_dict[exams][cum_type]['cumulative_type__name'],
                                'cum_type_id':cum_type,
                                'cum_type_alias':exam_cum_type_dict[exams][cum_type]['cumulative_type__alias'],
                            }
                            cum_data_list.append(temp_cum_data)
                        if cum_type == cum_id and \
                            student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__exam_schedule__exam_id'] == exams: 
                            temp_cum_data = {
                                'cum_id': student_cumulative_data_temp_mapping[cum_data]['exam_cumulative_id'],
                                'cum_marks': student_cumulative_data_temp_mapping[cum_data]['marks'],
                                'cum_max_marks':student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__max_marks'],
                                'cum_min_marks':student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__min_marks'],
                                'cum_type_name':'',
                                'cum_type_id':'',
                                'cum_type_alias':'',
                            }
                            for cum_data1 in student_cumulative_data_temp_mapping[cum_data]['cumulative_data_mapping']:
                                temp_cum_data['cum_type_name']+= cum_data1['cumulative_type_name']
                                temp_cum_data['cum_type_id']=cum_data1['cumulative_type_id']
                                temp_cum_data['cum_type_alias']+= cum_data1['cumulative_type_alias']
                            cum_data_list.append(temp_cum_data)
                            cum_total_marks += student_cumulative_data_temp_mapping[cum_data]['marks']
                            cum_total_max_marks += student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__max_marks']
                            cumulative_data_list.append(student_cumulative_data_temp_mapping[cum_data]['cumulative_data_mapping'][0]['cumulative_type_id'])
                            cum_total_min_marks += student_cumulative_data_temp_mapping[cum_data]['exam_cumulative__min_marks']
        # rem_cum_data = set(cumulative_type_ids) - set(cumulative_data_list)
        # for cum_type_id in rem_cum_data:
        #     cum_data_list.append({
        #         'cum_id':None,
        #         'cum_marks': None,
        #         'cum_max_marks': None,
        #         'cum_type_name':'',
        #         'cum_type_id':'',
        #     })
        temp_max_marks = markData['exam_schedule__max_marks'] if markData['exam_schedule__max_marks'] else 0
        temp_min_marks = markData['exam_schedule__min_marks'] if markData['exam_schedule__min_marks'] else 0
        total_max_marks=cum_total_max_marks+temp_max_marks
        total_min_marks=cum_total_min_marks+temp_min_marks
        if not markData['marks']:
            markData['marks'] = 0
        if markData['attendance_status']=='Present' and markData['exam_schedule__is_marks']:
            total_marks=cum_total_marks+markData['marks']
            total_written_marks +=markData['marks']
        else: # grade sub or absent in marks sub or absent in grade sub
            total_marks=cum_total_marks+0
            total_written_marks+=0
        studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
            markData['exam_schedule__exam']] = {
            'marks': markData['marks'], 'max_marks': markData['exam_schedule__max_marks'],'remark__name':markData['remark__name'],'marked_attendance_days':markData['marked_attendance_days'],
            'attendance_status': markData['attendance_status'],'grade':markData['grade'],'exam':markData['exam_schedule__exam'],'total_marks_cum':cum_total_marks,'total_marks_written':total_written_marks,
            'exam_name':markData['exam_schedule__exam__exam_type__name'],'cum_data':cum_data_list,'total_marks':total_marks,'total_max_marks':total_max_marks,'obtained_grade':'','part_type':markData['exam_schedule__subject__subject_part_type__code_name'],
            'exam_description':markData['exam_schedule__exam__description'],'total_min_marks':total_min_marks,'exam_schedule__is_marks':markData['exam_schedule__is_marks']
        }
        if studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][markData['exam_schedule__exam']]['grade']:
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][markData['exam_schedule__exam']]['obtained_grade'] = studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][markData['exam_schedule__exam']]['grade']
        # if markData['exam_schedule__subject__subject_part_type__code_name'] == 'part1':
        #     studentExamSchedule[markData['student']][markData['exam_schedule__exam']]['exam_part1_total_marks']+=total_marks
        #     studentExamSchedule[markData['student']][markData['exam_schedule__exam']]['exam_part1_total_max_marks']+=total_max_marks
        try:
            exam_grade_plan_obj = GradeExamScheduleMapping.objects.get(standard_section=markData['exam_schedule__standard_section'],exam=markData['exam_schedule__exam'])
            grade_list = Grade.objects.filter(grade_plan=exam_grade_plan_obj.grade_plan).values()
            total_grade_list = Grade.objects.filter(grade_plan=exam_grade_plan_obj.grade_plan_for_total).values()
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][markData['exam_schedule__exam']]['max_no_of_days_attendance'] = exam_grade_plan_obj.max_no_of_days_attendance
            grade_plan = exam_grade_plan_obj.grade_plan
            total_grade_plan = exam_grade_plan_obj.grade_plan_for_total
        except GradeExamScheduleMapping.DoesNotExist:
            grade_plan_obj = []
            grade_list = []
            total_grade_list = []
            grade_plan = None
            total_grade_plan = None
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][markData['exam_schedule__exam']]['max_no_of_days_attendance'] = None
        # studentExamSchedule[markData['student']][markData['exam_schedule__exam']]['exam_part1_obtained_grade'],studentExamSchedule[markData['student']][markData['exam_schedule__exam']]['exam_part1_obtained_percentage']=get_grade_for_marks(total_grade_list, studentExamSchedule[markData['student']][markData['exam_schedule__exam']]
        #                                                                                                                                ['exam_part1_total_marks'],studentExamSchedule[markData['student']][markData['exam_schedule__exam']]
        #                                                                                                                           ['exam_part1_total_max_marks'],total_grade_plan)
        if studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
            markData['exam_schedule__exam']]['exam_schedule__is_marks']:
            studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
                markData['exam_schedule__exam']]['obtained_grade'],studentSubjectSchedule[markData['student']][markData['exam_schedule__subject']][
                markData['exam_schedule__exam']]['percentage']= get_grade_for_marks(grade_list, float(total_marks), float(total_max_marks),grade_plan)
        # studentExamSchedule[markData['student']][markData['exam_schedule__exam']]['exam_part1_total_marks_in_words']=num2words(studentExamSchedule[markData['student']][markData['exam_schedule__exam']]['exam_part1_total_marks'], lang='en')
    if sectionConfiguration:
        studentExamFinalList = StudentExamFinalResult.objects.filter(student__in=list(set(studentIds)),
                                                                     result_config=configurationData['id']).values()
    for finalData in studentExamFinalList:
        studentFinalResult[finalData['student_id']] = finalData
        if finalData['is_announced']:
            isAnnounced = True
    # Fetch rank from StudentExamFinalResult using exam ID (same as working endpoint)
    # Get exam ID from final_result_section_approval_id (which is FinalResultConfiguration.id)
    exam_id_for_rank = None
    if final_result_section_approval_id:
        from apps.exams.models.final_result import FinalResultConfiguration
        try:
            final_result_config_obj = FinalResultConfiguration.objects.get(id=final_result_section_approval_id)
            exam_id_for_rank = final_result_config_obj.exam_id
        except FinalResultConfiguration.DoesNotExist:
            pass
    if exam_id_for_rank:
        finalResultDataForRank = StudentExamFinalResult.objects.filter(student__in=list(set(studentIds)), exam=exam_id_for_rank).values(
            'student', 'status', 'is_announced', 'section_rank', 'standard_rank')
        finalResultDataForRank = {temp['student']: temp for temp in finalResultDataForRank}
        for student_id in studentIds:
            if student_id in finalResultDataForRank:
                # Only update rank fields, don't overwrite existing studentFinalResult data
                if student_id not in studentFinalResult:
                    # If student not in studentFinalResult, create entry with all fields from rank data
                    studentFinalResult[student_id] = {
                        'status': finalResultDataForRank[student_id].get('status'),
                        'is_announced': finalResultDataForRank[student_id].get('is_announced', False),
                        'section_rank': finalResultDataForRank[student_id].get('section_rank'),
                        'standard_rank': finalResultDataForRank[student_id].get('standard_rank'),
                        'student_id': student_id
                    }
                else:
                    # Only update rank fields if they exist
                    if 'section_rank' in finalResultDataForRank[student_id]:
                        studentFinalResult[student_id]['section_rank'] = finalResultDataForRank[student_id]['section_rank']
                    if 'standard_rank' in finalResultDataForRank[student_id]:
                        studentFinalResult[student_id]['standard_rank'] = finalResultDataForRank[student_id]['standard_rank']
    # Collect remarks for each student from studentMarkData (same as working endpoint)
    # Store the first non-null remark encountered for each student
    student_remark_dict = {}
    for markData in studentMarkData:
        student_id = markData['student']
        if student_id not in student_remark_dict:
            # Store the first remark encountered for each student (similar to working endpoint)
            student_remark_dict[student_id] = {
                'remark': markData.get('remark'),
                'remark_name': markData.get('remark__name')
            }
        elif not student_remark_dict[student_id]['remark'] and markData.get('remark'):
            # If we haven't found a remark yet, update with first non-null remark
            student_remark_dict[student_id] = {
                'remark': markData.get('remark'),
                'remark_name': markData.get('remark__name')
            }
    
    updated_student_data = []
    for index, std_tmp in enumerate(studentData):
        student = copy.deepcopy(std_tmp)
        student['subject_list'] = []
        student['total_result'] = ''
        student['total_marks'] = 0
        student['obtained_marks'] = 0
        student['obtained_marks_round_off'] = 0
        student['part_wise_details'] = {}
        student['part1_obtained_marks'] = 0
        student['part1_obtained_marks_round_off'] = 0
        student['part1_total_marks'] = 0
        student['part1_total_config_marks_of_test'] = 0
        student['part1_total_config_max_marks_of_test'] = 0
        student['total_config_marks_of_test'] = 0
        student['total_config_max_marks_of_test'] = 0
        student['grade'] = ''
        student['total_percentage'] = 0
        student['part1_total_config_max_marks_of_exam'] = 0
        student['part1_total_config_marks_of_exam'] = 0
        student['subject_names_part_wise'] = {}
        student['obtained_marks_list_part_wise_graph'] = {}
        student['obtained_marks_list_part_wise'] = {}
        # student['total_max_no_of_days_attendance']=0
        # student['total_no_of_days_attendance']=0
        # student['total_no_of_days_attendance_absent']=0
        config_marks_of_test=0
        config_max_marks_of_test=0
        config_marks_of_exam=0
        config_max_marks_of_exam=0
        attendance_data = {}
        # y = {'2data':{'max_no_of_days_attendance':0,'no_of_days_attendance':0,'absent_days':0},'3data':{'max_no_of_days_attendance':0,'no_of_days_attendance':0,'absent_days':0}}
        if student['id'] in studentSubjectMapping:
            student['subject_list'] = studentSubjectMapping[student['id']]
        # Sort subject_list for graphs to show subjects in sorted order (same as subject_list_data)
        student['subject_list'] = sorted(
            student['subject_list'],
            key=lambda v: (
                0, int(v['subject__subject_code'])  # Numeric codes
            ) if v.get('subject__subject_code') and str(v['subject__subject_code']).isdigit() else (
                1, str(v.get('subject__subject_code') or '')  # Non-numeric or None
            )
        )
        # student['exam_details'] = studentExamSchedule[student['id']]
        for index1, subjectData in enumerate(student['subject_list']):
            student['total_max_no_of_days_attendance']=0
            student['total_no_of_days_attendance']=0
            student['total_no_of_days_attendance_absent']=0
            if student['subject_list'][index1]['subject_part_type_code_name'] not in student['part_wise_details']:
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']] ={}
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['part_name'] = student['subject_list'][index1]['subject_part_type_code_name']
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['obtained_marks'] = 0
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['obtained_marks_round_off_to_upper_number'] = 0
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['total_marks'] = 0
            # Initialize subject_names_part_wise and obtained_marks_list_part_wise_graph for graph generation
            part_type_code = student['subject_list'][index1]['subject_part_type_code_name']
            if part_type_code not in student['subject_names_part_wise']:
                student['subject_names_part_wise'][part_type_code] = []
            if part_type_code not in student['obtained_marks_list_part_wise_graph']:
                student['obtained_marks_list_part_wise_graph'][part_type_code] = []
            if part_type_code not in student['obtained_marks_list_part_wise']:
                student['obtained_marks_list_part_wise'][part_type_code] = []
            student['subject_names_part_wise'][part_type_code].append(subjectData.get('subject__name', ''))
            subject = subjectData['subject']
            if subject in subjectConfigurationMapping:
                student['subject_list'][index1]['id'] = subject
                if not subjectConfigurationMapping[subject]['max_marks']:
                    subjectConfigurationMapping[subject]['max_marks']=0
                if not subjectConfigurationMapping[subject]['min_marks']:
                    subjectConfigurationMapping[subject]['min_marks']=0
                student['subject_list'][index1]['max_marks'] = subjectConfigurationMapping[subject][
                    'max_marks']
                student['total_marks'] += subjectConfigurationMapping[subject][
                    'max_marks']
                student['subject_list'][index1]['min_marks'] = subjectConfigurationMapping[subject][
                    'min_marks']
                student['subject_list'][index1]['result'] = 'fail'
                student['subject_list'][index1]['marks'] = 0
                student['subject_list'][index1]['marks_round_off_to_upper_number'] = 0
                student['subject_list'][index1]['attendance_status'] = 'Present'
                student['subject_list'][index1]['unattended_schedule_marks'] = {}
                student['subject_list'][index1]['exam_details'] = {}
                student['subject_list'][index1]['merge_details'] = {}
                student['exam_details'] = {}
                tempTotalObtainedMarks = 0
                config_marks_of_test=0
                config_max_marks_of_test=0
                config_marks_of_exam=0
                config_max_marks_of_exam=0
                subject_config_list = []
                for index2,examSubject in enumerate(subjectConfigurationMapping[subject]['subject_exam_data']):
                    if examSubject['is_only_grade_for_config']:
                        examSubject['marks']
                    if student['id'] in studentSubjectSchedule:
                        if subject not in studentSubjectSchedule[student['id']] or examSubject['exam'] not in \
                                studentSubjectSchedule[student['id']][subject]:
                            if not student['subject_list'][index1]['unattended_schedule_marks'].get(examSubject['exam']):
                                student['subject_list'][index1]['unattended_schedule_marks'][examSubject['exam']] = {}
                            student['subject_list'][index1]['unattended_schedule_marks'][examSubject['exam']].update(
                                {'exam': examSubject['exam'], 'exam_description': examSubject['exam_description'], 'exam_name': examSubject['exam_name']}
                            )
                            continue
                        orignalMarkData = studentSubjectSchedule[student['id']][subject][examSubject['exam']]
                        
                        # for i in orignalMarkData:
                        #     print(i)
                        if orignalMarkData['exam'] not in attendance_data:
                            attendance_data[orignalMarkData['exam']]={
                                "max_no_of_days_attendance":0,
                                "no_of_days_attendance":0,
                                "absent_days":0
                            }
                        if 'exam' in orignalMarkData and orignalMarkData['exam'] not in student['subject_list'][index1]['exam_details']:
                            subject_config_list.append(orignalMarkData['exam'])
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]={}
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['exam']=orignalMarkData['exam']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['exam_description']=orignalMarkData['exam_description']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['exam_name']=orignalMarkData['exam_name']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['total_marks']=orignalMarkData['total_marks']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['total_marks_written']=orignalMarkData['total_marks_written']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['total_marks_cum']=orignalMarkData['total_marks_cum']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['total_max_marks']=orignalMarkData['total_max_marks']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['total_min_marks']=orignalMarkData['total_min_marks']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['obtained_grade']=orignalMarkData['obtained_grade']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['remark__name']=orignalMarkData['remark__name']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['marked_attendance_days']=orignalMarkData['marked_attendance_days']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['max_no_of_days_attendance']=orignalMarkData['max_no_of_days_attendance']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['attendance_status']=orignalMarkData['attendance_status']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['cumulative_data']=orignalMarkData['cum_data']
                            if 'percentage' in orignalMarkData:
                                student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['percentage']=orignalMarkData['percentage']
                            if orignalMarkData['exam'] not in studentExamSchedule[student['id']]:
                                studentExamSchedule[student['id']][orignalMarkData['exam']] = {
                                    'exam_description': '',
                                    'exam_part1_total_marks': 0,
                                    'exam_part1_total_max_marks': 0,
                                    'exam_part1_total_min_marks': 0,
                                    'exam_part1_obtained_percentage': 0,
                                    'exam_part1_obtained_grade': '',
                                    'exam_name':'',
                                    'exam_id':orignalMarkData['exam'],
                                }
                            if orignalMarkData['part_type'] == 'part1':
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_description'] = orignalMarkData['exam_description']
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_total_marks']+=orignalMarkData['total_marks']
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_total_max_marks']+=orignalMarkData['total_max_marks']
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_total_min_marks']+=orignalMarkData['total_min_marks']
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_name']=orignalMarkData['exam_name']
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_obtained_grade'],studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_obtained_percentage']=get_grade_for_marks(total_grade_list, studentExamSchedule[student['id']][orignalMarkData['exam']]
                                                                                                                                       ['exam_part1_total_marks'],studentExamSchedule[student['id']][orignalMarkData['exam']]
                                                                                                                                       ['exam_part1_total_max_marks'],total_grade_plan)
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_total_marks_in_words']=num2words(studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_part1_total_marks'], lang='en')
                                studentExamSchedule[student['id']][orignalMarkData['exam']]['exam_id'] = orignalMarkData['exam']
                            student['exam_details'] = studentExamSchedule[student['id']]
                            try:
                                student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['absent_days']=orignalMarkData['max_no_of_days_attendance']-orignalMarkData['marked_attendance_days']
                            except:
                                student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['absent_days']=orignalMarkData['max_no_of_days_attendance']
                            attendance_data[orignalMarkData['exam']]['max_no_of_days_attendance'] = orignalMarkData['max_no_of_days_attendance']
                            student['total_max_no_of_days_attendance']+=orignalMarkData['max_no_of_days_attendance'] if 'max_no_of_days_attendance' in orignalMarkData and orignalMarkData['max_no_of_days_attendance'] else 0
                            attendance_data[orignalMarkData['exam']]['no_of_days_attendance'] = orignalMarkData['marked_attendance_days']
                            student['total_no_of_days_attendance']+=orignalMarkData['marked_attendance_days'] if orignalMarkData['marked_attendance_days'] else 0
                            try:
                                attendance_data[orignalMarkData['exam']]['absent_days'] = attendance_data[orignalMarkData['exam']]['max_no_of_days_attendance'] -  attendance_data[orignalMarkData['exam']]['no_of_days_attendance']
                                student['total_no_of_days_attendance_absent'] += attendance_data[orignalMarkData['exam']]['max_no_of_days_attendance'] -  attendance_data[orignalMarkData['exam']]['no_of_days_attendance']
                            except:
                                attendance_data[orignalMarkData['exam']]['absent_days'] = attendance_data[orignalMarkData['exam']]['max_no_of_days_attendance']
                                student['total_no_of_days_attendance_absent'] += attendance_data[orignalMarkData['exam']]['max_no_of_days_attendance'] if 'max_no_of_days_attendance' in attendance_data[orignalMarkData['exam']] and attendance_data[orignalMarkData['exam']]['max_no_of_days_attendance'] else 0
                        if not orignalMarkData or orignalMarkData['attendance_status'] == 'Absent':
                            student['subject_list'][index1]['attendance_status'] = 'Absent'
                            tempTotalObtainedMarks = 0
                            orignalMarkData['marks'] =0
                            examSubject['config_marks'] = 0
                            if 'cum_data' in orignalMarkData and orignalMarkData['cum_data']:
                                orignalMarkData['total_marks']=orignalMarkData['total_marks']
                            else:
                                orignalMarkData['total_marks']=0
                            if 'cum_marks' in examSubject and examSubject['cum_marks']:
                                student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']+examSubject['cum_marks']
                            else:
                                student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']
                            if 'cum_data' in orignalMarkData and orignalMarkData['cum_data'] and 'cum_marks' in examSubject and examSubject['cum_marks']:
                                tempTotalObtainedMarks += ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                                examSubject['config_marks'] = ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                                student['subject_list'][index1]['config_marks'+str(index2)]=examSubject['config_marks']
                                student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']+examSubject['cum_marks']
                        elif examSubject['is_only_grade_for_config']:
                            tempTotalObtainedMarks=0
                            if orignalMarkData['grade']:
                                if orignalMarkData['exam'] == examSubject['exam']:
                                    student['subject_list'][index1]['grade'] = orignalMarkData['grade']
                                    examSubject['config_marks']=0
                        elif 'cum_marks' in examSubject and examSubject['cum_marks']:
                            tempTotalObtainedMarks += ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                            examSubject['config_marks'] = ((orignalMarkData['total_marks']*(examSubject['marks']+examSubject['cum_marks']))/orignalMarkData['total_max_marks'])
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['config_marks']=examSubject['config_marks']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['config_max_marks']=examSubject['marks']+examSubject['cum_marks']
                            student['subject_list'][index1]['config_marks'+str(index2)]=examSubject['config_marks']
                            student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']+examSubject['cum_marks']
                        else:
                            tempTotalObtainedMarks += (
                                    (orignalMarkData['marks'] * examSubject['marks']) / orignalMarkData['max_marks'])
                            examSubject['config_marks']=((orignalMarkData['marks'] * examSubject['marks']) / orignalMarkData['max_marks'])
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['config_marks']=examSubject['config_marks']
                            student['subject_list'][index1]['exam_details'][orignalMarkData['exam']]['config_max_marks']=examSubject['marks']
                            student['subject_list'][index1]['config_marks'+str(index2)]=examSubject['config_marks']
                            student['subject_list'][index1]['config_max_marks'+str(index2)]=examSubject['marks']
                        student['subject_list'][index1]['config_marks_of_test']=examSubject['config_marks'] if examSubject.get('config_marks') else 0
                        for merge in merge_list:
                            if examSubject['exam'] in merge['exams']:
                                if merge['merge_id'] not in student['subject_list'][index1]['merge_details']:
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']] = {}
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge']=merge['merge_id']
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge_name']=merge['merge_name']
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge_marks']=0
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge_max_marks']=0
                                student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge_marks']+=examSubject['config_marks'] if examSubject.get('config_marks') else 0
                                if examSubject['cum_marks']:
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge_max_marks']+=(examSubject['marks']+examSubject['cum_marks'])
                                else:
                                    student['subject_list'][index1]['merge_details'][merge['merge_id']]['merge_max_marks']+=examSubject['marks'] if examSubject['marks'] else 0
                        for merge in merge_list:
                            if examSubject['exam'] in merge['exams']:
                                config_marks_of_test+=examSubject['config_marks'] if examSubject.get('config_marks') else 0
                                if examSubject['cum_marks']:
                                    config_max_marks_of_test+=(examSubject['marks']+examSubject['cum_marks'])
                                else:
                                    config_max_marks_of_test+=examSubject['marks'] if examSubject['marks'] else 0
                            else:
                                config_marks_of_exam+=examSubject['config_marks'] if examSubject.get('config_marks') else 0
                                if examSubject['cum_marks']:
                                    config_max_marks_of_exam+=(examSubject['marks'] if examSubject['marks'] else 0+examSubject['cum_marks'])
                                else:
                                    config_max_marks_of_exam+=examSubject['marks'] if examSubject['marks'] else 0
                for merge in student['subject_list'][index1]['merge_details']:
                    student['subject_list'][index1]['merge_details'][merge]['merge_marks_round_off_to_upper_number'] = math.ceil(
                        student['subject_list'][index1]['merge_details'][merge]['merge_marks']
                    )
                    student['obtained_marks_round_off']+=subjectConfigurationMapping[subject]['max_marks']
                    student['subject_list'][index1]['marks_round_off_to_upper_number']+=student['subject_list'][index1]['merge_details'][merge]['merge_marks_round_off_to_upper_number']+config_marks_of_exam
                    student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['obtained_marks_round_off_to_upper_number'] +=student['subject_list'][index1]['merge_details'][merge]['merge_marks_round_off_to_upper_number']
                    if 'obtained_marks_of_merge'+str(merge) not in student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]:
                        student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['obtained_marks_of_merge'+str(merge)] =0
                        student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['max_marks_of_merge'+str(merge)] =0
                    student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['obtained_marks_of_merge'+str(merge)] += math.ceil(student['subject_list'][index1]['merge_details'][merge]['merge_marks'])
                    student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['max_marks_of_merge'+str(merge)] += student['subject_list'][index1]['merge_details'][merge]['merge_max_marks']
                student['subject_list'][index1]['config_marks_of_test']=math.ceil(config_marks_of_test)
                student['subject_list'][index1]['config_max_marks_of_test']=math.ceil(config_max_marks_of_test)
                config_test_grade = GradePlan.objects.filter(grade_type=1).first()
                config_test_grade_list = Grade.objects.filter(grade_plan=config_test_grade).values()
                student['subject_list'][index1]['config_marks_of_test_grade'],student['subject_list'][index1]['config_marks_of_test_percentage']=get_grade_for_marks(config_test_grade_list,student['subject_list'][index1]['config_marks_of_test'], student['subject_list'][index1]['config_max_marks_of_test'],config_test_grade)
                student['subject_list'][index1]['config_marks_of_exam']=config_marks_of_exam
                student['subject_list'][index1]['config_max_marks_of_exam']=config_max_marks_of_exam
                student['subject_list'][index1]['marks'] = tempTotalObtainedMarks
                student['subject_list'][index1]['marks_round_off_to_upper_number']=student['subject_list'][index1]['config_marks_of_test']+student['subject_list'][index1]['config_marks_of_exam']
                # Populate obtained_marks_list_part_wise_graph for graph generation
                part_type_code = student['subject_list'][index1]['subject_part_type_code_name']
                if part_type_code in student['obtained_marks_list_part_wise_graph']:
                    # print(part_type_code,'part_type_code')
                    if 'marks' in student['subject_list'][index1] and student['subject_list'][index1]['marks']:
                        student['obtained_marks_list_part_wise'][student['subject_list'][index1]['subject_part_type_code_name']].append(student['subject_list'][index1]['marks'])
                        student['obtained_marks_list_part_wise_graph'][student['subject_list'][index1]['subject_part_type_code_name']].append(student['subject_list'][index1]['marks'])
                    else:
                        student['obtained_marks_list_part_wise'][student['subject_list'][index1]['subject_part_type_code_name']].append(0)
                        if 'grade' in student['subject_list'][index1] and student['subject_list'][index1]['grade']:
                            grade_plans = Grade.objects.filter(grade_plan__grade_type=2,name = student['subject_list'][index1]['grade']).values().last()
                            # print(grade_plans,'grade_plans')
                            student['obtained_marks_list_part_wise_graph'][student['subject_list'][index1]['subject_part_type_code_name']].append(grade_plans['to_range'])
                        else:
                            student['obtained_marks_list_part_wise_graph'][student['subject_list'][index1]['subject_part_type_code_name']].append(0)
                    # student['obtained_marks_list_part_wise_graph'][part_type_code].append(tempTotalObtainedMarks if tempTotalObtainedMarks else 0)
                if student['subject_list'][index1]['marks']:
                    try:
                        grade_plan_obj_sub=FinalResultSectionMapping.objects.get(standard_section=standardSectionId,subject=subjectData['subject'],grade_plan__isnull=False)
                        grade_plan=grade_plan_obj_sub.grade_plan
                    except FinalResultSectionMapping.DoesNotExist:
                        try:
                            grade_plan_obj_sub=FinalResultSectionApproval.objects.get(standard_section=standardSectionId, final_result_config=final_result_section_approval_id)
                            grade_plan=grade_plan_obj_sub.grade_plan
                        except FinalResultSectionApproval.DoesNotExist:
                            grade_plan_obj_sub=None
                            grade_plan=None
                    grade_list = Grade.objects.filter(grade_plan=grade_plan).values()
                    student['subject_list'][index1]['percentage']=0
                    if 'grade' not in student['subject_list'][index1] or not student['subject_list'][index1]['grade']:
                        student['subject_list'][index1]['grade'],student['subject_list'][index1]['percentage']=get_grade_for_marks(grade_list,student['subject_list'][index1]['marks'], student['subject_list'][index1]['max_marks'],grade_plan)
                        student['subject_list'][index1]['grade_round_off_to_upper_number'],student['subject_list'][index1]['percentage_round_off_to_upper_number']=get_grade_for_marks(grade_list,student['subject_list'][index1]['marks_round_off_to_upper_number'], student['subject_list'][index1]['max_marks'],grade_plan)
                if student['subject_list'][index1]['subject_part_type_code_name']=='part1':
                    student['part1_obtained_marks']+=tempTotalObtainedMarks
                    student['part1_obtained_marks_round_off'] += (student['subject_list'][index1]['config_marks_of_test']+student['subject_list'][index1]['config_marks_of_exam'])
                    student['part1_total_marks'] += subjectConfigurationMapping[subject]['max_marks']
                    student['part1_total_config_marks_of_test']+=student['subject_list'][index1]['config_marks_of_test']
                    student['part1_total_config_max_marks_of_test']+=student['subject_list'][index1]['config_max_marks_of_test']
                    student['part1_total_config_max_marks_of_exam']+=student['subject_list'][index1]['config_max_marks_of_exam']
                    student['part1_total_config_marks_of_exam']+=student['subject_list'][index1]['config_marks_of_exam']
                    student['part1_obtained_marks_round_off_in_words']=num2words(student['part1_obtained_marks_round_off'], lang='en')
                if institute_objs.code == 'nandinividyanikethana' and self.request.GET.get('print_marks_card'):
                    try:
                        if "part1" in student['subject_names_part_wise'] and "part1" in student['obtained_marks_list_part_wise_graph']:
                            student['part1_graph'] = generate_chart(student['subject_names_part_wise']["part1"], student['obtained_marks_list_part_wise_graph']["part1"] ,'Scholastic', chart_type="bar", colors=None)
                        obtained_part_wise_graph = []
                        subject_names_part_wise =[]
                        if "parta" in student['obtained_marks_list_part_wise_graph']:
                            obtained_part_wise_graph+=student['obtained_marks_list_part_wise_graph']['parta']
                        if "partb" in student['obtained_marks_list_part_wise_graph']:
                            obtained_part_wise_graph+=student['obtained_marks_list_part_wise_graph']['partb']
                        if "partc" in student['obtained_marks_list_part_wise_graph']:
                            obtained_part_wise_graph+=student['obtained_marks_list_part_wise_graph']['partc']
                        if "parta" in student['subject_names_part_wise']:
                            subject_names_part_wise+=student['subject_names_part_wise']['parta']
                        if "partb" in student['subject_names_part_wise']:
                            subject_names_part_wise+=student['subject_names_part_wise']['partb']
                        if "partc" in student['subject_names_part_wise']:
                            subject_names_part_wise+=student['subject_names_part_wise']['partc']
                        if obtained_part_wise_graph and sum(obtained_part_wise_graph):
                            student['part2_graph'] = generate_chart(subject_names_part_wise, 
                                                    obtained_part_wise_graph ,'Co-Scholastic', chart_type="pie", colors=None)
                        if "char" in student['subject_names_part_wise'] and "char" in student['obtained_marks_list_part_wise_graph']:
                            student['part3_graph'] = generate_chart(student['subject_names_part_wise']["char"], student['obtained_marks_list_part_wise_graph']["char"] ,'Personality And Character', chart_type="line", colors=None)
                    except Exception as e:
                        print(f"Error generating graphs for student {student.get('id')}: {str(e)}")
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['obtained_marks'] +=tempTotalObtainedMarks
                student['part_wise_details'][student['subject_list'][index1]['subject_part_type_code_name']]['total_marks'] +=subjectConfigurationMapping[subject]['max_marks']
                student['obtained_marks'] += tempTotalObtainedMarks
                student['obtained_marks_round_off'] += (student['subject_list'][index1]['config_marks_of_test']+student['subject_list'][index1]['config_marks_of_exam'])
                student['total_config_marks_of_test']+=student['subject_list'][index1]['config_marks_of_test']
                student['total_config_max_marks_of_test']+=student['subject_list'][index1]['config_max_marks_of_test']
                if student['subject_list'][index1]['marks'] > \
                        student['subject_list'][index1]['min_marks']:
                    student['subject_list'][index1]['result'] = 'pass'
                if student['total_result'] != 'fail':
                    student['total_result'] = student['subject_list'][index1]['result']
            for exam in subjectConfigurationMapping['exam_list']:
                if exam not in subject_config_list:
                    student['subject_list'][index1]['exam_details'][exam] = subjectConfigurationMapping['exam_list'][exam]
            try:
                student['subject_list'][index1]['exam_details'] = dict(sorted(student['subject_list'][index1]['exam_details'].items(), key=lambda item: int(item[1]['exam_description'])))
            except:
                pass
            student['subject_list'][index1]['exam_details_list'] = student['subject_list'][index1]['exam_details'].values() #nikhil
            exam_keys_order = list(student['exam_details'].keys())
        student['subject_list'] = sorted(student['subject_list'], key=lambda v: (isinstance(v.get('subject__subject_code', "NA"), str), v.get('subject__subject_code', "NA")))
        student['exam_details'] = studentExamSchedule[student['id']] if student['id'] in studentExamSchedule else {}
        try:
            student['exam_details'] = dict(sorted(student['exam_details'].items(), key=lambda item: int(item[1]['exam_description']) if 'exam_description' in item[1] and item[1]['exam_description'] else 0))
        except:
            pass
        student['exam_details_list'] = student['exam_details'].values()

        try:                                                                              # for sorting subject in the markcard sep 4th nikil sir
            exam_order = sorted(
                student['exam_details'].items(),
                key=lambda item: int(item[1]['exam_description']) if 'exam_description' in item[1] and item[1]['exam_description'] else 0
            )
            exam_keys_order = [k for k, _ in exam_order]
        except:
            exam_keys_order = list(student['exam_details'].keys())
        student['exam_details'] = {k: student['exam_details'][k] for k in exam_keys_order}
        student['exam_details_list'] = list(student['exam_details'].values())

        # Step 3: Attach unattended (unscheduled) exams to each subject’s exam_details, then sort by exam_description.
        for index1, subjectData in enumerate(student['subject_list']):
            subject_exam_details = dict(subjectData.get('exam_details') or {})
            unattended = subjectData.get('unattended_schedule_marks')
            # `unattended_schedule_marks` is normally a dict: { exam_id: {exam, exam_name, ...} }.
            # Older code paths may use a list of {'subject': exam_id, 'exam': <config dict>}.
            if isinstance(unattended, dict):
                for exam_id, exam_meta in unattended.items():
                    if exam_id in subject_exam_details:
                        continue
                    if not isinstance(exam_meta, dict):
                        continue
                    subject_exam_details[exam_id] = {
                        'exam': exam_meta.get('exam', exam_id),
                        'exam_description': exam_meta.get('exam_description'),
                        'exam_name': exam_meta.get('exam_name'),
                    }
            elif isinstance(unattended, list):
                for item in unattended:
                    if not isinstance(item, dict):
                        continue
                    exam_id = item.get('subject')
                    exam_meta = item.get('exam')
                    if exam_id is None or exam_meta is None:
                        continue
                    if not isinstance(exam_meta, dict):
                        try:
                            exam_meta = dict(exam_meta)
                        except (TypeError, ValueError):
                            continue
                    if exam_id not in subject_exam_details:
                        subject_exam_details[exam_id] = {
                            'exam': exam_meta.get('exam', exam_id),
                            'exam_description': exam_meta.get('exam_description'),
                            'exam_name': exam_meta.get('exam_name'),
                        }
            try:
                subject_exam_details = dict(sorted(
                    subject_exam_details.items(),
                    key=lambda item: int(str(item[1].get('exam_description') or 0)),
                ))
            except (ValueError, TypeError):
                subject_exam_details = dict(sorted(
                    subject_exam_details.items(),
                    key=lambda item: str(item[1].get('exam_description') or ''),
                ))
            student['subject_list'][index1]['exam_details'] = subject_exam_details
            student['subject_list'][index1]['exam_details_list'] = list(subject_exam_details.values())
        # student['total_max_no_of_days_attendance']=y['2data']['max_no_of_days_attendance']+y['3data']['max_no_of_days_attendance']
        # try:
        #     student['total_no_of_days_attendance']=y['2data']['no_of_days_attendance']+y['3data']['no_of_days_attendance']
        # except:
        #     student['total_no_of_days_attendance']=0
        # student['total_no_of_days_attendance_absent']=y['2data']['absent_days']+y['3data']['absent_days']
        try:
            grade_plan_obj_config = FinalResultSectionApproval.objects.get(standard_section=standardSectionId,final_result_config=final_result_section_approval_id)
            grade_list_config = Grade.objects.filter(grade_plan=grade_plan_obj_config.total_grade_plan).values()
            total_grade_plan_config = grade_plan_obj_config.total_grade_plan
        except FinalResultSectionApproval.DoesNotExist:
            grade_plan_obj_config = []
            grade_list_config = []
            total_grade_plan_config =None
        student['grade'],student['total_percentage'] = get_grade_for_marks(grade_list_config, student['obtained_marks'], student['total_marks'],total_grade_plan_config)
        student['part1_grade'],student['part1_total_percentage'] = get_grade_for_marks(grade_list_config,student['part1_obtained_marks'],student['part1_total_marks'],total_grade_plan_config)
        student['grade_round_off'],student['total_percentage_round_off'] = get_grade_for_marks(grade_list_config, student['obtained_marks_round_off'], student['total_marks'],total_grade_plan_config)
        student['part1_grade_round_off'],student['part1_total_percentage_round_off'] = get_grade_for_marks(grade_list_config,student['part1_obtained_marks_round_off'],student['part1_total_marks'],total_grade_plan_config)
        student['part1_config_test_grade'],student['part1_config_test_total_percentage'] = get_grade_for_marks(config_test_grade_list,student['part1_total_config_marks_of_test'],student['part1_total_config_max_marks_of_test'],config_test_grade)
        # student['grade'] = get_grade_for_marks(gradeData, student['obtained_marks'], student['total_marks'], grade_plan_id)
        # nikhil    
        # student['subject_list'] = get_student_grade(student['subject_list'], academicYear, 'marks', 'max_marks')
        if student['id'] in studentFinalResult:
            if 'status' in studentFinalResult[student['id']]:
                student['total_result'] = studentFinalResult[student['id']]['status']
            if 'is_announced' in studentFinalResult[student['id']] and studentFinalResult[student['id']]['is_announced']:
                isAnnounced = True
                student['is_announced'] = True
            if 'id' in studentFinalResult[student['id']]:
                student['final_result_id'] = studentFinalResult[student['id']]['id']
            # Assign rank from final result (same as working endpoint)
            if 'section_rank' in studentFinalResult[student['id']]:
                student['section_rank'] = studentFinalResult[student['id']]['section_rank']
            if 'standard_rank' in studentFinalResult[student['id']]:
                student['standard_rank'] = studentFinalResult[student['id']]['standard_rank']
        updated_student_data.append(student)
    return {'studentData': updated_student_data, 'isAnnounced': isAnnounced}

def add_update_question_mark(self, return_data):
    response = {'Result': True, 'Reason': 'Data Added Successfully'}
    deletable_list = return_data['deletable_list']
    deletable_list = []
    standard_section_id = return_data['standard_section']
    data = return_data['question_mark_details']
    schedule_ids = set()
    exam_ids = set()
    student_subject_data = []
    student_unique_check = []
    student_mark_dict = {}
    question_wise_marks_exam = {}
    deletabe_data = StudentMarkQuestionWise.objects.filter(id__in=deletable_list)
    deletable_student_mark_data = list(
        deletabe_data.values_list('exam_schedule_question_mapping__exam_schedule__exam', flat=True))
    for mark_data in data:
        for subject_data in mark_data['subject_list']:
            schedule_ids.add(subject_data['exam_schedule_question_mapping'])
        if mark_data['student'] in student_unique_check:
            raise ValidationError('Duplicate Student Data Found')
        student_unique_check.append(mark_data['student'])
    student_mark_data = StudentMarkQuestionWise.objects.filter(exam_schedule_question_mapping__in=list(schedule_ids)).values()
    for student in student_mark_data:
        if student['student_id'] not in student_mark_dict:
            student_mark_dict[student['student_id']]={}
        if student['id'] not in student_mark_dict[student['student_id']]:
            student_mark_dict[student['student_id']][student['id']]=student
    exam_schedule_details = ExamScheduleQuestionmapping.objects.filter(id__in=list(schedule_ids)).values('exam_schedule','question_number','id','exam_schedule__subject',
                                                                                                         'sub_question_number','description','sequence','max_marks','min_marks',
                                                                                                         'course_outcome','group_name','exam_schedule__exam','option_link_id__group_name','option_link_id')
    exam_schedule_id_mapping = {}
    nowdate = datetime.now().date()
    nowtime = datetime.now().time()
    for exam_obj in exam_schedule_details:
        exam_ids.add(exam_obj['exam_schedule__exam'])
        # if nowdate <= exam_obj['fordate']:
        #     if nowdate != exam_obj['fordate'] or nowtime <= exam_obj['start_time']:
        #         raise ValidationError('Exam not yet started/ended')
    # grade_plan_data = Grade.objects.filter(
    #     grade_plan__grade_type=2, grade_plan__in=grade_plan_ids
    # ).values('name', 'grade_plan')
    # grade_plan_mapping = {}
    # for grade_plan in grade_plan_data:
    #     if grade_plan['grade_plan'] not in grade_plan_mapping:
    #         grade_plan_mapping[grade_plan['grade_plan']] = []
    #     grade_plan_mapping[grade_plan['grade_plan']].append(grade_plan['name'])
    exam_ids = set(list(exam_ids) + deletable_student_mark_data)
    if len(exam_ids) > 1:
        raise ValidationError('You can schedule only one exam at a time')
    if len(exam_ids) <= 0:
        response['Reason'] = 'No changes to update'
        return response
    exam_obj_d = Exam.objects.get(id=list(exam_ids)[0])
    ApprovalService.get_approval_status(self, exam_obj_d)
    for schedule in exam_schedule_details:
        exam_ids.add(schedule['exam_schedule__exam'])
        exam_schedule_id_mapping[schedule['id']] = schedule
    student_data = Student.get_student_for_standard(exam_obj_d.academic_year, \
                                                       [], [standard_section_id],
                                                       ['first_name', 'middle_name', 'last_name', 'id'])
    approved_standard_mark_list = get_approved_standard_section_list(exam_obj_d.id)
    if standard_section_id in approved_standard_mark_list:
        raise ValidationError('The Student marks is finalized you cant change.')
    student_data = {data['id']: data for data in student_data}
    data_to_save = []
    staff_id = 3
    group_name_option_group_name_mapping={}
    marks_obtained = 0
    student_mark=[]
    option_list =[]
    for mark_data in data:
        subject_list = []
        if mark_data['subject_list']:
            for subject_data in mark_data['subject_list']:
                schedule_id = subject_data['exam_schedule_question_mapping']
                if schedule_id not in exam_schedule_id_mapping:
                    raise ValidationError('Invalid exam schedule id')
                available_grades = []
                if not exam_schedule_id_mapping[schedule_id]['max_marks']:
                    raise ValidationError("Please update max_marks before entering marks.")
                # if is_marks and subject_data['attendance_status'] != "Absent" and (subject_data['marks'] == None or ('grade' in subject_data and subject_data['grade'])):
                #     raise ValidationError("Subject is not Grade only")
                if subject_data['attendance_status'] == "Absent":
                    subject_data['marks'] = None
                elif subject_data['marks'] and float(subject_data['marks']) > exam_schedule_id_mapping[schedule_id]['max_marks']:
                    raise ValidationError(
                        f'Please enter marks lesser than {exam_schedule_id_mapping[schedule_id]["max_marks"]}')
                if mark_data['student'] not in student_data:
                    raise ValidationError('Student not exist in given standard')
                subject_list.append(exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['exam_schedule__subject'])
                if (subject_data['attendance_status'] == 'Present' and subject_data['marks']) or (subject_data['attendance_status'] == "Absent"):
                    temp = {'marks': subject_data['marks'], 'exam_schedule_question_mapping': subject_data['exam_schedule_question_mapping'],
                            'student': mark_data['student'], 'staff': staff_id,'question_number':exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['question_number'],
                            'sub_question_number' : exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['sub_question_number'],'group_name':exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name'],
                            'option_link_id' : exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['option_link_id'],
                            'option_link_id__group_name' : exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['option_link_id__group_name'],
                            'attendance_status': subject_data['attendance_status'], 'exam_schedule': exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['exam_schedule'],
                            'is_active': True,
                            }
                    if exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name'] not in group_name_option_group_name_mapping:
                        group_name_option_group_name_mapping[exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name']] =''
                    if exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['option_link_id__group_name'] :
                        group_name_option_group_name_mapping[exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name']] = exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['option_link_id__group_name']
                        option_list.append(exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['option_link_id__group_name'])
                    if 'id' in subject_data:
                        temp['id'] = subject_data['id']
                    data_to_save.append(temp)
                    if mark_data['student'] not in question_wise_marks_exam:
                        question_wise_marks_exam[mark_data['student']]={}
                    if exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name'] not in question_wise_marks_exam[mark_data['student']]:
                        question_wise_marks_exam[mark_data['student']][exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name']] = {'marks_obtained':0}
                    question_wise_marks_exam[mark_data['student']][exam_schedule_id_mapping[subject_data['exam_schedule_question_mapping']]['group_name']]['marks_obtained']+= subject_data['marks']
            temp_validate_data = {'student': mark_data['student'], 'subject_list': subject_list}
            student_subject_data.append(temp_validate_data)
    for student in question_wise_marks_exam:
        marks_obtained = 0
        for group in group_name_option_group_name_mapping:
            if group not in option_list:
                if group_name_option_group_name_mapping[group] in question_wise_marks_exam[student]:
                    if question_wise_marks_exam[student][group]['marks_obtained'] < question_wise_marks_exam[student][group_name_option_group_name_mapping[group]]['marks_obtained']:
                        marks_obtained += question_wise_marks_exam[student][group_name_option_group_name_mapping[group]]['marks_obtained']
                    else:
                        marks_obtained += question_wise_marks_exam[student][group]['marks_obtained']
                else:
                    marks_obtained += question_wise_marks_exam[student][group]['marks_obtained']
        temp_student = {
            'attendance_status' : data_to_save[0]['attendance_status'],
            'exam_schedule': data_to_save[0]['exam_schedule'],
            'marks':marks_obtained,
            'student':student,
            'staff': 3,
            'is_active':1
        }
        student_mark.append(temp_student)
    check_subject_assigned_to_student(student_subject_data)
    with transaction.atomic(using=get_current_db_name()):
        for student in student_mark:
            try:
                mark_instance = StudentMark.objects.get(exam_schedule_id = data_to_save[0]['exam_schedule'],student_id=student['student'])
            except:
                mark_instance = None
            if mark_instance:
                serializer = StudentMarkSerializer(instance=mark_instance, data=student, partial=True)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            else:
                serializer = StudentMarkSerializer(data=student)
                serializer.is_valid(raise_exception=True)
                serializer.save()
        if len(deletable_list) > 0:
            deletabe_data.update(is_active=False)
        if data_to_save:
            response = SharedService.add_or_update_data(self, data_to_save)
    return response

def get_question_mark(self,data):
    response = {'data': {}}
    examId = self.request.GET.get('exam_id')
    standardSectionId = self.request.GET.get('standard_section_id')
    subject_id = self.request.GET.get('subject_id')
    student_ids = self.request.GET.get('student_id')
    approval_data = StudentMarkSectionWiseApproval.objects.filter(exam=examId,
                                                                standard_section=standardSectionId).values()
    institute_objs = Institute.get_institute(self)
    if self.request.GET.get('exam_result'):
        if not approval_data or approval_data[0]['approval_status'] != 1:
            raise ValidationError('Student marks are not approved for the section')
    exam_obj = Exam.objects.get(id=examId)
    standard_sec_obj = StandardSectionMapping.objects.get(id=standardSectionId)
    values_list = ['id', 'first_name', 'middle_name','last_name', 'current_reg_num', 'student_name']
    custom_annotate = {'student_name': Concat('first_name', V(' '), 'middle_name',V(' '), 'last_name')}
    if student_ids:
        student_data = Student.objects.filter(
            id__in=student_ids
        ).annotate(**custom_annotate).values(*values_list)
    else:
        student_data = Student.get_student_for_standard(None, None, [standardSectionId], values_list, custom_annotate)
    student_ids = []
    studentIdDict = {}
    part_type_list = SubjectPartType.objects.all().values()
    part_types = [{'name': s['name']} for s in part_type_list]
    for i in student_data:
        student_ids.append(i['id'])
        i['student'] = i['id']
        studentIdDict[i['id']] = i
    filterQuery = {
                'exam_schedule__exam': examId}
    scheduleQuery = {'exam': examId}
    scheduleQuery['standard_section'] = standardSectionId
    scheduleQuery['subject_id'] = subject_id
    scheduleQuery['sub_schedule_parent'] = None
    scheduleData = ExamSchedule.objects.filter(**scheduleQuery).values('id', 'subject', 'subject__name', 'min_marks',
                                                                    'max_marks', 'subject__subject_part_type__name',
                                                                        'subject__subject_part_type',
                                                                        'subject__subject_part_type__code_name',
                                                                        'subject__subject_code', 'is_marks', 'grade_plan',
                                                                        'grade_plan__name','subject__codename', 'schedule_sequence'
                                                                    )
    student_admission_form = get_student_admission_form_details(self, student_ids)
    scheduleIds = []
    subjectList = {}
    scheduleSubjectIds = []
    grade_plan_ids = set()
    for schedule in scheduleData:
        scheduleIds.append(schedule['id'])
        temp = {'subject': schedule['subject'], 'subject_name': schedule['subject__name'], 'subject_part_type': schedule['subject__subject_part_type__name'],
                'min_marks': schedule['min_marks'], 'max_marks': schedule['max_marks'], 'schedule': schedule['id'],
                'subject_part_type_id': schedule['subject__subject_part_type'],
                'subject_part_type_code_name': schedule['subject__subject_part_type__code_name'],
                'subject__subject_code': schedule['subject__subject_code'], 'is_marks': schedule['is_marks'],
                'grade_plan': schedule['grade_plan'], 'grade_plan_name': schedule['grade_plan__name'],'subject__codename':schedule['subject__codename']
                }
        grade_plan_ids.add(schedule['grade_plan'])
        subjectList[schedule['subject']] = temp
        scheduleSubjectIds.append(schedule['subject'])
    studentSubjectData = SubjectStudent.objects.filter(student__in=student_ids,
                                                    academic_year=exam_obj.academic_year,
                                                    subject__in=scheduleSubjectIds).values('subject', 'student','student_id',
                                                                                            'subject__name', 'subject__subject_code','subject__codename',
                                                                                            subject_name=F('subject__name'),
                                                                                            subject_part_type=F('subject__subject_part_type__name'),
                                                                                            subject_part_type_id=F('subject__subject_part_type'),
                                                                                            subject_part_type_code_name=F('subject__subject_part_type__code_name')
                                                                                            )
    studentSubjectMapping = {}
    student_subject = []
    for studentSubject in studentSubjectData:
        student_subject.append(studentSubject['student_id'])
        if studentSubject['student'] in studentSubjectMapping:
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
        else:
            studentSubjectMapping[studentSubject['student']] = []
            studentSubjectMapping[studentSubject['student']].append(studentSubject)
    filterQuery['exam_schedule__in'] = scheduleIds
    filterQuery['is_active'] = True
    student_mark_data = []
    student_dict = {}
    exam_schedule_question_dict = {}
    question_list = []
    temp_question_data = ExamScheduleQuestionmapping.objects.filter(**filterQuery).values('exam_schedule','question_number','sub_question_number','description',
                                                                                          'is_active','sequence','max_marks','min_marks','course_outcome','group_name',
                                                                                          'course_outcome__name','option_link_id','id')
    for questions in temp_question_data:
        question_list.append(questions['id'])
        if questions['id'] not in exam_schedule_question_dict:
            exam_schedule_question_dict[questions['id']] = {}
        exam_schedule_question_dict[questions['id']] = questions 
    student_question_details = {}
    question_wise_mark_data = StudentMarkQuestionWise.objects.filter(exam_schedule_question_mapping__in = question_list,is_active = True).values('exam_schedule_question_mapping','marks','grade',
                                                                                                                                                 'student','staff','attendance_status','is_active','id','exam_schedule_question_mapping__exam_schedule__subject',
                                                                                                                                                 'exam_schedule_question_mapping__exam_schedule__subject__name')
    for question_mark in question_wise_mark_data:
        if question_mark['student'] not in student_question_details:
            student_question_details[question_mark['student']] = {}
        if question_mark['exam_schedule_question_mapping'] not in student_question_details[question_mark['student']]:
            student_question_details[question_mark['student']][question_mark['exam_schedule_question_mapping']] = {}
        student_question_details[question_mark['student']][question_mark['exam_schedule_question_mapping']] = question_mark

    student_data = Student.objects.filter(id__in = student_subject).values('id','first_name','middle_name','last_name')
    for student in student_data:
        if student['id'] not in student_dict:
            student_dict[student['id']] = {}
        student_dict[student['id']] = student
        student_dict[student['id']]['student'] = student['id']
        student_dict[student['id']]['student_name'] = get_full_name(student['first_name'],student['middle_name'],student['last_name'])
        temp_question_details = {}
        for questions in exam_schedule_question_dict:
            if questions not in temp_question_details:
                temp_question_details[questions] = {}
            temp_question_details[questions] = exam_schedule_question_dict[questions]
            if student['id'] in student_question_details and questions in student_question_details[student['id']]:
                temp_question_details[questions]['marks'] = student_question_details[student['id']][questions]['marks']
        student_dict[student['id']]['question_list'] = list(temp_question_details.values())
    response['data']['standard_name'] = standard_sec_obj.standard.name
    response['data']['section_name'] = standard_sec_obj.section.name
    response['data']['standard_section'] = standard_sec_obj.id
    response['data']['standard'] = standard_sec_obj.standard.id
    response['data']['approval_status'] = approval_data[0]['approval_status'] if approval_data else '0'
    response['data']['academic_year_details'] = {
        'start_date': exam_obj.academic_year.start_date,
        'end_date': exam_obj.academic_year.end_date,
        'academic_year': exam_obj.academic_year.id
    }
    response['data']['term_details'] = exam_obj.term.name
    response['data']['exam_details'] = exam_obj.exam_type.name
    response['data']['exam_code_name'] = exam_obj.exam_type.code
    response['data']['subject_name'] = scheduleData.first()['subject__name']
    response['data']['student_list'] = list(student_dict.values())
    response['data']['question_list'] = temp_question_data
    return response

def get_exam_data_exceptions(self,request):
    is_for_question_paper = self.request.GET.get('is_for_question_paper')
    is_for_marks_entry = self.request.GET.get('is_for_marks_entry')
    exam_data = {}
    today = datetime.now().today()
    response_data = []
    approved_exams = ApproveStatus.objects.filter(content_type__app_label='exams',content_type__model='exam',approval_status = 1).values_list('object_id',flat=True)
    exam_details = Exam.objects.filter(id__in = approved_exams,is_active=True).values('id','exam_type_id','exam_type__name')
    for exam in exam_details:
        exam_data[exam['id']] = exam
    filter_queryset={
        'exam__in':exam_data.keys()
    }
    standard_section_exam_subject = {}
    if is_for_question_paper:
        filter_queryset['questionframe_lastdate__lte'] =today
    if is_for_marks_entry:
        filter_queryset['marksentry_lastdate__lte'] = today
    exam_schedule_data = ExamSchedule.objects.filter(**filter_queryset).values('exam_id','id','standard_section_id','subject_id','questionframe_lastdate','subject__name',
                                                                               'marksentry_lastdate','standard_section__section__name','standard_section__standard__name')
    exam_schedule_list=[]
    for exam_schedule in exam_schedule_data:
        exam_schedule_list.append(exam_schedule['id'])
    if is_for_question_paper:
        exam_schedule_question_data = ExamScheduleQuestionmapping.objects.filter(exam_schedule__in = exam_schedule_list).values_list('exam_schedule_id',flat=True)
        for schedule in exam_schedule_data:
            temp_data={}
            if schedule['id'] not in exam_schedule_question_data:
                temp_data = schedule
                temp_data.update(exam_data[schedule['exam_id']])
            if temp_data:
                response_data.append(temp_data)
    if is_for_marks_entry:
        approved_list = []
        exam_marks_approval = StudentMarkSectionSubjectWiseApproval.objects.filter(approval_status=1).values()
        for section_subject in exam_marks_approval:
            key = str(section_subject['standard_section_id'])+'_'+str(section_subject['subject_id'])+'_'+str(section_subject['exam_id'])
            approved_list.append(key)
        for schedule in exam_schedule_data:
            key = str(schedule['standard_section_id'])+'_'+str(schedule['subject_id'])+'_'+str(schedule['exam_id'])
            temp_data = {}
            if key not in approved_list:
                temp_data = schedule
                temp_data.update(exam_data[schedule['exam_id']])
            if temp_data:
                response_data.append(temp_data)
    staff_standard_section_subject_mapping = StaffHourSubjectMapping.objects.filter().values('subject',
        'subject__name','standard_section','standard_section__standard__name','standard_section__section__name',
        'staff_teaching_hour__staff','staff_teaching_hour__staff__first_name','staff_teaching_hour__staff__middle_name','staff_teaching_hour__staff__last_name',
    )
    standard_section_subject_data = {}
    for staff_subject in staff_standard_section_subject_mapping:
        key = str(staff_subject['standard_section'])+'_'+str(staff_subject['subject'])
        if key not in standard_section_subject_data:
            standard_section_subject_data[key] = staff_subject
    for data in response_data:
        key = str(data['standard_section_id'])+'_'+str(data['subject_id'])
        if key in standard_section_subject_data:
            data.update(standard_section_subject_data[key])
    return response_data

def get_amrita_report(self,request):
    exam_id = self.request.GET.get('exam_id').split(',')
    standard_section_id = self.request.GET.get('standard_section_id')
    subject_id = self.request.GET.get('subject_id')
    exam_question_marks_dict = {}
    student_question_marks = {}
    standard_section_details = StandardSectionMapping.objects.filter(id=standard_section_id).values('id','standard_id','standard__name','section','section__name').first()
    institute_details = Institute.get_institute(self)
    department_details = Standard.objects.filter(id = standard_section_details['standard_id']).values('id','name','branch','branch__name').first()
    subject_details_obj = SubjectDetails.objects.filter(subject_id=subject_id)
    subject_details = SubjectDetailsReadSerializer(subject_details_obj)
    subject_co_details = SubjectCourseOutcomeMapping.objects.filter(is_active=True,subject_id=subject_id).values('subject','subject_id','course_outcome','course_outcome_id','description','target','course_outcome__name')
    student_dict = {'internal_details':{},'co_details':{}}
    details_for_heading = {'internal_details':{},'co_details':{'internals':{},'assignment':{}}}
    CIE_Average = {}
    group_option_mapping = {}
    option_list = []
    question_marks_dict = {}
    subject_co_dict = {'total_no_students':0,'total':{
        'students_achived_target':0,'per_students_achived_target':0
    }}
    enrollment_data = (Enrollment.objects.filter(standard_section_id=standard_section_id).annotate(student_name=Concat('student__first_name',V(' '),'student__middle_name',V(' '),'student__last_name',)).values('id', 'standard_section_id', 'student_id', 'student_name'))
    for student in enrollment_data:
        subject_co_dict['total_no_students'] += 1
        if student['student_id'] not in student_dict['internal_details']:
            student_dict['internal_details'][student['student_id']] = {}
        if student['student_id'] not in student_dict['co_details']:
            student_dict['co_details'][student['student_id']] = {'total':0} 
        student_dict['internal_details'][student['student_id']] = student
        student_dict['co_details'][student['student_id']].update(student)
    exam_schedule_list = ExamSchedule.objects.filter(exam__in=exam_id,subject = subject_id,standard_section = standard_section_id).values_list('id',flat=True)
    exam_question = ExamScheduleQuestionmapping.objects.filter(exam_schedule__in = exam_schedule_list,is_active = True).values('id','exam_schedule','question_number','sub_question_number','description', 'exam_schedule__exam','exam_schedule__exam__exam_type__name',
                                                                                                                               'is_active','sequence','max_marks','min_marks','course_outcome','group_name','option_link_id__group_name','course_outcome__name',
                                                                                                                               'option_link_id')
    for course_outcome in subject_co_details:
        if course_outcome['course_outcome'] not in subject_co_dict:
            subject_co_dict[course_outcome['course_outcome']] = {'internals':{'co_max_marks':0,'students_achived_target':0,'total_students':subject_co_dict['total_no_students'],'attainment_level':0,'per_students_achived_target':0,'mim_marks':0},
                                                                 'assessment':{'co_max_marks':0,'students_achived_target':0,'total_students':subject_co_dict['total_no_students'],'attainment_level':0,'per_students_achived_target':0,'mim_marks':0,
                                                                    }}
        subject_co_dict[course_outcome['course_outcome']]['internals'].update(course_outcome)
        subject_co_dict[course_outcome['course_outcome']]['assessment'].update(course_outcome)
        if course_outcome['course_outcome'] not in CIE_Average:
            CIE_Average[course_outcome['course_outcome']] = {'asignment_attainment':0,'internals_attainment':0,'course_outcome_name':course_outcome['course_outcome__name'],'total_attainment':0}
    for question in exam_question:
        if question['exam_schedule__exam'] not in details_for_heading['internal_details']:
            details_for_heading['internal_details'][question['exam_schedule__exam']] = {}
        details_for_heading['internal_details'][question['exam_schedule__exam']]['exam_name'] = question['exam_schedule__exam__exam_type__name']
        if question['course_outcome'] not in details_for_heading['internal_details'][question['exam_schedule__exam']]:
            details_for_heading['internal_details'][question['exam_schedule__exam']][question['course_outcome']]={'max_marks':0}
        details_for_heading['internal_details'][question['exam_schedule__exam']][question['course_outcome']]['course_outcome_name'] = question['course_outcome__name']
        if question['group_name'] not in group_option_mapping:
            group_option_mapping[question['group_name']] = ''
        if question['option_link_id']:
            group_option_mapping[question['group_name']] = question['option_link_id__group_name']
        option_list.append(question['option_link_id__group_name'])
        exam_question_marks_dict[question['id']] = question
    for question in exam_question:
        if 'Assignment' in question['exam_schedule__exam__exam_type__name'] and question['group_name'] not in option_list:
            if question['course_outcome'] not in details_for_heading['co_details']['assignment']:
                details_for_heading['co_details']['assignment'][question['course_outcome']] = {'max_marks':0}
            details_for_heading['co_details']['assignment'][question['course_outcome']]['max_marks'] += question['max_marks']
            details_for_heading['co_details']['assignment'][question['course_outcome']]['course_outcome_name'] = question['course_outcome__name']
            subject_co_dict[question['course_outcome']]['assessment']['co_max_marks']+= question['max_marks']
            subject_co_dict[question['course_outcome']]['assessment']['mim_marks'] = (subject_co_dict[question['course_outcome']]['assessment']['target']*subject_co_dict[question['course_outcome']]['assessment']['co_max_marks'])/100
        elif 'Assignment' not in question['exam_schedule__exam__exam_type__name'] and question['group_name'] not in option_list:
            subject_co_dict[question['course_outcome']]['internals']['co_max_marks']+= question['max_marks']
            subject_co_dict[question['course_outcome']]['internals']['mim_marks'] = (subject_co_dict[question['course_outcome']]['internals']['target']*subject_co_dict[question['course_outcome']]['internals']['co_max_marks'])/100
            if question['course_outcome'] not in details_for_heading['co_details']['internals']:
                details_for_heading['co_details']['internals'][question['course_outcome']] = {'max_marks':0}
                details_for_heading['co_details']['internals'][question['course_outcome']]['course_outcome_name'] = question['course_outcome__name']
            details_for_heading['co_details']['internals'][question['course_outcome']]['max_marks'] += question['max_marks']
    
    exam_question_marks = StudentMarkQuestionWise.objects.filter(exam_schedule_question_mapping__in = exam_question_marks_dict.keys(),is_active=True).values('exam_schedule_question_mapping','marks','grade','student','staff','attendance_status','student_id','exam_schedule_question_mapping__max_marks')
    for marks in exam_question_marks:
        marks.update(exam_question_marks_dict[marks['exam_schedule_question_mapping']])
        if marks:
            if marks['student_id'] not in student_question_marks:
                student_question_marks[marks['student_id']] = {}
            if marks['group_name'] not in student_question_marks[marks['student_id']]:
                student_question_marks[marks['student_id']][marks['group_name']] = {}
            if marks['sub_question_number'] not in student_question_marks[marks['student_id']][marks['group_name']]:
                student_question_marks[marks['student_id']][marks['group_name']][marks['sub_question_number']] = {}
            student_question_marks[marks['student_id']][marks['group_name']][marks['sub_question_number']] = marks
            if marks['student_id'] not in question_marks_dict:
                question_marks_dict[marks['student_id']] = {}
            if marks['group_name'] not in question_marks_dict[marks['student_id']]:
                question_marks_dict[marks['student_id']][marks['group_name']]={'marks':0}
            question_marks_dict[marks['student_id']][marks['group_name']]['marks']+=marks['marks']
    students_counted_for_assessment = {}
    students_counted_for_internals = {}
    for student in student_question_marks:
        for group in student_question_marks[student]:
            if group not in option_list:
                if group in group_option_mapping and group_option_mapping[group] in question_marks_dict[student] and question_marks_dict[student][group]['marks'] < question_marks_dict[student][group_option_mapping[group]]['marks']:
                    for sub_question in student_question_marks[student][group_option_mapping[group]]:
                        key1 = str(student_question_marks[student][group_option_mapping[group]][sub_question]['exam_schedule__exam'])+'_'+str(student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome'])
                        if key1 not in student_dict['internal_details'][student]:
                            student_dict['internal_details'][student][key1] = 0
                        if 'total_internal' not in student_dict['internal_details'][student]:
                            student_dict['internal_details'][student]['total_internal'] = 0
                        student_dict['internal_details'][student][key1] += student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                        student_dict['internal_details'][student]['total_internal'] += student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                        details_for_heading['internal_details'][student_question_marks[student][group_option_mapping[group]][sub_question]['exam_schedule__exam']][student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['max_marks'] =student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                        if 'Assignment' in student_question_marks[student][group_option_mapping[group]][sub_question]['exam_schedule__exam__exam_type__name']:
                            # details_for_heading['co_details']['assignment'][student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['max_marks'] =student_question_marks[student][group_option_mapping[group]][sub_question]['exam_schedule_question_mapping__max_marks']
                            key2 = 'assignment'+str(student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome'])
                            if key2 not in student_dict['co_details'][student]:
                                student_dict['co_details'][student][key2] = 0
                            student_dict['co_details'][student][key2] += student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                            student_dict['co_details'][student]['total'] += student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                            co_id = student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']
                            student_id = student

                            if student_dict['co_details'][student][key2] >= subject_co_dict[co_id]['assessment']['mim_marks']:
                                if co_id not in students_counted_for_assessment:
                                    students_counted_for_assessment[co_id] = set()

                                if student_id not in students_counted_for_assessment[co_id]:
                                    subject_co_dict[co_id]['assessment']['students_achived_target'] += 1
                                    students_counted_for_assessment[co_id].add(student_id)
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target']=(subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['students_achived_target']/subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['total_students'])*100
                                if subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >= 60:
                                    subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['attainment_level']=3
                                elif subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >= 50:
                                    subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['attainment_level']=2
                                elif subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >= 40:
                                    subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['attainment_level']=1
                                CIE_Average[student_question_marks[student][group][sub_question]['course_outcome']]['asignment_attainment']=subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['attainment_level']
                            if student_dict['co_details'][student]['total'] >= subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['mim_marks']:
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['students_achived_target']+=1
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target']=(subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['students_achived_target']/subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['total_students'])*100
                                if subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >= 60:
                                    subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['attainment_level']=3
                                elif subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >= 50:
                                    subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['attainment_level']=2
                                elif subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >= 40:
                                    subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['assessment']['attainment_level']=1
                                CIE_Average[student_question_marks[student][group][sub_question]['course_outcome']]['asignment_attainment']=subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['attainment_level']
                        else:
                            # details_for_heading['co_details']['internals'][student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['max_marks'] =student_question_marks[student][group_option_mapping[group]][sub_question]['exam_schedule_question_mapping__max_marks']
                            key2 = 'internals'+str(student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome'])
                            if key2 not in student_dict['co_details'][student]:
                                student_dict['co_details'][student][key2] = 0
                            student_dict['co_details'][student][key2] += student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                            student_dict['co_details'][student]['total'] += student_question_marks[student][group_option_mapping[group]][sub_question]['marks']
                            co_id = student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']
                            student_id = student

                            if student_dict['co_details'][student][key2] >= subject_co_dict[co_id]['internals']['mim_marks']:
                                if co_id not in students_counted_for_internals:
                                    students_counted_for_internals[co_id] = set()

                                if student_id not in students_counted_for_internals[co_id]:
                                    subject_co_dict[co_id]['internals']['students_achived_target'] += 1
                                    students_counted_for_internals[co_id].add(student_id)

                            # if student_dict['co_details'][student][key2] >= subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['mim_marks']:
                            #     subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['students_achived_target']+=1
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['per_students_achived_target']=(subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['students_achived_target']/subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['total_students'])*100
                            if subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['per_students_achived_target'] >= 60:
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['attainment_level']=3
                            elif subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['per_students_achived_target'] >= 50:
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['attainment_level']=2
                            elif subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['per_students_achived_target'] >= 40:
                                subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['attainment_level']=1
                            if student_dict['co_details'][student]['total'] >= subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['mim_marks']:
                                subject_co_dict['total']['students_achived_target']+=1
                                subject_co_dict['total']['per_students_achived_target']=(subject_co_dict['total']['students_achived_target']/subject_co_dict['total_no_students'])*100
                            if subject_co_dict['total']['per_students_achived_target'] >= 60:
                                subject_co_dict['total']['attainment_level']=3
                            elif subject_co_dict['total']['per_students_achived_target'] >= 50:
                                subject_co_dict['total']['attainment_level']=2
                            elif subject_co_dict['total']['per_students_achived_target'] >= 40:
                                subject_co_dict['total']['attainment_level']=1
                                CIE_Average[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals_attainment']=subject_co_dict[student_question_marks[student][group_option_mapping[group]][sub_question]['course_outcome']]['internals']['attainment_level']
                else:
                    for sub_question in student_question_marks[student][group]:
                        key1 = str(student_question_marks[student][group][sub_question]['exam_schedule__exam'])+'_'+str(student_question_marks[student][group][sub_question]['course_outcome'])
                        if key1 not in student_dict['internal_details'][student]:
                            student_dict['internal_details'][student][key1] = 0
                        student_dict['internal_details'][student][key1] += student_question_marks[student][group][sub_question]['marks']
                        if 'total_internal' not in student_dict['internal_details'][student]:
                            student_dict['internal_details'][student]['total_internal'] = 0
                        student_dict['internal_details'][student]['total_internal'] += student_question_marks[student][group][sub_question]['marks']
                        details_for_heading['internal_details'][student_question_marks[student][group][sub_question]['exam_schedule__exam']][student_question_marks[student][group][sub_question]['course_outcome']]['max_marks'] = student_question_marks[student][group][sub_question]['marks']
                        if 'Assignment' in student_question_marks[student][group][sub_question]['exam_schedule__exam__exam_type__name']:
                            # details_for_heading['co_details']['assignment'][student_question_marks[student][group][sub_question]['course_outcome']]['max_marks'] =student_question_marks[student][group][sub_question]['exam_schedule_question_mapping__max_marks']
                            key2 = 'assignment'+str(student_question_marks[student][group][sub_question]['course_outcome'])
                            if key2 not in student_dict['co_details'][student]:
                                student_dict['co_details'][student][key2] = 0
                            student_dict['co_details'][student][key2] += student_question_marks[student][group][sub_question]['marks']
                            student_dict['co_details'][student]['total'] += student_question_marks[student][group][sub_question]['marks']
                            co_id = student_question_marks[student][group][sub_question]['course_outcome']
                            student_id = student

                            if student_dict['co_details'][student][key2] >= subject_co_dict[co_id]['assessment']['mim_marks']:
                                if co_id not in students_counted_for_assessment:
                                    students_counted_for_assessment[co_id] = set()

                                if student_id not in students_counted_for_assessment[co_id]:
                                    subject_co_dict[co_id]['assessment']['students_achived_target'] += 1
                                    students_counted_for_assessment[co_id].add(student_id)
                            # if student_dict['co_details'][student][key2] >= subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['mim_marks']:
                            #     subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['students_achived_target']+=1
                                subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['per_students_achived_target']=(subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['students_achived_target']/subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['total_students'])*100
                                if subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >=60:
                                    subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['attainment_level']=3
                                elif subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >=50:
                                    subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['attainment_level']=2
                                elif subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['per_students_achived_target'] >=40:
                                    subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['attainment_level']=1
                                CIE_Average[student_question_marks[student][group][sub_question]['course_outcome']]['asignment_attainment']=subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['assessment']['attainment_level']
                        else:
                            # details_for_heading['co_details']['internals'][student_question_marks[student][group][sub_question]['course_outcome']]['max_marks'] =student_question_marks[student][group][sub_question]['exam_schedule_question_mapping__max_marks']
                            key2 = 'internals'+str(student_question_marks[student][group][sub_question]['course_outcome'])
                            if key2 not in student_dict['co_details'][student]:
                                student_dict['co_details'][student][key2] = 0
                            student_dict['co_details'][student][key2] += student_question_marks[student][group][sub_question]['marks']
                            student_dict['co_details'][student]['total'] += student_question_marks[student][group][sub_question]['marks']
                            co_id = student_question_marks[student][group][sub_question]['course_outcome']
                            student_id = student

                            if student_dict['co_details'][student][key2] >= subject_co_dict[co_id]['internals']['mim_marks']:
                                if co_id not in students_counted_for_internals:
                                    students_counted_for_internals[co_id] = set()

                                if student_id not in students_counted_for_internals[co_id]:
                                    subject_co_dict[co_id]['internals']['students_achived_target'] += 1
                                    students_counted_for_internals[co_id].add(student_id)

                            # if student_dict['co_details'][student][key2] >= subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['mim_marks']:
                            #     subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['students_achived_target']+=1
                                subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['per_students_achived_target']=(subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['students_achived_target']/subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['total_students'])*100
                                if subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['per_students_achived_target'] >= 60:
                                    subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['attainment_level']=3
                                elif subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['per_students_achived_target'] >= 50:
                                    subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['attainment_level']=2
                                elif subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['per_students_achived_target'] >= 60:
                                    subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['attainment_level']=1
                                CIE_Average[student_question_marks[student][group][sub_question]['course_outcome']]['internals_attainment']=subject_co_dict[student_question_marks[student][group][sub_question]['course_outcome']]['internals']['attainment_level']
    co_po_matrix_column = {'co':{
        'name':'CO','value':'course_outcome_name'
    }}
    co_po_matrix_data = read_subject_course_outcome_program_mapping_matrix(self,True)
    co_po_dict = {}
    co_po_total_data = {}
    co_po_matrix_summary = {}
    for co_po in co_po_matrix_data['co_po_mapping']:
        key1='_po'+str(co_po['program_outcome'])
        if co_po['course_outcome'] not in co_po_dict:
            co_po_dict[co_po['course_outcome']] = {}
        co_po_dict[co_po['course_outcome']]['course_outcome_name'] = co_po['course_outcome_name']
        co_po_dict[co_po['course_outcome']][key1] = co_po['value']
        if key1 not in co_po_total_data:
            co_po_total_data[key1]={'total':0,'no_of_data':0,'avg':0}
        co_po_total_data[key1]['total']+=co_po['value']
        co_po_total_data[key1]['no_of_data']+=1
        co_po_total_data[key1]['avg']=round(co_po_total_data[key1]['total']/co_po_total_data[key1]['no_of_data'],2)
        if co_po['program_outcome_name'] not in co_po_matrix_column:
            co_po_matrix_column[co_po['program_outcome_name']] = {'name':co_po['program_outcome_name'],'value':key1}
    for co_po in co_po_matrix_data['co_pso_mapping']:
        key1='_pso'+str(co_po['program_specific_outcome'])
        if co_po['course_outcome'] not in co_po_dict:
            co_po_dict[co_po['course_outcome']] = {}
        co_po_dict[co_po['course_outcome']]['course_outcome_name'] = co_po['course_outcome_name']
        co_po_dict[co_po['course_outcome']][key1] = co_po['value']
        if key1 not in co_po_total_data:
            co_po_total_data[key1]={'total':0,'no_of_data':0,'avg':0}
        co_po_total_data[key1]['total']+=co_po['value']
        co_po_total_data[key1]['no_of_data']+=1
        co_po_total_data[key1]['avg']=round(co_po_total_data[key1]['total']/co_po_total_data[key1]['no_of_data'],2)
        if co_po['program_specific_outcome_name'] not in co_po_matrix_column:
            co_po_matrix_column[co_po['program_specific_outcome_name']] = {'name':co_po['program_specific_outcome_name'],'value':key1}
    for co_po in co_po_matrix_data['co_peo_mapping']:
        key1='_peo'+str(co_po['program_educational_objectives'])
        if co_po['course_outcome'] not in co_po_dict:
            co_po_dict[co_po['course_outcome']] = {}
        co_po_dict[co_po['course_outcome']]['course_outcome_name'] = co_po['course_outcome_name']
        co_po_dict[co_po['course_outcome']][key1] = co_po['value']
        if key1 not in co_po_total_data:
            co_po_total_data[key1]={'total':0,'no_of_data':0,'avg':0}
        co_po_total_data[key1]['total']+=co_po['value']
        co_po_total_data[key1]['no_of_data']+=1
        co_po_total_data[key1]['avg']=round(co_po_total_data[key1]['total']/co_po_total_data[key1]['no_of_data'],2)
        if co_po['program_educational_objectives_name'] not in co_po_matrix_column:
            co_po_matrix_column[co_po['program_educational_objectives_name']] = {'name':co_po['program_educational_objectives_name'],'value':key1}
    level = {}
    metrics = [
        ('% Target(Set Your Own Target)', 'target'),
        ('Mim. Marks required to achive the target', 'mim_marks'),
        ('No. of students achived the target', 'students_achived_target'),
        ('Total no of students', 'total_students'),
        ('% of students achived the target', 'per_students_achived_target'),
        ('Attainment Level', 'attainment_level'),
    ]
    for co, co_data in subject_co_dict.items():
        if co in ('total_no_students', 'total'):
            continue

        for name, key in metrics:
            level.setdefault(name, {'name': name})
            level[name][f'internals{co}'] = co_data['internals'].get(key, None)

    for co, co_data in subject_co_dict.items():
        if co in ('total_no_students', 'total'):
            continue

        for name, key in metrics:
            level[name][f'assignment{co}'] = co_data['assessment'].get(key, None)
    attainment_level = [
        {'name':'60per and More' ,'value': "3"},
        {'name':'50per and More' ,'value': "2"},
        {'name':'40per and More' ,'value': "1"},
    ]
    ada_data = {}
    total = 0
    num_of_co = 0

    for key, data in subject_co_dict.items():
        if key not in ('total', 'total_no_students'):
            co_id = data['internals']['course_outcome_id']
            co_name = data['internals']['course_outcome__name']
            co_ia = data['internals']['attainment_level']
            co_ass = data['assessment']['attainment_level']

            cie_avg = (co_ia + co_ass) / 2
            see = 0
            direct = (see * 0.5) + (cie_avg * 0.5)
            indirect = 2
            co_attainment = (direct * 0.8) + (indirect * 0.2)

            ada_data[co_id] = {
                'co_name': co_name,
                'co_ia': co_ia,
                'co_ass': co_ass,
                'see': see,
                'cie_avg': cie_avg,
                'direct': direct,
                'indirect': indirect,
                'co_attainment': co_attainment,
            }

            total += co_attainment
            num_of_co += 1
    summary = total / num_of_co if num_of_co > 0 else 0
    co_attainment_summary = [
        {'name':'CO ATTAINMENT OF THE COURSE','value':summary}
    ]
    for avg in co_po_total_data:
        if 'total' not in co_po_dict:
            co_po_dict['total'] = {}
        co_po_dict['total']['course_outcome_name'] = 'Average'
        co_po_dict['total'][avg] = co_po_total_data[avg]['avg']
        if 'total' not in co_po_matrix_summary:
            co_po_matrix_summary['total'] = {}
        co_po_matrix_summary['total']['course_outcome_name'] = 'Average Maping(M)'
        co_po_matrix_summary['total'][avg] = co_po_total_data[avg]['avg']
        if 'po_attainment' not in co_po_matrix_summary:
            co_po_matrix_summary['po_attainment'] = {}
        co_po_matrix_summary['po_attainment']['course_outcome_name'] = 'PO/PSO Attainment Level'
        co_po_matrix_summary['po_attainment'][avg] = round((co_po_total_data[avg]['avg']*summary)/3,2)
    # response = {'subject_co_dict':subject_co_dict,'student_dict':student_dict,'details_for_heading':details_for_heading,'level':level.values()}
    response = {'details_for_heading':details_for_heading,'level':level.values(),'attainment_level':attainment_level,
                'ada_data':ada_data.values(),'co_po_matrix':co_po_matrix_column.values()}
    json_response = json_for_excel_amrita(response)
    options={}
    options['title'] = 'Amrita'
    options['description'] = 'marks'
    options['examname']= ''
    options['standardname']=''
    options['sectionname']=''
    options['institute_name']=''
    options['extraWorksheet'] = False
    options['Data'] = {
        'SDA':{
            'sheet_name':'SDA',
            'student_list':student_dict['internal_details'].values(),
            'columns':json_response['column_data_sad'],
            'headings':[{
                "value":"https://production-edubricz.s3.amazonaws.com/113/aims.jpg",
                "type":"image",
                "merge":"till_last_column",
            },
            {
                "value":department_details['branch__name'],
                "merge":"till_last_column"
            }]
        },
        'DA':{
            'sheet_name':'DA',
            'student_list':student_dict['co_details'].values(),
            'columns':json_response['column_data_da'],
            'summary':[{
                "columns":json_response['column_data_da2'],
                "data": level.values(),
                "start_column":2
            },
            {
                "columns":json_response['column_data_da3'],
                "data": attainment_level,
                "start_column":2
            }
            ]
        },
        'IA':{
            'sheet_name':'IA',
            'student_list':[{'co':'CO1','3':0,'2':36,'1':0,'IA':2},
                            {'co':'CO2','3':0,'2':36,'1':0,'IA':2},
                            {'co':'CO3','3':0,'2':36,'1':0,'IA':2},
                            {'co':'CO4','3':0,'2':36,'1':0,'IA':2},
                            {'co':'CO5','3':0,'2':36,'1':0,'IA':2}],
            'columns':json_response['column_data_ia'],
            'headings':[{
                "value":"COURSE END SURVEY ANALYSIS",
                "merge":"till_last_column",
            }]
        },
        'ADA':{
            'sheet_name':'ADA',
            'student_list':ada_data.values(),
            'columns':json_response['column_data_ada'],
            'headings':[
            {
                "value":institute_details.name,
                "merge":"till_last_column",
            },
            {
                "value":"Course Name :"+str(department_details['branch__name']),
                # +"Department :"+str(department_details['branch__name']),
                "merge":"till_last_column",
            },
            # {
            #     "value":"Course Code :"+str(department_details['branch__name'])+"Section :"+str(department_details['branch__name'])+"Teacher :",
            #     "merge":"till_last_column",
            # },
            {
                "value":"CIE Average",
                "merge":"till_last_column",
            }
            ]
        },
        'co_attainment':{
            'sheet_name':'CO Attainment',
            'student_list':ada_data.values(),
            'columns':json_response['column_data_co_attainment'],
            'summary':[{
                "columns":json_response['column_data_da3'],
                "data": co_attainment_summary,
                "start_column":3
            }
            ]
        },
        'co_po_mapping':{
            'sheet_name':'CO PO Mapping',
            'student_list':co_po_dict.values(),
            'columns':json_response['column_data_matrix'],
            'summary':[{
                "columns":json_response['column_data_matrix'],
                "data": co_po_matrix_summary.values(),
                "start_column":1
            }
            ]
        }

    }
    options['extraWorksheetData'] = dict()
    options['columns'] = json_response['column_data_da2']
    return write_to_excel_multiple_tabs_amrita(self, options, {},{})