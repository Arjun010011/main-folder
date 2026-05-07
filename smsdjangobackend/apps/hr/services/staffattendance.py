import os
from django.db.models import F, Value as V
from django.db.models.functions import Concat
from rest_framework import exceptions
from datetime import datetime, timedelta
from django.db import transaction

from apps.general.models import HolidayCalender
from django.contrib.auth.models import Group
from apps.hr.models import AssignShift, StaffAttendance, Day, Shift, ShiftSchedule
from apps.hr.serializers import StaffAttendanceSerializer
from apps.hr.services.default_varialbes import get_lop_attendance_list
from apps.hr.services.shift import getshift_details
from apps.institutes.models import FinancialYear
from apps.payroll.models.payroll import SalaryEmployeeMonthPlan
from apps.shared.services import SharedService,PDFService
from apps.shared.services_shared.common import get_full_name, get_selected_template
from apps.staffs.models import Staff
from apps.institutes.models import Institute
from apps.tenants.services.middlewares import get_current_db_name
from apps.shared.services import SharedService, NotificationBodyTemplate
from apps.users.models import User
from apps.notification.services.notification_service import send_notification

#throw_unassigned_error -> dont throw error when machine attendance hits
def staff_attendance_add(self, data, throw_unassigned_error=True, marked_from='web', is_send_notification=True):
    for_date = data['for_date']
    if SharedService.date_to_obj(for_date) > datetime.today().date():
        raise exceptions.ValidationError('Marking attendance for future date.')
    staff_ids = data['staff_ids']
    data_to_save = []
    unassigned_staff_list = AssignShift.get_unassigned_shift_for_staff(staff_ids, for_date)
    if unassigned_staff_list and throw_unassigned_error:
        unassigned_staff_list = Staff.get_all_staff_full_name(self, unassigned_staff_list)
        error_message = 'Staffs '
        for i in unassigned_staff_list:
            error_message += i['staff_name'] + ', '
        error_message += 'Not assigned to any shift'
        raise exceptions.ValidationError(error_message)
    temp = {'marked_from': marked_from, 'marked_by_user': self.request.user.id}
    # Check if status is being manually set (not from machine/biometric)
    is_manual_status = 'status' in data and marked_from != 'machine_attendance'
    
    # Statuses that require no times
    no_time_statuses = ['absent', 'unmarked', 'shiftnotassigned', 'nonworkingday', 'holiday', 'leave_applied', 'lop_attendance']
    # Statuses that require times (will use shift times if available) - only for "present" when manually set
    time_required_statuses = ['present', 'late', 'halfday', 'lateandhalfday', 'halfdaylate', 'halfdayandlate', 
                               'first_ses_leave_sec_sess_half', 'first_ses_half_sec_sess_leave']
    
    if 'status' in data:
        # Status is explicitly provided - handle based on status type
        status = data['status']
        if status in no_time_statuses:
            for staff_id in staff_ids:
                data_to_save.append({
                    'staff': staff_id, 'for_date': for_date, 'in_time': None, 'out_time': None,
                    'status': status,
                    'is_status_manually_set': is_manual_status,
                    'status_changed_by': self.request.user.id if is_manual_status else None,
                    'status_changed_at': datetime.now() if is_manual_status else None,
                    **temp
                })
        elif status == 'checkinmarked':
            # For checkinmarked, use provided in_time or get from shift
            staff_shift_details = get_staff_with_shift_details(self, staff_ids, for_date)
            for staff_id in staff_ids:
                if data.get('in_time'):
                    in_time = data['in_time']
                elif staff_id in staff_shift_details and 'schedules' in staff_shift_details[staff_id]:
                    in_time = for_date + ' ' + staff_shift_details[staff_id]['schedules']['start_time'].strftime('%H:%M:%S')
                else:
                    in_time = None
                data_to_save.append({
                    'staff': staff_id, 'for_date': for_date, 'in_time': in_time, 'out_time': None,
                    'status': 'checkinmarked',
                    'is_status_manually_set': is_manual_status,
                    'status_changed_by': self.request.user.id if is_manual_status else None,
                    'status_changed_at': datetime.now() if is_manual_status else None,
                    **temp
                })
        elif status == 'present' and is_manual_status:
            # Only "present" status gets times automatically when manually set
            staff_shift_details = get_staff_with_shift_details(self, staff_ids, for_date)
            for staff_id in staff_ids:
                if staff_id in staff_shift_details and 'schedules' in staff_shift_details[staff_id]:
                    in_time = for_date + ' ' + staff_shift_details[staff_id]['schedules']['start_time'].strftime('%H:%M:%S')
                    out_time = for_date + ' ' + staff_shift_details[staff_id]['schedules']['end_time'].strftime('%H:%M:%S')
                else:
                    # If shift not found, use provided times or None
                    in_time = data.get('in_time', None)
                    out_time = data.get('out_time', None)
                data_to_save.append({
                    'staff': staff_id, 'for_date': for_date, 'in_time': in_time, 'out_time': out_time,
                    'status': status,
                    'is_status_manually_set': is_manual_status,
                    'status_changed_by': self.request.user.id if is_manual_status else None,
                    'status_changed_at': datetime.now() if is_manual_status else None,
                    **temp
                })
        else:
            # For manually selected statuses (other than present), set without times
            # Explicitly set times to None when status is manually selected
            for staff_id in staff_ids:
                data_to_save.append({
                    'staff': staff_id, 'for_date': for_date, 
                    'in_time': None if is_manual_status else data.get('in_time', None), 
                    'out_time': None if is_manual_status else data.get('out_time', None),
                    'status': status,
                    'is_status_manually_set': is_manual_status,
                    'status_changed_by': self.request.user.id if is_manual_status else None,
                    'status_changed_at': datetime.now() if is_manual_status else None,
                    **temp
                })
    elif not data.get('out_time') and data.get('in_time'):
        # Only checkin marked - status is hardcoded to checkinmarked
        for staff_id in staff_ids:
            data_to_save.append({
                'staff': staff_id, 'for_date': for_date, 'in_time': data['in_time'], 'out_time': None,
                'status': 'checkinmarked',
                'is_status_manually_set': is_manual_status,
                'status_changed_by': self.request.user.id if is_manual_status else None,
                'status_changed_at': datetime.now() if is_manual_status else None,
                **temp
            })
    else:
        # Calculate status from times, but respect manual status if set
        result = get_staff_attendance_status(self, staff_ids, for_date, data.get('in_time'),
                                                                    data.get('out_time'))
        for staff_id in staff_ids:
            data_to_save.append({
                'staff': staff_id, 'for_date': for_date, 'in_time': data.get('in_time'), 'out_time': data.get('out_time'),
                'status': result[staff_id] if staff_id in result else 'shiftnotassigned',
                'is_status_manually_set': is_manual_status,
                'status_changed_by': self.request.user.id if is_manual_status else None,
                'status_changed_at': datetime.now() if is_manual_status else None,
                **temp
            })
    if len(data_to_save) == 1 and 'id' in data and data['id']:#update when it comes from machine attendance
        data_to_save[0]['id'] = data['id']
    self.serializer_class = StaffAttendanceSerializer
    self.queryset = StaffAttendance.objects.all()
    with transaction.atomic(using=get_current_db_name()):
        for row_data in data_to_save:
            if 'id' in row_data:
                # This is an update - check if status was manually set
                existing_instance = StaffAttendance.objects.get(id=row_data['id'])
                # If status was manually set and this is a machine/biometric update, preserve the status
                if existing_instance.is_status_manually_set and marked_from == 'machine_attendance':
                    # Don't change status, only update times
                    row_data.pop('status', None)  # Remove status from update data
                    # Also preserve the manual status flags
                    row_data.pop('is_status_manually_set', None)
                    row_data.pop('status_changed_by', None)
                    row_data.pop('status_changed_at', None)
                response = SharedService.update_data(self, row_data, **{'customObjectData': existing_instance})
            else:
                response = SharedService.add_data(self, row_data, False)
    staff_id=str(staff_id)
    if is_send_notification:
        SharedService.custom_thread(staff_attendance_notification, self, data_to_save, staff_id)
    return response

def staff_attendance_notification(self, data_to_save, staff_id):
    if not isinstance(staff_id, list):
        staff_id = [staff_id]
    users = User.objects.filter(staff__in=staff_id)
    customizedData = list()
    notification_obj = NotificationBodyTemplate('staff_attendance_create')
    for data in data_to_save:
        for staff in users:
            if staff.staff:
                name = staff.staff.first_name
            elif staff.student:
                name = staff.student.first_name
            else:
                name =staff.username
            duration = 0
            if data['out_time']:
                in_time = datetime.strptime(data['in_time'], "%Y-%m-%d %H:%M:%S")
                out_time = datetime.strptime(data['out_time'], "%Y-%m-%d %H:%M:%S")
                duration = round((out_time - in_time).total_seconds() / 3600, 2)
            temp = {
                'fordate':SharedService.date_to_obj(data["for_date"]).strftime("%d/%m/%Y"),
                'staffname':name,'outtime':data['out_time'],
                'intime':data['in_time'],'status':data['status'],
                'duration': duration
            }
            body_email = notification_obj.select_template('email', temp)
            body_push = notification_obj.select_template('push', temp)
            body_email_for_other_user = notification_obj.select_template_for_other_user('email',temp)
            body_push_for_other_user = notification_obj.select_template_for_other_user('push',temp)
            if staff.staff.email:
                customizedData.append({'email': staff.staff.email, 'user_id': staff.pk, 'email_subject': None,
                                    'email_body': body_email,'email_notification':1,'email_body_for_others':body_email_for_other_user})
            customizedData.append(
                {'push_subject': None, 'push_body': body_push, 'push_notification': 1, 'user_id': staff.pk, 'extra_params': {},
                'push_body_for_others':body_push_for_other_user})
    send_notification('staff_attendance_create', body=None, customizedData=customizedData)

def validate_bulk_attendance_upload(self, staff_data):
    staff_ids = set()
    staff_wise_from_date_to_date = {}
    data_to_save = []
    from_date_in_given_dates = None
    to_date_in_given_dates = None
    temp = {'marked_from': 'bulk_add', 'marked_by_user': self.request.user.id}
    for staff in staff_data:
        staff_id = staff['staff']
        staff_ids.add(staff_id)
        if staff_id not in staff_wise_from_date_to_date:
            staff_wise_from_date_to_date[staff_id] = {'from_date': '', 'to_date': ''}
        for attendance_data in staff['attendance_data']:
            in_time = attendance_data['in_time'] if 'in_time' in attendance_data else None
            out_time = attendance_data['out_time'] if 'out_time' in attendance_data else None
            status = attendance_data.get('status', 'unmarked')
            # Bulk marking is always manual, so set the flag
            data_to_save.append({
                'staff': staff_id, 
                'for_date': attendance_data['for_date'], 
                'in_time': in_time, 
                'out_time': out_time,
                'status': status,
                'is_status_manually_set': True,  # Bulk marking is always manual
                'status_changed_by': self.request.user.id,
                'status_changed_at': datetime.now(),
                **temp
            })
            if not staff_wise_from_date_to_date[staff_id]['from_date'] or staff_wise_from_date_to_date[staff_id]['from_date'] > attendance_data['for_date']: #finds the lowest date
                staff_wise_from_date_to_date[staff_id]['from_date'] = attendance_data['for_date']
            if not staff_wise_from_date_to_date[staff_id]['to_date'] or attendance_data['for_date'] > staff_wise_from_date_to_date[staff_id]['to_date']:
                staff_wise_from_date_to_date[staff_id]['to_date'] = attendance_data['for_date']
            if not from_date_in_given_dates or from_date_in_given_dates > attendance_data['for_date']:
                from_date_in_given_dates = attendance_data['for_date']
            if not to_date_in_given_dates or attendance_data['for_date'] > to_date_in_given_dates:
                to_date_in_given_dates = attendance_data['for_date']
    shift_data = AssignShift.objects.filter(staff__in=list(staff_ids)).values()
    error_data = []
    for staff_id in staff_wise_from_date_to_date:
        staff_data = staff_wise_from_date_to_date[staff_id]
        is_shift_exists = False
        for table_data in shift_data:
            if ((table_data['fromdate'].strftime('%Y-%m-%d') <= staff_data['from_date'] <= table_data['todate'].strftime('%Y-%m-%d'))
                    or (table_data['fromdate'].strftime('%Y-%m-%d') <= staff_data['to_date'] <= table_data['todate'].strftime(
                        '%Y-%m-%d'))):
                is_shift_exists = True
            if ((staff_data['from_date'] <= table_data['fromdate'].strftime('%Y-%m-%d') <= staff_data['to_date'])
                    or (staff_data['from_date'] <= table_data['todate'].strftime('%Y-%m-%d') <= staff_data['to_date'])):
                is_shift_exists = True
        if not is_shift_exists:
            name = get_staff_full_name(self, staff_id)
            error_data.append(
                f'{name}'
            )
    existing_attendance = StaffAttendance.objects.filter(staff__in=staff_ids, for_date__range=(from_date_in_given_dates, to_date_in_given_dates)).values(
        'staff', 'id', 'for_date'
    )
    staff_attendance_map = {}
    for existing in existing_attendance:
        for_date = existing['for_date'].strftime('%Y-%m-%d')
        if existing['staff'] not in staff_attendance_map:
            staff_attendance_map[existing['staff']] = {}
        if for_date not in staff_attendance_map[existing['staff']]:
            staff_attendance_map[existing['staff']][for_date] = {}
        staff_attendance_map[existing['staff']][for_date] = existing
    for attendance in data_to_save:
        if attendance['staff'] in  staff_attendance_map and attendance['for_date'] in staff_attendance_map[attendance['staff']]:
            attendance['id'] =  staff_attendance_map[attendance['staff']][attendance['for_date']]['id']
    if error_data:
        error = ', '.join(error_data)
        raise exceptions.ValidationError(f'{error} not assigned to any shifts')
    return {'data_to_save' : data_to_save, 'staff_ids': staff_ids}


def staff_attendance_bulk(self, data):
    staff_data = data['staff_list']
    return_data = validate_bulk_attendance_upload(self, staff_data)
    data_to_save = return_data['data_to_save']
    response = {'data': {}}
    with transaction.atomic(using=get_current_db_name()):
        for row_data in data_to_save:
            if 'id' in row_data:
                response = SharedService.update_data(self, row_data, **{'customObjectData': StaffAttendance.objects.get(id=row_data['id'])})
            else:
                response = SharedService.add_data(self, row_data, False)
    return response

def bulk_update_attendance_status(self, data):
    """
    Bulk update attendance status for multiple records by their IDs or staff_id + for_date
    Expected data format: {
        'attendance_ids': [1, 2, 3, ...],  # List of attendance record IDs
        'staff_attendance_pairs': [{'staff_id': 1, 'for_date': '2026-01-07'}, ...],  # For records without IDs
        'status': any valid status from StaffAttendance.attendanceStatus,
        'for_date': 'YYYY-MM-DD' (optional, for validation)
    }
    """
    from apps.hr.models.staffAttendance import StaffAttendance
    attendance_ids = data.get('attendance_ids', [])
    staff_attendance_pairs = data.get('staff_attendance_pairs', [])
    status = data.get('status')
    for_date = data.get('for_date')
    
    if not attendance_ids and not staff_attendance_pairs:
        raise exceptions.ValidationError('Either attendance_ids or staff_attendance_pairs is required')
    if not status:
        raise exceptions.ValidationError('status is required')
    
    # Validate status is one of the allowed choices
    valid_statuses = [choice[0] for choice in StaffAttendance.attendanceStatus]
    if status not in valid_statuses:
        raise exceptions.ValidationError(f'status must be one of: {", ".join(valid_statuses)}')
    
    # Collect all record IDs to fetch
    all_record_ids = list(attendance_ids) if attendance_ids else []
    
    # Get or create records for staff_attendance_pairs and collect their IDs
    if staff_attendance_pairs:
        for pair in staff_attendance_pairs:
            staff_id = pair.get('staff_id')
            pair_for_date = pair.get('for_date')
            if staff_id and pair_for_date:
                for_date_obj = SharedService.date_to_obj(pair_for_date)
                # Try to get existing active record first
                attendance_record = StaffAttendance.objects.filter(
                    staff_id=staff_id,
                    for_date=for_date_obj,
                    is_active=True
                ).first()
                
                if not attendance_record:
                    # If no active record exists, check for inactive one
                    inactive_record = StaffAttendance.objects.filter(
                        staff_id=staff_id,
                        for_date=for_date_obj,
                        is_active=False
                    ).first()
                    
                    if inactive_record:
                        # Reactivate the existing record
                        inactive_record.is_active = True
                        inactive_record.status = 'unmarked'
                        inactive_record.save()
                        attendance_record = inactive_record
                    else:
                        # Create new record
                        attendance_record = StaffAttendance.objects.create(
                            staff_id=staff_id,
                            for_date=for_date_obj,
                            status='unmarked',
                            is_active=True
                        )
                
                if attendance_record.id not in all_record_ids:
                    all_record_ids.append(attendance_record.id)
    
    # Get all attendance records by collected IDs
    if not all_record_ids:
        raise exceptions.ValidationError('No valid attendance records found')
    
    attendance_records = StaffAttendance.objects.filter(id__in=all_record_ids, is_active=True)
    
    if not attendance_records.exists():
        raise exceptions.ValidationError('No valid attendance records found')
    
    # Validate that all records are for the same date if for_date is provided
    if for_date:
        for_date_obj = SharedService.date_to_obj(for_date)
        invalid_records = attendance_records.exclude(for_date=for_date_obj)
        if invalid_records.exists():
            raise exceptions.ValidationError('All attendance records must be for the same date')
    
    # Check salary payment restrictions
    updated_count = 0
    data_to_save = []
    
    # Statuses that require no times
    no_time_statuses = ['absent', 'unmarked', 'shiftnotassigned', 'nonworkingday', 'holiday', 'leave_applied', 'lop_attendance']
    # Statuses that require times (will use shift times if available)
    time_required_statuses = ['present', 'late', 'halfday', 'lateandhalfday', 'halfdaylate', 'halfdayandlate', 
                               'first_ses_leave_sec_sess_half', 'first_ses_half_sec_sess_leave']
    
    with transaction.atomic(using=get_current_db_name()):
        for instance in attendance_records:
            first_day_of_month = instance.for_date - timedelta(days=int(instance.for_date.strftime("%d")) - 1)
            if SalaryEmployeeMonthPlan.objects.filter(staff=instance.staff_id, salary_month=first_day_of_month).exists():
                continue  # Skip records where salary is already paid
            
            # Update status and times based on status type
            if status in no_time_statuses:
                instance.status = status
                instance.in_time = None
                instance.out_time = None
            elif status in time_required_statuses:
                # For statuses that require times, try to get from shift
                staff_shift_details = get_staff_with_shift_details(self, [instance.staff_id], instance.for_date.strftime('%Y-%m-%d'))
                if instance.staff_id in staff_shift_details and 'schedules' in staff_shift_details[instance.staff_id]:
                    in_time = instance.for_date.strftime('%Y-%m-%d') + ' ' + staff_shift_details[instance.staff_id]['schedules']['start_time'].strftime('%H:%M:%S')
                    out_time = instance.for_date.strftime('%Y-%m-%d') + ' ' + staff_shift_details[instance.staff_id]['schedules']['end_time'].strftime('%H:%M:%S')
                    instance.in_time = in_time
                    instance.out_time = out_time
                instance.status = status
            elif status == 'checkinmarked':
                # For checkinmarked, keep existing in_time if available, or set from shift
                if not instance.in_time:
                    staff_shift_details = get_staff_with_shift_details(self, [instance.staff_id], instance.for_date.strftime('%Y-%m-%d'))
                    if instance.staff_id in staff_shift_details and 'schedules' in staff_shift_details[instance.staff_id]:
                        in_time = instance.for_date.strftime('%Y-%m-%d') + ' ' + staff_shift_details[instance.staff_id]['schedules']['start_time'].strftime('%H:%M:%S')
                        instance.in_time = in_time
                instance.out_time = None
                instance.status = status
            else:
                # For any other status, just update the status, keep existing times
                instance.status = status
            
            # Mark as manually set
            instance.is_status_manually_set = True
            instance.status_changed_by = self.request.user if hasattr(self, 'request') else None
            instance.status_changed_at = datetime.now()
            instance.marked_by_user = self.request.user if hasattr(self, 'request') else None
            
            instance.save()
            updated_count += 1
            
            data_to_save.append({
                'staff': instance.staff_id,
                'for_date': instance.for_date.strftime('%Y-%m-%d'),
                'in_time': instance.in_time,
                'out_time': instance.out_time,
                'status': instance.status
            })
    
    # Send notifications
    if data_to_save:
        staff_ids = list(set([d['staff'] for d in data_to_save]))
        SharedService.custom_thread(staff_attendance_notification, self, data_to_save, staff_ids)
    
    return {'Result': True, 'Reason': f'Successfully updated {updated_count} attendance record(s)'}


def bulk_update_attendance_details(self, data):
    """
    Bulk update attendance details (date, time, status) for multiple records by their IDs or staff_id + for_date
    Expected data format: {
        'attendance_ids': [1, 2, 3, ...],  # List of attendance record IDs
        'staff_attendance_pairs': [{'staff_id': 1, 'for_date': '2026-01-07'}, ...],  # For records without IDs
        'for_date': 'YYYY-MM-DD' (optional),
        'in_time': 'HH:mm:ss' or 'HH:mm' (optional),
        'out_time': 'HH:mm:ss' or 'HH:mm' (optional),
        'status': 'present' or other status (optional, will be calculated if not provided and times are set)
    }
    """
    from apps.hr.models.staffAttendance import StaffAttendance
    attendance_ids = data.get('attendance_ids', [])
    staff_attendance_pairs = data.get('staff_attendance_pairs', [])
    for_date = data.get('for_date')
    in_time = data.get('in_time')
    out_time = data.get('out_time')
    status = data.get('status')
    
    if not attendance_ids and not staff_attendance_pairs:
        raise exceptions.ValidationError('Either attendance_ids or staff_attendance_pairs is required')
    
    # Collect all record IDs to fetch
    all_record_ids = list(attendance_ids) if attendance_ids else []
    
    # Get or create records for staff_attendance_pairs and collect their IDs
    if staff_attendance_pairs:
        for pair in staff_attendance_pairs:
            staff_id = pair.get('staff_id')
            pair_for_date = pair.get('for_date')
            if staff_id and pair_for_date:
                for_date_obj = SharedService.date_to_obj(pair_for_date)
                # Try to get existing active record first
                attendance_record = StaffAttendance.objects.filter(
                    staff_id=staff_id,
                    for_date=for_date_obj,
                    is_active=True
                ).first()
                
                if not attendance_record:
                    # If no active record exists, check for inactive one
                    inactive_record = StaffAttendance.objects.filter(
                        staff_id=staff_id,
                        for_date=for_date_obj,
                        is_active=False
                    ).first()
                    
                    if inactive_record:
                        # Reactivate the existing record
                        inactive_record.is_active = True
                        inactive_record.status = 'unmarked'
                        inactive_record.save()
                        attendance_record = inactive_record
                    else:
                        # Create new record
                        attendance_record = StaffAttendance.objects.create(
                            staff_id=staff_id,
                            for_date=for_date_obj,
                            status='unmarked',
                            is_active=True
                        )
                
                if attendance_record.id not in all_record_ids:
                    all_record_ids.append(attendance_record.id)
    
    # Get all attendance records by collected IDs
    if not all_record_ids:
        raise exceptions.ValidationError('No valid attendance records found')
    
    attendance_records = StaffAttendance.objects.filter(id__in=all_record_ids, is_active=True)
    
    if not attendance_records.exists():
        raise exceptions.ValidationError('No valid attendance records found')
    
    # Check salary payment restrictions
    updated_count = 0
    data_to_save = []
    
    with transaction.atomic(using=get_current_db_name()):
        for instance in attendance_records:
            first_day_of_month = instance.for_date - timedelta(days=int(instance.for_date.strftime("%d")) - 1)
            if SalaryEmployeeMonthPlan.objects.filter(staff=instance.staff_id, salary_month=first_day_of_month).exists():
                continue  # Skip records where salary is already paid
            
            # Update date if provided
            if for_date:
                for_date_obj = SharedService.date_to_obj(for_date)
                instance.for_date = for_date_obj
            
            # Update times if provided
            # Ensure date is set before combining with time
            if not instance.for_date:
                raise exceptions.ValidationError('for_date must be set before updating times')
            
            if in_time is not None:
                if in_time:
                    # Combine date with time
                    date_str = instance.for_date.strftime('%Y-%m-%d')
                    try:
                        # Try parsing with seconds first
                        instance.in_time = datetime.strptime(f"{date_str} {in_time}", '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        # If that fails, try without seconds
                        try:
                            instance.in_time = datetime.strptime(f"{date_str} {in_time}", '%Y-%m-%d %H:%M')
                        except ValueError:
                            raise exceptions.ValidationError(f'Invalid in_time format: {in_time}. Expected HH:mm:ss or HH:mm')
                else:
                    instance.in_time = None
            
            if out_time is not None:
                if out_time:
                    # Combine date with time
                    date_str = instance.for_date.strftime('%Y-%m-%d')
                    try:
                        # Try parsing with seconds first
                        instance.out_time = datetime.strptime(f"{date_str} {out_time}", '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        # If that fails, try without seconds
                        try:
                            instance.out_time = datetime.strptime(f"{date_str} {out_time}", '%Y-%m-%d %H:%M')
                        except ValueError:
                            raise exceptions.ValidationError(f'Invalid out_time format: {out_time}. Expected HH:mm:ss or HH:mm')
                else:
                    instance.out_time = None
            
            # Update status
            if status:
                # Status is explicitly provided
                valid_statuses = [choice[0] for choice in StaffAttendance.attendanceStatus]
                if status not in valid_statuses:
                    raise exceptions.ValidationError(f'status must be one of: {", ".join(valid_statuses)}')
                instance.status = status
                # Mark as manually set
                instance.is_status_manually_set = True
                instance.status_changed_by = self.request.user if hasattr(self, 'request') else None
                instance.status_changed_at = datetime.now()
            elif in_time is not None or out_time is not None:
                # If times are provided but no status, calculate status from times
                # But only if status wasn't manually set before
                # If status was manually set, keep it unchanged even when times are updated
                if not instance.is_status_manually_set:
                    # Only calculate status if both in_time and out_time are available
                    # get_staff_attendance_status requires both times
                    if instance.in_time and instance.out_time:
                        staff_ids = [instance.staff_id]
                        try:
                            calculated_status = get_staff_attendance_status(
                                self, 
                                staff_ids, 
                                instance.for_date.strftime('%Y-%m-%d'),
                                instance.in_time.strftime('%Y-%m-%d %H:%M:%S'),
                                instance.out_time.strftime('%Y-%m-%d %H:%M:%S')
                            )
                            if instance.staff_id in calculated_status:
                                instance.status = calculated_status[instance.staff_id]
                            else:
                                # If calculation didn't return a status for this staff, default to present
                                instance.status = 'present'
                        except Exception as e:
                            # Log the error for debugging but don't fail the update
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.error(f"Error calculating status for staff {instance.staff_id}: {str(e)}")
                            # If status calculation fails, try to set a reasonable default
                            if instance.in_time and not instance.out_time:
                                instance.status = 'checkinmarked'
                            elif instance.in_time and instance.out_time:
                                # Both times exist but calculation failed - default to present
                                instance.status = 'present'
                            # Otherwise, keep the existing status
                    elif instance.in_time and not instance.out_time:
                        # Only in_time provided, set to checkinmarked
                        instance.status = 'checkinmarked'
                    # If only out_time is provided without in_time, keep existing status
                # If status was manually set, don't change it even when times are updated
            
            instance.marked_by_user = self.request.user if hasattr(self, 'request') else None
            
            instance.save()
            updated_count += 1
            
            data_to_save.append({
                'staff': instance.staff_id,
                'for_date': instance.for_date.strftime('%Y-%m-%d'),
                'in_time': instance.in_time.strftime('%Y-%m-%d %H:%M:%S') if instance.in_time else None,
                'out_time': instance.out_time.strftime('%Y-%m-%d %H:%M:%S') if instance.out_time else None,
                'status': instance.status
            })
    
    # Send notifications
    if data_to_save:
        staff_ids = list(set([d['staff'] for d in data_to_save]))
        SharedService.custom_thread(staff_attendance_notification, self, data_to_save, staff_ids)
    
    return {'Result': True, 'Reason': f'Successfully updated {updated_count} attendance record(s)'}


def mark_absent_for_unmarked_in_month(self, staffId, monthStartDate, monthEndDate):
    monthStartDate = SharedService.date_to_obj(monthStartDate)
    monthEndDate = SharedService.date_to_obj(monthEndDate)
    datesInGivenDateRange = SharedService.get_for_date_from_date_range(monthStartDate, monthEndDate)
    staffAttendList = StaffAttendance.objects.filter(for_date__in=datesInGivenDateRange, staff=staffId).values_list(
        'for_date', flat=True)
    holidayDateRangeList = HolidayCalender.get_upcoming_holidays(self, monthStartDate, monthEndDate, True)
    workingDays = Day.get_staff_working_day(self)
    dataToSave = []
    response = {'Result': True}
    for date in datesInGivenDateRange:
        day = SharedService.get_day_for_date(date.strftime('%Y-%m-%d'))
        if date not in staffAttendList and day in workingDays and date not in holidayDateRangeList:
            dataToSave.append(
                {
                    'for_date': date, 'status': 'absent', 'is_active': True, 'staff': staffId
                }
            )
    if dataToSave:
        self.queryset = StaffAttendance
        self.serializer_class = StaffAttendanceSerializer
        response = SharedService.add_data(self, dataToSave)
    return response


def update_staff_attendance(self, data,is_send_notification=True):
    data_to_save = []
    instance = self.get_queryset().get(id=self.kwargs['pk'])
    first_day_of_month = instance.for_date - timedelta(days=int(instance.for_date.strftime("%d")) - 1)
    if SalaryEmployeeMonthPlan.objects.filter(staff=self.kwargs['pk'], salary_month=first_day_of_month).exists():
        raise exceptions.ValidationError('Not able to mark attendance Salary Paid for this month')
    
    # Check if this is a biometric/machine update
    is_biometric_update = instance.marked_from == 'machine_attendance' or (hasattr(self, 'request') and hasattr(self.request, 'data') and self.request.data.get('marked_from') == 'machine_attendance')
    
    # Check if status is being manually changed
    is_manual_status_change = 'status' in data and instance.marked_from != 'machine_attendance'
    
    # If status was manually set previously AND this is a biometric checkout update, don't change status
    if instance.is_status_manually_set and is_biometric_update and 'out_time' in data and not 'status' in data:
        # Only update checkout time, keep status unchanged
        instance.out_time = data.get('out_time')
        instance.in_time = data.get('in_time', instance.in_time)
    elif 'status' in data and data['status'] == 'absent':
        # Manual status change to absent
        instance.status = 'absent'
        instance.in_time = None
        instance.out_time = None
        instance.is_status_manually_set = True
        instance.status_changed_by = self.request.user if hasattr(self, 'request') else None
        instance.status_changed_at = datetime.now()
    elif 'status' in data:
        # Manual status change to any other status
        instance.status = data['status']
        instance.is_status_manually_set = True
        instance.status_changed_by = self.request.user if hasattr(self, 'request') else None
        instance.status_changed_at = datetime.now()
        # Update times if provided, but respect the manually set status
        if 'in_time' in data:
            instance.in_time = data['in_time']
        if 'out_time' in data:
            instance.out_time = data['out_time']
    else:
        # No status provided - only update times
        # Only auto-calculate status if:
        # 1. Status was NOT manually set, AND
        # 2. This is NOT a biometric update, AND
        # 3. Both in_time and out_time are provided
        if not instance.is_status_manually_set and not is_biometric_update and data.get('out_time'):
            result = get_staff_attendance_status(self, [instance.staff_id],
                                                                        instance.for_date.strftime('%Y-%m-%d'),
                                                                        data.get('in_time', instance.in_time), data['out_time'])
            instance.status = result[instance.staff_id]
        elif not instance.is_status_manually_set and not data.get('out_time') and data.get('in_time'):
            # Only checkin provided - set to checkinmarked
            instance.status = 'checkinmarked'
        elif not data.get('out_time') and not data.get('in_time'):
            instance.status = 'unmarked'
        
        # Update times
        if 'in_time' in data:
            instance.in_time = data['in_time']
        if 'out_time' in data:
            instance.out_time = data['out_time']
    
    data_to_save.append({
        'staff' : instance.staff_id,
        'for_date':instance.for_date.strftime('%Y-%m-%d'),
        'in_time':instance.in_time,
        'out_time' : instance.out_time,
        'status':instance.status
    })
    response = SharedService.add_or_update_data(self, [instance.__dict__])
    if is_send_notification:
        SharedService.custom_thread(staff_attendance_notification, self,data_to_save, instance.staff_id)
    return response


def get_staff_attendance_status(self, staffIds, fordate, inTime, outTime,status=''):
    if status=='lop_attendance':
        return status
    inTime = datetime.strptime(inTime, "%Y-%m-%d %H:%M:%S")
    outTime = datetime.strptime(outTime, "%Y-%m-%d %H:%M:%S")
    for_date_obj = datetime.strptime(fordate, '%Y-%m-%d')
    if outTime < inTime:
        raise exceptions.ValidationError('Out time should greater than in time')
    if (outTime - inTime).days > 1:
        raise exceptions.ValidationError('Out time and in time difference should be less than one day')
    staff_shift_details = get_staff_with_shift_details(self, staffIds, fordate)
    data_to_save = {}
    for staff_id, staff_data in staff_shift_details.items():
        if 'schedules' not in staff_data:
            data_to_save[staff_id] = 'nonworkingday'
            continue
        shift_in_time = staff_data['schedules']['start_time'].strftime('%H:%M:%S')
        shift_out_time = staff_data['schedules']['end_time'].strftime('%H:%M:%S')
        if shift_in_time > shift_out_time:
            shift_out_time = (for_date_obj + timedelta(days=1)).strftime('%Y-%m-%d') + ' ' + shift_out_time
        else:
            shift_out_time = (for_date_obj).strftime('%Y-%m-%d') + ' ' + shift_out_time
        shift_in_time = (for_date_obj).strftime('%Y-%m-%d') + ' ' + shift_in_time
        if str(inTime.date()) != shift_in_time.split(' ')[0] and str(inTime.date()) != shift_out_time.split(' ')[0]:
            raise exceptions.ValidationError('Intime date and fordate should be same.')
        buffer_time = staff_data['schedules']['buffer_time']
        late_buffer_time = staff_data['schedules']['late_buffer_time']
        session1_end_time = staff_data['schedules']['first_session_end_time'].strftime('%H:%M:%S') if staff_data['schedules']['first_session_end_time'] else None
        session_2_start_time = staff_data['schedules']['second_session_start_time'].strftime('%H:%M:%S') if staff_data['schedules']['second_session_start_time'] else None
        
        if session1_end_time:
            if shift_in_time > session1_end_time:
                session1_end_time = (for_date_obj).strftime('%Y-%m-%d') + ' ' + session1_end_time
                session1_end_time = datetime.strptime(session1_end_time, '%Y-%m-%d %H:%M:%S')
            else:
                session1_end_time = (for_date_obj + timedelta(days=1)).strftime('%Y-%m-%d') + ' ' + session1_end_time
                session1_end_time = datetime.strptime(session1_end_time, '%Y-%m-%d %H:%M:%S')
        
        if session_2_start_time:
            if shift_in_time > session_2_start_time:
                session_2_start_time = (for_date_obj).strftime('%Y-%m-%d') + ' ' + session_2_start_time
                session_2_start_time = datetime.strptime(session_2_start_time, '%Y-%m-%d %H:%M:%S')
            else:
                session_2_start_time = (for_date_obj + timedelta(days=1)).strftime('%Y-%m-%d') + ' ' + session_2_start_time
                session_2_start_time = datetime.strptime(session_2_start_time, '%Y-%m-%d %H:%M:%S')

        status = 'absent'
        if shift_out_time > shift_in_time and inTime > outTime:
            raise exceptions.ValidationError('In time should be less than out Time')


        checkin_status = 'absent'
        checkout_status = 'absent'

        """ Find the checkin status """

        shift_in_time = datetime.strptime(shift_in_time, '%Y-%m-%d %H:%M:%S')
        shift_in_time_plus_buffer = shift_in_time + timedelta(0, ((late_buffer_time) * 60))
        shift_in_time_plus_later_buffer = shift_in_time + timedelta(0, ((buffer_time) * 60))

        shift_out_time = datetime.strptime(shift_out_time, '%Y-%m-%d %H:%M:%S')


        if session1_end_time:
            session1_end_time_minus_late_buffer = session1_end_time - + timedelta(0, ((buffer_time) * 60))
        else: #handle for halfday
            session1_end_time_minus_late_buffer = None
        if session_2_start_time:
            session_2_start_time_plus_buffer = session_2_start_time + timedelta(0, ((late_buffer_time) * 60))
            session_2_start_time_plus_late_buffer = session_2_start_time + timedelta(0, ((buffer_time) * 60))
        else:#handle for halfday
            session_2_start_time_plus_buffer = None
            session_2_start_time_plus_late_buffer = None
        session_2_end_time_minus_buffer = shift_out_time - + timedelta(0, ((late_buffer_time) * 60))
        session_2_end_time_minus_late_buffer = shift_out_time - + timedelta(0, ((buffer_time) * 60))
        if inTime.strftime('%Y-%m-%d %H:%M:%S') <= shift_in_time_plus_buffer.strftime('%Y-%m-%d %H:%M:%S'):
            checkin_status = 'present'
        elif SharedService.time_is_between(self, inTime.strftime('%Y-%m-%d %H:%M:%S'),
            shift_in_time_plus_buffer.strftime('%Y-%m-%d %H:%M:%S'),
            shift_in_time_plus_later_buffer.strftime('%Y-%m-%d %H:%M:%S') #attend in late time
        ):
            checkin_status = 'late'
        elif session_2_start_time_plus_buffer and session_2_start_time_plus_late_buffer and SharedService.time_is_between(self, inTime.strftime('%Y-%m-%d %H:%M:%S'),
            session_2_start_time_plus_buffer.strftime('%Y-%m-%d %H:%M:%S'),
            session_2_start_time_plus_late_buffer.strftime('%Y-%m-%d %H:%M:%S') #attend in late time
        ):
            checkin_status = 'attending_session_2_but_late'
        elif session_2_start_time_plus_buffer and SharedService.time_is_between(self, inTime.strftime('%Y-%m-%d %H:%M:%S'),
            shift_in_time_plus_later_buffer.strftime('%Y-%m-%d %H:%M:%S'),
            session_2_start_time_plus_buffer.strftime('%Y-%m-%d %H:%M:%S')
        ):
            checkin_status = 'attending_session_2_present'
        elif session_2_start_time_plus_late_buffer and inTime > session_2_start_time_plus_late_buffer:
            checkin_status = 'absent'

        """ Find the checkout status ordering is important"""

        if outTime >= session_2_end_time_minus_buffer:
            checkout_status = 'present'
        elif SharedService.time_is_between(self, outTime.strftime('%Y-%m-%d %H:%M:%S'),
            session_2_end_time_minus_late_buffer.strftime('%Y-%m-%d %H:%M:%S'),
            shift_out_time.strftime('%Y-%m-%d %H:%M:%S')
        ):
            checkout_status = 'secondofflate' #leaving the  school in second off but leaving early
        elif session1_end_time and outTime >= session1_end_time:
            checkout_status = 'checkout_in_second_off_before_end' #checkouts after the session1
        elif session1_end_time_minus_late_buffer and outTime >= session1_end_time_minus_late_buffer:
            checkout_status = 'firsthalflate'
        else:
            checkout_status = 'absent'

        if checkin_status == 'absent' or checkout_status == 'absent':
            status = 'absent'
        elif checkin_status  == 'present' and checkout_status == 'present':
            status = 'present'
        elif checkin_status == 'present' and checkout_status == 'secondofflate':
            status = 'late'
        elif checkin_status == 'present' and checkout_status == 'checkout_in_second_off_before_end':
            status = 'halfday'
        elif checkin_status == 'present' and checkout_status == 'firsthalflate':
            status = 'halfdaylate'
        elif checkin_status == 'late' and checkout_status == 'present':
            status = 'late'
        elif checkin_status == 'late' and checkout_status == 'secondofflate':
            status = 'late'
        elif checkin_status == 'late' and checkout_status == 'checkout_in_second_off_before_end':
            status = 'lateandhalfday'
        elif checkin_status == 'late' and checkout_status == 'firsthalflate':
            status = 'halfdayandlate'
        elif checkin_status == 'attending_session_2_but_late' and checkout_status == 'present':
            status = 'halfdayandlate'
        elif checkin_status == 'attending_session_2_but_late' and checkout_status == 'secondofflate':
            status = 'halfdayandlate'
        elif checkin_status == 'attending_session_2_but_late' and checkout_status == 'checkout_in_second_off_before_end':
            status = 'absent'
        elif checkin_status == 'attending_session_2_but_late' and checkout_status == 'firsthalflate':
            status = 'absent'
        elif checkin_status == 'attending_session_2_present' and checkout_status == 'present':
            status = 'halfday'
        elif checkin_status == 'attending_session_2_present' and checkout_status == 'secondofflate':
            status = 'halfdayandlate'
        elif checkin_status == 'attending_session_2_present' and checkout_status == 'checkout_in_second_off_before_end':
            status = 'absent'
        elif checkin_status == 'attending_session_2_present' and checkout_status == 'firsthalflate':
            status = 'absent'
        data_to_save[staff_id] = status
    return data_to_save

def get_staff_full_name(self, staffId):
    staffObj = Staff.objects.get(id=staffId)
    name = get_full_name(staffObj.first_name , staffObj.middle_name, staffObj.last_name)
    return name

def get_staff_with_shift_details(self, staff_ids, fordate):
    queryset = Shift.objects.filter(assign_shift_shift__staff__in=staff_ids, assign_shift_shift__fromdate__lte=fordate,
                                    assign_shift_shift__todate__gte=fordate, is_active=True).annotate(
        staff=F('assign_shift_shift__staff')).values('id',
                                                     'deduction_days', 'late_attempt_per_month', 'name', 'staff')
    shift_dict_obj = {shiftdata['id']: shiftdata for shiftdata in queryset}
    day_name = SharedService.get_day_for_date(fordate)
    shift_schedule_data = ShiftSchedule.objects.filter(shift__in=shift_dict_obj.keys(), is_active=True,
                                                     day__name=day_name).values(
        'buffer_time', 'day_id', 'end_time', 'first_session_end_time', 'late_buffer_time',
        'second_session_start_time', 'shift_id', 'start_time', 'day__name'
    )
    return_data = {}
    for schedule in shift_schedule_data:
        shift_dict_obj[schedule['shift_id']]['schedules'] = schedule
    for staff_data in queryset:
        return_data[staff_data['staff']] = shift_dict_obj[staff_data['id']]
    return return_data

def json_staff_daily_attendence_report(today):
    column_data=[
        {
            'column': 'SI No', 'required': False, 'schemacolumn': 'sl_no'
        }]
    column_data.append({
                'column': 'STAFF NAME', 'required': False, 'schemacolumn':'staff_name', 'group':'staff_group_id'
        })
    column_data.append({
                'column': 'In time', 'required': False, 'schemacolumn':'in_time' , 'colour':'status_details_in_time'
        })
    column_data.append({
                'column': 'Out time', 'required': False, 'schemacolumn':'out_time', 'colour':'status_details_out_time'
        })
    column_data.append({
                'column': 'Status', 'required': False, 'schemacolumn':'status'
        })
    return column_data

def staff_attendence_daily_report(self,data,staffIds,today_str):
    staff_shift_details = get_staff_with_shift_details(self, staffIds, today_str)
    shift_in_time = None
    shift_out_time = None
    for staff_id, staff_data in staff_shift_details.items():
        shift_in_time = staff_data['schedules']['start_time']
        shift_out_time = staff_data['schedules']['end_time']
    if not shift_in_time or not shift_out_time:
        raise exceptions.ValidationError(f"shift not assigned to staff {shift_in_time} {shift_out_time}")
    staff_row_data={}
    staff_row_data2={}
    staff_total_absent=0
    staff_total_present=0
    groups = Group.objects.all().values('id','name')
    for group in groups:
        staff_row_data[group['name']]=[]
        staff_row_data2[group['name']]=[]
        for staff in data:
            staff_name = data[staff]['staff_details']['name']
            staff_group_id = data[staff]['staff_details']['staff_group']
            staff_group_name = data[staff]['staff_details']['staff_group_name']
            if staff_group_id == group['id']:
                staff_row_data[group['name']] = {'staff_name': staff_name, 'staff_group_id' : staff_group_id, 'staff_group_name' : staff_group_name}
                for dates in data[staff]['day_list_status']:
                    staff_row_data[group['name']][dates]=dates
                    in_time=data[staff]['day_list_status'][dates]['in_time'].strftime('%H:%M:%S') if 'in_time' in data[staff]['day_list_status'][dates] and \
                        data[staff]['day_list_status'][dates]['in_time'] else ''
                    out_time=data[staff]['day_list_status'][dates]['out_time'].strftime('%H:%M:%S') if 'out_time' in data[staff]['day_list_status'][dates] and \
                        data[staff]['day_list_status'][dates]['out_time'] else ''
                    status=data[staff]['day_list_status'][dates]['status'] if 'status' in data[staff]['day_list_status'][dates] else ''
                    staff_row_data[group['name']]['in_time'] = in_time if in_time else ''
                    staff_row_data[group['name']]['out_time'] = out_time if out_time else ''
                    staff_row_data[group['name']]['status'] = status
                    if status=='unmarked':
                        staff_total_absent +=1
                    else:
                        staff_total_present+=1
                    staff_row_data[group['name']]['status_details_in_time']=''
                    staff_row_data[group['name']]['status_details_out_time']=''
                    if in_time=='' or in_time>=(shift_in_time.strftime('%H:%M:%S')):
                        staff_row_data[group['name']]['status_details_in_time']='late'
                    elif in_time<=((datetime.strptime(shift_in_time.strftime('%H:%M:%S'), '%H:%M:%S')-timedelta(minutes=15)).strftime('%H:%M:%S')):
                        staff_row_data[group['name']]['status_details_in_time']='extra_miles'
                    if out_time=='' or out_time<=(shift_out_time.strftime('%H:%M:%S')):
                        staff_row_data[group['name']]['status_details_out_time']='late'
                    elif out_time>=((datetime.strptime(shift_out_time.strftime('%H:%M:%S'), '%H:%M:%S')+timedelta(minutes=15)).strftime('%H:%M:%S')):
                        staff_row_data[group['name']]['status_details_out_time']='extra_miles'
                    in_time = datetime.strptime(in_time,'%H:%M:%S') if in_time != '' else ''
                    staff_row_data[group['name']]['in_time'] = in_time.strftime('%I:%M %p') if in_time != '' else ''
                    out_time = datetime.strptime(out_time,'%H:%M:%S') if out_time != '' else ''
                    staff_row_data[group['name']]['out_time'] = out_time.strftime('%I:%M %p') if out_time != '' else ''
                staff_row_data2[group['name']].append(staff_row_data[group['name']])
    return staff_row_data2,staff_total_present,staff_total_absent

def get_staff_attendance(self, request):
    queryset = self.get_queryset()
    financial_year = request.GET.get('financial_year', None)
    from_date = request.GET.get('from_date', None)
    to_date = request.GET.get('to_date', None)
    staff_id = request.GET.get('staff', None)

    # get full staff list
    if staff_id:
        staff_ids = [int(staff_id)]
    else:
        staff_ids = list(
            Staff.objects.filter(is_active=True).values_list('id', flat=True)
        )

    # resolve date range
    if financial_year:
        date_range_obj = FinancialYear.objects.get(id=financial_year)
        if not date_range_obj:
            raise exceptions.ValidationError("Invalid Financial Year")

        from_date = date_range_obj.start_date
        to_date = date_range_obj.end_date

        queryset = queryset.filter(for_date__range=(from_date, to_date))

    elif from_date and to_date:
        queryset = queryset.filter(for_date__range=(from_date, to_date))

    if staff_id:
        queryset = queryset.filter(staff=staff_id)

    serializer = self.get_serializer(queryset, many=True)
    marked_attendance = serializer.data

    # Generate date range
    date_range_list = SharedService.get_for_date_from_date_range(
        datetime.strptime(str(from_date), "%Y-%m-%d"),
        datetime.strptime(str(to_date), "%Y-%m-%d")
    )

    # Build lookup map
    attendance_map = {}
    for item in marked_attendance:
        sid = item["staff"]
        date = item["for_date"]
        if sid not in attendance_map:
            attendance_map[sid] = {}
        attendance_map[sid][date] = item

    final_flat_list = []

    marked_staff_set = set()
    unmarked_staff_set = set()
    checkedin_staff_set = set()
    late_staff_set = set()
    late_comes_set = set()

    # Build final flat response
    for sid in staff_ids:
        staff_obj = Staff.objects.filter(id=sid).first()
        if not staff_obj:
            continue

        has_marked = False
        has_checkin = False
        has_late = False

        for d in date_range_list:
            d_str = d.strftime("%Y-%m-%d")

            if sid in attendance_map and d_str in attendance_map[sid]:
                att = attendance_map[sid][d_str]
                final_flat_list.append(att)

                if att["status"] != "unmarked":
                    has_marked = True
                if att["status"] == "checkinmarked":
                    has_checkin = True
                if att["status"] == "late":
                    has_late = True
                if att["status"] in ["late", "lateandhalfday", "halfdaylate", "halfdayandlate"]:
                    late_comes_set.add(sid)

            else:
                unmarked_att = {
                    "staff": sid,
                    "staff_name": staff_obj.first_name + " " + (staff_obj.last_name or ""),
                    "profile_pic_details": None,
                    "for_date": d_str,
                    "status": "unmarked",
                    "in_time": None,
                    "out_time": None,
                    "is_active": True,
                    "marked_from": None,
                    "created": None,
                    "modified": None,
                    "marked_by_user": None
                }
                final_flat_list.append(unmarked_att)

        if has_marked:
            marked_staff_set.add(sid)
        else:
            unmarked_staff_set.add(sid)

        if has_checkin:
            checkedin_staff_set.add(sid)

        if has_late:
            late_staff_set.add(sid)

    # ------------------------------------
    # SORT: marked first, unmarked last
    # ------------------------------------
    def sort_key(x):
        status = x.get("status")

        if status == "unmarked":
            return (2, )  # bottom
        elif status == "checkinmarked":
            return (1, )  # after present
        else:
            return (0, )  # present/late/halfday first

    final_flat_list = sorted(final_flat_list, key=sort_key)

    from apps.hr.services.default_varialbes import get_lop_attendance_list
    status_list = get_lop_attendance_list()
    
    response = {
        "data": final_flat_list,
        "summary": {
            "number_of_staff_marked": len(marked_staff_set),
            "number_of_staff_unmarked": len(unmarked_staff_set),
            "only_check_in_marked": len(checkedin_staff_set),
            "number_of_late_comers": len(late_staff_set)
        },
        "status_list": status_list
    }

    if self.request.GET.get('return_unmarked'):
        response["unmarked_attendance"] = get_unmarked_attendance_for_date(self, request)

    return response

def get_intime_outtime(self,lop_data):
    for staff_list in lop_data['staff_list']:
        for date in lop_data['staff_list'][staff_list]['day_list_status']:
            if lop_data['staff_list'][staff_list]['day_list_status'][date]['status'] != 'unmarked' and lop_data['staff_list'][staff_list]['day_list_status'][date]['status'] != 'nonworkingday':    
                intime_outtime=StaffAttendance.objects.filter(is_active=True,staff=staff_list,for_date=date).values('in_time','out_time').first()
                if intime_outtime:
                    lop_data['staff_list'][staff_list]['day_list_status'][date]['in_time']=intime_outtime['in_time']
                    lop_data['staff_list'][staff_list]['day_list_status'][date]['out_time']=intime_outtime['out_time']
    return lop_data

from apps.bdu.services.write_to_excel import write_to_excel_new

def get_staff_attendance_with_intime_outtime_detailed(self, request):
    from apps.hr.services.staffleave import get_lop_count_and_date_status
    from_date = self.request.GET.get('from_date')
    to_date = self.request.GET.get('to_date')
    financial_year = self.request.GET.get('financial_year')
    return_only_day_list_status = self.request.GET.get('return_only_day_list_status', False)
    staff_ids=[]
    if not from_date and not to_date and financial_year:
        financial_year_obj = FinancialYear.objects.get(id=financial_year)
        from_date = datetime.strftime(financial_year_obj.start_date, '%Y-%m-%d')
        to_date = datetime.strftime(financial_year_obj.end_date, '%Y-%m-%d')
    if self.request.GET.get('staff_ids'):
        staff_id = self.request.GET.get('staff_ids').split(',')
    else:
        staff_id = list(Staff.objects.filter(is_active=True, date_joined__lte=to_date).values_list('id', flat=True))
    for staff in staff_id:
        staff_ids.append(int(staff))
    lop_data = get_lop_count_and_date_status(self, from_date, to_date, staff_ids, False)
    lop_data = get_intime_outtime(self,lop_data)
    date_range_list = SharedService.get_for_date_from_date_range(
            datetime.strptime(from_date, '%Y-%m-%d'), datetime.strptime(to_date, '%Y-%m-%d')
    )
    if return_only_day_list_status: #works good only for one staff because of pagination
        return {'data':{
            'status_list': {
                lop_data_row['staff_details']['staff_id']:
                    lop_data_row['day_list_status'] for lop_data_row in lop_data['staff_list'].values()}
        }}
    data=lop_data['staff_list']
    options={}
    options['Data'] = staff_attendence_report(data)
    if self.request.GET.get('download_pdf'):
        if len(staff_ids)>1:
            default = 'multiple_staff_report.html'
        else:
            default = 'single_staff_report.html'
        selected_template, number_of_copies = get_selected_template(self, 'staff_attendance_report', 'pdf', default)
        # options['Data'] = Institute.get_institute(self)
        path = 'staff_attendance_report/'+selected_template
        data1={}
        data1['staff']=options['Data']
        data1['institute']=Institute.get_institute(self)
        data1['date_range_list']=date_range_list
        total_staff = len(options['Data'])
        # Build number_of_tens_plus_ten for pagination (first page: 1-10, subsequent: 11-20, 21-30, ...)
        first_page_size = 10
        other_page_size = 10
        number_of_tens_plus_ten = {}
        empty_spaces_per_page = {}
        start = first_page_size
        while start < total_staff:
            end = start + other_page_size
            if end > total_staff:
                end = total_staff
            number_of_tens_plus_ten[start] = end
            # Calculate empty spaces for this page (should have 10 rows total)
            staff_count_in_page = end - start
            empty_spaces = other_page_size - staff_count_in_page
            empty_spaces_per_page[start] = empty_spaces if empty_spaces > 0 else 0
            start = end
        data1['number_of_tens_plus_ten'] = number_of_tens_plus_ten
        data1['empty_spaces_per_page'] = empty_spaces_per_page
        # Calculate empty spaces for first page only if needed
        if total_staff < 10:
            data1['number_of_empty_spaces'] = 10 - total_staff
        else:
            data1['number_of_empty_spaces'] = 0
        response = PDFService.receipt_new(self, data1, 'staff_attendance_report', path, False)
        return response
    if self.request.GET.get('download_excel'):
        options['extraWorksheetData'] = dict()
        options['columns'] = json_staff_attendence_report(date_range_list)
        options['title']='Staff attendence Report'
        return write_to_excel_new(self,options,{},{})

def get_staff_attendance_detailed(self, request):
    from apps.hr.services.staffleave import get_lop_count_and_date_status
    from_date = self.request.GET.get('from_date')
    to_date = self.request.GET.get('to_date')
    financial_year = self.request.GET.get('financial_year')
    return_only_day_list_status = self.request.GET.get('return_only_day_list_status', False)
    if not from_date and not to_date and financial_year:
        financial_year_obj = FinancialYear.objects.get(id=financial_year)
        from_date = datetime.strftime(financial_year_obj.start_date, '%Y-%m-%d')
        to_date = datetime.strftime(financial_year_obj.end_date, '%Y-%m-%d')
    if self.request.GET.get('staff_ids'):
        staff_ids = self.request.GET.get('staff_ids').split(',')
    else:
        staff_ids = list(Staff.objects.filter(is_active=True, date_joined__lte=to_date).values_list('id', flat=True))
    lop_data = get_lop_count_and_date_status(self, from_date, to_date, staff_ids, False)
    if return_only_day_list_status: #works good only for one staff because of pagination
        return {'data':{
            'status_list': {
                lop_data_row['staff_details']['staff_id']:
                    lop_data_row['day_list_status'] for lop_data_row in lop_data['staff_list'].values()}
        }}
    return {'data': {
        'staff_list': lop_data['staff_list'].values(),
        'status_list': get_lop_attendance_list(),
        'day_list': lop_data['day_list']
    }}

def staff_attendence_report(data):
    staff_row_data2=[]
    for staff in data:
        staff_name = data[staff]['staff_details']['name']
        staff_row_data = {'staff_name': staff_name,'day_list':[]}
        for dates in data[staff]['day_list_status']:
            datetime_object = datetime.strptime(dates,'%Y-%m-%d')
            staff_row_data[dates]=datetime_object.strftime('%d-%m-%Y')
            in_time=data[staff]['day_list_status'][dates]['in_time'] if 'in_time' in data[staff]['day_list_status'][dates] else ''
            out_time=data[staff]['day_list_status'][dates]['out_time'] if 'out_time' in data[staff]['day_list_status'][dates] else ''
            if in_time and out_time:
                staff_row_data[str(dates)+'in_out_time'] = in_time.strftime('%I:%M')+os.linesep+out_time.strftime('%I:%M')
            elif in_time:
                staff_row_data[str(dates)+'in_out_time'] = in_time.strftime('%I:%M')
            else:
                staff_row_data[str(dates)+'in_out_time'] = ''
            day_list={'in_time':in_time.strftime('%I:%M %p') if in_time else '' ,
                      'out_time':out_time.strftime('%I:%M %p') if out_time else '','date_obj':datetime_object,'date':staff_row_data[dates]}
            staff_row_data['day_list'].append(day_list)
        staff_row_data['total_unmarked']=data[staff]['status_report']['unmarked']['count'] if 'unmarked' in data[staff]['status_report'] else 0
        staff_row_data['total_present']=data[staff]['status_report']['present']['count'] if 'present' in data[staff]['status_report'] else 0
        staff_row_data['total_checkinmarked']=data[staff]['status_report']['checkinmarked']['count'] if 'checkinmarked' in data[staff]['status_report'] else 0
        staff_row_data['total_halfday']=data[staff]['status_report']['halfday']['count'] if 'halfday' in data[staff]['status_report'] else 0
        staff_row_data['total_late']=data[staff]['status_report']['late']['count'] if 'late' in data[staff]['status_report'] else 0
        staff_row_data2.append(staff_row_data)
    if len(staff_row_data2) > 1:
        return staff_row_data2
    else:
        return staff_row_data

def json_staff_attendence_report(date_range_list):
    column_data=[
        {
            'column': date_range_list[0].strftime("%B %Y"), 'required': False, 'schemacolumn': 'staff_name'
        }]
    for dates in date_range_list:
        column_data.append({
                'column': dates.strftime("%d"), 'required': False, 'schemacolumn': str(dates.date())+'in_out_time'
        })
    column_data.append({
                'column': 'TOTALUNMARKED', 'required': False, 'schemacolumn': 'total_unmarked'
            })
    column_data.append({
                'column': 'TOTALPRESENT', 'required': False, 'schemacolumn': 'total_present'
            })
    column_data.append({
                'column': 'TOTALCHECKINMARKED', 'required': False, 'schemacolumn': 'total_checkinmarked'
            })
    column_data.append({
                'column': 'TOTALHALFDAY', 'required': False, 'schemacolumn': 'total_halfday'
            })
    column_data.append({
                'column': 'TOTALLATE', 'required': False, 'schemacolumn': 'total_late'
            })
    return column_data

def get_unmarked_attendance_for_date(self, request):
    from apps.hr.services.staffleave import get_staff_leave_list
    forDate = request.GET.get('fordate')
    if not forDate:
        raise exceptions.ValidationError('For date is mandatory')
    assignedStaffs = AssignShift.get_assigned_shift_for_staff(forDate)
    assignedStaffs = {staff['staff']: staff for staff in assignedStaffs}
    staffIds = list(get_staff_leave_list(forDate,forDate))
    staffData = Staff.objects.filter(id__in=assignedStaffs.keys(),is_active=True).exclude(id__in=self.get_queryset(). \
        filter(for_date=forDate).values_list(
        'staff', flat=True)).exclude(id__in=staffIds).annotate(staff_name=Concat('first_name', V(' '), 'middle_name', V(' '), 'last_name')) \
        .values('staff_name', 'id')
    for staff in staffData:
        staff.update(assignedStaffs[staff['id']])
    return {'data': staffData}


"""
    {'2020-10-10': 1, '2020-10-11': 2} 1 -> halfday 2-> fullday
    return lop count in session
"""

def get_attendance_with_deductable_count(from_date, to_date, staff_ids):
    attendance_data = StaffAttendance.objects.filter(is_active=True, staff__in=staff_ids).values(
        'for_date', 'staff', 'in_time', 'out_time', 'status'
    )

    attendance_list = {}
    response = {}
    for attendance in attendance_data:
        if attendance['staff'] not in attendance_list:
            attendance_list[attendance['staff']] = {}
        if attendance['for_date'] not in attendance_list[attendance['staff']]:
            attendance_list[attendance['staff']][attendance['for_date']] = attendance
    fromdate = SharedService.date_to_obj(from_date)
    todate = SharedService.date_to_obj(to_date)
    attendance_dates = {}
    lop_attendance_list = get_lop_attendance_list()
    shift_data = AssignShift.objects.filter(staff__in=staff_ids).values('staff', 'shift', 'fromdate', 'todate',
                                                                 'shift__deduction_days',
                                                                 'shift__late_attempt_per_month')
    active_shifts = {}
    late_datas = {}  # base on dateRange
    date_range_list = SharedService.get_for_date_from_date_range(
            datetime.strptime(from_date, '%Y-%m-%d'), datetime.strptime(to_date, '%Y-%m-%d'),return_format_date=True
    )
    for date_row in shift_data:
        if date_row['staff'] not in active_shifts:
            active_shifts[date_row['staff']] = {}
        if ((date_row['fromdate'].strftime('%Y-%m-%d') <= fromdate.strftime('%Y-%m-%d') <= date_row['todate'].strftime(
                '%Y-%m-%d'))
                or (date_row['fromdate'].strftime('%Y-%m-%d') <= todate.strftime('%Y-%m-%d') <= date_row[
                    'todate'].strftime(
                    '%Y-%m-%d'))):
            temp = date_row['fromdate'].strftime('%Y-%m-%d') + '###' + date_row['todate'].strftime('%Y-%m-%d')
            active_shifts[date_row['staff']][temp] = date_row
        elif ((fromdate.strftime('%Y-%m-%d') <= date_row['fromdate'].strftime('%Y-%m-%d') <= todate.strftime('%Y-%m-%d'))
              or (fromdate.strftime('%Y-%m-%d') <= date_row['todate'].strftime('%Y-%m-%d') <= todate.strftime(
                    '%Y-%m-%d'))):
            temp = date_row['fromdate'].strftime('%Y-%m-%d') + '###' + date_row['todate'].strftime('%Y-%m-%d')
            active_shifts[date_row['staff']][temp] = date_row
    for staff_id in attendance_list:
        for attendance_date in attendance_list[staff_id]:
            if attendance_date in date_range_list:
                attendance_row_data = attendance_list[staff_id][attendance_date]
                if staff_id in active_shifts:
                    for i, date_data in active_shifts[staff_id].items():
                        if date_data['fromdate'] <= attendance_date <= date_data['todate']:
                            if attendance_list[staff_id][attendance_date]['status'] in ['late', 'lateandhalfday',
                                                                                            'halfdayandlate']:
                                temp = from_date + '###' + to_date
                                if staff_id not in late_datas:
                                    late_datas[staff_id] = {}
                                if temp in late_datas[staff_id]:
                                    late_datas[staff_id][temp]['count'] += 1
                                    late_datas[staff_id][temp]['fordate'] = attendance_date
                                else:
                                    late_datas[staff_id][temp] = {'count': 1, 'fordate': attendance_date,
                                                    'late_attempt': date_data['shift__late_attempt_per_month'],
                                                    'deduction_days': date_data['shift__deduction_days']}
                    if staff_id not in attendance_dates:
                        attendance_dates[staff_id] = {}
                    if attendance_row_data['status'] in lop_attendance_list:
                        attendance_dates[staff_id][attendance_date.strftime('%Y-%m-%d')] = lop_attendance_list[attendance_row_data['status']]
                        attendance_dates[staff_id][attendance_date.strftime('%Y-%m-%d')]['in_time'] = attendance_row_data['in_time']
                        attendance_dates[staff_id][attendance_date.strftime('%Y-%m-%d')]['out_time'] = attendance_row_data['out_time']
                    else:
                        raise exceptions.ValidationError(f"Unhandled exceptions {attendance_row_data['status']}")
    deductable_days = {}
    for staff_id in late_datas:
        if staff_id not in deductable_days:
            deductable_days[staff_id] = 0
        for late_day in late_datas[staff_id]:
            if late_datas[staff_id][late_day]['late_attempt'] and late_datas[staff_id][late_day]['count'] >= late_datas[staff_id][late_day]['late_attempt']:
                if late_datas[staff_id][late_day]['deduction_days']:
                    noofdays = 0
                    noofdays = int(late_datas[staff_id][late_day]['count'] / late_datas[staff_id][late_day]['late_attempt'])
                    deductable_days[staff_id] += int(late_datas[staff_id][late_day]['deduction_days']) * noofdays
    for staff in staff_ids:
        response[staff] = {
            'attendance_dates': attendance_dates[staff] if staff in attendance_dates else {},
            'deductableLateDays': deductable_days[staff] if staff in deductable_days else 0
        }
    return response


def is_staff_attendance_exist(self, staff_id, for_date):
    machine_data = StaffAttendance.objects.filter(staff=staff_id, for_date=for_date).values()
    if machine_data:
        return True, machine_data[0]
    return False, {}

def custom_bulk_attendance_edit(self, data):
    
    attendance_ids = data.get('attendance_ids', [])
    staff_attendance_pairs = data.get('staff_attendance_pairs', [])
    new_status = data.get('status', None)
    new_in_time = data.get('in_time', None)
    new_out_time = data.get('out_time', None)
    reason = data.get('reason', '')

    if not attendance_ids and not staff_attendance_pairs:
        raise exceptions.ValidationError('attendance_ids or staff_attendance_pairs is required')

    if not new_status and not new_in_time and not new_out_time:
        raise exceptions.ValidationError('At least one of status, in_time, or out_time must be provided')

    records_to_update = []
    if attendance_ids:
        records_to_update = list(StaffAttendance.objects.filter(id__in=attendance_ids, is_active=True))
    elif staff_attendance_pairs:
        for pair in staff_attendance_pairs:
            record = StaffAttendance.objects.filter(
                staff_id=pair['staff_id'],
                for_date=pair['for_date'],
                is_active=True
            ).first()
            if record:
                records_to_update.append(record)

    if not records_to_update:
        raise exceptions.ValidationError('No matching attendance records found')

    error_records = []
    for record in records_to_update:
        first_day_of_month = record.for_date - timedelta(days=int(record.for_date.strftime("%d")) - 1)
        if SalaryEmployeeMonthPlan.objects.filter(staff=record.staff_id, salary_month=first_day_of_month).exists():
            staff_name = get_staff_full_name(self, record.staff_id)
            error_records.append(f'{staff_name} (date: {record.for_date.strftime("%Y-%m-%d")})')
    if error_records:
        raise exceptions.ValidationError(
            f'Cannot update attendance, salary already paid for: {", ".join(error_records)}'
        )

    updated_count = 0
    self.serializer_class = StaffAttendanceSerializer
    self.queryset = StaffAttendance.objects.all()

    with transaction.atomic(using=get_current_db_name()):
        for record in records_to_update:
            update_data = {}

            if new_status:
                if new_status in ('absent', 'lop_attendance'):
                    update_data['status'] = new_status
                    update_data['in_time'] = None
                    update_data['out_time'] = None
                else:
                    update_data['status'] = new_status

            if new_in_time:
                for_date_str = record.for_date.strftime('%Y-%m-%d')
                update_data['in_time'] = datetime.strptime(for_date_str + ' ' + new_in_time, '%Y-%m-%d %H:%M:%S')
            if new_out_time:
                for_date_str = record.for_date.strftime('%Y-%m-%d')
                update_data['out_time'] = datetime.strptime(for_date_str + ' ' + new_out_time, '%Y-%m-%d %H:%M:%S')

            if new_in_time and new_out_time and new_status not in ('absent', 'lop_attendance', None):
                for_date_str = record.for_date.strftime('%Y-%m-%d')
                try:
                    result = get_staff_attendance_status(
                        self, [record.staff_id], for_date_str,
                        for_date_str + ' ' + new_in_time,
                        for_date_str + ' ' + new_out_time
                    )
                    if record.staff_id in result:
                        update_data['status'] = result[record.staff_id]
                except Exception:
                    pass

            if reason:
                update_data['reason'] = reason
            update_data['is_status_manually_set'] = True
            update_data['status_changed_by_id'] = self.request.user.id
            update_data['status_changed_at'] = datetime.now()

            for key, value in update_data.items():
                setattr(record, key, value)
            record.save()
            updated_count += 1

    return {
        'message': f'Successfully updated {updated_count} attendance record(s)',
        'updated_count': updated_count,
        'Reason': 'Data updated Successfully!'
    }


def get_department_wise_attendance_report(self, request):
    
    from apps.hr.services.staffleave import get_lop_count_and_date_status
    from apps.staffs.models.department import DepartmentStaffMapping
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse

    from_date = self.request.GET.get('from_date')
    to_date = self.request.GET.get('to_date')
    financial_year = self.request.GET.get('financial_year')
    download_type = self.request.GET.get('download_department_wise_type', 'pdf')

    staff_ids = []
    if not from_date and not to_date and financial_year:
        financial_year_obj = FinancialYear.objects.get(id=financial_year)
        from_date = datetime.strftime(financial_year_obj.start_date, '%Y-%m-%d')
        to_date = datetime.strftime(financial_year_obj.end_date, '%Y-%m-%d')

    if self.request.GET.get('staff_ids'):
        staff_id = self.request.GET.get('staff_ids').split(',')
    else:
        staff_id = list(Staff.objects.filter(is_active=True, date_joined__lte=to_date).values_list('id', flat=True))

    for staff in staff_id:
        staff_ids.append(int(staff))

    # Fetch attendance data using existing functions (untouched)
    lop_data = get_lop_count_and_date_status(self, from_date, to_date, staff_ids, False)
    lop_data = get_intime_outtime(self, lop_data)
    date_range_list = SharedService.get_for_date_from_date_range(
        datetime.strptime(from_date, '%Y-%m-%d'), datetime.strptime(to_date, '%Y-%m-%d')
    )
    data = lop_data['staff_list']
    report_data = staff_attendence_report(data)

    # Ensure report_data is always a list
    if isinstance(report_data, dict):
        report_data = [report_data]

    # Build staff_id -> department_name mapping from DepartmentStaffMapping
    dept_mappings = DepartmentStaffMapping.objects.filter(
        is_active=True, staff_id__in=staff_ids
    ).select_related('department').values('staff_id', 'department__name')

    staff_dept_map = {}
    for mapping in dept_mappings:
        staff_dept_map[mapping['staff_id']] = mapping['department__name']

    # Build staff_name -> staff_id lookup from lop_data
    staff_name_to_id = {}
    for staff_key in data:
        sid = data[staff_key]['staff_details']['staff_id']
        sname = data[staff_key]['staff_details']['name']
        staff_name_to_id[sname] = sid

    # Group report_data by department
    department_groups = {}
    for staff_row in report_data:
        staff_name = staff_row.get('staff_name', '')
        sid = staff_name_to_id.get(staff_name)
        dept_name = staff_dept_map.get(sid, 'Unassigned')
        if dept_name not in department_groups:
            department_groups[dept_name] = []
        department_groups[dept_name].append(staff_row)

    # Sort departments alphabetically, but keep 'Unassigned' last
    sorted_depts = sorted([d for d in department_groups if d != 'Unassigned'])
    if 'Unassigned' in department_groups:
        sorted_depts.append('Unassigned')

    if download_type == 'excel':
        return _generate_department_wise_excel(self, department_groups, sorted_depts, date_range_list)
    else:
        return _generate_department_wise_pdf(self, department_groups, sorted_depts, date_range_list)


def _generate_department_wise_pdf(self, department_groups, sorted_depts, date_range_list):
    """Generate PDF with department-wise grouping using a dedicated template."""
    departments_data = []
    for dept_name in sorted_depts:
        staff_list = department_groups[dept_name]
        total_staff = len(staff_list)
        # Pagination: 10 staff per page
        number_of_tens_plus_ten = {}
        empty_spaces_per_page = {}
        start = 10
        while start < total_staff:
            end = min(start + 10, total_staff)
            number_of_tens_plus_ten[start] = end
            empty_spaces = 10 - (end - start)
            empty_spaces_per_page[start] = empty_spaces if empty_spaces > 0 else 0
            start = end

        departments_data.append({
            'department_name': dept_name,
            'staff': staff_list,
            'number_of_tens_plus_ten': number_of_tens_plus_ten,
            'empty_spaces_per_page': empty_spaces_per_page,
            'number_of_empty_spaces': 10 - total_staff if total_staff < 10 else 0,
        })

    template_data = {
        'departments': departments_data,
        'institute': Institute.get_institute(self),
        'date_range_list': date_range_list,
    }

    path = 'staff_attendance_report/department_wise_report.html'
    response = PDFService.receipt_new(self, template_data, 'staff_attendance_dept_report', path, False)
    return response


def _generate_department_wise_excel(self, department_groups, sorted_depts, date_range_list):
    """Generate Excel with one worksheet per department."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse

    columns = json_staff_attendence_report(date_range_list)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import datetime as dt_cls
    filename = f'Staff_Attendance_Department_Wise-{dt_cls.today().date()}.xlsx'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'

    workbook = Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    dark_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    for dept_name in sorted_depts:
        staff_list = department_groups[dept_name]
        # Worksheet name max 31 chars
        sheet_title = dept_name[:31] if dept_name else 'Unassigned'
        worksheet = workbook.create_sheet(title=sheet_title)

        # Write header row
        for col_index, col_def in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_index)
            cell.value = col_def['column']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = dark_border
            worksheet.column_dimensions[cell.column_letter].width = max(len(col_def['column']) + 4, 10)

        worksheet.freeze_panes = worksheet['A2']

        # Write data rows
        for row_num, row_data in enumerate(staff_list, 2):
            for col_num, col_def in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = ''
                if col_def['schemacolumn'] in row_data:
                    value = row_data[col_def['schemacolumn']]
                if 'in_out_time' in col_def['schemacolumn']:
                    worksheet.row_dimensions[row_num].height = 30
                cell.value = value
                cell.border = dark_border

    workbook.save(response)
    return response