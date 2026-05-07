import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from django.http import HttpResponse

from apps.shared.services_shared.common import format_date_range
from apps.shared.services_shared.common import get_full_name

from apps.classes.models import StandardSectionMapping
from apps.classes.services.subject_attendance import get_subject_attendance_detail
from apps.institutes.models import Institute
from apps.classes.models.attendance import SubjectAttendance, Attendance
from apps.institutes.models import AcademicYear
from django.db.models import Count, Case, When, IntegerField, Q

from apps.institutes.models import Institute
from apps.classes.models import Subject
from apps.classes.models.enrollment import Enrollment
from apps.students.models import Student
from apps.classes.services.attendance import get_working_dates
from calendar import monthrange
from datetime import datetime

from apps.shared.services_shared.store_api_result import store_long_running_process, start_long_running_process
from apps.shared.services import UploadTypeService
from rest_framework import exceptions

    
def get_std_subject_attendance(self):
    try:
        from_date = self.request.GET.get("from_date")  
        to_date = self.request.GET.get("to_date")
        download_excel = self.request.GET.get("download_excel")
        # download_all_stansec_sub = self.request.GET.get("download_all_standsec_sub")
        subject = self.request.GET.get("subject")
        
        academic_year_qs = AcademicYear.objects.filter(is_active=True).order_by('start_date')
        year_obj = academic_year_qs.get(start_date__lte=from_date, end_date__gte=to_date)
        
        filter_query = {'academic_year_id': year_obj}
        values_list = ["id"]
        standard_section_student_details = StandardSectionMapping.objects.filter(**filter_query).values(*values_list)
        standard_section_ids = [data['id'] for data in standard_section_student_details]
        
        subject_attendance_filter = {
            'standard_section_id__in': standard_section_ids,
            'for_date__range': (from_date, to_date)
        }
        
        sub_value = [
            'standard_section__standard__id',
            'standard_section__section__id',
            'standard_section__section__name',
            'standard_section__standard__name',
            'student_id','student__first_name',
            'student__middle_name', 'student__last_name',
            'student__mobile_num',
            'student__current_reg_num',
            'subject_id','subject__name',
            'subject__codename',
            'marked_by__staff__first_name',
            'marked_by__staff__middle_name',
            'marked_by__staff__last_name',
            'marked_by__staff__designation'
        ]
        
        subject_attendance = SubjectAttendance.objects.filter(**subject_attendance_filter).values(*sub_value).annotate(
            total_classes=Count('id'),
            total_present=Count(Case(When(status='Present', then=1), output_field=IntegerField())),
            total_absent=Count(Case(When(status='Absent', then=1), output_field=IntegerField())),
        ).order_by('standard_section__standard__id','subject__id')
        
        attendance_data = list(subject_attendance) 
        output_json = build_section_attendance_json(attendance_data)
        institute_data = list(Institute.objects.filter(pk=1).values('name','code','trust_name','board_name'))
        year_response = format_date_range(from_date, to_date)
        if institute_data:
            institute_data[0]['date_range'] = year_response
        output_json["institute_details"] = institute_data
        file_name = 'Attendance_Report'
        response = None
        
        # Pass all required parameters when calling helper functions
        if not response and subject:
            file_name += 'subject_wise'+'_' + year_obj.start_date.strftime('%Y') + '_' + year_obj.end_date.strftime('%Y') + '.xlsx'
            response = get_single_section_attendance(self, file_name, from_date, to_date)
        elif not response:
            file_name+= '_' + year_obj.start_date.strftime('%Y') + '_' + year_obj.end_date.strftime('%Y') + '.xlsx'
            response = write_to_excel_subattend(self, output_json, file_name)
        
        if self.request.GET.get('long_running_process'):
            if response and hasattr(response, 'status_code') and response.status_code == 200:
                with open(file_name, 'wb') as file:
                    file.write(response.content)
                filename = file_name
                url = UploadTypeService.upload_local_file(filename, path='Attendance_Report')
                if download_excel:
                    if os.path.exists(file_name):
                        os.remove(file_name)
                    transaction_id = self.request.GET.get('transaction_id')
                    user_id = self.request.user.id
                    store_long_running_process(self, transaction_id, {'url': url})
                else:
                    return response
        return response

    except Exception as e:
        if self.request.GET.get('long_running_process'):
            transaction_id = self.request.GET.get('transaction_id')
            user_id = self.request.user.id
            store_long_running_process(self, transaction_id, {'error': str(e)[:250]})
        else:
            raise e


def addvalue(present_data:int, class_host:int) -> int:
    if not class_host or class_host==0:
        return " Not Hosted"
    return int(present_data*100/class_host)

def build_section_attendance_json(queryset_data):
    result = {"data": {"standards": {}}}  

    for row in queryset_data:
        std_id = row["standard_section__standard__id"]
        std_name = row["standard_section__standard__name"]

        section_id = row.get("standard_section__section__id")
        section_name = row.get("standard_section__section__name")

        # --- Standard level ---
        if std_id not in result["data"]["standards"]:
            result["data"]["standards"][std_id] = {
                "standard_id": std_id,
                "standard_name": std_name,
                "sections": {}
            }

        # --- Section level ---
        std_sections = result["data"]["standards"][std_id]["sections"]
        if section_id not in std_sections:
            std_sections[section_id] = {
                "section_id": section_id,
                "section_name": section_name,
                "students": {}
            }

        # --- Student level ---
        
        section_students = std_sections[section_id]["students"]
        student_id = row["student_id"]
        student_name = row["student__first_name"]
        student_reg_num = row["student__current_reg_num"] 
        student_mid_name = row["student__middle_name"]
        student_last_name = row["student__last_name"]
        student_mobile = row["student__mobile_num"]
        full_name  = get_full_name(student_name, student_mid_name, student_last_name)
        
        if student_id not in section_students:
            section_students[student_id] = {
                "student_id": student_id,
                "student_name": full_name,
                "student_reg_num":student_reg_num if student_reg_num else "Not Updated",
                "student_mobile_num": student_mobile if student_mobile else "Not Updated",
                "attendance": {}
            }

        # --- Subject attendance (object not list) ---
        teacher = get_full_name(row["marked_by__staff__first_name"],row["marked_by__staff__middle_name"],row["marked_by__staff__last_name"])
        teacher_designation = row["marked_by__staff__designation"]
        section_students[student_id]["attendance"][row["subject_id"]] = {
            "subject_id": row["subject_id"],
            "subject_name": row["subject__name"],
            "total_classes": row["total_classes"],
            "present_count": row["total_present"],
            "absent_count": row["total_absent"],
            "percent": addvalue(row["total_present"], row["total_classes"]),
            "marked_by":teacher,
            "teacher_designation":teacher_designation
        }

    return result

def write_to_excel_subattend(self, query_json,file_name):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{file_name}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    
    wb = Workbook()

    bold_center = Font(bold=True, name='Calibri', size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    standards = query_json["data"]["standards"]
    
    if not query_json or "data" not in query_json or not standards:
        ws = wb.create_sheet(title="Attendance_Not_Updated ")
        ws.merge_cells('B2:M2')
        ws['B2'] = "No attendance data available to export.\nPlease check the Attendance or data source."
        ws['B2'].font = Font(bold=True, color='FF0000', size=14)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        wb.save(response)
        return response
        
        
    for std_key, std_value in standards.items():
        std_name = std_value["standard_name"]
        for sec_key, section in std_value["sections"].items():
            section_name = section.get("section_name")

            import re

            def clean_sheet_title(title: str) -> str:
                # Remove invalid characters
                title = re.sub(r'[:\\/*?\[\]]', '-', title)
                # Trim length to max 31
                return title[:31]

            sheet_title = clean_sheet_title(f"{std_name}-{section_name}")
            ws = wb.create_sheet(title=sheet_title)
            # ws = wb.create_sheet(title=f"{std_name}-{section_name}")
            header_details = query_json.get('institute_details', [{}])[0]
            trust_name = header_details.get("trust_name", "")
            if trust_name:
                trust_name.upper()
            institute_name = header_details.get("name", "INSTITUTE")
            if institute_name:
                institute_name.upper()
            date_range = header_details.get("date_range", "")

            if trust_name:
                ws.merge_cells('A1:U1')
                ws['A1'] = trust_name
                ws['A1'].font = Font(bold=True, name='Times New Roman', size=10)
                ws['A1'].alignment = center_align

            ws.merge_cells('A2:U2')
            ws['A2'] = institute_name
            ws['A2'].font = Font(bold=True, size=18)
            ws['A2'].alignment = center_align

            ws.merge_cells('A3:U3')
            ws['A3'] = "ATTENDANCE"
            ws['A3'].font = Font(bold=True, name='Times New Roman', size=12)
            ws['A3'].alignment = center_align

            ws.merge_cells('A4:U4')
            ws['A4'] = f"{std_name}-{section_name} {date_range}"
            ws['A4'].font = Font(bold=True, name='Times New Roman')
            ws['A4'].alignment = center_align

            # Collect subjects and marked_by names
            all_subjects = {}
            sub_lec = {}
            for student in section["students"].values():
                for sub_id, sub in student.get("attendance", {}).items():
                    sub_name = sub["subject_name"]
                    all_subjects[sub_id] = sub_name
                    if sub_name not in sub_lec:  # Only once per subject
                        sub_lec[sub_name] = sub["marked_by"]

            # Table headers
            ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1)
            ws.cell(row=6, column=1, value="SL NO")
            ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=2)
            ws.cell(row=6, column=2, value="REG NO").alignment = center_align
            ws.merge_cells(start_row=6, start_column=3, end_row=7, end_column=3)
            ws.cell(row=6, column=3, value="STUDENT NAME").alignment = center_align

            col = 4
            for sub_id, sub_name in all_subjects.items():
                ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+2)
                ws.cell(row=6, column=col, value=sub_name)
                ws.cell(row=7, column=col, value="CH")
                ws.cell(row=7, column=col + 1, value="CA")
                ws.cell(row=7, column=col + 2, value="Total %")
                col += 3

            for row in ws.iter_rows(min_row=6, max_row=7, min_col=1, max_col=col-1):
                for cell in row:
                    cell.font = bold_center
                    cell.alignment = center_align
                    cell.border = thin_border

            ws.column_dimensions['B'].width = 16
            ws.column_dimensions['C'].width = 25

            # Student data
            row_num = 8
            for index, (student_key, student_value) in enumerate(section["students"].items(), start=1):
                ws.cell(row=row_num, column=1, value=index).alignment = center_align
                ws.cell(row=row_num, column=2, value=student_value.get("student_reg_num")).alignment = center_align
                ws.cell(row=row_num, column=3, value=student_value.get("student_name")).alignment = center_align
                col = 4
                for sub_id in all_subjects.keys():
                    att = student_value.get("attendance", {}).get(sub_id, {})
                    ws.cell(row=row_num, column=col, value=att.get("total_classes", 0)).alignment = center_align
                    ws.cell(row=row_num, column=col + 1, value=att.get("present_count", 0)).alignment = center_align
                    ws.cell(row=row_num, column=col + 2, value=att.get("percent", "NA")).alignment = center_align
                    ws.cell(row=row_num, column=col).border = thin_border
                    ws.cell(row=row_num, column=col + 1).border = thin_border
                    ws.cell(row=row_num, column=col + 2).border = thin_border
                    col += 3
                for c in range(1, 4):
                    ws.cell(row=row_num, column=c).border = thin_border
                row_num += 1

            # ---- Add footer at the end of table ----
            footer_start = row_num + 2

            # Add subject and marked by (name) table
            # Example: subject_code | marked_by
            ws.cell(row=footer_start, column=2, value="Subject").font = bold_center
            ws.cell(row=footer_start, column=3, value="Marked By").font = bold_center
            ws.cell(row=footer_start, column=2).alignment = center_align
            ws.cell(row=footer_start, column=3).alignment = center_align

            last_footer_row = footer_start
            for i, (sub_name, teacher) in enumerate(sub_lec.items()):
                ws.cell(row=footer_start + i + 1, column=2, value=sub_name).alignment = center_align
                ws.cell(row=footer_start + i + 1, column=3, value=teacher).alignment = center_align
                last_footer_row = footer_start + i + 1
            # Style footer cells
            for r in range(footer_start, last_footer_row + 1):
                ws.cell(row=r, column=2).font = Font(bold=True, size=11)
                ws.cell(row=r, column=3).font = Font(bold=True, size=11)
                ws.cell(row=r, column=2).border = thin_border
                ws.cell(row=r, column=3).border = thin_border

            # Signature row, similar to your image
            sign_row = last_footer_row - 3
            ws.merge_cells(start_row=sign_row, start_column=5, end_row=sign_row, end_column=7)
            ws.cell(row=sign_row, column=5, value="Attendance Co-ordinator")
            ws.cell(row=sign_row, column=5).font = Font(bold=True, size=12)
            ws.cell(row=sign_row, column=5).alignment = Alignment(horizontal="center", vertical="center")

            last_col = ws.max_column
            ws.merge_cells(start_row=sign_row, start_column=last_col-3, end_row=sign_row, end_column=last_col)
            ws.cell(row=sign_row, column=last_col-3, value="Principal")
            ws.cell(row=sign_row, column=last_col-3).font = Font(bold=True, size=12)
            ws.cell(row=sign_row, column=last_col-3).alignment = Alignment(horizontal="center", vertical="center")

    # Remove default sheet "Sheet" if present
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response

def get_single_section_attendance(self,file_name,from_date,to_date): 
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{file_name}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    data = get_subject_attendance_detail(self)
    students = list(data["data"]["student"])
    days = data["data"]["days"] or 0

    institute = Institute.objects.filter(pk=1).first()
    if institute:
        institute_name = institute.name
    else:
        institute_name = "Edubricz Institute"

    subject_id = self.request.GET.get("subject")
    subject = Subject.objects.get(id=subject_id)
    report_title = f"ATTENDANCE-{subject.name.upper()}"
    date_range = format_date_range(from_date,to_date)
    academic_info = f"{data['data']['standard_name']} {data['data']['section_name']} {date_range}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add merged headers
    last_col = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

    ws.cell(row=1, column=1, value=institute_name).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=report_title).font = Font(bold=True, size=12)
    ws.cell(row=3, column=1, value=academic_info).font = Font(bold=True, size=11)

    for r in range(1, 4):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    headers = ["SL NO", "REG NO", "STUDENT NAME", "CH", "CA", "Total %"]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        # Set column width for better display
        if header == "STUDENT NAME":
            ws.column_dimensions[cell.column_letter].width = 25
        elif header == "REG NO":
            ws.column_dimensions[cell.column_letter].width = 15
        else:
            ws.column_dimensions[cell.column_letter].width = 12

    # Fill data rows
    for idx, student in enumerate(students, start=1):
        row = 4 + idx
        reg_no = student.get("current_reg_num") or ""
        name = student.get("name") or ""
        present = student.get("present") or 0
        total_percent = 0 if days == 0 else round(present / days * 100)

        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=reg_no).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=name).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=days).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=present).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6, value=total_percent).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=6).border = thin_border

    # Signature rows
    last_row = ws.max_row + 3
        
    ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=3)
    ws.cell(row=last_row, column=2, value="Attendance Co-ordinator")
    ws.cell(row=last_row, column=2).font = Font(bold=True, size=11)
    ws.cell(row=last_row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=last_row, start_column=4, end_row=last_row, end_column=5)
    ws.cell(row=last_row, column=4, value="Principal")
    ws.cell(row=last_row, column=4).font = Font(bold=True, size=11)
    ws.cell(row=last_row, column=4).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    wb.save(response)
    return response

def get_sats_attendance(self):
    try:
        from_date = self.request.GET.get("from_date")
        to_date = self.request.GET.get("to_date")
        standard_section_id = self.request.GET.get("standard_section")
        standard_ids = self.request.GET.get("standard_ids")  # Comma-separated standard IDs
        year_id = self.request.GET.get("year")
        download_excel = self.request.GET.get("download_excel")
        
        if not from_date or not to_date:
            raise exceptions.ValidationError("from_date and to_date are required")
        
        # Get academic year
        if year_id:
            year_obj = AcademicYear.objects.filter(id=year_id, is_active=True).first()
        else:
            academic_year_qs = AcademicYear.objects.filter(is_active=True).order_by('start_date')
            year_obj = academic_year_qs.filter(start_date__lte=from_date, end_date__gte=to_date).first()
        
        if not year_obj:
            raise exceptions.ValidationError("No active academic year found for the given date range")
        
        # Get institute details
        institute = Institute.objects.filter(pk=1).first()
        if not institute:
            raise exceptions.ValidationError("Institute not found")
        
        institute_name = institute.name or "School"
        institute_code = institute.code or ""
        
        # Get students based on filters
        if standard_section_id:
            # Single standard section
            enrollments = Enrollment.objects.filter(
                standard_section_id=standard_section_id,
                student__is_active=True
            ).select_related('student', 'student__student_medium', 'student__student_parent__parent', 'standard_section__standard', 'standard_section__section')
        elif standard_ids:
            # Multiple standards - get all standard sections for selected standards
            standard_id_list = [int(sid.strip()) for sid in standard_ids.split(',') if sid.strip()]
            standard_sections = StandardSectionMapping.objects.filter(
                academic_year=year_obj,
                standard_id__in=standard_id_list
            )
            enrollments = Enrollment.objects.filter(
                standard_section__in=standard_sections,
                student__is_active=True
            ).select_related('student', 'student__student_medium', 'student__student_parent__parent', 'standard_section__standard', 'standard_section__section')
        else:
            # Get all students in the academic year
            standard_sections = StandardSectionMapping.objects.filter(academic_year=year_obj)
            enrollments = Enrollment.objects.filter(
                standard_section__in=standard_sections,
                student__is_active=True
            ).select_related('student', 'student__student_medium', 'student__student_parent__parent', 'standard_section__standard', 'standard_section__section')
        
        # Group by standard for the report
        students_data = {}
        for enrollment in enrollments:
            standard_obj = enrollment.standard_section.standard if enrollment.standard_section else None
            standard_name = standard_obj.name if standard_obj else ""
            if standard_name not in students_data:
                students_data[standard_name] = []
            
            student = enrollment.student
            # Get father name from StudentParentMapping -> ParentDetail relationship
            father_name = ""
            if hasattr(student, 'student_parent') and student.student_parent:
                if student.student_parent.parent:
                    father_name = student.student_parent.parent.father_name or ""
            # Use medium from student's medium, fallback to standard's medium, then default to English
            medium = student.student_medium.name if student.student_medium else (standard_obj.sats_medium if standard_obj and standard_obj.sats_medium else "English")
            # Use STS number from student table
            enrollment_no = student.sts or ""
            
            students_data[standard_name].append({
                'medium': medium,
                'standard': standard_name,
                'enrollment_no': enrollment_no,
                'name': get_full_name(student.first_name, student.middle_name, student.last_name),
                'father_name': father_name,
                'student_id': student.id,
                'standard_section_id': enrollment.standard_section.id
            })
        
        # Generate months in date range - always use full months
        start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        
        # Always start from the first day of the month containing from_date
        month_start = start_date.replace(day=1)
        # Always end at the last day of the month containing to_date
        month_end = end_date.replace(day=monthrange(end_date.year, end_date.month)[1])
        
        months = []
        current_date = month_start
        while current_date <= month_end:
            month_name = current_date.strftime('%B-%Y')
            # Always use full month (first day to last day of the month)
            month_first_day = current_date.replace(day=1)
            month_last_day = current_date.replace(day=monthrange(current_date.year, current_date.month)[1])
            
            months.append({
                'name': month_name,
                'year': current_date.year,
                'month': current_date.month,
                'start': month_first_day,  # Always first day of month
                'end': month_last_day      # Always last day of month
            })
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        # Calculate working days for each month
        for month_data in months:
            working_dates = get_working_dates(
                self,
                month_data['start'].strftime('%Y-%m-%d'),
                month_data['end'].strftime('%Y-%m-%d')
            )
            month_data['working_days'] = len(working_dates)
        
        # Generate Excel
        file_name = f'SATS_Attendance_{year_obj.start_date.year}_{year_obj.end_date.year}.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SATS Attendance"
        
        # Header rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Calculate total columns needed (5 base columns + number of months)
        total_cols = 5 + len(months)
        
        # Row 1: Headers as per image format
        # Column A: "School Name (School Code)"
        ws.cell(row=1, column=1, value="School Name (School Code)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=11)
        
        # Column B: School name and code
        ws.cell(row=1, column=2, value=f"{institute_name} ({institute_code})")
        ws.cell(row=1, column=2).font = Font(bold=True, size=11)
        
        # Column C: "Academic Year"
        ws.cell(row=1, column=3, value="Academic Year")
        ws.cell(row=1, column=3).font = Font(bold=True, size=11)
        
        # Column D: Academic year value
        ws.cell(row=1, column=4, value=f"{year_obj.start_date.year}-{year_obj.end_date.year}")
        ws.cell(row=1, column=4).font = Font(bold=True, size=11)
        
        # Column E: "Enter Working Days"
        ws.cell(row=1, column=5, value="Enter Working Days")
        ws.cell(row=1, column=5).font = Font(bold=True, size=11)
        
        # Column F onwards: For each month, row 1 has working days value, row 2 has month name
        col_num = 6
        for month in months:
            # Row 1: Working days value for this month
            ws.cell(row=1, column=col_num, value=month['working_days'])
            ws.cell(row=1, column=col_num).font = Font(bold=True, size=11)
            col_num += 1
        
        # Row 2: Sub-headers as per image format
        # Column A: "Medium"
        ws.cell(row=2, column=1, value="Medium")
        ws.cell(row=2, column=1).font = Font(bold=True, size=11)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=1).border = thin_border
        
        # Column B: "Standard"
        ws.cell(row=2, column=2, value="Standard")
        ws.cell(row=2, column=2).font = Font(bold=True, size=11)
        ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=2).border = thin_border
        
        # Column C: "Enrollment no"
        ws.cell(row=2, column=3, value="Enrollment no")
        ws.cell(row=2, column=3).font = Font(bold=True, size=11)
        ws.cell(row=2, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=3).border = thin_border
        
        # Column D: "Name"
        ws.cell(row=2, column=4, value="Name")
        ws.cell(row=2, column=4).font = Font(bold=True, size=11)
        ws.cell(row=2, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=4).border = thin_border
        
        # Column E: "Father's Name"
        ws.cell(row=2, column=5, value="Father's Name")
        ws.cell(row=2, column=5).font = Font(bold=True, size=11)
        ws.cell(row=2, column=5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=5).border = thin_border
        
        # Column F onwards: Month names in row 2
        col_num = 6
        for month in months:
            ws.cell(row=2, column=col_num, value=month['name'])
            ws.cell(row=2, column=col_num).font = Font(bold=True, size=11)
            ws.cell(row=2, column=col_num).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=2, column=col_num).border = thin_border
            col_num += 1
        
        # Add student data starting from row 3
        row_num = 3
        for standard_name, students in sorted(students_data.items()):
            for student in sorted(students, key=lambda x: x['name']):
                # Column A: Medium
                ws.cell(row=row_num, column=1, value=student['medium'])
                ws.cell(row=row_num, column=1).border = thin_border
                
                # Column B: Standard
                ws.cell(row=row_num, column=2, value=student['standard'])
                ws.cell(row=row_num, column=2).border = thin_border
                
                # Column C: Enrollment no (STS number)
                ws.cell(row=row_num, column=3, value=student['enrollment_no'])
                ws.cell(row=row_num, column=3).border = thin_border
                
                # Column D: Name
                ws.cell(row=row_num, column=4, value=student['name'])
                ws.cell(row=row_num, column=4).border = thin_border
                
                # Column E: Father's Name
                ws.cell(row=row_num, column=5, value=student['father_name'])
                ws.cell(row=row_num, column=5).border = thin_border
                
                # Column F onwards: Attendance for each month
                col_num = 6
                for month in months:
                    # Get attendance for this student in this month
                    attendance_records = Attendance.objects.filter(
                        student_id=student['student_id'],
                        for_date__gte=month['start'],
                        for_date__lte=month['end'],
                        status='present'
                    ).count()
                    
                    ws.cell(row=row_num, column=col_num, value=attendance_records/2)
                    ws.cell(row=row_num, column=col_num).border = thin_border
                    col_num += 1
                
                row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # Medium
        ws.column_dimensions['B'].width = 15  # Standard
        ws.column_dimensions['C'].width = 20  # Enrollment no
        ws.column_dimensions['D'].width = 30  # Name
        ws.column_dimensions['E'].width = 30  # Father's Name
        # Month columns
        for i in range(6, total_cols + 1):
            col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + ((i - 1) % 26) + 1)
            ws.column_dimensions[col_letter].width = 15
        
        wb.save(response)
        
        if self.request.GET.get('long_running_process'):
            if response and hasattr(response, 'status_code') and response.status_code == 200:
                with open(file_name, 'wb') as file:
                    file.write(response.content)
                filename = file_name
                url = UploadTypeService.upload_local_file(filename, path='SATS_Attendance_Report')
                if download_excel:
                    if os.path.exists(file_name):
                        os.remove(file_name)
                    transaction_id = self.request.GET.get('transaction_id')
                    store_long_running_process(self, transaction_id, {'url': url})
                return {'Result': True}
        
        return response
        
    except Exception as e:
        raise exceptions.ValidationError(f"Error generating SATS attendance report: {str(e)}")

