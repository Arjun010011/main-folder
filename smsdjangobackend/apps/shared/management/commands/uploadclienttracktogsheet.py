import boto3
from openpyxl import load_workbook
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime,date,timedelta
from apps.shared.services import UploadTypeService
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from apps.institutes.models import Institute
from datetime import datetime,timedelta,date
from apps.institutes.models.institute import InstitutePocMapping
from apps.institutes.models import AcademicYear
from apps.finance.models import FeeCollection
from apps.classes.models.attendance import Attendance
from apps.classes.models.standard import StandardSectionMapping
from apps.diary.models.diary import Diary
from apps.exams.models.schedule import Exam
from apps.exams.models.marks import StudentMark
from apps.hr.models.staffAttendance import StaffAttendance
from apps.notification.models.notification import BulkNotification
from apps.hr.models.timeTable import TimeTableDateRange
from apps.store.models.dataEntry import ItemSold
from apps.finance.models.miscellaneous import MiscellaneousPayment
from apps.expenditure.models.expense import Expense
from apps.shared.services import SharedService
from django.core.management.base import BaseCommand
from django.conf import settings
from apps.shared.services import UploadTypeService
AWS_STORAGE_BUCKET_NAME = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
AWS_REGION_NAME = getattr(settings, 'AWS_REGION_NAME', None)
AWS_ACCESS_KEY_ID = getattr(settings, 'AWS_ACCESS_KEY_ID', None)
AWS_SECRET_ACCESS_KEY = getattr(settings, 'AWS_SECRET_ACCESS_KEY', None)
AWS_S3_CUSTOM_DOMAIN = getattr(settings, 'AWS_S3_CUSTOM_DOMAIN', None)

class Command(BaseCommand):

    def upload_gsheet(self):

        data_to_save = {}
        sheet_name = datetime.today().strftime("%d-%m-%Y")
        # sheet_name ="pooja"
        institute = Institute.get_institute(self)
        data_to_save['institute_name'] = institute.name
        institute_poc_mapping = InstitutePocMapping.objects.filter(institute=institute).values().first()
        data_to_save['institute_poc']=''
        if institute_poc_mapping:
            data_to_save['institute_poc']=institute_poc_mapping['poc']
        data=[]
        data_start=[]
        standard_section_lists = []
        attendance_used_section_list = []
        diary_used_section_list = []
        exam_marks_used_list =[]
        attendance_notused_section_list=[]
        data_to_save['attendance_used_percentage']=0.0
        diary_notused_section_list=[]
        data_to_save['diary_used_percentage']=0.0
        last_exam_not_scheduled_sections_list=[]
        data_to_save['last_exam_used_percentage']=0.0
        exam_marks_notused_list=[]
        data_to_save['exam_marks_used_percentage']=0.0
        academicYear = AcademicYear.get_academic_year_for_date(self, datetime.today(), True)
        standard_section_ids = StandardSectionMapping.objects.filter(academic_year=academicYear).values()
        for section in standard_section_ids:
            if section['id'] not in standard_section_lists:
                standard_section_lists.append(section['id'])
        standard_section_lists = list(set(standard_section_lists))
        fee_collection = FeeCollection.objects.filter(is_active=True,payment_detail__fee_plan__standard_fee__academic_year=academicYear).order_by('-created').first()
        if academicYear:
            student_attendance_last_used = Attendance.objects.filter(for_date__gte=academicYear.start_date).order_by('-created').first()
            if student_attendance_last_used:
                student_attendance = Attendance.objects.filter(created__date=student_attendance_last_used.created.date()).values()
                data_to_save['student_attendance_last_used_value']= student_attendance_last_used.created.strftime("%d-%m-%Y")
                for attendance in student_attendance:
                    if attendance['standard_section_id'] not in attendance_used_section_list:
                        attendance_used_section_list.append(attendance['standard_section_id'])
                attendance_used_section_list=list(set(attendance_used_section_list))
                attendance_notused_section_list = list(set(standard_section_lists)-set(attendance_used_section_list))
                data_to_save['attendance_used_percentage'] = (len(attendance_used_section_list)/len(standard_section_lists))*100
            else:
                data_to_save['student_attendance_last_used_value']= 'Never Used'
            student_diary_last_used = Diary.objects.filter(is_active=True,due_date__gte=academicYear.start_date).order_by('-created').first()
            if student_diary_last_used:
                student_diary = Diary.objects.filter(created__date=student_diary_last_used.created.date()).values('title','description','subject','is_student','is_active','marks','due_date','status','created_user','is_student_can_update','created','modified','diary_standard__standard_section')
                data_to_save['student_diary_last_used_value'] = student_diary_last_used.created.strftime("%d-%m-%Y")
                for diary in student_diary:
                    if diary['diary_standard__standard_section'] not in diary_used_section_list:
                        diary_used_section_list.append(diary['diary_standard__standard_section'])
                diary_used_section_list = list(set(diary_used_section_list))
                diary_notused_section_list = list(set(standard_section_lists)-set(diary_used_section_list))
                data_to_save['diary_used_percentage'] = (len(diary_used_section_list)/len(standard_section_lists))*100
            else:
                data_to_save['student_diary_last_used_value'] = 'Never Used'
        else:
            data_to_save['student_attendance_last_used_value']= 'Never Used'
            data_to_save['student_diary_last_used_value'] = 'Never Used'
        staff_attendance_last_marked = StaffAttendance.objects.filter(is_active=True).order_by('-created').first()
        bulknotification_last_created = BulkNotification.objects.filter(academic_year=academicYear).order_by('-created').first()
        exam_schedule = Exam.objects.filter(is_active=True,academic_year=academicYear).order_by('-created').first()
        if exam_schedule:
            last_exam_scheduled_section_list = [int(num) for num in exam_schedule.standard_section_ids.split(',')]
            last_exam_scheduled_section_list=list(set(last_exam_scheduled_section_list))
            last_exam_not_scheduled_sections_list = list(set(standard_section_lists)-set(last_exam_scheduled_section_list))
            data_to_save['last_exam_used_percentage'] = (len(last_exam_scheduled_section_list)/len(standard_section_lists))*100
            data_to_save['exam_schedule_last_used_value'] = exam_schedule.created.strftime("%d-%m-%Y")
            exam_marks = StudentMark.objects.filter(exam_schedule__exam = exam_schedule.id).values('exam_schedule__standard_section')
            if exam_marks:
                for mark in exam_marks:
                    if mark['exam_schedule__standard_section'] not in exam_marks_used_list:
                        exam_marks_used_list.append(mark['exam_schedule__standard_section'])
                exam_marks_used_list = list(set(exam_marks_used_list))
                exam_marks_notused_list=list(set(standard_section_lists)-set(exam_marks_used_list))
                data_to_save['exam_marks_used_percentage'] = (len(exam_marks_used_list)/(len(standard_section_lists)))*100
                data_to_save['exam_marks_last_used_value'] = exam_schedule.created.strftime("%d-%m-%Y")
            else:
                data_to_save['exam_marks_last_used_value'] = 'Never Used'
        else:
            data_to_save['exam_schedule_last_used_value'] = 'Never Used'
            data_to_save['exam_marks_last_used_value'] = 'Never Used'
        timetable = TimeTableDateRange.objects.filter(academic_year=academicYear)
        if academicYear:
            item_sold = ItemSold.objects.filter(is_active=True,created__gte=academicYear.start_date).order_by('-created').first()
        else:
            item_sold = None
        miscellaneous = MiscellaneousPayment.objects.filter(miscellaneous__is_active=True,misc__academic_year=academicYear).order_by('-miscellaneous__created').first()
        expenses = Expense.objects.filter(is_active=True).order_by('-created').first()
        if fee_collection:
            data_to_save['fee_collection_last_used_value']=fee_collection.created.strftime("%d-%m-%Y")
        else:
            data_to_save['fee_collection_last_used_value']='Never Used'
        if staff_attendance_last_marked:
            data_to_save['staff_attendance_last_marked_value'] = staff_attendance_last_marked.created.strftime("%d-%m-%Y")
        else:
            data_to_save['staff_attendance_last_marked_value'] = 'Never Used'
        if bulknotification_last_created:
            data_to_save['bulknotification_last_marked_value'] = bulknotification_last_created.created.strftime("%d-%m-%Y")
        else:
            data_to_save['bulknotification_last_marked_value']='Never Used'
        if timetable:
            data_to_save['timetable_marked_value'] = 'Yes Using'
        else:
            data_to_save['timetable_marked_value']='Never Used'
        if item_sold:
            data_to_save['item_sold_last_marked_value'] = item_sold.created.strftime("%d-%m-%Y")
        else:
            data_to_save['item_sold_last_marked_value'] = 'Never Used'
        if miscellaneous:
            data_to_save['miscellaneous_last_marked_value'] = miscellaneous.miscellaneous.created.strftime("%d-%m-%Y")
        else:
            data_to_save['miscellaneous_last_marked_value'] = 'Never Used'
        if expenses:
            data_to_save['expenses_last_marked_value'] = expenses.created.strftime("%d-%m-%Y")
        else:
            data_to_save['expenses_last_marked_value'] = 'Never Used'
        data_to_save['attendance_not_used_standard_section_str']=''
        data_to_save['diary_not_used_standard_section_str']=''
        data_to_save['exam_not_used_standard_section_str']=''
        data_to_save['exam_marks_not_used_standard_section_str']=''
        attendance_not_used_standard_section=list(SharedService.standard_section_name_using_standard_section_mappingid(attendance_notused_section_list).values())
        for num in attendance_not_used_standard_section:
            data_to_save['attendance_not_used_standard_section_str']+=num+','
        diary_not_used_standard_section = list(SharedService.standard_section_name_using_standard_section_mappingid(diary_notused_section_list).values())
        for num in diary_not_used_standard_section:
            data_to_save['diary_not_used_standard_section_str']+=num+','
        exam_not_used_standard_section = list(SharedService.standard_section_name_using_standard_section_mappingid(last_exam_not_scheduled_sections_list).values())
        for num in exam_not_used_standard_section:
            data_to_save['exam_not_used_standard_section_str']+=num+','
        exam_marks_not_used_standard_section = list(SharedService.standard_section_name_using_standard_section_mappingid(exam_marks_notused_list).values())
        for num in exam_marks_not_used_standard_section:
            data_to_save['exam_marks_not_used_standard_section_str']+=num+','
        column_alias={'institute_name':'Institute Name','fee_collection_last_used_value':'Fee Collection last used','student_attendance_last_used_value':'Attendance last used date',
        'attendance_used_percentage':'Attendance %', 'student_diary_last_used_value':'Homework last used','diary_used_percentage':'Homework used %',
        'staff_attendance_last_marked_value':'Staff Attendance','bulknotification_last_marked_value':'Bulk Notification','timetable_marked_value':'Timetable',
        'exam_schedule_last_used_value':'Exam last scheduled','last_exam_used_percentage':'Exam using %','exam_marks_last_used_value':'Exam Marks',
        'exam_marks_used_percentage':'Exam Marks %','item_sold_last_marked_value':'Store','miscellaneous_last_marked_value':'Miscellaneous',
        'expenses_last_marked_value':'Expense','attendance_not_used_standard_section_str':'Attendance Not used','diary_not_used_standard_section_str':'Homework Not used',
        'exam_not_used_standard_section_str': 'Exam Not Used','exam_marks_not_used_standard_section_str':'Exam marks not used','institute_poc':'POC'}
        data_start.append({'data':[{'institute_name':data_to_save['institute_name'],'fee_collection_last_used_value':data_to_save['fee_collection_last_used_value'],
                    'student_attendance_last_used_value':data_to_save['student_attendance_last_used_value'],'attendance_used_percentage':round(data_to_save['attendance_used_percentage'],0),
                    'student_diary_last_used_value':data_to_save['student_diary_last_used_value'],'diary_used_percentage':round(data_to_save['diary_used_percentage'],0),
                    'staff_attendance_last_marked_value':data_to_save['staff_attendance_last_marked_value'],'bulknotification_last_marked_value':data_to_save['bulknotification_last_marked_value'],
                    'timetable_marked_value':data_to_save['timetable_marked_value'],
                    'exam_schedule_last_used_value':data_to_save['exam_schedule_last_used_value'],'last_exam_used_percentage':round(data_to_save['last_exam_used_percentage'],0),
                    'exam_marks_last_used_value':data_to_save['exam_marks_last_used_value'],'exam_marks_used_percentage':round(data_to_save['exam_marks_used_percentage'],0),
                    'item_sold_last_marked_value':data_to_save['item_sold_last_marked_value'],'miscellaneous_last_marked_value':data_to_save['miscellaneous_last_marked_value'],
                    'expenses_last_marked_value':data_to_save['expenses_last_marked_value'],
                    'attendance_not_used_standard_section_str':data_to_save['attendance_not_used_standard_section_str'],
                    'diary_not_used_standard_section_str':data_to_save['diary_not_used_standard_section_str'],
                    'exam_not_used_standard_section_str':data_to_save['exam_not_used_standard_section_str'],
                    'exam_marks_not_used_standard_section_str':data_to_save['exam_marks_not_used_standard_section_str'],'institute_poc':data_to_save['institute_poc']}],'tab_name':sheet_name})
        options={}
        options['title'] = 'ClientReport'
        options['description'] = 'ClientReport'
        options['extraWorksheet'] = False
        options['Data'] = []
        options['extraWorksheetData'] = dict()
        options['Data']=data_start
        options['columns'] = []
        for data in data_start[0]['data'][0].keys():
            options['columns'].append(
                {
                    'column': column_alias[data], 'required': False, 'schemacolumn': data
                }
            )
        s3_file_data = {}
        folderName = 'temp'
        try:
            UploadTypeService.location = f'{folderName}'
        except Exception as e:
            raise exceptions.ValidationError('Unable to setup AWS S3 Bucket!')
        try:
            s3 = boto3.client(
                's3',
                aws_access_key_id=AWS_ACCESS_KEY_ID,
                aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
                region_name=AWS_REGION_NAME
            )
            bucket_name = AWS_STORAGE_BUCKET_NAME
            # try:
            #     UploadTypeService.location = f'{folderName}'
            # except Exception as e:
            #     raise exceptions.ValidationError('Unable to setup AWS S3 Bucket!')
            # object_key = UploadTypeService.location+'/ClientReport'
            # object_key = f'https://{AWS_S3_CUSTOM_DOMAIN}/{UploadTypeService.location}'
            object_key = 'temp/ClientReport.xlsx'
            local_file = 'ClientReport.xlsx'
            s3.download_file(bucket_name, object_key, local_file)
            s3_file_data['wb'] = load_workbook(local_file)
            s3_file_data['local_file'] = local_file
        except:
            pass
        download_report = self.write_to_excel_client_track(options,s3_file_data)
        # if download_report.status_code == 200:
        with open(local_file, 'wb') as file:
            file.write(download_report.content)
        filename = local_file
        url = UploadTypeService.upload_local_file(filename)
        # else:
        #     pass
        print(url)
        return url
        
    def handle(self, *args, **options):
        self.upload_gsheet()

    def write_to_excel_client_track(self,options,s3_file_data):
        if options is None:
            options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
        columns = list()
        for column in options['columns']:
            if 'required' in column and column['required']:
                column['column'] = '*' + column['column']
            columns.append(column['column'])
        if not s3_file_data:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'ClientReport.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            workbook = Workbook()
        else:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = 'ClientReport.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            workbook = s3_file_data['wb']
        num=0
        for standard in options['Data']:
            sheet_name=standard['tab_name']
            sheet_name=sheet_name.replace('/','_')
            if str(sheet_name) in workbook.sheetnames:
                sheet_name = workbook[str(sheet_name)]
            else:
                sheet_name=workbook.create_sheet(str(sheet_name),num)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col_num, column_title in enumerate(columns, 1):
                sheet_name.freeze_panes = sheet_name['A2']
                cell = sheet_name.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.fill = header_fill
                column_letter = cell.column_letter
                cell.alignment = header_alignment
                sheet_name.column_dimensions[column_letter].width = max(len(column_title) + 4, 10)  # Adjust column width
                dark_border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
                for row in sheet_name.iter_rows(min_row=1, max_row=sheet_name.max_row, min_col=1, max_col=sheet_name.max_column):
                    for cell in row:
                        cell.border = dark_border
            for row in sheet_name.iter_rows(min_row=2, max_row=sheet_name.max_row, values_only=True):
                if any(cell is not None for cell in row):  # If any cell in the row has data
                    row_num=sheet_name.max_row+1
                else:
                    row_num=sheet_name.max_row
            num+=1
            sl_no = 0
            for row_data in standard['data']:
                sl_no += 1
                for col_num, col_data in enumerate(options['columns'], 1):
                    value = ''
                    if row_data:
                        if col_data['schemacolumn'] == 'sl_no':
                            value = sl_no
                    if col_data['schemacolumn'] in row_data:
                        value = row_data[col_data['schemacolumn']]
                    if col_data['schemacolumn'] == 'institute_name':
                        for row in sheet_name.iter_rows(min_row=1, max_row=sheet_name.max_row, min_col=1, max_col=1, values_only=False):
                            cell = row[0]  # Value in the first column
                            cell_value = cell.value
                            row_index = cell.row
                            if cell_value == row_data[col_data['schemacolumn']]:
                                row_num = row_index
                    cell = sheet_name.cell(row=row_num, column=col_num)
                    cell.value = value
                    total_usage = 0.0
                    percentage_used=0.0
                    colour = 'white'
                    value_date=self.is_date_string(value)
                    if value_date:
                        week_date = date.today() - timedelta(weeks=1)
                        three_days = date.today() - timedelta(days=3)
                        month_date = date.today() - timedelta(days=30)
                        if col_data['schemacolumn'] == 'exam_schedule_last_used_value' or col_data['schemacolumn'] == 'exam_marks_last_used_value':
                            if month_date>=value_date:
                                colour ='red'
                                total_usage +=50.0
                            else:
                                total_usage +=100.0
                        else:
                            if week_date>=value_date:
                                colour ='red'
                                total_usage +=50.0
                            elif three_days>=value_date:
                                colour ='orange'
                                total_usage += 75.0
                            else:
                                total_usage +=100.0
                    elif value == 'Never Used':
                        colour = 'red'
                        total_usage +=0.0
                    elif isinstance(value, float):
                        if value<=50.0:
                            colour='red'
                            total_usage+=50.0
                        elif value<=75.0:
                            colour='orange'
                            total_usage+=75.0
                        else:
                            total_usage+=100.0
                    if colour =='red':
                        cell.fill = red_fill
                    elif colour == 'orange':
                        cell.fill = orange_fill
                percentage_used = total_usage/len(standard['data'][0])
                if percentage_used <=50.0:
                    start_cell = sheet_name.cell(row=row_num, column=1)
                    start_cell.fill = red_fill
        workbook.save(response)
        return response
    
    def is_date_string(self,date_string, date_format="%d-%m-%Y"):
        try:
            date = datetime.strptime(date_string, date_format)
            return date.date()
        except:
            return False