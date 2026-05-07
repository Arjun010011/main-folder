import math
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment,Font
from apps.shared.services import UploadTypeService
from openpyxl.styles import Alignment,Font
from openpyxl.styles import PatternFill
from apps.users.models import User
from apps.notification.services.notification_service import send_notification
from apps.shared.services import NotificationBodyTemplate
from apps.shared.services import SharedService
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from apps.shared.services_shared.store_api_result import store_long_running_process
from openpyxl.drawing.image import Image
import requests
from io import BytesIO

"""
    summary_data shows {'columns' : [], 'rows': []}
    total_column_data = {
        'amount_paid':{'value':data['amount_paid'], 'is_auto_calculate': True}, 'pending_amount': {'value': 'Total Pending Amount'}
    }
"""

def write_to_excel_new(self, options=None, summary_data=[], total_column_data={}):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = options['title']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_index, column in enumerate(options['columns'], start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.value = column['column']
        cell.font = header_font
        cell.fill = header_fill
        column_letter = cell.column_letter
        cell.alignment = header_alignment

        worksheet.column_dimensions[column_letter].width = max(len(column['column']) + 4, 10)  # Adjust column width
    worksheet.freeze_panes = worksheet['A2']
    dark_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = dark_border

    last_row_count = 2
    auto_calculated_columns = {}
    for key in total_column_data:
        if 'is_auto_calculate' in total_column_data[key] and total_column_data[key]['is_auto_calculate']:
            auto_calculated_columns[key] = total_column_data[key]
    sl_no = 0
    for row_num, row_data in enumerate(options['Data'], 2):
        sl_no += 1
        last_row_count = row_num
        for col_num, col_data in enumerate(options['columns'], 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            value = ''
            if row_data:
                if col_data['schemacolumn'] == 'sl_no':
                    value = sl_no
            if col_data['schemacolumn'] in row_data:
                if 'mode_of_payment_list' in row_data and len(row_data['mode_of_payment_list']) > 1:
                    worksheet.row_dimensions[row_num].height = 30
                if 'in_out_time' in col_data['schemacolumn']:
                    worksheet.row_dimensions[row_num].height = 30
                value = row_data[col_data['schemacolumn']]
            cell.value = value
            if col_data['schemacolumn'] in auto_calculated_columns:
                if 'value' not in total_column_data[col_data['schemacolumn']] or not total_column_data[col_data['schemacolumn']]['value']:
                    total_column_data[col_data['schemacolumn']]['value'] = 0
                if col_data['schemacolumn'] in total_column_data and col_data['schemacolumn'] in row_data and not isinstance(row_data[col_data['schemacolumn']], str):
                    total_column_data[col_data['schemacolumn']]['value'] += row_data[col_data['schemacolumn']]
            cell.border = dark_border
    if total_column_data:
        last_row_count = last_row_count + 1
        for col_num, col_data in enumerate(options['columns'], 1):
            if col_data['schemacolumn'] in total_column_data:
                cell = worksheet.cell(row=last_row_count, column=col_num)
                cell.value = total_column_data[col_data['schemacolumn']]['value']
                cell.border = dark_border
    if summary_data:
        for col_num, column_data in enumerate(summary_data['columns'], 1):
            cell = worksheet.cell(row=last_row_count + 1, column=col_num)
            cell.value = column_data['column']
        for row_data in summary_data['rows']:
            last_row_count = last_row_count + 1
            for col_num, col_data in enumerate(summary_data['columns'], 1):
                cell = worksheet.cell(row=last_row_count, column=col_num)
                value = ''
                if col_data['schemacolumn'] in row_data:
                    value = row_data[col_data['schemacolumn']]
                cell.value = value
                cell.border = dark_border
    workbook.save(response)
    return response


def write_to_excel_new_fee_collection_group_wise(self, options=None, summary_data=[],total_column_data={},total_column_data_local={}):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = options['title']
    last_row_count = 2
    auto_calculated_columns = {}
    for key in total_column_data:
        if 'is_auto_calculate' in total_column_data[key] and total_column_data[key]['is_auto_calculate']:
            auto_calculated_columns[key] = total_column_data[key]
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    sl_no = 0
    for row_num, row_data in enumerate(options['Data'], 2):
        sl_no += 1
        last_row_count = row_num
        for col_num, col_data in enumerate(options['columns'], 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            value = ''
            if row_data:
                if col_data['schemacolumn'] == 'sl_no':
                    value = sl_no
            if col_data['schemacolumn'] in row_data:
                value = row_data[col_data['schemacolumn']]
            cell.value = value
            if col_data['schemacolumn'] in auto_calculated_columns:
                if 'value' not in total_column_data[col_data['schemacolumn']] or not total_column_data[col_data['schemacolumn']]['value']:
                    total_column_data[col_data['schemacolumn']]['value'] = 0
                if col_data['schemacolumn'] in total_column_data and col_data['schemacolumn'] in row_data and not isinstance(row_data[col_data['schemacolumn']], str):
                    total_column_data[col_data['schemacolumn']]['value'] += row_data[col_data['schemacolumn']]
    if total_column_data:
        last_row_count = last_row_count + 1
        for col_num, col_data in enumerate(options['columns'], 1):
            if col_data['schemacolumn'] in total_column_data:
                cell = worksheet.cell(row=last_row_count, column=col_num)
                cell.value = total_column_data[col_data['schemacolumn']]['value']
    if summary_data:
        for col_num, column_data in enumerate(summary_data['columns'], 1):
            cell = worksheet.cell(row=last_row_count+1, column=col_num)
            cell.value = column_data['column']
        for row_data in summary_data['rows']:
            last_row_count = last_row_count+1
            for col_num, col_data in enumerate(summary_data['columns'], 1):
                cell = worksheet.cell(row=last_row_count, column=col_num)
                value = ''
                if col_data['schemacolumn'] in row_data:
                    value = row_data[col_data['schemacolumn']]
                cell.value = value
    workbook.save(response)
    return response



def write_to_excel_new_consolidation(self, options=None, summary_data=[], total_column_data={}):
    if options is None:
        options = {'title': 'Consolidated Marks report', 'description': '', 'columns': {}}
    columns = list()
    parent_columns = []
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        if 'parent' in column and column['parent']:
            parent_columns.append(column['parent'])
        else:
            parent_columns.append({'schemacolumn':'', 'column': ''})
        columns.append(column['column'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = options['title']
    row_num=1
    if options['institute_name'] != '':
        worksheet.column_dimensions['A'].width = 20
    else:
        worksheet.column_dimensions['B'].width = 20
    last_col=len(columns)
    if options['institute_name'] != '':
        worksheet.merge_cells(start_row=row_num,start_column=1,end_row=row_num,end_column=last_col)
        cell=worksheet.cell(row=row_num,column=1)
        institute_name  = options['institute_name']
        cell.value = institute_name if options['institute_name'] else 'JNANA JYOTHI SCHOOL'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calbri',size=18)
        row_num+=2
        worksheet.merge_cells(start_row=row_num,start_column=1,end_row=row_num,end_column=last_col)
        cell=worksheet.cell(row=row_num,column=1)
        cell.value=options['examname'].upper()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calbri',size=18)
        row_num+=1
        cell=worksheet.cell(row=row_num,column=1)
        cell.value='CLASS:'+' '+options['standardname']+' '+options['sectionname']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calbri',size=16)
        row_num+=1
    else:
        worksheet.merge_cells(start_row=row_num,start_column=1,end_row=row_num,end_column=2)
        cell=worksheet.cell(row=row_num,column=1)
        cell.value='CLASS:'+' '+options['standardname']+options['sectionname']
        worksheet.merge_cells(start_row=row_num,start_column=3,end_row=row_num,end_column=last_col)
        cell=worksheet.cell(row=row_num,column=3)
        cell.value='CONSOLIDATE MARKS FOR '+options['examname'].upper()+'-2023-24'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Calbri',size=16)
        row_num+=1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num+1, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True,name='Calbri')
    temp_col_num = 1
    for col_num, column_title in enumerate(parent_columns, 1):
            if 'number_of_cells' not in column_title:
                cell_value = worksheet.cell(row=row_num+1,column=temp_col_num).value
                cell = worksheet.cell(row=row_num,column=temp_col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True,name='Calbri')
                cell_child = worksheet.cell(row=row_num+1,column=temp_col_num)
                cell_child.value = ''
                temp_col_num += 1
            elif column_title['number_of_cells'] > 0:
                incremental = temp_col_num+column_title['number_of_cells']
                worksheet.merge_cells(start_row=row_num,start_column=temp_col_num,end_row=row_num,end_column=incremental-1)
                cell = worksheet.cell(row=row_num,column=temp_col_num)
                cell.value = column_title['column']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True,name='Calbri')
                temp_col_num = incremental
    sl_no = 0
    for row, row_data in enumerate(options['Data'], row_num+2):
        row_num = row
        sl_no += 1
        for col_num, col_data in enumerate(options['columns'], 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            value = ''
            if col_data['schemacolumn'] == 'sl_no':
                value = sl_no
            if col_data['schemacolumn'] in row_data:
                value = row_data[col_data['schemacolumn']]
            cell.value = value
            if col_data['schemacolumn'] != 'student_name':
                cell.alignment = Alignment(horizontal='center', vertical='center')
    if 'total_summary_data' in options:
        row_num=row_num+2
        for col_num, column_title in enumerate(options['total_summary_data'], 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title['column']
            row_num+=1
        for row, row_data in enumerate(options['total_summary_details'], row_num):
            sl_no += 1
            row_num = row
            for col_num, col_data in enumerate(options['total_summary_data'], 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = ''
                if col_data['schemacolumn'] in row_data:
                    value = row_data[col_data['schemacolumn']]
                cell.value = value
                
    if options['institute_code'] == 'sbvshr':
        row_num += 2

        subject_marks_summary = options.get('subject_marks_summary') or {}
        subjects = [s for s in subject_marks_summary if isinstance(subject_marks_summary.get(s), dict)]

        grade_ranges = subject_marks_summary.get('subject_total_grade') or []
        grades_display = []
        percentage_display = []

        for gr in grade_ranges:
            if gr is None:
                continue
            grades_display.append(gr.get('name', ''))

            range_text = gr.get('from_to_range')
            if not range_text:  
                from_val = gr.get('from_range', '')
                to_val = gr.get('to_range', '')
                if from_val != '' and to_val != '':
                    try:
                        f, t = float(from_val), float(to_val)
                        range_text = f"{int(max(f, t))}-{int(min(f, t))}"
                    except Exception:
                        range_text = f"{from_val}-{to_val}"
                else:
                    range_text = ''
            percentage_display.append(range_text)

        all_grades = set(grades_display)
        for s in subjects:
            sub_summary = subject_marks_summary.get(s) or {}
            grade_list = sub_summary.get('subject_grades_list') or {}
            all_grades.update(grade_list.keys())

        for g in sorted(all_grades):
            if g not in grades_display:
                grades_display.append(g)
                percentage_display.append('')

        grades_display.append("ABSENT")
        percentage_display.append("ABSENT")
        grades_display.append("TOTAL")
        percentage_display.append("TOTAL")

        worksheet.cell(row_num, 1).value = "METRIC"
        for col_idx, subject in enumerate(subjects, start=2):
            worksheet.cell(row_num, col_idx).value = subject.upper()
            worksheet.cell(row_num, col_idx).font = Font(bold=True)
        row_num += 1

        efficiency_metrics = [
            ("SUBJECT WISE TOTAL", "total_obtained_marks"),
            ("SUBJECT WISE MAXIMUM MARKS", "max_marks_conducted"),
            ("SUBJECT WISE PERCENTAGE", "subject_percentage"),
            ("SUBJECT WISE POSITION", "subject_rank"),
        ]

        for metric_name, metric_key in efficiency_metrics:
            worksheet.cell(row_num, 1).value = metric_name
            worksheet.cell(row_num, 1).font = Font(bold=True)
            for col_idx, subject in enumerate(subjects, start=2):
                value = subject_marks_summary[subject].get(metric_key, '')
                if metric_key == 'subject_percentage' and isinstance(value, (float, int)):
                    value = math.ceil(value)
                worksheet.cell(row_num, col_idx).value = value
            row_num += 1

        row_num += 2

        worksheet.cell(row_num, 1).value = "PERCENTAGE RANGE"
        worksheet.cell(row_num, 2).value = "GRADE"
        for col_idx, subject in enumerate(subjects, start=3):
            worksheet.cell(row_num, col_idx).value = subject.upper()
            worksheet.cell(row_num, col_idx).font = Font(bold=True)
        row_num += 1

        for i, grade in enumerate(grades_display):
            worksheet.cell(row_num, 1).value = percentage_display[i] or ''
            worksheet.cell(row_num, 2).value = grade or ''
            for col_idx, subject in enumerate(subjects, start=3):
                summary = subject_marks_summary.get(subject) or {}
                grades = summary.get('subject_grades_list') or {}
                attendance = summary.get('attendance') or {}
                absent = attendance.get('Absent_Count', 0) or 0
                total = sum(attendance.values())
                if grade == "ABSENT":
                    value = absent
                elif grade == "TOTAL":
                    value = total
                else:
                    value = grades.get(grade, 0)
                worksheet.cell(row_num, col_idx).value = value or 0
            row_num += 1

        row_num += 2

        class_eff = subject_marks_summary.get('class_efficiency')
        if isinstance(class_eff, (int, float)):
            worksheet.cell(row_num, 1).value = f"CLASS EFFICIENCY: {class_eff:.2f}%"
            worksheet.cell(row_num, 1).font = Font(bold=True)

    workbook.save(response)
    return response

def write_to_excel_new_attendance(self,options=None, summary_data=[], total_column_data={}):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = options['title']
    worksheet.merge_cells(start_row=1,start_column=2,end_row=1,end_column=6)
    cell=worksheet.cell(row=1,column=2)
    cell.value=options['institute_name']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.column_dimensions['B'].width = 20
    cell=worksheet.cell(row=2,column=2)
    cell.value='Reported Date'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell=worksheet.cell(row=2,column=4)
    cell.value=options['date']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    last_row_count = 2
    auto_calculated_columns = {}
    for key in total_column_data:
        if 'is_auto_calculate' in total_column_data[key] and total_column_data[key]['is_auto_calculate']:
            auto_calculated_columns[key] = total_column_data[key]
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = column_title
    row_num = 3
    for group in options['Data']:
        if options['Data'][group]:
            row_num+=1
            worksheet.merge_cells(start_row=row_num,start_column=2,end_row=row_num,end_column=6)
            cell=worksheet.cell(row=row_num,column=2)
            cell.value = group
            cell.alignment = Alignment(horizontal='center', vertical='center')
        sl_no = 0
        for row_data in options['Data'][group]:
            sl_no += 1
            row_num+=1
            last_row_count = row_num
            for col_num, col_data in enumerate(options['columns'], 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = ''
                if row_data:
                    if col_data['schemacolumn'] == 'sl_no':
                        value = sl_no
                if col_data['schemacolumn'] in row_data:
                    value = row_data[col_data['schemacolumn']]
                if 'colour' in col_data:
                    if col_data['colour'] in row_data:
                        colour_value = row_data[col_data['colour']]
                        if colour_value == 'late':
                            worksheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                            worksheet.cell(row=row_num, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        elif colour_value == 'extra_miles':
                            worksheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type = 'solid')
                            worksheet.cell(row=row_num, column=2).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type = 'solid')
                cell.value = value
                if col_data['schemacolumn'] in auto_calculated_columns:
                    if 'value' not in total_column_data[col_data['schemacolumn']] or not total_column_data[col_data['schemacolumn']]['value']:
                        total_column_data[col_data['schemacolumn']]['value'] = 0
                    if col_data['schemacolumn'] in total_column_data and col_data['schemacolumn'] in row_data:
                        total_column_data[col_data['schemacolumn']]['value'] += row_data[col_data['schemacolumn']]
    loop_variable=[{
        'label':'Total no of teachers in ROLL',
        'value':sl_no
        },
        {
        'label':'No of teachers Present',
        'value':options['present_num']
        },
        {
        'label':'No of teachers Absent',
        'value':options['absent_num']
        },
    ]
    last_row_count=last_row_count+2
    for labels in loop_variable:
        last_row_count+=1
        cell=worksheet.cell(row = last_row_count,column=2)
        cell.value = labels['label']
        cell=worksheet.cell(row = last_row_count,column=3)
        cell.value = labels['value']
    workbook.save(filename)
    if options['institute_code'] == 'bluebell':
        email_id_list=['bluebellschool59@gmail.com','edubricz@gmail.com']
    elif options['institute_code'] == 'shiksha':
        email_id_list=['shikshaintacademy@gmail.com','edubricz@gmail.com']
    url=UploadTypeService.upload_local_file(filename,path='StaffAttendance')
    customizedData = list()
    notification_obj = NotificationBodyTemplate('staff_daily_attendance_report_create')
    body_email = notification_obj.select_template('email',{})
    for email_id in email_id_list:
        customizedData.append({'email':  email_id, 'user_id':None,'email_subject': None,
                                   'email_body': body_email,'email_notification':1,'attachmentLinks':[{'url': url, 'file_name': filename.split('.')[0]}]})
    send_notification('staff_daily_attendance_report_create', body=None, customizedData=customizedData)

def write_to_excel_new_student_attendence(self, standard_section_ids,options=None, summary_data=[], total_column_data={}):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    num=0
    workbook = Workbook()
    standard_dict=SharedService.get_standard_and_section_name_using_standard_section(self,standard_section_ids)
    for standard_section in standard_dict:
        sheet_name=standard_dict[standard_section]['standard__name']+'-'+standard_dict[standard_section]['section__name']
        sheet_name=workbook.create_sheet(str(sheet_name),num)
        num+=1
        sheet_name.merge_cells(start_row=1,start_column=2,end_row=1,end_column=6)
        cell=sheet_name.cell(row=1,column=2)
        cell.value=options['institute_name']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell=sheet_name.cell(row=2,column=2)
        cell.value='Reported Date'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell=sheet_name.cell(row=2,column=4)
        cell.value=options['date']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell=sheet_name.cell(row=3,column=4)
        cell.value=standard_dict[standard_section]['standard__name']+'-'+standard_dict[standard_section]['section__name']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet_name.cell(row=4, column=col_num)
            cell.value = column_title
        last_row_count = 2
        sl_no = 0
        for section in options['Data']:
            if section == standard_section:
                if not options['Data'][section]:
                    sheet_name.merge_cells(start_row=5,start_column=2,end_row=5,end_column=6)
                    cell=sheet_name.cell(row=5,column=2)
                    cell.value='Attendance not marked'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    continue
                for row_num, row_data in enumerate(options['Data'][section], 5):
                    sl_no += 1
                    last_row_count = row_num
                    for col_num, col_data in enumerate(options['columns'], 1):
                        cell = sheet_name.cell(row=row_num, column=col_num)
                        value = ''
                        if row_data:
                            if col_data['schemacolumn'] == 'sl_no':
                                value = sl_no
                        if col_data['schemacolumn'] in row_data:
                            value = row_data[col_data['schemacolumn']]
                        if col_data['schemacolumn'] == 'status_session1' or col_data['schemacolumn'] == 'status_session2':
                            if value=='absent':
                                sheet_name.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                                sheet_name.cell(row=row_num, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                        cell.value = value
        for section in options['summary_data']:
            if section == standard_section:
                if options['Data'][section]:
                    for row_num, row_data in enumerate(summary_data, 2):
                        cell = sheet_name.cell(row=last_row_count+2, column=2)
                        value=row_data['label']
                        cell.value=value
                        cell = sheet_name.cell(row=last_row_count+2, column=3)
                        value = ''
                        if row_data['value'] in options['summary_data'][section]:
                            value = options['summary_data'][section][row_data['value']]
                        cell.value = value
                        last_row_count+=1
    workbook.save(filename)
    if options['institute_code'] == 'bluebell':
        email_id_list=['bluebellschool59@gmail.com','edubricz@gamil.com']
    elif options['institute_code'] == 'shiksha':
        email_id_list=['shikshaintacademy@gmail.com','edubricz@gmail.com']
    if email_id_list:
        url=UploadTypeService.upload_local_file(filename,path='StudentAttendance')
        customizedData = list()
        notification_obj = NotificationBodyTemplate('student_daily_attendance_report_create')
        body_email = notification_obj.select_template('email',{})
        for email_id in email_id_list:
            customizedData.append({'email':  email_id, 'user_id':None,'email_subject': None,
                                   'email_body': body_email,'email_notification':1,'attachmentLinks':[{'url': url, 'file_name': filename.split('.')[0]}]})
        send_notification('student_daily_attendance_report_create', body=None, customizedData=customizedData)

def write_to_excel_multiple_tabs(self,options=None, summary_data=[], total_column_data={}):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    column_data = options['columns']
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    num=0
    workbook = Workbook()
    dark_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    for index, standard in enumerate(options['Data']):
        if ('report_name' in options and options['report_name'] == 'fee_report_custom_division'):
            if index == 0:
                sheet_name = 'A_Summary'
                sheet_name = workbook.create_sheet(str(sheet_name),num)
                last_row_count = 2
        else:
            sheet_name=options['Data'][standard]['standard_name'].replace(' ','_')
            sheet_name=sheet_name.replace('/','_')
            sheet_name=workbook.create_sheet(str(sheet_name),num)
            last_row_count = 2
        auto_calculated_columns = {}
        last_col=len(columns)
        for key in total_column_data:
            if key != "sl_no":
                total_column_data[key]['value'] = 0
            if 'is_auto_calculate' in total_column_data[key] and total_column_data[key]['is_auto_calculate']:
                auto_calculated_columns[key] = total_column_data[key]
        row_start=last_row_count
        if ('report_name' in options and options['report_name'] == 'fee_report_custom_division' and index == 0) or\
            ('report_name' in options and options['report_name'] != 'fee_report_custom_division') or\
                ('report_name' not in options):
            if 'date' in options and options['date']:
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value = options['date']
                row_start+=1
            if 'heading_one' in options and options['heading_one']:
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.value = options['heading_one']
                row_start+=1
            if 'heading_two' in options and options['heading_two']:
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.value = options['heading_two']
                row_start+=1
            if 'institute_details' in options:
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value=options['institute_details']['name']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri',size=18)
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value=options['institute_details']['name_two']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri',size=18)
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value=options['institute_details']['Affiliation']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value=options['institute_details']['Affiliaction_num']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value=options['institute_details']['Address']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value=options['institute_details']['Address_two']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value='GRADE : '+str(options['standardname'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                row_start+=1
                sheet_name.merge_cells(start_row=row_start,start_column=1,end_row=row_start,end_column=int(last_col/2))
                cell=sheet_name.cell(row=row_start,column=1)
                cell.value='SUBJECT : '+str(options['Data'][standard].get('subject_name','Not Updated'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                sheet_name.merge_cells(start_row=row_start,start_column=int(last_col/2)+1,end_row=row_start,end_column=last_col)
                cell=sheet_name.cell(row=row_start,column=int(last_col/2)+1)
                cell.value='SECTION : '+str(options['sectionname'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calbri')
                row_start+=1
        if standard == 'fee_type_summary' and 'fee_type_summary_column' in options:
            for col_num, column_title in enumerate(options['fee_type_summary_column'], 1):
                cell = sheet_name.cell(row=row_start, column=col_num)
                cell.value = column_title['column']
                options['columns'] = options['fee_type_summary_column']
        elif standard == 'student_group_wise_summary' and 'student_group_wise_summary_column' in options:
            for col_num, column_title in enumerate(options['student_group_wise_summary_column'], 1):
                cell = sheet_name.cell(row=row_start, column=col_num)
                cell.value = column_title['column']
                options['columns'] = options['student_group_wise_summary_column']
        elif ('report_name' in options and options['report_name'] == 'fee_report_custom_division' and index == 1) or\
            ('report_name' in options and options['report_name'] != 'fee_report_custom_division') or\
                ('report_name' not in options):
            for col_num, column_title in enumerate(columns, 1):
                cell = sheet_name.cell(row=row_start, column=col_num)
                if 'academic_year' in options:
                    cell.value = column_title.replace('{academic_year}', options['academic_year'])
                else:
                    cell.value = column_title
                options['columns'] = column_data
        row_start+=1
        num+=1
        sl_no = 0
        for row_num, row_data in enumerate(options['Data'][standard]['student_list'], row_start):
            sl_no += 1
            last_row_count = row_num
            for col_num, col_data in enumerate(options['columns'], 1):
                cell = sheet_name.cell(row=row_num, column=col_num)
                value = ''
                if row_data:
                    if col_data['schemacolumn'] == 'sl_no':
                        value = sl_no
                if col_data['schemacolumn'] in row_data:
                    value = row_data[col_data['schemacolumn']]
                cell.value = value
                if col_data['schemacolumn'] in auto_calculated_columns:
                    if 'value' not in total_column_data[col_data['schemacolumn']] or not total_column_data[col_data['schemacolumn']]['value']:
                        total_column_data[col_data['schemacolumn']]['value'] = 0
                    if col_data['schemacolumn'] in total_column_data and col_data['schemacolumn'] in row_data and not isinstance(row_data[col_data['schemacolumn']], str):
                        total_column_data[col_data['schemacolumn']]['value'] += row_data[col_data['schemacolumn']]
        if total_column_data and (options['Data'][standard]['standard_name'] != 'A Summary'):
            last_row_count = last_row_count + 1
            for col_num, col_data in enumerate(options['columns'], 1):
                if col_data['schemacolumn'] in total_column_data:
                    cell = sheet_name.cell(row=last_row_count, column=col_num)
                    cell.value = total_column_data[col_data['schemacolumn']]['value']
                    cell.border = dark_border
            last_row_count += 2
        row_start = last_row_count+2
        if 'abstract_data' in options:
            for col_num, column_title in enumerate(options['abstract_data'], 2):
                cell = sheet_name.cell(row=row_start, column=col_num)
                cell.value = column_title['column']
            row_start+=1
            for row_num, row_data in enumerate(options['Data'][standard]['abstract_data'], row_start):
                sl_no += 1
                last_row_count = row_num
                for col_num, col_data in enumerate(options['abstract_data'], 2):
                    cell = sheet_name.cell(row=row_num, column=col_num)
                    value = ''
                    if col_data['schemacolumn'] in row_data:
                        value = row_data[col_data['schemacolumn']]
                    cell.value = value
                    if col_data['schemacolumn'] in auto_calculated_columns:
                        if 'value' not in total_column_data[col_data['schemacolumn']] or not total_column_data[col_data['schemacolumn']]['value']:
                            total_column_data[col_data['schemacolumn']]['value'] = 0
                        if col_data['schemacolumn'] in total_column_data and col_data['schemacolumn'] in row_data and not isinstance(row_data[col_data['schemacolumn']], str):
                            total_column_data[col_data['schemacolumn']]['value'] += row_data[col_data['schemacolumn']]
        row_start = last_row_count+2
        if 'summary' in options:
            for col_num, column_title in enumerate(options['summary'], 2):
                cell = sheet_name.cell(row=row_start, column=col_num)
                cell.value = column_title['column']
            row_start+=1
            for row_num, row_data in enumerate(options['Data'][standard]['summary'], row_start):
                sl_no += 1
                last_row_count = row_num
                for col_num, col_data in enumerate(options['summary'], 2):
                    cell = sheet_name.cell(row=row_num, column=col_num)
                    value = ''
                    if col_data['schemacolumn'] in row_data:
                        value = row_data[col_data['schemacolumn']]
                    cell.value = value
                    if col_data['schemacolumn'] in auto_calculated_columns:
                        if 'value' not in total_column_data[col_data['schemacolumn']] or not total_column_data[col_data['schemacolumn']]['value']:
                            total_column_data[col_data['schemacolumn']]['value'] = 0
                        if col_data['schemacolumn'] in total_column_data and col_data['schemacolumn'] in row_data and not isinstance(row_data[col_data['schemacolumn']], str):
                            total_column_data[col_data['schemacolumn']]['value'] += row_data[col_data['schemacolumn']]
        # if total_column_data[standard]:
        #     last_row_count = last_row_count + 1
        #     for col_num, col_data in enumerate(options['columns'], 1):
        #         if col_data['schemacolumn'] in total_column_data[standard]:
        #             cell = sheet_name.cell(row=last_row_count, column=col_num)
        #             cell.value = total_column_data[standard][col_data['schemacolumn']]['value']
    workbook.save(response)
    return response

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
import requests


def write_to_excel_multiple_tabs_amrita(self, options=None, summary_data=[], total_column_data={}):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'

    num = 0
    workbook = Workbook()

    # Style definitions
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # yellow-orange
    summary_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # soft yellow
    summary_header_fill = PatternFill(start_color="FDB913", end_color="FDB913", fill_type="solid")  # deep orange
    data_fill_even = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")  # light red
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_alignment = Alignment(horizontal='center', vertical='center')

    for sheet in options['Data']:
        sheet_name = options['Data'][sheet]['sheet_name'].replace(' ', '_').replace('/', '_')
        sheet_name = workbook.create_sheet(str(sheet_name), num)
        last_row_count = 2
        auto_calculated_columns = {}

        # Keep your auto-calc logic
        for key in total_column_data:
            if 'is_auto_calculate' in total_column_data[key] and total_column_data[key]['is_auto_calculate']:
                auto_calculated_columns[key] = total_column_data[key]

        row_start = 1

        #  HEADINGS + IMAGE SUPPORT
        if 'headings' in options['Data'][sheet] and options['Data'][sheet]:
            for headings in options['Data'][sheet]['headings']:
                if 'type' in headings and headings['type'] == 'image':
                    if 'merge' in headings and headings['merge'] == 'till_last_column':
                        last_col = len(options['Data'][sheet]['columns'])
                    else:
                        last_col = headings['merge']

                    sheet_name.merge_cells(
                        start_row=row_start,
                        start_column=1,
                        end_row=row_start + 2,
                        end_column=last_col
                    )
                    cell = sheet_name.cell(row=row_start, column=1)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    response_file = requests.get(headings['value']) 
                    img_file = BytesIO(response_file.content) 
                    img = Image(img_file) 
                    row_height_px = 20 
                    col_width_px = 14 
                    img.height = 3 * row_height_px # 3 rows tall 
                    img.width = last_col * col_width_px * 8 # slightly scale columns width (adjust factor 8 for your font) 
                    anchor_cell = f"{get_column_letter(1)}{row_start}" 
                    sheet_name.add_image(img, anchor_cell) 
                    row_start += 3
                else:
                    if 'merge' in headings and headings['merge'] == 'till_last_column':
                        last_col = len(options['Data'][sheet]['columns'])
                    else:
                        last_col = headings['merge']
                    sheet_name.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=last_col)
                    cell = sheet_name.cell(row=row_start, column=1)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Calibri', size=12, bold=True, color="D83B01")  # bold orange title
                    cell.value = headings['value']
                    row_start += 1

        # PARENT + CHILD COLUMN HEADER LOGIC
        columns = []
        parent_columns = []
        for column in options['Data'][sheet]['columns']:
            if 'required' in column and column['required']:
                column['column'] = '*' + column['column']
            if 'parent' in column and column['parent']:
                parent_columns.append(column['parent'])
            else:
                parent_columns.append({'schemacolumn': '', 'column': ''})
            columns.append(column['column'])

        temp_col_num = 1
        for col_num, column_title in enumerate(parent_columns, 1):
            if 'number_of_cells' not in column_title:
                cell_value = sheet_name.cell(row=row_start + 1, column=temp_col_num).value
                cell = sheet_name.cell(row=row_start, column=temp_col_num)
                cell.value = cell_value
                cell.alignment = header_alignment
                cell.font = Font(bold=True, name='Calibri')
                cell_child = sheet_name.cell(row=row_start + 1, column=temp_col_num)
                cell_child.value = ''
                temp_col_num += 1
            elif column_title['number_of_cells'] > 0:
                incremental = temp_col_num + column_title['number_of_cells']
                sheet_name.merge_cells(start_row=row_start, start_column=temp_col_num, end_row=row_start, end_column=incremental - 1)
                cell = sheet_name.cell(row=row_start, column=temp_col_num)
                cell.value = column_title['column']
                cell.alignment = header_alignment
                cell.font = Font(bold=True, name='Calibri', color="7030A0")  # purple for parent
                temp_col_num = incremental
        row_start += 1

        # COLUMN HEADERS
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet_name.cell(row=row_start, column=col_num)
            cell.value = column_title
            cell.alignment = header_alignment
            cell.font = Font(bold=True, name='Calibri', color="FFFFFF")
            cell.fill = header_fill
            cell.border = thin_border
            column_letter = cell.column_letter
            sheet_name.column_dimensions[column_letter].width = max(len(column_title) + 4, 15)
        row_start += 1
        num += 1

        # DATA ROWS
        sl_no = 0
        for row_num, row_data in enumerate(options['Data'][sheet]['student_list'], row_start):
            sl_no += 1
            last_row_count = row_num
            for col_num, col_data in enumerate(options['Data'][sheet]['columns'], 1):
                cell = sheet_name.cell(row=row_num, column=col_num)
                value = ''
                if row_data:
                    if col_data['schemacolumn'] == 'sl_no':
                        value = sl_no
                    elif col_data['schemacolumn'] in row_data:
                        value = row_data[col_data['schemacolumn']]

                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Alternate color fill
                if sl_no % 2 == 0:
                    cell.fill = data_fill_even

        row_start = last_row_count + 2

        # SUMMARY SECTION (unchanged)
        if 'summary' in options['Data'][sheet]:
            for summary in options['Data'][sheet]['summary']:
                for col_num, column_title in enumerate(summary['columns'], summary['start_column']):
                    cell = sheet_name.cell(row=row_start, column=col_num)
                    cell.value = column_title['column']
                    cell.font = Font(bold=True, name='Calibri', color="000000")
                    cell.fill = summary_header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                row_start += 1
                for row_num, row_data in enumerate(summary['data'], row_start):
                    for col_num, col_data in enumerate(summary['columns'], summary['start_column']):
                        cell = sheet_name.cell(row=row_num, column=col_num)
                        value = row_data.get(col_data['schemacolumn'], '')
                        cell.value = value
                        cell.font = Font(bold=True, color="C00000", name='Calibri')
                        cell.fill = summary_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    row_start = row_num
                row_start = row_start + 2

    workbook.remove(workbook["Sheet"])
    workbook.save(response)
    return response