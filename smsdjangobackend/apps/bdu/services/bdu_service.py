import io
import random

import openpyxl
from django.conf import settings
from django.db import transaction

from apps.bdu.models import BduColumn, BduValidationClass
from apps.bdu.serializers import (BduColumnSerializer, BduValidationSerializer, BduUpdateColumnSerializer,
                                  BduValidationClassSerializer)
from apps.notification.services.notification_service import send_notification
from apps.shared.services import SharedService, UploadTypeService
from apps.tenants.services.middlewares import get_current_db_name

EMAIL_HOST_USER = getattr(settings, 'EMAIL_HOST_USER', None)
AWS_ACCESS_KEY_ID = getattr(settings, 'AWS_ACCESS_KEY_ID', None)
AWS_SECRET_ACCESS_KEY = getattr(settings, 'AWS_SECRET_ACCESS_KEY', None)
AWS_STORAGE_BUCKET_NAME = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)


def bdu_add_or_update_data(self, data, *args, **kwargs):
    SharedService.duplicate_list_one_object(data['columns'], 'schema_column')
    SharedService.duplicate_list_one_object(data['columns'], 'alias')
    # for column in data['columns']:
    #     for validation in column['validations']:
    #         if not validation['validation_value']:
    #             raise exceptions.ValidationError('validation_value is required!')
    with transaction.atomic(using=get_current_db_name()):
        if data['bdu'].get('id'):
            self.kwargs['pk'] = bduId = data['bdu']['id']
            SharedService.update_data(self, data['bdu'], **kwargs)
        else:
            data['bdu']['transaction_id'] = random.randint(0, 9)
            response = SharedService.add_data(self, data['bdu'], False)
            bduId = response['data']['id']
            # bduId = self.kwargs['pk']
        if 'deletable_ids' in data and data['deletable_ids']:
            BduColumn.objects.filter(id__in=data['deletable_ids']).delete()
        self.queryset = BduColumn.objects.all()
        for column in data['columns']:
            column['bdu'] = bduId
            if 'id' in column:
                if column['id'] in data['deletable_ids']:
                    continue
                self.kwargs['pk'] = column['id']
                partial = kwargs.pop('partial', False)
                instance = BduColumn.objects.get(id=column['id'])
                serializer = BduUpdateColumnSerializer(instance=instance, data=column, partial=partial)
            else:
                column_serializer = BduColumnSerializer(data=column)
                column_serializer.is_valid(raise_exception=True)
                bdu_column = column_serializer.save()
                for validation in column['bdu_validation_column']:
                    validation.update({'bdu_column': bdu_column.pk})
                serializer = BduValidationSerializer(data=column['bdu_validation_column'], many=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return {'Reason': 'Data Updated Successfully!'}


def get_bdu(self):
    res = SharedService.read_data(self)
    queryset = BduValidationClass.objects.all()
    serializer = BduValidationClassSerializer(queryset, many=True)
    return {'data': {'columns': res['data'].pop('bdu_column_bdu'), 'bdu': res['data'],
                     'bdu_validation_class': serializer.data}}


def send_email(self, response):
    if self.request.user.staff:
        staff = self.request.user.staff
        firstName = staff.first_name
        to = staff.email
    else:
        firstName = self.request.user.username
        to = EMAIL_HOST_USER
    if response['error']:
        body = f'Bulk Data Upload is Failed for the file {response["S3File"]["file_name"]}. Resolve the error(s) and upload it again. {response["error"]}'
    else:
        body = f'Bulk Data Upload is successful for the file {response["S3File"]["file_name"]}. Please review the data.'
    bodyFormat = f'Hi {firstName},<br/><br/>{body}<br/><br/>PFB for sheet download Link:<br/>  {response["S3File"]["file"]}<br/><br/>Thanks,<br/>Edubricz'

    customizedData = [
        {'email': to, 'user_id': self.request.user.pk, 'email_body': '','email_notification': 1,
         'email_subject': f'Bulk Data Upload status | {response["bdu_name"]}'}
    ]
    send_notification('bduupload_update', bodyFormat, touserIds=[], customizedData=customizedData)


def write_to_s3_file(self, response):
    if response['error']:
        global column
        S3File_NAME = response['S3File']['file'].split('/')[-1]
        s3 = UploadTypeService()
        s3.set_bucket_folder_path('bdu')
        file = s3.open(name=S3File_NAME)
        contents = file.read()
        wb = openpyxl.load_workbook(filename=(io.BytesIO(contents)), data_only=True)
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1)):
            column = len(row) + 1
            cell = ws.cell(row=1, column=column)
            cell.value = 'Status'
            cell = ws.cell(row=1, column=column + 1)
            cell.value = 'Error(s) Detail'
        for idx1, row2 in enumerate(ws.iter_rows(min_row=2), start=2):
            cell1 = ws.cell(row=idx1, column=column)
            cell2 = ws.cell(row=idx1, column=column + 1)
            if idx1 in response['Reason']:
                cell1.value = 'Error'
                cell2.value = str(response['Reason'][idx1])
            else:
                cell1.value = 'No Error'
        wb.save(S3File_NAME)
        UploadTypeService.upload_local_file(S3File_NAME)
    send_email(self, response)