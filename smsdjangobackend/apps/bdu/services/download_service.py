import re

from datetime import datetime
from django.db.models import F, Value
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from rest_framework import exceptions
from apps.bdu.services.upload_service import get_detail, supportedExtension
from apps.finance.models.feeCollection import PaymentDetail
from apps.finance.services import fee_collection
from apps.students.models.student import Student
from apps.students.models.studentDetail import GuardianDetail, ParentDetail, StudentAddress, StudentDetails, StudentParentMapping


def bdu_download_file(self, request, *args, **kwargs):
    if self.request.GET.get('extn', 'xls') not in supportedExtension:
        raise exceptions.ValidationError('Please select xls or xlsx file extension.')
    options = dict()
    bdu_detail = get_detail(self, True, False, True)
    data = bdu_detail['data']
    options['title'] = data['bdu']['name']
    options['description'] = data['bdu']['description']
    options['extraWorksheet'] = False
    options['extraWorksheetData'] = dict()
    options['columns'] = list()
    for column_definition in data['columns']:
        if (not column_definition['ignored']) and (not column_definition['exclude_from_view']):
            options['columns'].append({'column': column_definition['alias'], 'required': column_definition['required'],
                                       'schemacolumn': column_definition['schema_column']})
        elif column_definition['update_allowed'] and data['bdu']['upload_type'] != 'insert':
            options['columns'].append({'column': column_definition['alias'], 'required': False,
                                       'schemacolumn': column_definition['schema_column']})
    # set the id value as mandatory in case of "update/both".
    table_name = data['bdu']['primary_table'].split('.')[1]
    strippedtable_name = table_name + '_id'
    if options['columns'][0]['column'] == 'id':
        options['columns'][0]['column'] = strippedtable_name
    if options['columns'][0]['column'] and options['columns'][0]['column'] == strippedtable_name and data['bdu'][
        'upload_type'] != 'insert':
        options['columns'][0]['required'] = True
    if data['bdu']['upload_type'] == 'update':
        student_data = Student.objects.values()
        column_names = dict()
        student_ids = []
        for column_name in options['columns']:
            column_names[column_name['column']] = []
        for student in student_data:
            student_ids.append(student['id'])
            for key,item in student.items():
                if key not in column_names:
                    column_names[key] = []
                column_names[key].append(item)
        student_details = {stu['student_id']:stu for stu in StudentDetails.objects.filter(student__in=student_ids).values() }
        student_parent_mapping ={student_parent['student_id']: {'parent_id': student_parent['parent_id'], 'guardian_id': student_parent['guardian_id']}for student_parent in StudentParentMapping.objects.filter(student__in=student_ids).values() } 
        parent_ids = [parent_guardian['parent_id'] for parent_guardian in student_parent_mapping.values()]
        guardian_ids = [parent_guardian['guardian_id'] for parent_guardian in student_parent_mapping.values()]
        parent_details = {parent['id']:parent for parent in ParentDetail.objects.filter(id__in=parent_ids).values()}
        guardian_details ={guardian['id']:guardian for guardian in GuardianDetail.objects.filter(id__in=guardian_ids).values()}
        student_address = {stud['student_id']:stud for stud in  StudentAddress.objects.filter(student_id__in=student_ids).values()}
        for key,student_detail in student_details.items():
            if key in student_parent_mapping:
                parent_id = student_parent_mapping[key]['parent_id']
                guardian_id = student_parent_mapping[key]['guardian_id']
                if parent_id in parent_details:
                    student_detail.update(parent_details[parent_id])
                if guardian_id in guardian_details:
                    student_detail.update(guardian_details[guardian_id])
            if key in student_address:
                student_detail.update(student_address[key])
            for key,detail in student_detail.items():
                if key not in column_names:
                    column_names[key] = []
                column_names[key].append(detail)
        for column_info in options['columns']:
            if column_info['column'] in column_names:
                column_info['values'] = column_names[column_info['column']]
    return write_to_excel(self, options, data['bdu']['upload_type'])

def write_to_excel(self, options=None, uploadType=None):
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = options['title']
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    for col_num, column_info in enumerate(options['columns'], 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_info['column']

        if 'values' in column_info:
            for row_num, value in enumerate(column_info['values'], start=2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
    workbook.save(response)
    return response

def get_group_details(standard_id):
    standard_grouping = {
        "school": {"group_code": "school", "group_name": "School", "standard_list": [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]},
        "pu": {"group_code": "pu","group_name": "Pu", "standard_list": [16,17,46,47,48,49,50,51,52,53]},
        "degree": {"group_code": "degree","group_name": "Degree", "standard_list": [18,19,20,21,22,23,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45]},
        "evening_degree": {"group_code": "evening_degree","group_name": "Evening Degree", "standard_list": [24,25,26,27,28,29,58,59]},
        "mcom": {"group_code": "mcom", "group_name": "Mcom", "standard_list": [42,43,44,45]}
    }
    group_data = {}
    for group in standard_grouping.values():
        for standard in group['standard_list']:
            if standard ==  standard_id:
                group_data = group
    return group_data

def download_excel_for_jnanajyothi(self, options, from_date, to_date ):
    start_receipt = {}
    end_receipt = {}
    evening_degree_cash = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    mcom_cash = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    degree_cash =  {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    school_cash = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    pu_cash = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    evening_degree_online = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    degree_online =  {'data_list': [], 'summary': {'total': 0, 'total_deleted':0, 'col_index': 5}}
    school_online = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    pu_online = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    mcom_online = {'data_list': [], 'summary': {'total': 0,'total_deleted':0, 'col_index': 5}}
    misc_data = {'data_list': [], 'summary': {'total': 0, 'col_index': 5}, 'columns_variable': 'misc_columns'}
    fee_groups_to_include = [1,5,6,7,8,9] #we have to show only few feetype data in the summary
    fee_groups_to_include_m_com = [1,2,5,6,7,8,9]
    if options is None:
        options = {'title': 'Edubricz BDU File', 'description': '', 'columns': {}}
    columns = list()
    misc_columns = list()
    for column in options['columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        columns.append(column['column'])
    for column in options['misc_columns']:
        if 'required' in column and column['required']:
            column['column'] = '*' + column['column']
        misc_columns.append(column['column'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{options["title"]}-{datetime.today().date()}.{self.request.GET.get("extn", "xls")}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = options['title']
    last_row_count = 2
    amount_col_index = 0
    total_amount_paid = 0
    deleted_total_amount = 0
    total_discount_amount = 0
    discount_col_index = 0
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    options['Data'] = sorted(options['Data'], key=lambda d: d['created'])
    for row_num, row_data in enumerate(options['Data'], 2):
        row_data['sl_no'] = row_num-1
        last_row_count = row_num
        amount_col_index = 0
        discount_col_index = 0
        total_amount_paid += row_data['amount_paid'] if not 'is_active' in row_data or row_data['is_active'] else 0
        deleted_total_amount +=row_data['amount_paid'] if 'is_active' in row_data and not row_data['is_active'] else 0
        total_discount_amount += row_data['discount'] if 'discount' in row_data and row_data['discount'] else 0
        for col_num, col_data in enumerate(options['columns'], 1):
            if col_data['schemacolumn'] == 'amount_paid':
                amount_col_index = col_num
            cell = worksheet.cell(row=row_num, column=col_num)
            value = ''
            if col_data['schemacolumn'] == 'discount':
                discount_col_index = col_num
            cell = worksheet.cell(row=row_num, column=col_num)
            value = ''
            if col_data['schemacolumn'] in row_data:
                value = row_data[col_data['schemacolumn']]
            if 'is_active' in row_data and not row_data['is_active']:
                worksheet.row_dimensions[row_num].height = 30
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = "solid")
                if col_data['schemacolumn'] == 'full_name':
                    value=value+'(deleted)'
            cell.value = value
        receipt_group = "".join(re.findall("[a-zA-Z]+", row_data['receipt_num']))
        if receipt_group not in start_receipt:
            start_receipt[receipt_group] = {'receipt_num': row_data['receipt_num']}
        if receipt_group not in end_receipt:
            end_receipt[receipt_group] = {'receipt_num': row_data['receipt_num']}
        existing_rec = int(''.join(filter(str.isdigit, start_receipt[receipt_group]['receipt_num'])))
        given_rec = int(''.join(filter(str.isdigit, row_data['receipt_num'])))
        existing_rec_end = int(''.join(filter(str.isdigit, end_receipt[receipt_group]['receipt_num'])))
        given_rec_end = int(''.join(filter(str.isdigit, row_data['receipt_num'])))
        if given_rec < existing_rec:
            start_receipt[receipt_group]['receipt_num'] = row_data['receipt_num']
        if given_rec_end > existing_rec_end:
            end_receipt[receipt_group]['receipt_num'] = row_data['receipt_num']
        if 'local_module' in row_data and row_data['local_module'] == 'fee_collection' and \
            'standard' in row_data and 'fee_group' in row_data:
            if row_data['fee_group'] in fee_groups_to_include:
                group_details = get_group_details(row_data['standard'])
                if 'mode_of_payment' in row_data and row_data['mode_of_payment'] == 'Cash':
                    if group_details['group_code'] == 'evening_degree':
                        evening_degree_cash['data_list'].append(row_data)
                        evening_degree_cash['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        evening_degree_cash['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        evening_degree_cash['summary']['col_index'] = amount_col_index
                    elif group_details['group_code'] == 'degree':
                        degree_cash['data_list'].append(row_data)
                        degree_cash['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        degree_cash['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        degree_cash['summary']['col_index'] = amount_col_index
                    elif group_details['group_code'] == 'pu':
                        pu_cash['data_list'].append(row_data)
                        pu_cash['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        pu_cash['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        pu_cash['summary']['col_index'] = amount_col_index
                    elif group_details['group_code'] == 'school':
                        school_cash['data_list'].append(row_data)
                        school_cash['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        school_cash['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        school_cash['summary']['col_index'] = amount_col_index
            if row_data['fee_group'] in fee_groups_to_include_m_com:
                group_details = get_group_details(row_data['standard'])
                if 'mode_of_payment' in row_data and row_data['mode_of_payment'] == 'Cash':
                    if group_details['group_code'] == 'mcom':
                        mcom_cash['data_list'].append(row_data)
                        mcom_cash['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        mcom_cash['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        mcom_cash['summary']['col_index'] = amount_col_index
            if row_data['fee_group'] in fee_groups_to_include:
                group_details = get_group_details(row_data['standard'])
                if 'mode_of_payment' in row_data and row_data['mode_of_payment'] in ['Cheque', 'CreditCard', 'DebitCard', 'UPIPayments', 'Online', 'NetBanking']:
                    if group_details['group_code'] == 'evening_degree':
                        evening_degree_online['data_list'].append(row_data)
                        evening_degree_online['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        evening_degree_online['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        evening_degree_online['summary']['col_index'] = amount_col_index
                    elif group_details['group_code'] == 'degree':
                        degree_online['data_list'].append(row_data)
                        degree_online['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        degree_online['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        degree_online['summary']['col_index'] = amount_col_index
                    elif group_details['group_code'] == 'pu':
                        pu_online['data_list'].append(row_data)
                        pu_online['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        pu_online['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        pu_online['summary']['col_index'] = amount_col_index
                    elif group_details['group_code'] == 'school':
                        school_online['data_list'].append(row_data)
                        school_online['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        school_online['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        school_online['summary']['col_index'] = amount_col_index
            if row_data['fee_group'] in fee_groups_to_include_m_com:
                group_details = get_group_details(row_data['standard'])
                if 'mode_of_payment' in row_data and row_data['mode_of_payment'] in ['Cheque', 'CreditCard', 'DebitCard', 'UPIPayments', 'Online', 'NetBanking']:
                    if group_details['group_code'] == 'mcom':
                        mcom_online['data_list'].append(row_data)
                        mcom_online['summary']['total'] += row_data['amount_paid'] if row_data['is_active'] else 0
                        mcom_online['summary']['total_deleted'] +=row_data['amount_paid'] if not row_data['is_active'] else 0
                        mcom_online['summary']['col_index'] = amount_col_index
        elif 'local_module' in row_data and row_data['local_module'] == 'misc_collection':
            misc_data['data_list'].append(row_data)
            misc_data['summary']['total'] += row_data['amount_paid']
            misc_data['summary']['col_index'] = amount_col_index
    last_row_count = last_row_count + 1
    cell = worksheet.cell(row=last_row_count, column=amount_col_index-1)
    cell.value = 'Total'
    cell = worksheet.cell(row=last_row_count, column=amount_col_index)
    cell.value = total_amount_paid
    cell = worksheet.cell(row=last_row_count, column=discount_col_index-1)
    cell.value = 'Total Discount'
    cell = worksheet.cell(row=last_row_count, column=discount_col_index)
    cell.value = total_discount_amount
    last_row_count = last_row_count + 1
    cell = worksheet.cell(row=last_row_count, column=amount_col_index-1)
    cell.value = 'Total Deleted Amount'
    cell = worksheet.cell(row=last_row_count, column=amount_col_index)
    cell.value = deleted_total_amount
    last_row_count = last_row_count + 1
    cell = worksheet.cell(row=last_row_count, column=amount_col_index-1)
    """starting index"""
    last_row_count = last_row_count + 2
    cell = worksheet.cell(row=last_row_count, column=2)
    cell.value = 'STARTING RECEIPT NO.'
    col_index = 3
    for receipt_rw in start_receipt:
        cell = worksheet.cell(row=last_row_count, column=col_index)
        cell.value = start_receipt[receipt_rw]['receipt_num']
        col_index += 1
    last_row_count += 1
    col_index = 3
    cell = worksheet.cell(row=last_row_count, column=2)
    cell.value = 'Ending Receipt No'
    for receipt_rw in end_receipt:
        cell = worksheet.cell(row=last_row_count, column=col_index)
        cell.value = end_receipt[receipt_rw]['receipt_num']
        col_index += 1

    loop_variable = [
        {
            'label': 'EVENING DEGREE CASH',
            'variable': evening_degree_cash
        },{
            'label': 'MCOM CASH',
            'variable': mcom_cash
        },{
            'label': 'DEGREE CASH',
            'variable': degree_cash
        },{
            'label': 'PU CASH',
            'variable': pu_cash
        },{
            'label': 'PS AND HS CASH',
            'variable': school_cash
        },{
            'label': 'EVENING DEGREE ONLINE',
            'variable': evening_degree_online
        },{
            'label': 'MCOM ONLINE',
            'variable': mcom_online
        },{
            'label': 'DEGREE ONLINE',
            'variable': degree_online
        },{
            'label': 'PU ONLINE',
            'variable': pu_online
        },{
            'label': 'PS AND HS ONLINE',
            'variable': school_online
        },{
            'label': 'Miscellaneous',
            'variable': misc_data,
            'columns_variable': 'misc_columns'
        }
    ]

    for loop in loop_variable:
        last_row_count += 2
        cell = worksheet.cell(row=last_row_count, column=2)
        cell.value = loop['label']
        col_index = 1
        last_row_count += 1
        column_varialbe = loop['columns_variable'] if 'columns_variable' in loop else 'columns'
        temp = {'columns': columns if column_varialbe == 'columns' else misc_columns} #for now supporting misc
        for col_num, column_title in enumerate(temp['columns'], 1):
            cell = worksheet.cell(row=last_row_count, column=col_num)
            cell.value = column_title
        for row_num, row_data in enumerate(loop['variable']['data_list'], 2):
            last_row_count +=1
            for col_num, col_data in enumerate(options[column_varialbe], 1):
                cell = worksheet.cell(row=last_row_count, column=col_num)
                if 'is_active' in row_data and not row_data['is_active']:
                    worksheet.row_dimensions[row_num].height = 30
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = "solid")
                cell.value = row_data[col_data['schemacolumn']]
        last_row_count +=1
        previous_index = loop['variable']['summary']['col_index']-1
        if loop['variable']['data_list']:
            cell = worksheet.cell(row=last_row_count, column=previous_index)
            cell.value = 'Total'
            cell = worksheet.cell(row=last_row_count, column=loop['variable']['summary']['col_index'])
            cell.value = loop['variable']['summary']['total']
            if 'total_deleted' in loop['variable']['summary'] and loop['variable']['summary']['total_deleted']:
                last_row_count+=1
                cell = worksheet.cell(row=last_row_count, column=previous_index)
                cell.value = 'Total Deleted'
                cell = worksheet.cell(row=last_row_count, column=loop['variable']['summary']['col_index'])
                cell.value = loop['variable']['summary']['total_deleted']
                last_row_count+=1
        else:
            cell = worksheet.cell(row=last_row_count, column=previous_index)
            cell.value = ''
    last_row_count+=1
    cell=worksheet.cell(row=last_row_count,column=2)
    cell.value='Cash Payment Summary'
    final_summary = [
       {
            'label': 'Evening Degree',
            'variable': evening_degree_cash['summary']['total']
        },{
           'label': 'Degree',
           'variable': degree_cash['summary']['total']
        },{
           'label': 'Pu',
           'variable': pu_cash['summary']['total']
        },{
           'label': 'PS AND HS CASH',
           'variable': school_cash['summary']['total']
        },{
           'label': 'MCOM',
           'variable': mcom_cash['summary']['total']
        }
    ]
    last_row_count +=1
    total = 0
    for final in final_summary:
        last_row_count +=1
        cell = worksheet.cell(row=last_row_count, column=2)
        cell.value = final['label']
        cell = worksheet.cell(row=last_row_count, column=3)
        cell.value = final['variable']
        total += final['variable']
    last_row_count +=1
    cell = worksheet.cell(row=last_row_count, column=2)
    cell.value = 'Total'
    cell = worksheet.cell(row=last_row_count, column=3)
    cell.value = total

    last_row_count+=2
    cell=worksheet.cell(row=last_row_count,column=2)
    cell.value='Online Payment Summary'

    final_summary = [
        {
            'label': 'Evening Degree',
            'variable': evening_degree_online['summary']['total']
        },{
           'label': 'Degree',
           'variable': degree_online['summary']['total']
        },{
           'label': 'Pu',
           'variable': pu_online['summary']['total']
        },{
           'label': 'PS AND HS ONLINE',
           'variable': school_online['summary']['total']
        },{
           'label': 'MCOM',
           'variable': mcom_online['summary']['total']
        }
    ]
    last_row_count +=1
    total_on = 0
    for final in final_summary:
        last_row_count +=1
        cell = worksheet.cell(row=last_row_count, column=2)
        cell.value = final['label']
        cell = worksheet.cell(row=last_row_count, column=3)
        cell.value = final['variable']
        total_on += final['variable']
    last_row_count +=1
    cell = worksheet.cell(row=last_row_count, column=2)
    cell.value = 'Total'
    cell = worksheet.cell(row=last_row_count, column=3)
    cell.value = total_on
    
    deleted_payment_data = PaymentDetail.objects.filter(fee_collection__transaction_date__range=(from_date, to_date), fee_collection__is_active=False,
                                                ).order_by('created').annotate(fee_type_name=F('fee_plan__standard_fee__fee_type__name'),
                                                    date=F('fee_collection__transaction_date'),
                                                    fee_collection_receipt_num=F('fee_collection__receipt_num'),
                                                    user=F('fee_collection__user'),
                                                    student_first_name=F('fee_collection__student__first_name'),
                                                    student_middle_name=F('fee_collection__student__middle_name'),
                                                    student_last_name=F('fee_collection__student__last_name'),
                                                    reg_num=F('fee_collection__student__current_reg_num'),
                                                    student=F('fee_collection__student'),
                                                    local_module=Value('fee_collection'),
                                                    online_payment=F('fee_collection__online_payment__id'),
                                                    mode_of_payment=F('fee_collection__mode_of_payment'),
                                                    payment_note=F('fee_collection__payment_note'),
                                                    fee_group_name=F('fee_plan__standard_fee__fee_group__name'),
                                                    fee_group=F('fee_plan__standard_fee__fee_group'),
                                                    standard=F('fee_plan__standard_fee__standard'),
                                                    ref_number=F('fee_collection__payment_ref_num'),
                                                    standard_name=F('fee_plan__standard_fee__standard__name')).values(
            'fee_type_name', 'amount_paid', 'date', 'receipt_num', 'user', 'student_first_name', 'student_middle_name',
            'student_last_name', 'reg_num', 'standard_name', 'student', 'online_payment', 'fee_collection_receipt_num', 'local_module', 'standard',
            'mode_of_payment', 'fee_collection__payment_note', 'fee_group_name', 'ref_number', 'fee_collection', 'fee_group',
            'fee_collection__fee_collection_delete_tracking_fee_collection__reason')
    last_row_count += 2
    cell = worksheet.cell(row=last_row_count, column=2)
    cell.value = 'DELETED PAYMENT DATA'

    deleted_total = 0
    cell = worksheet.cell(row=last_row_count, column=col_num)
    last_row_count += 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=last_row_count, column=col_num)
        cell.value = column_title
    cell = worksheet.cell(row=last_row_count, column=col_num+1)
    cell.value = 'Reason'
    last_row_count += 1
    for row_num, row_data in enumerate(deleted_payment_data, 2):
        row_data['sl_no'] = row_num-1
        last_row_count += 1
        for col_num, col_data in enumerate(options['columns'], 1):
            cell = worksheet.cell(row=last_row_count, column=col_num)
            temp = row_data[col_data['schemacolumn']] if col_data['schemacolumn'] in row_data else ''
            cell.value = temp
        cell = worksheet.cell(row=last_row_count, column=col_num+1)
        cell.value = row_data['fee_collection__fee_collection_delete_tracking_fee_collection__reason']
        deleted_total += row_data['amount_paid']
    last_row_count += 1
    cell = worksheet.cell(row=last_row_count, column=amount_col_index-1)
    cell.value = 'Total'
    cell = worksheet.cell(row=last_row_count, column=amount_col_index)
    cell.value = deleted_total
    workbook.save(response)
    return response